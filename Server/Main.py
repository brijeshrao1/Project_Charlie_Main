import time
import pandas as pd
from fastapi import FastAPI, HTTPException, Request, Body, Query, Form, UploadFile, File, status, BackgroundTasks, Depends
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
import tempfile
from typing import Dict, List, Optional, Tuple, Set, Any
from pathlib import Path
from pydantic import BaseModel, Json, Field, HttpUrl, ConfigDict
import json
import logging
import shutil
import uuid
import math
import numpy as np
import openpyxl
from io import BytesIO
import re
from datetime import datetime, timedelta # Ensure datetime is imported
import os
from fastapi.middleware.cors import CORSMiddleware
import requests
import base64 # Import base64 for file decoding
from fastapi.staticfiles import StaticFiles # Import StaticFiles
import zipfile
import dateutil.parser
from dotenv import load_dotenv, find_dotenv
from collections import defaultdict 
import io
import httpx
import base64
import xml.etree.ElementTree as ET
from dotenv import dotenv_values
import time
import asyncio
import threading
import shutil
from pathlib import Path
import sys
import importlib.util
# Import GenAI client robustly (support multiple install/import variants).
try:
    from google import genai
except Exception:
    try:
        import google.genai as genai
    except Exception as e:
        genai = None
        # Only warn if the user has configured a GEMINI_API_KEY — otherwise
        # silently disable Gemini features to avoid noisy startup logs.
        if os.environ.get("GEMINI_API_KEY"):
            logging.warning(
                "GenAI client import failed (%s). To enable Gemini features install: pip install google-genai",
                e,
            )
        else:
            logging.debug("GenAI client not installed; Gemini features disabled.")
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
try:
    import polars as _pl_check
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False
    logger.warning("Polars not installed. Large-scale validation endpoint will be unavailable. Install with: pip install polars")


# Load environment variables
load_dotenv()
ORACLE_USERNAME = os.getenv("ORACLE_USERNAME")
ORACLE_PASSWORD = os.getenv("ORACLE_PASSWORD")
ORACLE_ENV = os.getenv("ORACLE_ENV") 

app = FastAPI(
    title="Excel Transformation API",
    description="API to transform and download Excel files.",
    version="1.0.0"
)

# Configure CORS from environment for production safety. If not set, default to localhost dev origins.
_allowed_origins = os.getenv("ALLOWED_ORIGINS", "").strip()
if _allowed_origins:
    try:
        ALLOWED_ORIGINS = [o.strip() for o in _allowed_origins.split(",") if o.strip()]
    except Exception:
        ALLOWED_ORIGINS = ["http://localhost:3000"]
else:
    # default dev-friendly origins
    ALLOWED_ORIGINS = ["http://localhost:3000", "http://127.0.0.1:3000"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=[
        "X-Bundle-Filename",
        "Content-Disposition",
        "X-Transform-Rules-Applied",
        "X-Transform-Cells-Changed",
        "X-Transform-Columns-Changed",
        "X-Transform-Total-Rules",
    ]
)

def _startup_env_check():
    """Fail-fast checks for required environment variables in production.

    If APP_ENV=production, ensure critical values are present.
    """
    app_env = os.getenv("APP_ENV", "development").strip().lower()
    missing = []
    if missing:
        logging.critical("Missing required environment variables for production: %s", ",".join(missing))
        raise SystemExit(1)


# run startup checks early
try:
    _startup_env_check()
except SystemExit:
    raise
except Exception as e:
    logging.warning("Startup env check failed unexpectedly: %s", e)

def _is_public_path(path: str) -> bool:
    if path in {"/", "/docs", "/redoc", "/openapi.json", "/api/health"}:
        return True
    if path.startswith("/home"):
        return True
    return False


# Authentication middleware removed; no token interception in this deployment.


UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
VALIDATION_RESULTS_DIR = UPLOAD_DIR / "validation_results"
VALIDATION_RESULTS_DIR.mkdir(parents=True, exist_ok=True)

COMPLETED_FOLDER = Path("validation/completed/Excel_Files")
COMPLETED_FOLDER.mkdir(parents=True, exist_ok=True) 
EXCEL_FILE_PATH = Path("Required_files/HDL_BO_Hierarchy_All_Objects_Charlie.xlsx")
TRANSFORMATION_ATTRIBUTES_FILE_PATH = Path("Required_files/Transformation - Common Attributes v3 2.xlsx")

# Configure logging
LOG_FILE_PATH = "server.log"    
logging.basicConfig(
    level=logging.DEBUG,  # <-- Only INFO and above will be logged
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH, mode='a', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


os.makedirs("static", exist_ok=True)
# "/home" for the static folder index.html and all those
app.mount("/home", StaticFiles(directory="static"), name="home")
@app.get("/home")
def read_root():
    return FileResponse("static/index.html")

@app.get("/api/health")
def health():
    return (
        {"status": "healthy"}
    )


pass_df = pd.DataFrame()
fail_df = pd.DataFrame()
def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=['Level-1']) 
    df = df.fillna("")
    return df

def extract_customer_instance_names_from_env(env_path: Path):
    combos = set()
    if env_path.exists():
        with env_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _ = line.split("=", 1)
                key = key.strip().strip('"').strip("'")
                m = re.match(r"([A-Za-z0-9]+)_([A-Za-z0-9]+)_", key)
                if m:
                    cust = m.group(1).capitalize()
                    inst = m.group(2).capitalize()
                    combos.add((cust, inst))
    return sorted(list(combos))

def read_and_normalize_excel(excel_path: Path):
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")
    xls = pd.ExcelFile(excel_path)
    df = None
    for sheet in xls.sheet_names:
        tmp = pd.read_excel(excel_path, sheet_name=sheet, dtype=str)
        tmp.columns = [str(c).strip() for c in tmp.columns]
        if tmp.shape[0] > 0 and tmp.dropna(how='all').shape[0] > 0:
            df = tmp
            break
    if df is None:
        df = pd.read_excel(excel_path, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
    df = df.fillna("")
    col_key_map = {}
    for col in df.columns:
        key = re.sub(r'[^a-z0-9]', '', str(col).lower())
        col_key_map[key] = col
    return df, col_key_map

def build_hierarchy_from_df(df: pd.DataFrame, col_map: dict, combos: list):
    """
    Build hierarchical tree from the cleaned dataframe + normalized col_map.

    - col_map is expected to be the normalized-key -> original-column-name map
      created by `read_and_normalize_excel`.
    - combos is list of (customer, instance) pairs (from .env).
    """
    import re
    # ---------- helpers ----------
    def normalize_key(s: str) -> str:
        if not s:
            return ""
        return re.sub(r'[^a-z0-9]', '', str(s).lower())

    def resolve_col(col_map_local: dict, *variants: str):
        """Try multiple human variants and return the original column name (or None)."""
        for v in variants:
            k = normalize_key(v)
            if k in col_map_local and col_map_local[k]:
                return col_map_local[k]
        return None


    def parse_mandatory(raw) -> bool:
        if raw is None:
            return False
        s = str(raw).strip().lower()
        if s in ("1", "1.0", "true", "t", "yes", "y", "required", "mandatory"):
            return True
        if s in ("0", "0.0", "false", "f", "no", "n", ""):
            return False
        # try numeric fallback
        try:
            return float(s) != 0.0
        except Exception:
            logging.warning(f"Unexpected Mandatory_Objects value: {raw!r}; defaulting to False")
            return False

    # ---------- resolve columns robustly ----------
    level_cols = {
        i: resolve_col(col_map,
                      f"Level-{i}", f"Level {i}", f"level{i}", f"level_{i}")
        for i in range(3, 11)
    }
    file_col = resolve_col(col_map, "File Name", "File", "Filename", "file")
    template_col = resolve_col(col_map, "Template Name", "Template", "templatename")
    mandatory_col = resolve_col(col_map, "Mandatory Objects", "Mandatory_Objects", "Mandatory")
    required_helper_col = resolve_col(col_map,
                                     "Required - Helper Text",
                                     "Required Helper Text",
                                     "RequiredHelperText",
                                     "Required_Helper_Text",
                                     "Required Helper",
                                     "Required")
    supported_helper_col = resolve_col(col_map,
                                      "Supported Action - Helper Text",
                                      "Supported Action Helper Text",
                                      "SupportedActionHelperText",
                                      "Supported_Helper",
                                      "Supported Action",
                                      "SupportedAction")

    logging.debug("Resolved column mapping in build_hierarchy_from_df: "
                  f"file={file_col}, template={template_col}, mandatory={mandatory_col}, "
                  f"required_helper={required_helper_col}, supported_helper={supported_helper_col}, "
                  f"level_cols={level_cols}")

    # ---------- collect variants ----------
    variants = []
    for _, row in df.iterrows():
        # Collect levels 3..10 values (use original column names if resolved)
        levels = []
        for i in range(3, 11):
            colname = level_cols.get(i)
            val = ""
            if colname:
                # row.get works on Series; fallback to empty string
                raw = row.get(colname, "")
                val = "" if pd.isna(raw) else str(raw).strip()
                if val.lower() == "nan":
                    val = ""
            levels.append(val)

        if not any(levels):
            continue

        file_val = ""
        if file_col:
            raw = row.get(file_col, "")
            file_val = "" if pd.isna(raw) else str(raw).strip()

        template_val = ""
        if template_col:
            raw = row.get(template_col, "")
            template_val = "" if pd.isna(raw) else str(raw).strip()

        mandatory_val = False
        if mandatory_col:
            mandatory_val = parse_mandatory(row.get(mandatory_col, ""))

        required_helper_val = ""
        if required_helper_col:
            raw = row.get(required_helper_col, "")
            required_helper_val = "" if pd.isna(raw) else str(raw).strip()

        supported_helper_val = ""
        if supported_helper_col:
            raw = row.get(supported_helper_col, "")
            supported_helper_val = "" if pd.isna(raw) else str(raw).strip()

        variants.append({
            "levels": levels,
            "file": file_val,
            "dat_template": template_val,
            "Mandatory_Objects": mandatory_val,
            "Required - Helper Text": required_helper_val,
            "Supported Action - Helper Text": supported_helper_val,
        })

    # ---------- tree builder with safe stamping ----------
    folder_roots = {}

    def ensure_all_fields(node: dict, extra_fields: dict, set_file_fields: bool = False):
        """Ensure the node has the expected structural and helper keys without clobbering good data."""
        keys = [
            "file", "dat_template",
            "level_1", "level_2", "level_3", "level_4",
            "level_5", "level_6", "level_7", "level_8", "level_9", "level_10",
            "Required - Helper Text", "Supported Action - Helper Text",
        ]
        for k in keys:
            # prefer explicit values from extra_fields (only set when non-empty),
            # but ensure the key exists on the node (default to empty string or False for Mandatory_Objects).
            incoming = extra_fields.get(k) if extra_fields is not None else None

            if k in ("file", "dat_template"):
                if set_file_fields and incoming:
                    node[k] = incoming
                elif k not in node:
                    node[k] = ""
            else:
                # text fields / levels / helper texts
                if incoming is not None and incoming != "":
                    node[k] = incoming
                else:
                    # if missing, ensure key exists but do not overwrite existing non-empty values
                    if k not in node:
                        node[k] = ""
        # Handle Mandatory_Objects separately
        if "Mandatory_Objects" not in node:
            node["Mandatory_Objects"] = False
        return node

    def get_or_create_node(parent_collection, node_name, is_root=False, extra_fields=None, set_file_fields=False):
        # parent_collection is either a dict (for roots) or list (for children)
        if is_root:
            if node_name not in parent_collection:
                parent_collection[node_name] = {"name": node_name, "children": []}
            if extra_fields:
                ensure_all_fields(parent_collection[node_name], extra_fields, set_file_fields)
            return parent_collection[node_name]
        else:
            for child in parent_collection:
                if child.get("name") == node_name:
                    if extra_fields:
                        ensure_all_fields(child, extra_fields, set_file_fields)
                    return child
            new_node = {"name": node_name, "children": []}
            if extra_fields:
                ensure_all_fields(new_node, extra_fields, set_file_fields)
            parent_collection.append(new_node)
            return new_node

    # ---------- assemble tree ----------
    for cust, inst in combos:
        for v in variants:
            try:
                last_nonempty_idx = max(idx for idx, nm in enumerate(v["levels"]) if nm)
            except ValueError:
                continue

            extra = {
                "file": v["file"],
                "dat_template": v["dat_template"],
                "level_1": cust,
                "level_2": inst,
                "level_3": v["levels"][0],
                "level_4": v["levels"][1],
                "level_5": v["levels"][2],
                "level_6": v["levels"][3],
                "level_7": v["levels"][4],
                "level_8": v["levels"][5],
                "level_9": v["levels"][6],
                "level_10": v["levels"][7],
                "Required - Helper Text": v["Required - Helper Text"],
                "Supported Action - Helper Text": v["Supported Action - Helper Text"],
                # We will handle Mandatory_Objects separately at the leaf
            }

            root = get_or_create_node(folder_roots, cust, is_root=True, extra_fields=extra)
            inst_node = get_or_create_node(root["children"], inst, extra_fields=extra)
            current = inst_node

            for i, name in enumerate(v["levels"]):
                if not name:
                    continue
                is_leaf = (i == last_nonempty_idx)
                current = get_or_create_node(
                    current["children"], name,
                    extra_fields=extra,
                    set_file_fields=is_leaf
                )
                if is_leaf:
                    # Explicitly set Mandatory_Objects only on the leaf node
                    if "Mandatory_Objects" in v:
                        current["Mandatory_Objects"] = bool(v["Mandatory_Objects"])
                    # Helper texts: only set if non-empty (do not overwrite existing helper text with empty)
                    req_ht = v.get("Required - Helper Text", "")
                    if req_ht:
                        current["Required - Helper Text"] = req_ht
                    sup_ht = v.get("Supported Action - Helper Text", "")
                    if sup_ht:
                        current["Supported Action - Helper Text"] = sup_ht

    # Return list of root nodes (preserves the structure expected by frontend)
    return list(folder_roots.values())

@app.get("/")
def root():
    """
    Root endpoint providing a welcome message.
    """
    return {"message": "Hit /api/utils/menu-items to get the 8-level hierarchy tree."}


USER_EXCEL_FILE_PATH = Path("Required_files/Users.xlsx")
def load_user_data(file_path: Path):
    try:
        # Assuming it's an Excel file that pandas can read, or a CSV named .xlsx
        # If it's truly an .xlsx file, ensure openpyxl is installed (`pip install openpyxl`)
        # If it's actually a CSV that happens to be named .xlsx, pandas will often handle it.
        # If it's a CSV, you might want to use pd.read_csv instead depending on actual file format
        df = pd.read_excel(file_path) # Changed to read_excel based on user's original filename
        
        users = {}
        for index, row in df.iterrows():
            if 'UserName' in row and 'Password' in row and 'UserType' in row:
                users[row['UserName']] = {
                    "password": row['Password'],
                    "usertype": row['UserType']
                }
            else:
                logging.warning(f"Row {index} in {file_path} is missing expected columns (UserName, Password, UserType).")
        return users
    except FileNotFoundError:
        logging.error(f"Error: User file not found at {file_path}")
        return {}
    except Exception as e:
        logging.error(f"Error loading user data from {file_path}: {e}")
        return {}
    
# Load user data when the application starts
USER_DB = load_user_data(USER_EXCEL_FILE_PATH)

class UserLogin(BaseModel):
    username: str
    password: str

@app.post("/api/utils/login-access")
async def login_access(user_login: UserLogin):
    """
    Login page code will fetch the details from the Users.xlsx file and check and send it back to the frontend.
    """
    username = user_login.username
    password = user_login.password

    if username in USER_DB and USER_DB[username]["password"] == password:
        user_type = USER_DB[username]["usertype"]
        return JSONResponse(
            content={
                "message": "Login successful",
                "username": username,
                "user_type": user_type
            },
            status_code=status.HTTP_200_OK
        )
    else:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid username or password"
        )

ENV_PATH = Path(".env")
@app.get("/api/utils/hdl/menu-items")
def get_hierarchy_api():
    # Validate Excel path
    if not EXCEL_FILE_PATH.exists():
        raise HTTPException(status_code=500, detail=f"Excel file not found at {EXCEL_FILE_PATH}")

    try:
        # Read Excel and get cleaned DataFrame + column key map
        df_cleaned, col_map = read_and_normalize_excel(EXCEL_FILE_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Excel file: {str(e)}")

    # Extract customer-instance combos from .env
    combos = extract_customer_instance_names_from_env(ENV_PATH)
    if not combos:
        raise HTTPException(status_code=500, detail="No customer-instance combos found in .env")

    # Build hierarchy
    try:
        hierarchy_data = build_hierarchy_from_df(df_cleaned, col_map, combos)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to build hierarchy: {str(e)}")

    return {"hierarchy": hierarchy_data}


@app.get("/api/utils/hdl/stats")
def get_dashboard_stats():
    """
    Returns dashboard statistics:
    - Total uploads (files in uploads/Excel_Files directory)
    - Validated files (files in validation_results directory)
    - Transformed files (transformed files)
    - Active customers (from .env customer-instance combos)
    """
    try:
        excel_files_dir = Path("uploads/Excel_Files")
        validation_results_dir = Path("validation_results")
        
        # Count total uploads
        total_uploads = 0
        if excel_files_dir.exists():
            total_uploads = len(list(excel_files_dir.rglob("*.xlsx")))
        
        # Count validated files
        validated_files = 0
        if validation_results_dir.exists():
            validated_files = len(list(validation_results_dir.glob("*.xlsx"))) + len(list(validation_results_dir.glob("*.dat")))
        
        # Count transformed files (files with "transformed" in name)
        transformed_files = 0
        if validation_results_dir.exists():
            transformed_files = len([f for f in validation_results_dir.glob("*") if "transformed" in f.name.lower()])
        
        # Count active customers from .env
        combos = extract_customer_instance_names_from_env(ENV_PATH)
        active_customers = len(set([customer for customer, instance in combos]))
        
        return {
            "stats": {
                "totalUploads": total_uploads,
                "validatedFiles": validated_files,
                "transformedFiles": transformed_files,
                "activeCustomers": active_customers,
            }
        }
    except Exception as e:
        logger.error(f"Error retrieving dashboard stats: {e}")
        return {
            "stats": {
                "totalUploads": 0,
                "validatedFiles": 0,
                "transformedFiles": 0,
                "activeCustomers": 0,
            }
        }


@app.get("/api/utils/system-status")
def get_system_status():
    """
    Returns system status for Backend API, Oracle Database, and NLP Service
    """
    status_data = {
        "backendAPI": "online",  # Current service is always online
        "oracleDB": "online",    # Simplified - would normally check connectivity
        "nlpService": "online",  # Simplified - would normally check connectivity
    }
    
    # Check Backend API connectivity (this service itself)
    try:
        # If this endpoint is responding, backend is online
        status_data["backendAPI"] = "online"
    except Exception:
        status_data["backendAPI"] = "offline"
    
    # Check NLP Service connectivity (if running on localhost:9000)
    try:
        response = requests.head("http://localhost:9000/validate", timeout=2)
        status_data["nlpService"] = "online" if response.status_code < 500 else "offline"
    except Exception:
        status_data["nlpService"] = "offline"
    
    # Check Oracle Database connectivity (if credentials are set)
    try:
        if ORACLE_USERNAME and ORACLE_PASSWORD:
            # Simplified status - in production, make actual DB connection
            status_data["oracleDB"] = "online"
        else:
            status_data["oracleDB"] = "offline"
    except Exception:
        status_data["oracleDB"] = "offline"
    
    return {"status": status_data}


def get_columns_from_dat(file_bytes: bytes) -> Tuple[List[str], List[str], List[str]]:
    try:
        content = file_bytes.decode("utf-8-sig").splitlines()
        if not content:
            raise ValueError("The provided .dat file is empty.")

        header_line = content[0]
        all_columns = [col.strip() for col in header_line.split("|")]

        skipped_columns_list = []
        non_skipped_columns = []
        last_source_system_id_index = -1

        # Reverted to a simple substring check for "SourceSystemId"
        # This will correctly identify both "SourceSystemId" and "PersonId(SourceSystemId)"
        for i, col_name in enumerate(all_columns):
            if "(SourceSystemId)" in col_name:
                last_source_system_id_index = i
            if "SourceSystemId" in col_name:
                last_source_system_id_index = i  

        if last_source_system_id_index != -1:
            skipped_columns_list = all_columns[:last_source_system_id_index + 1]
            non_skipped_columns = all_columns[last_source_system_id_index + 1:]
        else:
            skipped_columns_list = []
            non_skipped_columns = all_columns

        return all_columns, non_skipped_columns, skipped_columns_list
    except Exception as e:
        logging.error(f"Failed to parse .dat file: {e}", exc_info=True)
        raise ValueError(f"Failed to parse .dat file: {str(e)}. Please ensure it's a valid pipe-separated file with a header row.")

@app.post("/api/hdl/upload-dat")
async def upload_dat_file(datFile: UploadFile = File(...)):
    if not datFile.filename.endswith(".dat"):
        raise HTTPException(status_code=400, detail="Invalid file format. Only .dat files are accepted.")

    dat_path = UPLOAD_DIR / "uploaded_dat.dat"

    try:
        dat_bytes = await datFile.read()
        all_columns, non_skipped_columns, skipped_columns = get_columns_from_dat(dat_bytes)

        with open(dat_path, "wb") as f:
            f.write(dat_bytes)
        # Ensure the response keys are consistent with the variable names
        return JSONResponse(content={
            "message": "DAT file uploaded successfully.",
            "datFileName": datFile.filename,
            "all_columns": all_columns,
            "non_skipped_columns": non_skipped_columns, # Renamed for clarity and consistency
            "skipped_columns": skipped_columns
        })

    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except Exception as e:
        logging.error(f"DAT file upload failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"DAT file upload failed: An unexpected server error occurred. Please try again.")

@app.post("/api/hdl/upload-excel")
async def upload_excel_file(excelFile: UploadFile = File(...)):
    """
    Uploads an .xlsx file and saves it.
    The file is saved as 'uploaded_excel.xlsx' in the UPLOAD_DIR.
    """
    if not excelFile.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Upload an .xlsx file.")

    excel_path = UPLOAD_DIR / "uploaded_excel.xlsx"

    try:
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excelFile.file, f)

        return JSONResponse(content={
            "message": "Excel file uploaded successfully.",
            "excelFileName": excelFile.filename,
        })

    except Exception as e:
        logging.error(f"Excel file upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Excel file upload failed: {str(e)}")

@app.post("/api/hdl/upload")
async def upload_hdl_files(datFile: UploadFile = File(...), excelFile: UploadFile = File(...)):
    """
    Uploads DAT and Excel files, saves them, and returns .dat column names.
    This endpoint expects both files simultaneously. Use /api/hdl/upload-dat
    and /api/hdl/upload-excel for separate uploads.
    """
    if not datFile.filename.endswith(".dat") or not excelFile.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Upload a .dat and .xlsx file.")

    dat_path = UPLOAD_DIR / "uploaded_file.dat"
    excel_path = UPLOAD_DIR / "uploaded_file.xlsx"

    try:
        dat_bytes = await datFile.read()
        columns = get_columns_from_dat(dat_bytes)
        with open(dat_path, "wb") as f:
            f.write(dat_bytes)

        # Save Excel file
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excelFile.file, f)

        return JSONResponse(content={
            "message": "Files uploaded successfully.",
            "datFileName": datFile.filename,
            "excelFileName": excelFile.filename,
            "columns_from_dat": columns
        })

    except Exception as e:
        logging.error(f"Combined upload failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Combined upload failed: {str(e)}")

@app.post("/api/hdl/{component}/save")
async def save_hdl_component(component: str, request: Request):
    """
    Save HDL component data sent from the frontend into a JSON file.
    Stores it in a temp folder under 'uploads/user/{component}/timestamped.json'.
    """
    try:
        payload = await request.json()

        # Create directories
        user_dir = UPLOAD_DIR / "user" / component
        user_dir.mkdir(parents=True, exist_ok=True)

        # Generate timestamped file name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = user_dir / f"{component}_{timestamp}.json"

        # Save payload as JSON
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)

        return {"message": f"Component '{component}' saved successfully.", "file": str(file_path)}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving component: {str(e)}")


def fetch_key_values(bo: str, attributes: list):
    """
    Fetches key values (Yes/No as boolean) for given component (bo) and attributes
    from the Key Values Mapping.xlsx file (single global file).
    """
    try:
        file_path = Path(__file__).parent / "Required_files" / "Key Values Mapping.xlsx"

        if not file_path.exists():
            logger.error(f"{file_path.name} not found in Required_files directory.")
            return {}

        df = pd.read_excel(file_path)

        # Normalize headers
        df.columns = df.columns.str.strip()
        column_map = {
            "BO HDL File Name": "GLOBAL_BO",
            "Component Name": "PARENT_BO",
            "HDL Attribute Name": "HDL_ATTRIBUTE_NAME",
            "Key_Values": "KEY_VALUES"
        }
        df = df.rename(columns={k: v for k, v in column_map.items() if k in df.columns})

        # Normalize columns
        for col in ["GLOBAL_BO", "PARENT_BO", "HDL_ATTRIBUTE_NAME", "KEY_VALUES"]:
            if col in df.columns:
                df[col] = df[col].fillna("").astype(str).str.strip()

        # Filter for this BO + attributes
        filtered = df[
            (df["PARENT_BO"].str.lower() == bo.lower()) &
            (df["HDL_ATTRIBUTE_NAME"].isin(attributes))
        ]

        # Build mapping: attr -> boolean
        key_values_map = {}
        for attr in attributes:
            row = filtered.loc[filtered["HDL_ATTRIBUTE_NAME"] == attr, "KEY_VALUES"]
            if not row.empty and row.iloc[0].lower() == "yes":
                key_values_map[attr] = True
            else:
                key_values_map[attr] = False

        return key_values_map

    except Exception as e:
        logger.error(f"Key values fetch failed: {str(e)}", exc_info=True)
        return {}


@app.post("/api/hdl/mandatory/batch")
def get_required_batch(
    bo: str = Body(..., alias="componentName"),
    attributes: list = Body(..., embed=True),
    customerName: str = Body(None, alias="customerName"),
    instanceName: str = Body(None, alias="instanceName"),
):
    """
    Retrieves mandatory attributes, helper text, and key values
    for a given component and a list of attributes.
    """
    try:
        if not customerName or not instanceName:
            return JSONResponse(
                status_code=400,
                content={"error": "customerName and instanceName are required."}
            )

        # Mandatory fields Excel
        file_path = Path(__file__).parent / "Required_files" / f"{customerName}_{instanceName}_MandatoryFields.xlsx"
        if not file_path.exists():
            return JSONResponse(
                status_code=404,
                content={"error": f"{file_path.name} not found."}
            )

        mandate = pd.read_excel(file_path)
        mandate.columns = mandate.columns.str.strip()
        column_map = {
            "BO HDL File Name": "GLOBAL_BO",
            "Component Name": "PARENT_BO",
            "HDL Attribute Name": "HDL_ATTRIBUTE_NAME",
            "Required": "REQUIRED",
            "Helper_Text": "HELPER_TEXT",
            "Data Type": "DATA_TYPE"
        }
        mandate = mandate.rename(columns={k: v for k, v in column_map.items() if k in mandate.columns})
        for col in ["GLOBAL_BO", "PARENT_BO", "HDL_ATTRIBUTE_NAME"]:
            if col in mandate.columns:
                mandate[col] = mandate[col].fillna("").astype(str).str.strip()
        if "HELPER_TEXT" in mandate.columns:
            mandate["HELPER_TEXT"] = mandate["HELPER_TEXT"].fillna("").astype(str).str.strip()
        else:
            mandate["HELPER_TEXT"] = ""

        if "DATA_TYPE" in mandate.columns:
            mandate["DATA_TYPE"] = mandate["DATA_TYPE"].fillna("").astype(str).str.strip()
        else:
            mandate["DATA_TYPE"] = ""
        if "REQUIRED" in mandate.columns:
            mandate["REQUIRED"] = mandate["REQUIRED"].fillna("").astype(str).str.strip()
        else:
            mandate["REQUIRED"] = "No"
        # Filter rows
        filtered = mandate[
            (mandate["PARENT_BO"].str.lower() == bo.lower()) &
            (mandate["HDL_ATTRIBUTE_NAME"].isin(attributes))
        ]

        # Fetch key values for this BO + attributes
        key_values_map = fetch_key_values(bo, attributes)

        # Build response dict
        mandatory_dict = {}
        for attr in attributes:
            row = filtered[filtered["HDL_ATTRIBUTE_NAME"] == attr]
            mandatory_dict[attr] = {
                "mandatory": row["REQUIRED"].iloc[0].strip().lower() == "yes" if not row.empty else False,
                "helper_text": row["HELPER_TEXT"].iloc[0] if not row.empty else "",
                "data_type": row["DATA_TYPE"].iloc[0] if not row.empty and "DATA_TYPE" in row.columns else "",
                "key_values": key_values_map.get(attr, [])
            }

        return {"mandatory": mandatory_dict}

    except Exception as e:
        logger.error(f"Mandatory fetch failed: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={"error": f"Mandatory fetch failed: {str(e)}"}
        )

@app.post("/api/hdl/lookup/batch")
def robust_lookup(
    bo: str = Body(..., alias="componentName"),
    global_bo: str = Body(None, alias="globalComponentName"),
    attributes: List[str] = Body(..., embed=True, alias="Attributes"),
    transaction: Optional[bool] = Body(False, alias="transaction"),
    customerName: str = Body(None, alias="customerName"),
    instanceName: str = Body(None, alias="instanceName"),
):
    """
    Retrieves lookup data for a given component and a list of attributes,
    with strict or fallback matching based on provided parameters.
    If transaction is True, always append results for only attribute name as well.
    """
    try:
        
        # Changed to read Excel since the provided file is an Excel sheet
        lookup_df = pd.read_excel(Path(__file__).parent/ "Required_files"/f"{customerName}_{instanceName}_LookupData.xlsx")

        logger.info(f"Lookup fetching started with global_bo='{global_bo}', bo='{bo}', attributes={attributes}, transaction={transaction}")

        # Normalize relevant columns for comparison
        for col in ["BO_NAME", "COMP_NAME", "HDL_Attribute_Name", "CODE_Name"]:
            if col in lookup_df.columns:
                # Explicitly fill NaN with empty string before other string operations
                lookup_df[col] = lookup_df[col].fillna('').astype(str).str.strip().str.lower()
            else:
                # Ensure the column exists, even if empty, to prevent KeyError
                lookup_df[col] = "" # Initialize with empty string for consistent comparison

        bo_name = global_bo.strip().lower() if global_bo else ""
        comp_name = bo.strip().lower() if bo else ""
        normalized_attrs = [attr.strip().lower() for attr in attributes]

        # --- Initial Filtering Logic ---
        # Strictest: all three columns
        if bo_name and comp_name:
            filtered_df = lookup_df[
                (lookup_df["BO_NAME"] == bo_name) &
                (lookup_df["COMP_NAME"] == comp_name) &
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs))
            ]
        # Fallback: COMP_NAME + HDL_Attribute_Name
        elif comp_name:
            filtered_df = lookup_df[
                (lookup_df["COMP_NAME"] == comp_name) &
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs))
            ]
        # Fallback: BO_NAME + HDL_Attribute_Name
        elif bo_name:
            filtered_df = lookup_df[
                (lookup_df["BO_NAME"] == bo_name) &
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs))
            ]
        # Fallback: only HDL_Attribute_Name (when no BO or COMP name is provided)
        else:
            filtered_df = lookup_df[
                lookup_df["HDL_Attribute_Name"].isin(normalized_attrs)
            ]
        
        # --- Transaction Logic ---
        # Ensure transaction is a boolean
        if not isinstance(transaction, bool):
            raise ValueError("Transaction parameter must be a boolean value.")

        # If transaction is True, always append results where BO_NAME and COMP_NAME are empty
        if transaction:
            # Filter for transaction-specific lookups (where BO_NAME and COMP_NAME are empty)
            transaction_df = lookup_df[
                (lookup_df["HDL_Attribute_Name"].isin(normalized_attrs)) &
                (lookup_df["BO_NAME"] == "") &
                (lookup_df["COMP_NAME"] == "")
            ]
            logger.debug(f"Transaction DF before concat for attributes {normalized_attrs}: \n{transaction_df}") # Debug log
            
            # Concatenate only if transaction_df is not empty to avoid unnecessary operations
            if not transaction_df.empty:
                # Use pd.concat to combine the dataframes and drop duplicates
                # This ensures that if a row from transaction_df is already in filtered_df, it's not duplicated.
                filtered_df = pd.concat([filtered_df, transaction_df]).drop_duplicates().reset_index(drop=True)
        
        logger.debug(f"Final filtered DF before processing: \n{filtered_df}") # Debug log

        if filtered_df.empty:
            logger.info("No lookups found for the given criteria.")
            return {"lookups": {}, "default_code_names": {}}

        lookups = {}
        default_code_names = {}
        # Create a mapping from normalized attribute names back to their original case
        norm_to_orig = {attr.strip().lower(): attr for attr in attributes}

        for norm_attr in normalized_attrs:
            # Filter for rows corresponding to the current normalized attribute
            attr_rows = filtered_df[filtered_df["HDL_Attribute_Name"] == norm_attr]
            
            # Convert relevant columns to a list of dictionaries, filling NaN with empty string
            lookup_list = attr_rows[["CODE_Name", "Value", "Meaning", "Enabled_Flag", "Effective_Date"]].fillna("").to_dict(orient="records")
            
            if lookup_list:
                # Use the original attribute name for the key in the response
                lookups[norm_to_orig[norm_attr]] = lookup_list
                # Set the default_code_name from the first entry in the lookup list
                default_code_names[norm_to_orig[norm_attr]] = lookup_list[0]["CODE_Name"]

        logger.info("Lookup fetching completed successfully.")
        return {"lookups": lookups, "default_code_names": default_code_names}

    except FileNotFoundError:
        logger.error(f"{customerName}_{instanceName}_LookupData.xlsx not found. Please ensure it's in the 'Required_files' directory.", exc_info=True)
        return JSONResponse(status_code=404, content={"error": f"{customerName}_{instanceName}_LookupData.xlsx not found."})
    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}", exc_info=True)
        return JSONResponse(status_code=400, content={"error": f"Validation error: {str(ve)}"})
    except Exception as e:
        logger.error(f"Lookup failed: {str(e)}", exc_info=True)
        return JSONResponse(status_code=500, content={"error": f"Lookup failed: {str(e)}"})

# =========================================================
# Request Model
# =========================================================
class GetAttributesRequest(BaseModel):
    componentName: str = Field(..., min_length=1, description="DAT template file name")


# =========================================================
# Response Model
# =========================================================
class GetAttributesResponse(BaseModel):
    attributes: list[str]
    count: int


# =========================================================
# Endpoint
# =========================================================
@app.post("/api/hdl/get-attributes", response_model=GetAttributesResponse, tags=["HDL"])
def get_attributes(request: GetAttributesRequest):
    """
    Fetch attributes from DAT file header.
    Parses pipe-delimited first line and returns attribute list.
    """

    component_name = request.componentName.strip()

    try:
        dat_file_path = Path("Required_files/Dat_Files") / f"{component_name}.dat"

        # ---- File check ----
        if not dat_file_path.exists():
            logger.error(f"DAT file not found: {component_name}")
            raise HTTPException(
                status_code=404,
                detail=f"DAT file not found for component '{component_name}'"
            )

        # ---- Read header ----
        with dat_file_path.open("r", encoding="utf-8") as f:
            first_line = f.readline().strip()

        if not first_line:
            raise HTTPException(
                status_code=400,
                detail="DAT file header is empty"
            )

        # ---- Parse attributes ----
        attributes = [col.strip() for col in first_line.split("|") if col.strip()]

        if not attributes:
            raise HTTPException(
                status_code=400,
                detail="No attributes found in DAT header"
            )

        logger.info(f"{component_name}: extracted {len(attributes)} attributes")

        return GetAttributesResponse(
            attributes=attributes,
            count=len(attributes)
        )

    except HTTPException:
        raise

    except Exception as e:
        logger.exception("DAT parsing failed")
        raise HTTPException(
            status_code=500,
            detail="Internal error while parsing DAT file"
        )

class ColumnRequest(BaseModel):
    file_id: str
    columns: list[str]

@app.post("/api/get-excel-columns")
async def get_excel_columns(file: UploadFile = File(...), user_id: str = Form(...)):
    try:
        file_id = str(uuid.uuid4())
        file_path = os.path.join(UPLOAD_DIR, f"{file_id}.xlsx")
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

        logger.info(f"Saved file to: {file_path}")
        df = pd.read_excel(file_path, engine="openpyxl")
        logger.info(f"Excel columns: {df.columns.tolist()}")
        return {
            "columns": df.columns.tolist(),
            "file_id": file_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@app.post("/api/hdl/apply-transformation-and-download", summary="Apply transformation and download Excel")
async def apply_transformation_and_download(
    attribute_column: str = Form(..., description="Name of the column to use for transformation (e.g., 'Suffix')"),
    raw_excel_file: UploadFile = File(..., description="Raw Excel file for transformation"),
):
    logger.info(f"Received request for file: {raw_excel_file.filename} with attribute column: {attribute_column}")

    try:
        contents = await raw_excel_file.read()
        logger.debug(f"Read {len(contents)} bytes from uploaded file.")
        excel_file_bytes = BytesIO(contents)

        try:
            df = pd.read_excel(excel_file_bytes, engine='openpyxl')
            logger.info(f"Successfully read Excel into DataFrame. Shape: {df.shape}")
            if df.empty:
                logger.warning("Uploaded Excel file is empty.")
                raise HTTPException(status_code=400, detail="Uploaded Excel file is empty or contains no data.")
        except pd.errors.EmptyDataError:
            logger.error("Pandas could not read data from the Excel file. It might be empty or malformed.")
            raise HTTPException(status_code=400, detail="Could not read data from Excel. File might be empty or corrupted.")
        except Exception as e:
            logger.exception("Error reading Excel file with pandas.")
            raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

        df_transformed = df.copy() 
        logger.debug("Starting data transformation.")

        for col in df_transformed.columns:
            if df_transformed[col].dtype == 'object':
                if pd.isna(df_transformed[col]).any():
                    df_transformed[col] = df_transformed[col].fillna('')
                    logger.debug(f"Replaced NaN with empty string in column: {col}")
            elif pd.api.types.is_numeric_dtype(df_transformed[col]):
                if pd.isna(df_transformed[col]).any() or np.isinf(df_transformed[col]).any():
                    df_transformed[col] = df_transformed[col].replace({np.nan: 0, np.inf: 1e308, -np.inf: -1e308})
                    logger.debug(f"Handled NaN/Inf in numeric column: {col}")

        if attribute_column in df_transformed.columns:
            def transform_value(value):
                clean_value = str(value).strip().lower() 
                if clean_value == 'senior':
                    return 'Sr'
                elif clean_value == 'junior':
                    return 'Jr'
                elif clean_value == '': 
                    return '' 
                else:
                    return value 
            
            # Apply the transformation function directly to the specified column
            df_transformed[attribute_column] = df_transformed[attribute_column].apply(transform_value)
            logger.info(f"Applied specific transformation directly to column: '{attribute_column}'.")
        else:
            logger.warning(f"Attribute column '{attribute_column}' not found in the Excel file. No in-place transformation performed.")
            df_transformed['Info'] = f"Attribute column '{attribute_column}' not found for in-place transformation."

        # --- End of specific transformation logic ---

        output_excel_file = BytesIO()
        try:
            df_transformed.to_excel(output_excel_file, index=False, engine='openpyxl')
            output_excel_file.seek(0)
            file_size = output_excel_file.getbuffer().nbytes
            logger.info(f"Successfully saved transformed DataFrame to BytesIO. File size: {file_size} bytes.")
            if file_size == 0:
                logger.error("Generated Excel file is empty (0 bytes).")
                raise HTTPException(status_code=500, detail="Generated Excel file is empty. Transformation might have failed.")
        except Exception as e:
            logger.exception("Error saving DataFrame to Excel BytesIO.")
            raise HTTPException(status_code=500, detail=f"Failed to generate Excel file: {e}")

        original_filename = raw_excel_file.filename
        base_filename = original_filename.split('/')[-1].split('\\')[-1]
        filename = f"transformed_{base_filename}"
        
        headers = {
            "Content-Disposition": f"attachment; filename=\"{filename}\"",
            "Content-Length": str(file_size)
        }
        
        logger.info(f"Sending transformed Excel file: {filename}")
        return StreamingResponse(
            output_excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.exception("An unhandled error occurred during file transformation.")
        raise HTTPException(status_code=500, detail=f"An unexpected server error occurred: {e}. Please try again or contact support.")


@app.get("/api/transform/get-mapping")
async def get_attribute_mapping(attribute: str = Query(..., description="The attribute for which to retrieve the mapping.")):
    """
    Retrieves Oracle Value as source and Customer Value as target for a given attribute from 'Transformation - Common Attributes v1.xlsx'.
    This endpoint is designed to provide data for React Flow to visualize mappings.
    Null values (NaN, None) will be returned as empty strings.
    If all retrieved mappings have empty source_value and target_value, an empty JSON object {} is returned.

    """
    try:
        if not TRANSFORMATION_ATTRIBUTES_FILE_PATH.exists():
            logging.error(f"Transformation mapping file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}")
            raise HTTPException(status_code=500, detail=f"Transformation mapping file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}")

        df = pd.read_excel(TRANSFORMATION_ATTRIBUTES_FILE_PATH, engine="openpyxl")
        df.columns = [col.strip() for col in df.columns] # Clean column names

        # Expected columns for filtering and output
        filter_col = 'Attributes for Transformation'
        source_col = 'Customer Value'
        target_col = 'Oracle Value'

        # Validate that the necessary columns exist in the Excel file
        required_cols = [filter_col, source_col, target_col]
        if not all(col in df.columns for col in required_cols):
            raise HTTPException(
                status_code=500,
                detail=f"Required columns '{filter_col}', '{source_col}', or '{target_col}' not found in "
                       f"'{TRANSFORMATION_ATTRIBUTES_FILE_PATH}'. Please ensure the Excel file has these columns."
            )

        # Filter for the specific attribute (case-insensitive and trimmed)
        filtered_df = df[df[filter_col].astype(str).str.strip().str.lower() == attribute.strip().lower()]

        # Prepare the list of mappings for the frontend
        mappings = []
        for index, row in filtered_df.iterrows():
            # Get source and target values
            source_val = row[source_col]
            target_val = row[target_col]

            # Convert NaN or None to empty string
            processed_source_val = ""
            if source_val is not None and (not (isinstance(source_val, float) and math.isnan(source_val))):
                processed_source_val = str(source_val)

            processed_target_val = ""
            if target_val is not None and (not (isinstance(target_val, float) and math.isnan(target_val))):
                processed_target_val = str(target_val)
            
            mappings.append({
                "source_value": processed_source_val,
                "target_value": processed_target_val
            })
        
        # Check if the mappings list is empty or if all entries have empty source and target values
        if not mappings or all(item["source_value"] == "" and item["target_value"] == "" for item in mappings):
            return JSONResponse(content={}, status_code=200)
        
        return JSONResponse(content=mappings, status_code=200)

    except Exception as e:
        logging.error(f"Error fetching attribute mapping for '{attribute}': {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching attribute mapping: {str(e)}")


# Define the path to your transformation attributes Excel file
TRANSFORMATION_ATTRIBUTES_FILE_PATH = Path("./Required_files/Transformation - Common Attributes v3 2.xlsx")

# Ensure the Required_files directory exists
TRANSFORMATION_ATTRIBUTES_FILE_PATH.parent.mkdir(parents=True, exist_ok=True)


# --- Pydantic Model for Request Body ---
class BulkTransformationPayload(BaseModel):
    componentName: str
    transformations: Dict[str, str] # Dictionary for 'Attribute for transformation' -> 'Transformation map name'
    # NEW FIELD: Dictionary for 'Customer Value' -> 'Oracle Value' replacements
    customer_oracle_replacements: Optional[Dict[str, str]] = {}

# --- API Endpoint ---

# --- Updated Payload Model ---
class BulkTransformationPayload(BaseModel):
    """
    Payload for bulk data transformation mappings and triggering customer value transformations.
    """
    componentName: str
    transformations: Dict[str, str] # e.g., {"Attribute A": "Map1", "Attribute B": "Map2"}
    # New field: List of attribute names for which Customer Value should be replaced by Oracle Value
    attributes_for_customer_value_transformation: Optional[List[str]] = None

@app.post("/api/hdl/transform-customer-excel")
async def transform_customer_excel(
    raw_excel_file: UploadFile = File(..., description="Customer Excel file to be transformed")
):
    """
    Receives a customer's Excel file, applies transformations based on the 
    'Transformation - Common Attributes v3 2.xlsx' file, and returns the
    modified Excel file for download.
    """
    try:
        # 1. Load the transformation rules from the master Excel file on the server
        # It's good practice to ensure this file exists before trying to read it
        if not TRANSFORMATION_ATTRIBUTES_FILE_PATH.exists():
            logging.error(f"Transformation rules file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}")
            raise HTTPException(
                status_code=500, 
                detail=f"Transformation rules file not found at {TRANSFORMATION_ATTRIBUTES_FILE_PATH}. Please ensure it's in the 'Required_files' directory."
            )
        
        rules_df = pd.read_excel(TRANSFORMATION_ATTRIBUTES_FILE_PATH, engine='openpyxl')
        rules_df.columns = [col.strip() for col in rules_df.columns]
        
        # Ensure required columns exist in the rules file
        required_rule_cols = ['Attributes for Transformation', 'Customer Value', 'Oracle Value']
        if not all(col in rules_df.columns for col in required_rule_cols):
            raise HTTPException(status_code=500, detail="Transformation rules file is missing required columns. Expected: 'Attributes for Transformation', 'Customer Value', 'Oracle Value'.")

        # 2. Load the uploaded customer Excel file into a pandas DataFrame
        # FIX: Read the content into BytesIO first
        contents = await raw_excel_file.read()
        excel_in_memory = BytesIO(contents)
        
        # First read without header to check number of rows
        temp_df = pd.read_excel(excel_in_memory, engine='openpyxl', header=None)
        if len(temp_df) < 2:  # Check if file has at least 2 rows (header + data)
            raise HTTPException(
                status_code=400,
                detail="Excel file must contain at least 2 rows: one header row and at least one data row."
            )
        
        # Reset file pointer and read with proper header
        excel_in_memory.seek(0)
        customer_df = pd.read_excel(excel_in_memory, engine='openpyxl', header=1)
        # Ensure column names are strings before stripping
        customer_df.columns = [str(col).strip() for col in customer_df.columns]
        logging.info(f"Successfully read customer Excel file. Shape: {customer_df.shape}")

        # 3. Apply transformations column by column
        for column_to_transform in customer_df.columns:
            # Find the transformation rules for the current column
            specific_rules = rules_df[rules_df['Attributes for Transformation'] == column_to_transform]
            
            if not specific_rules.empty:
                # Create a mapping dictionary: { 'Customer Value': 'Oracle Value' }
                # Drop rows where 'Customer Value' is empty to avoid incorrect mapping
                specific_rules = specific_rules.dropna(subset=['Customer Value'])
                
                # Convert both keys and values to string before creating Series/dict
                # This handles cases where Excel might interpret values as numbers
                value_map = pd.Series(
                    specific_rules['Oracle Value'].astype(str).values,
                    index=specific_rules['Customer Value'].astype(str)
                ).to_dict()

                if value_map:
                    # Apply the mapping to the column in the customer's DataFrame
                    # Ensure values in customer_df column are also strings for consistent replacement
                    customer_df[column_to_transform] = customer_df[column_to_transform].astype(str).replace(value_map)
                    logging.info(f"Applied transformation to column: '{column_to_transform}' using map: {value_map}")
                else:
                    logging.info(f"No valid value mappings found for column '{column_to_transform}' in the transformation rules.")
            else:
                logging.debug(f"No specific transformation rules found for column: '{column_to_transform}'. Skipping.")


        # 4. Save the transformed DataFrame to an in-memory Excel file
        output_excel_file = BytesIO()
        try:
            customer_df.to_excel(output_excel_file, index=False, engine='openpyxl')
            output_excel_file.seek(0) # Rewind the buffer to the beginning
            file_size = output_excel_file.getbuffer().nbytes
            logging.info(f"Successfully created transformed Excel file in memory. Size: {file_size} bytes.")
            if file_size == 0:
                logging.warning("Generated transformed Excel file is empty (0 bytes).")
                raise HTTPException(status_code=500, detail="Transformed Excel file is empty. Transformation might have resulted in no data.")
        except Exception as save_error:
            logging.error(f"Error saving transformed DataFrame to in-memory Excel: {save_error}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to generate transformed Excel file: {save_error}")


        # 5. Return the in-memory file as a downloadable response
        original_filename = raw_excel_file.filename
        # Sanitize filename to prevent directory traversal or other issues
        safe_original_filename = Path(original_filename).name
        output_filename = f"transformed_{safe_original_filename}"
        
        headers = {
            "Content-Disposition": f"attachment; filename=\"{output_filename}\"",
            "Content-Length": str(file_size) # Set Content-Length header
        }
        
        logging.info(f"Sending transformed Excel file: {output_filename}")
        return StreamingResponse(
            output_excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException as http_exc:
        # Re-raise HTTPExceptions as they contain specific error details and status codes
        raise http_exc
    except Exception as e:
        # Catch any other unexpected errors and log them
        logging.error(f"An unexpected error occurred during transform_customer_excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"An unexpected server error occurred during file transformation: {str(e)}")


@app.post("/api/hdl/nlp/validate")
async def proxy_nlp_validate(request: Request):
    """
    Proxies the request to the Flask NLP validation service at /validate.
    Accepts multipart/form-data with 'csv_file' and 'validation_file'.
    """
    try:
        # Read the incoming form data
        form = await request.form()
        files = {}
        for key in form:
            file = form[key]
            if hasattr(file, 'filename'):
                files[key] = (file.filename, await file.read(), file.content_type)
        # Proxy to Flask app (assume running on localhost:9000)
        flask_url = "http://localhost:9000/validate"
        response = requests.post(flask_url, files=files)
        return JSONResponse(status_code=response.status_code, content=response.json())
    except Exception as e:
        logger.error(f"Error proxying to NLP validation service: {e}")
        raise HTTPException(status_code=500, detail=f"Proxy error: {e}")

@app.post("/api/hdl/save_code")
async def save_code(request: Request):
    try:
        payload = await request.json()
        code = payload.get("code", "")
        componentName = payload.get("component_name", "")
        rules = payload.get("rules", {})
        conditions = payload.get("conditions", [])
        customerName = payload.get("customerName", "")
        instanceName = payload.get("instanceName", "")

        if not componentName:
            return {"success": False, "error": "component_name missing"}

        # Filename pattern
        filename = f"{customerName}_{instanceName}_{componentName}.py"

        save_dir = UPLOAD_DIR / "saved_code"
        save_dir.mkdir(parents=True, exist_ok=True)
        file_path = save_dir / filename

        # Save python code
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(code)

        # Load existing JSON
        json_path = Path("Required_files/Available_NLP.json")
        data = []
        if json_path.exists():
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except:
                data = []

        # Update or insert
        found = False
        for item in data:
            if item.get("file_name") == filename:
                item["rules"] = rules
                item["conditions"] = conditions
                found = True
                break

        if not found:
            data.append({
                "component": componentName,
                "file_name": filename,
                "rules": rules,
                "conditions": conditions
            })

        # Save JSON
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return {
            "success": True,
            "message": "Code and rules stored successfully",
            "file": str(file_path)
        }

    except Exception as e:
        logger.error(f"Save Code Error: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


JSON_PATH = Path("Required_files/Available_NLP.json")


@app.get("/api/hdl/get_rules")
async def get_rules(customerName: str, instanceName: str, componentName: str):
    try:
        json_path = Path("Required_files/Available_NLP.json")
        if not json_path.exists():
            return {"rules": []}

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        # Match stored filename pattern
        expected_filename = f"{customerName}_{instanceName}_{componentName}.py"

        for item in data:
            if item.get("file_name") == expected_filename:
                rules_obj = item.get("rules", {})

                # Convert rule object → frontend expected array format
                normalized_rules = [
                    {"nlr": rule_data.get("rule", ""), "column": col}
                    for col, rule_data in rules_obj.items()
                ]

                return {"rules": normalized_rules}

        return {"rules": []}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load rules: {str(e)}")


@app.post("/api/hdl/nlp/batch")
def get_nlr_rules_batch(
    attributes: List[str] = Body(..., embed=True, alias="attributes")
):
    """
    Retrieves NLR rules for a list of attributes from the Excel file 'Required_files/Available_NLP.xlsx'.
    Returns a dictionary mapping each attribute to its rules and conditions, and a 'has_rules' boolean for each attribute.
    """
    try:
        import openpyxl
        excel_path = Path("Required_files/Available_NLP.xlsx")
        if not excel_path.exists():
            return JSONResponse(status_code=404, content={"error": "NLR rules file not found."})

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # Build a dict of all available rules
        rules_dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            attribute = str(row[0]).strip() if row[0] else ""
            if attribute:
                rules = str(row[1]).strip() if row[1] else ""
                conditions = str(row[2]).strip() if row[2] else ""
                rules_list = rules.split(", ") if rules else []
                conditions_list = conditions.split(", ") if conditions else []
                rules_dict[attribute] = {
                    "rules": rules_list,
                    "conditions": conditions_list
                }

        # Filter for requested attributes only and add has_rules
        filtered = {}
        for attr in attributes:
            entry = rules_dict.get(attr, {"rules": [], "conditions": []})
            filtered[attr] = {
                **entry,
                "has_rules": bool(entry["rules"] and any(r.strip() for r in entry["rules"]))
            }

        return {
            "nlr_rules": filtered,
            "available_attributes": list(rules_dict.keys())
        }

    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"NLR rules batch lookup failed: {str(e)}"})

  

# DIR for bulk excel upload (redefined for clarity, or can use UPLOAD_DIR)
DIR = Path("uploads/Excel_Files")
DIR.mkdir(parents=True, exist_ok=True) # Ensure this directory exists at startup

def populate_actual_termination_date_from_resignation(
    file_path: Path,
    header_row_num: int = 2,
    save_as_new_file: bool = False,
    hireActions: List[str] = [],
    globalTransfers: List[str] = [],
    termAction: List[str] = []
):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[header_row_num]]

        combined_hire_glb = [x.upper() for x in hireActions + globalTransfers]
        termAction = [x.upper() for x in termAction]

        # Ensure ActualTerminationDate exists
        if "ActualTerminationDate" not in headers:
            headers.append("ActualTerminationDate")
            ws.cell(row=header_row_num, column=len(headers)).value = "ActualTerminationDate"

        person_idx = headers.index("PersonNumber")
        action_idx = headers.index("ActionCode")
        eff_start_idx = headers.index("EffectiveStartDate")
        actual_term_idx = headers.index("ActualTerminationDate")  # <-- Correct column reference

        # Group rows per employee
        person_rows = {}
        for i, row in enumerate(ws.iter_rows(min_row=header_row_num+1), start=header_row_num+1):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue
            person_rows.setdefault(person, []).append((i, row))

        for person, rows in person_rows.items():
            # Sort rows by EffectiveStartDate (ignore infinite dates)
            def safe_date(r):
                val = r[1][eff_start_idx].value
                if isinstance(val, datetime) and val.year >= 4000:  # Oracle Infinite Date
                    return datetime.max
                return val or datetime.max

            sorted_rows = sorted(rows, key=safe_date)

            open_cycle_start = None

            for idx, row in sorted_rows:
                action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""
                eff_date = row[eff_start_idx].value
                # If no EffectiveStartDate → make ATD blank
                if eff_date is None:
                    ws.cell(row=idx, column=actual_term_idx+1).value = None
                    continue

                # Ignore invalid future boundary date (Oracle infinite)
                if isinstance(eff_date, datetime) and eff_date.year >= 4000:
                    ws.cell(row=idx, column=actual_term_idx+1).value = None
                    continue

                if action in combined_hire_glb:
                    open_cycle_start = idx

                elif action in termAction and open_cycle_start:
                    resignation_date = eff_date
                    termination_date = resignation_date - timedelta(days=1)

                    # Apply termination date only to rows in this cycle
                    for sub_idx, _ in sorted_rows:
                        if open_cycle_start <= sub_idx < idx:
                            ws.cell(row=sub_idx, column=actual_term_idx+1).value = termination_date

                    # Reset after closing cycle
                    open_cycle_start = None

        # Save output
        if save_as_new_file:
            new_file = file_path.parent / f"{file_path.stem}_updated.xlsx"
            wb.save(new_file)
            return new_file

        wb.save(file_path)
        return file_path

    except Exception as e:
        logger.error(f"[ERROR] ❌ {e}")
        return None

    
def parse_excel_date(val):
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).split(' ')[0], '%Y-%m-%d')
    except Exception:
        return None


def validate_termination_date(file_path: Path, header_row_num: int = 2, save_as_new_file: bool = True, TermActions: List[str] = [], HireActions: List[str] = []):
    if TermActions is None:
        logger.info("[INFO] No Termination Actions provided")
        return
    if HireActions is None:
        logger.info("[INFO] No Hire Actions provided")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        logger.info(f"[VALIDATION] 🚦 Checking termination logic on {file_path.name}")

        headers = [cell.value for cell in ws[header_row_num]]
        person_idx = headers.index("PersonNumber")
        action_idx = headers.index("ActionCode")
        eff_start_idx = headers.index("EffectiveStartDate")

        start_row = header_row_num + 1
        rows_to_delete = set()
        person_rows = {}

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue
            if person not in person_rows:
                person_rows[person] = []
            person_rows[person].append((i, row))

        for person, rows in person_rows.items():
            hire_dates = []
            termination_rows = []

            for idx, row in rows:
                action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""
                date_val = parse_excel_date(row[eff_start_idx].value)
                if action in HireActions and date_val:
                    hire_dates.append(date_val)
                if action in TermActions and date_val:
                    termination_rows.append((idx, date_val))

            if termination_rows:
                if not hire_dates:
                    logger.info(f"[INVALID] ❌ {person} has termination but no hire record. Removing all rows.")
                    rows_to_delete.update(idx for idx, _ in rows)
                    continue

                earliest_hire = min(hire_dates)
                for term_idx, term_date in termination_rows:
                    if term_date < earliest_hire:
                        logger.info(f"[INVALID] ❌ {person} termination date {term_date} < hire date {earliest_hire}. Removing all rows.")
                        rows_to_delete.update(idx for idx, _ in rows)
                        break


        if rows_to_delete:
            logger.info(f"[CLEANUP] 🧹 Total Rows Marked for Deletion: {len(rows_to_delete)}")
        else:
            logger.info("[CLEANUP] ✅ No invalid termination records found.")

        deleted_count = 0
        for row_num in sorted(rows_to_delete, reverse=True):
            if row_num > header_row_num:
                ws.delete_rows(row_num)
                deleted_count += 1

        logger.info(f"[DELETED] 🗑️ Total Rows Deleted: {deleted_count}")

        if save_as_new_file:
            out_path = file_path.parent / f"{file_path.stem}_validated.xlsx"
            wb.save(out_path)
            logger.info(f"[DONE] ✅ Saved as {out_path.name}")
            return out_path
        else:
            wb.save(file_path)
            logger.info(f"[DONE] ✅ Overwritten {file_path.name}")
            return file_path

    except Exception as e:
        logger.error(f"[ERROR] ❌ Validation failed: {e}")
        return None

def validate_workrelationship_sheet(file_path: Path, header_row_num: int = 2, save_as_new_file: bool = True, hireActions: List = [], globalTransfers: List = []):
    """
    Filters the Excel file to keep:
    ✅ ONLY HIRE / REHIRE / GLB_TRANSFER (aka GLOBALTRANSFER)
    ✅ The earliest HIRE
    ✅ The latest REHIRE or GT in each cycle
    ❌ All other actions (PROMOTION, ASG_CHANGE, RESIGNATION, etc) are discarded
    """

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[header_row_num]]
        logger.info(f"[INFO] Headers found in row {header_row_num}: {headers}")

        try:
            person_idx = headers.index("PersonNumber")
            action_idx = headers.index("ActionCode")
            date_idx = headers.index("EffectiveStartDate")
        except ValueError:
            logger.info(f"[WARN] Required headers missing in {file_path.name}")
            return

        start_row = header_row_num + 1
        person_rows = {}

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue
            if person not in person_rows:
                person_rows[person] = []
            person_rows[person].append((i, row))

        allowed_starts = hireActions + globalTransfers
        keep_rows = set()

        for person, rows in person_rows.items():
            sorted_rows = sorted(rows, key=lambda r: r[1][date_idx].value)
            current_start = None
            latest_action_row = None

            for idx, row in sorted_rows:
                action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""

                if action not in allowed_starts:
                    continue

                # New cycle detected
                if action in hireActions or globalTransfers:
                    if current_start is not None:
                        keep_rows.add(current_start)
                        if latest_action_row and latest_action_row != current_start:
                            keep_rows.add(latest_action_row)
                    current_start = idx
                    latest_action_row = idx
                else:
                    # Should never hit this due to earlier if guard
                    continue

            # End of loop – commit the final cycle
            if current_start:
                keep_rows.add(current_start)
                if latest_action_row and latest_action_row != current_start:
                    keep_rows.add(latest_action_row)

        logger.info(f"[INFO] Keeping strictly HIRE/REHIRE/GT only rows: {sorted(keep_rows)}")

        # Delete non-matching rows bottom-up for all that person numbers
        for row_num in range(ws.max_row, header_row_num, -1):
            if row_num not in keep_rows:
                ws.delete_rows(row_num)

        # Save
        if save_as_new_file:
            output_path = file_path.parent / f"{file_path.stem}.xlsx"
            wb.save(output_path)
            logger.info(f"[DONE] ✅ Saved as: {output_path.name}")
            return output_path
        else:
            wb.save(file_path)
            logger.info(f"[DONE] ✅ Overwritten: {file_path.name}")
            return file_path

    except Exception as e:
        logger.info(f"[ERROR] ❌ Failed while strictly filtering WorkRelationship: {e}")

def Assignment_type_Code(file_path: Path, assignment_status_rules: list, header_row_num: int = 2, save_as_new_file: bool = True):
    """
    Validates AssignmentStatusTypeCode using regex to extract first word.
    Logs all first words found. Matches based on assignment_status_rules.
    """
    try:
        logger.info(f"[INFO] Processing Assignment_type_Code for file: {file_path.name}")
        logger.info(f"[INFO] Assignment status rules received: {assignment_status_rules}")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[header_row_num]]
        logger.info(f"[INFO] Headers found at row {header_row_num}: {headers}")

        try:
            person_idx = headers.index("PersonNumber")
            action_idx = headers.index("ActionCode")
            assignment_type_idx = headers.index("AssignmentStatusTypeCode")
        except ValueError as ve:
            logger.error(f"[ERROR] Required columns missing in {file_path.name}: {ve}")
            return None

        start_row = header_row_num + 1
        valid_rows = []
        first_words_logged = set()

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            person = str(row[person_idx].value).strip() if row[person_idx].value else None
            if not person:
                continue

            action = str(row[action_idx].value).strip().upper() if row[action_idx].value else ""
            assignment_status_full = str(row[assignment_type_idx].value).strip().upper() if row[assignment_type_idx].value else ""

            # Extract first word using regex (match till space, dash, or non-word char)
            match = re.match(r'^([A-Z0-9]+)', assignment_status_full)
            assignment_status_first_word = match.group(1) if match else ""

            first_words_logged.add(assignment_status_first_word)

            matched = False
            for rule in assignment_status_rules:
                rule_key = rule.get("key", "").strip().lower()
                rule_value = rule.get("value", "").upper()
                rule_result = rule.get("result", "").upper()

                if rule_key == "else":
                    # ELSE rule when nothing matched before
                    expected_status = rule_result
                    if assignment_status_first_word == expected_status:
                        valid_rows.append([cell.value for cell in row])
                    else:
                        logger.info(f"[REMOVED][ELSE] Row {i}: Person={person}, Action={action}, AssignmentType={assignment_status_full} — Expected '{expected_status}' as first word")
                    matched = True
                    break
                else:
                    action_list = [val.strip().upper() for val in rule_value.split(",")]
                    if action in action_list:
                        expected_status = rule_result
                        if assignment_status_first_word == expected_status:
                            valid_rows.append([cell.value for cell in row])
                        else:
                            logger.info(f"[REMOVED] Row {i}: Person={person}, Action={action}, AssignmentType={assignment_status_full} — Expected '{expected_status}' as first word")
                        matched = True
                        break

            if not matched:
                logger.info(f"[REMOVED] Row {i}: Person={person}, Action={action} didn't match any rule and no ELSE provided.")

        logger.info(f"[INFO] Unique first words of AssignmentStatusTypeCode found: {sorted(first_words_logged)}")

        # Clear old rows & save valid ones
        ws.delete_rows(start_row, ws.max_row - header_row_num)

        for idx, row_data in enumerate(valid_rows, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=idx, column=col_idx, value=value)

        out_path = file_path.parent / f"{file_path.stem}.xlsx" if save_as_new_file else file_path
        wb.save(out_path)
        logger.info(f"[DONE] ✅ Assignment Type Code validated and saved as: {out_path.name}")
        return out_path

    except Exception as e:
        logger.error(f"[ERROR] ❌ Exception during Assignment_type_Code: {e}")
        return None

def validate_LegalEmployer_change(
    file_content: BytesIO,
    original_filename: str,
    hire_action_codes: Optional[str] = None,
    termination_action_codes: Optional[str] = None,
    allowed_le_change_action_codes: Optional[str] = None
):
    logger.info(f"Validating file: {original_filename}")
    inconsistent_records = []
    person_numbers_with_errors = set()

    if not original_filename.lower().endswith(('.xls', '.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel (.xls, .xlsx) or CSV files are allowed.")

    try:
        file_content.seek(0)
        if original_filename.lower().endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file_content, header=1)
        else:
            df = pd.read_csv(file_content, header=1)

        df.columns = df.columns.str.strip().str.replace(' ', '').str.upper()
        required_columns = ['PERSONNUMBER', 'ACTIONCODE', 'LEGALEMPLOYERNAME', 'EFFECTIVESTARTDATE']
        for col in required_columns:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"Missing required column: {col}")

        original_df = df.copy()
        df.dropna(subset=required_columns, inplace=True)

        dropped_rows = original_df[~original_df.index.isin(df.index)]
        for _, row in dropped_rows.iterrows():
            pn = str(row.get('PERSONNUMBER', 'N/A')).strip()
            missing_cols = [col for col in required_columns if pd.isna(row.get(col)) or str(row.get(col)).strip() == '']
            inconsistent_records.append({
                'PersonNumber': pn,
                'Scenario': 'Missing Required Data',
                'Status': f"Missing columns: {', '.join(missing_cols)}"
            })
            if pn != 'N/A':
                person_numbers_with_errors.add(pn)

        if df.empty:
            return _make_response(inconsistent_records, person_numbers_with_errors, df)

        df['ACTIONCODE'] = df['ACTIONCODE'].astype(str).str.strip().str.upper()
        df['LEGALEMPLOYERNAME'] = df['LEGALEMPLOYERNAME'].astype(str).str.strip()
        df['PERSONNUMBER'] = df['PERSONNUMBER'].astype(str).str.strip()
        df['EFFECTIVESTARTDATE'] = pd.to_datetime(df['EFFECTIVESTARTDATE'], errors='coerce')

        seen_inconsistencies = set()

        def add_inconsistency(rec):
            rec_key = tuple(sorted(rec.items()))
            if rec_key not in seen_inconsistencies:
                inconsistent_records.append(rec)
                person_numbers_with_errors.add(rec['PersonNumber'])
                seen_inconsistencies.add(rec_key)

        # Invalid dates
        invalid_dates = df[df['EFFECTIVESTARTDATE'].isnull()]
        for _, row in invalid_dates.iterrows():
            add_inconsistency({
                'PersonNumber': row['PERSONNUMBER'],
                'EffectiveStartDate': 'Invalid Date',
                'ActionCode': row['ACTIONCODE'],
                'LegalEmployerName': row['LEGALEMPLOYERNAME'],
                'Scenario': 'Invalid Effective Start Date',
                'Status': 'Date could not be parsed'
            })
        df = df[df['EFFECTIVESTARTDATE'].notnull()]

        # Missing LE
        missing_le = df[df['LEGALEMPLOYERNAME'] == '']
        for _, row in missing_le.iterrows():
            add_inconsistency({
                'PersonNumber': row['PERSONNUMBER'],
                'EffectiveStartDate': row['EFFECTIVESTARTDATE'].strftime('%Y-%m-%d'),
                'ActionCode': row['ACTIONCODE'],
                'LegalEmployerName': '',
                'Scenario': 'Missing Legal Employer Name',
                'Status': 'Legal Employer is empty'
            })
        df = df[df['LEGALEMPLOYERNAME'] != '']

        hire_actions = [code.strip().upper() for code in hire_action_codes.split(',')] if hire_action_codes else []
        term_actions = [code.strip().upper() for code in termination_action_codes.split(',')] if termination_action_codes else []
        allowed_le_change_actions = [code.strip().upper() for code in allowed_le_change_action_codes.split(',')] if allowed_le_change_action_codes else []

        for pn, group in df.groupby('PERSONNUMBER'):
            group = group.sort_values(by='EFFECTIVESTARTDATE').reset_index(drop=True)
            current_le = None
            last_was_termination = False

            for _, row in group.iterrows():
                action = row['ACTIONCODE']
                le = row['LEGALEMPLOYERNAME']
                eff_date = row['EFFECTIVESTARTDATE']

                if action in hire_actions:
                    if current_le and not last_was_termination and current_le != le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'PreviousLegalEmployer': current_le,
                            'Scenario': 'Hire without prior Termination with LE change',
                            'Status': f"LE changed to '{le}' without termination from '{current_le}'"
                        })
                    current_le = le
                    last_was_termination = False

                elif action in term_actions:
                    if not current_le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'Scenario': 'Termination without prior Hire',
                            'Status': 'Termination found without Hire'
                        })
                    if current_le and current_le != le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'PreviousLegalEmployer': current_le,
                            'Scenario': 'Termination with mismatched LE',
                            'Status': f"Termination LE '{le}' does not match '{current_le}'"
                        })
                    current_le = None
                    last_was_termination = True

                else:
                    if not current_le:
                        add_inconsistency({
                            'PersonNumber': pn,
                            'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                            'ActionCode': action,
                            'LegalEmployerName': le,
                            'Scenario': 'Action without Hire',
                            'Status': f"Action '{action}' outside of employment period"
                        })
                    elif current_le != le:
                        if action in allowed_le_change_actions:
                            current_le = le
                        else:
                            add_inconsistency({
                                'PersonNumber': pn,
                                'EffectiveStartDate': eff_date.strftime('%Y-%m-%d'),
                                'ActionCode': action,
                                'LegalEmployerName': le,
                                'PreviousLegalEmployer': current_le,
                                'Scenario': 'LE changed mid-employment without valid action',
                                'Status': f"LE changed to '{le}' with action '{action}'"
                            })
                            current_le = le
                    last_was_termination = False

        if person_numbers_with_errors:
            df = df[~df['PERSONNUMBER'].isin(person_numbers_with_errors)]
            logger.info(f"[INFO] Removed inconsistent PersonNumbers: {person_numbers_with_errors}")

        for pn, group in df.groupby('PERSONNUMBER'):
            hires = group[group['ACTIONCODE'].isin(hire_actions)].sort_values(by='EFFECTIVESTARTDATE')
            terms = group[group['ACTIONCODE'].isin(term_actions)].sort_values(by='EFFECTIVESTARTDATE')

            if not hires.empty and not terms.empty:
                if hires.iloc[0]['LEGALEMPLOYERNAME'] != terms.iloc[0]['LEGALEMPLOYERNAME']:
                    add_inconsistency({
                        'PersonNumber': pn,
                        'Scenario': 'First Hire vs First Termination LE mismatch',
                        'FirstHireLE': hires.iloc[0]['LEGALEMPLOYERNAME'],
                        'FirstTerminationLE': terms.iloc[0]['LEGALEMPLOYERNAME'],
                        'Status': 'Mismatch between hire and termination LE'
                    })
            elif hires.empty and not terms.empty:
                add_inconsistency({
                    'PersonNumber': pn,
                    'Scenario': 'Termination without any Hire',
                    'Status': 'Termination action found but no hire action'
                })

        return _make_response(inconsistent_records, person_numbers_with_errors, df)

    except Exception as e:
        logger.exception(f"Error while validating file {original_filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Internal Error: {str(e)}")

def _make_response(inconsistent_records, person_numbers_with_errors, df):
    def safe_cast_errors(errors: List[dict]):
        for record in errors:
            for key, value in record.items():
                if isinstance(value, (np.integer, np.int64)):
                    record[key] = int(value)
                elif isinstance(value, (np.floating, np.float64)):
                    record[key] = float(value)
                elif isinstance(value, pd.Timestamp):
                    record[key] = value.strftime('%Y-%m-%d')
        return errors

    def df_to_dict(df: pd.DataFrame):
        df_copy = df.copy()
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d')
            elif pd.api.types.is_numeric_dtype(df_copy[col]):
                df_copy[col] = df_copy[col].astype(float)
        return df_copy.to_dict(orient='records')

    return {
        "inconsistencies": safe_cast_errors(inconsistent_records),
        "person_numbers_with_errors": list(person_numbers_with_errors),
        "cleaned_data": df_to_dict(df.reset_index(drop=True))
    }




@app.post("/api/hdl/bulk-excel-upload")
async def bulk_excel_upload(
    parent_name: str = Form(...),
    excelFile: UploadFile = File(...),
    Mandatory_Objects: str = Form(...),
    assignment_status_rules: str = Form(...),
    TermActions: str = Form(...),
    HireActions: str = Form(...),
    glbTransfers: str = Form(...),
    all_mandatory_objects: str = Form(...),
    all_non_mandatory_objects: str = Form(...),
    # Add customerName and InstanceName as form parameters
    customerName: str = Form(...),
    InstanceName: str = Form(...)
):
    if not excelFile.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an .xlsx file.")

    # Construct the new parent_folder path: uploads/customerName/InstanceName/parent_name
    customer_folder = DIR / customerName
    instance_folder = customer_folder / InstanceName
    parent_folder = instance_folder / parent_name

    # Create the nested directories if they don't exist
    parent_folder.mkdir(parents=True, exist_ok=True)
    logger.info(f"Created directory structure: {parent_folder}")

    try:
        # ✅ Parse JSON strings into lists
        term_actions_list = json.loads(TermActions)
        hire_actions_list = json.loads(HireActions)
        glb_transfer_list = json.loads(glbTransfers)
        assignment_status_rules_list = json.loads(assignment_status_rules)
        mandatory_objects_list = json.loads(all_mandatory_objects)
        all_non_mandatory_objects_list = json.loads(all_non_mandatory_objects) # Added this as it's passed from frontend

        logger.info(f"✅ Parsed Actions: Term={term_actions_list}, Hire={hire_actions_list}, GT={glb_transfer_list}")

        contents = await excelFile.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        saved_files = []
        mandatory_person_numbers = set()

        # === 1️⃣ First Pass — Collect PersonNumbers from Mandatory Sheets ===
        for sheet_name in wb.sheetnames:
            if sheet_name.strip() in mandatory_objects_list:
                ws = wb[sheet_name]
                # Assuming headers are in the 3rd row (index 2 for 0-indexed list)
                headers = [cell.value for cell in ws[3]]
                if not headers:
                    continue
                try:
                    person_idx = headers.index("PersonNumber")
                except ValueError:
                    continue

                for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                    if person_idx is not None and person_idx < len(row) and row[person_idx]:
                        mandatory_person_numbers.add(str(row[person_idx]).strip())

        # === 2️⃣ Second Pass — Filter Rows and Save Each Sheet ===
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Assuming headers are in the 3rd row (index 2 for 0-indexed list)
            headers = [cell.value for cell in ws[3]]
            if not headers:
                continue

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = sheet_name
            # Append the original headers
            new_ws.append(headers)

            try:
                person_idx = headers.index("PersonNumber")
            except ValueError:
                person_idx = None

            is_mandatory = sheet_name.strip() in mandatory_objects_list
            rows_added = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
                if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
                    continue # Skip entirely empty rows

                if person_idx is not None:
                    if person_idx < len(row) and row[person_idx]:
                        person_number = str(row[person_idx]).strip()
                        # If it's a non-mandatory sheet, and the person number is NOT in the mandatory set, skip this row
                        if not is_mandatory and person_number not in mandatory_person_numbers:
                            continue
                    elif is_mandatory:
                        # If it's a mandatory sheet and PersonNumber is missing for a row, log a warning and skip
                        logger.warning(f"⚠️ Mandatory sheet '{sheet_name}' row {row_idx} missing PersonNumber. Skipping row.")
                        continue

                new_ws.append(row)
                rows_added += 1

            if rows_added > 0:
                safe_sheet_name = "_".join(sheet_name.strip().split())
                out_path = parent_folder / f"{safe_sheet_name}.xlsx"
                new_wb.save(out_path)
                saved_files.append({
                    "sheet": sheet_name,
                    # Ensure the relative path is correct from the base upload directory
                    "file": str(out_path.relative_to(DIR)),
                    "parent": parent_name,
                    "child": sheet_name
                })
                logger.info(f"[✅] Saved sheet '{sheet_name}' to '{out_path.name}'")

        # === 3️⃣ Post-Save Validations (assuming these functions are defined elsewhere) ===
        errors = []

        # Placeholder for validation functions (you need to define these or import them)
        def populate_actual_termination_date_from_resignation(file_path, hireActions, globalTransfers, termAction):
            logger.info(f"Running populate_actual_termination_date_from_resignation for {file_path}")
            # Implement your logic here
            pass

        def validate_workrelationship_sheet(file_path, hireActions, globalTransfers):
            logger.info(f"Running validate_workrelationship_sheet for {file_path}")
            # Implement your logic here
            pass

        def Assignment_type_Code(file_path, assignment_status_rules, save_as_new_file):
            logger.info(f"Running Assignment_type_Code for {file_path}")
            # Implement your logic here
            pass

        def validate_termination_date(file_path, save_as_new_file, TermActions, HireActions):
            logger.info(f"Running validate_termination_date for {file_path}")
            # Implement your logic here
            pass

        def validate_LegalEmployer_change(file_content, original_filename, hire_action_codes, termination_action_codes, allowed_le_change_action_codes):
            logger.info(f"Running validate_LegalEmployer_change for {original_filename}")
            # Implement your logic here, return a list of errors if any
            return []


        for file_meta in saved_files:
            sheet_name_lower = file_meta["sheet"].strip().lower()
            # Ensure the path is absolute for internal operations
            current_file_path = DIR / file_meta["file"]

            if sheet_name_lower == "workrelationship":
                populate_actual_termination_date_from_resignation(
                    current_file_path, hireActions=hire_actions_list,
                    globalTransfers=glb_transfer_list, termAction=term_actions_list
                )

                validate_workrelationship_sheet(
                    current_file_path, hireActions=hire_actions_list,
                    globalTransfers=glb_transfer_list
                )

                assignment_file = parent_folder / "Assignment.xlsx"
                if assignment_file.exists():
                    Assignment_type_Code(
                        assignment_file, assignment_status_rules=assignment_status_rules_list, save_as_new_file=False
                    )
                    validate_termination_date(
                        assignment_file, save_as_new_file=False, TermActions=term_actions_list, HireActions=hire_actions_list
                    )
                else:
                    logger.warning("⚠️ Assignment.xlsx not found — skipping validations.")

                workterms_file = parent_folder / "WorkTerms.xlsx"
                if workterms_file.exists():
                    with open(workterms_file, "rb") as f:
                        workterms_content = BytesIO(f.read())

                    le_validation_results = validate_LegalEmployer_change(
                        file_content=workterms_content,
                        original_filename="WorkTerms.xlsx",
                        hire_action_codes=",".join(hire_actions_list),
                        termination_action_codes=",".join(term_actions_list),
                        allowed_le_change_action_codes=",".join(glb_transfer_list)
                    )

                    if le_validation_results:
                        errors.extend(le_validation_results)
                        logger.warning(f"⚠️ Legal Employer validation found inconsistencies: {le_validation_results}")
                    else:
                        logger.info("✅ Legal Employer validation passed.")
                else:
                    logger.warning("⚠️ WorkTerms.xlsx not found — skipping LE validation.")

                break  # Only validating WorkRelationship cycle

        return {"parent": parent_name, "files": saved_files, "errors": errors}

    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.critical(f"❌ Unhandled error during bulk upload for {excelFile.filename}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error during file processing: {str(e)}")




# Mount the static directory for Dat_Files. Ensure the directory exists.
DAT_FILES_DIR = Path("Required_files/Dat_Files")
DAT_FILES_DIR.mkdir(parents=True, exist_ok=True) # Ensure this directory exists at startup
app.mount("/static", StaticFiles(directory=DAT_FILES_DIR), name="static_dat_files")

@app.post("/api/hdl/data-transformation")
async def get_data_transformation_mapping(Attributes: List[str] = Body(..., embed=True)):
    """
    Retrieves existing data transformation mappings for a list of attributes from the Excel file.
    """
    try:
        # Define the path to the mapping file (using the Excel file name)
        mapping_file_path = TRANSFORMATION_ATTRIBUTES_FILE_PATH
        
        # Load the Excel file into a DataFrame
        df = pd.read_excel(mapping_file_path, engine="openpyxl")

        # Clean column names by stripping whitespace
        df.columns = [col.strip() for col in df.columns]

        # Filter the DataFrame to include only the correct column names
        mapping_df = df[['Attributes for Transformation', 'Transformation Map Name']].copy()

        # Create a dictionary from the filtered DataFrame
        existing_mappings = mapping_df.set_index('Attributes for Transformation')['Transformation Map Name'].fillna('').to_dict()
        existing_mappings = {k: v.strip() for k, v in existing_mappings.items()}
        filtered_mappings = {attr: existing_mappings.get(attr, "") for attr in Attributes}        
        return {"mapping": filtered_mappings}

    except FileNotFoundError:
        logging.error(f"Transformation file not found at {mapping_file_path}")
        raise HTTPException(status_code=500, detail=f"Transformation file not found at {mapping_file_path}")
    except Exception as e:
        logging.error(f"Error retrieving data transformation mapping: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving data transformation mapping: {str(e)}")

class TransformationPayload(BaseModel):
    attribute: str
    newMapping: str # Changed from Json[dict] to str for simple string mapping
    componentName: str # Added componentName

class ApplyTransformationPayload(BaseModel):
    mapping_json_file: str
    raw_excel_file: str
    attribute_column: str
    
logger = logging.getLogger(__name__)

# Helper function to apply transformations to a DataFrame
def apply_data_transformation(df: pd.DataFrame, transformations: Dict[str, str]) -> pd.DataFrame:
    """
    Applies a set of transformation rules to the DataFrame.
    `transformations` is a dict where key is attribute name and value is the transformation string.
    """
    df_transformed = df.copy()
    for col in df_transformed.columns:
        if df_transformed[col].dtype == 'object':
            if pd.isna(df[col]).any(): # Check original df for NaN
                df_transformed[col] = df_transformed[col].fillna('')
                logger.debug(f"Replaced NaN with empty string in column: {col}")
            elif pd.api.types.is_numeric_dtype(df_transformed[col]):
                if pd.isna(df[col]).any() or np.isinf(df[col]).any(): # Check original df for NaN/Inf
                    df_transformed[col] = df_transformed[col].replace({np.nan: 0, np.inf: 1e308, -np.inf: -1e308})
                    logger.debug(f"Handled NaN/Inf in numeric column: {col}")

    for attr, transformation_str in transformations.items():
        if attr in df_transformed.columns:
            if attr.lower() == 'suffix':
                def transform_value(value):
                    clean_value = str(value).strip().lower()
                    if clean_value == 'senior':
                        return 'Sr'
                    elif clean_value == 'junior':
                        return 'Jr'
                    elif clean_value == '':
                        return ''
                    else:
                        return value 
                df_transformed[attr] = df_transformed[attr].apply(transform_value)
                logger.info(f"Applied specific 'Suffix' transformation to column: '{attr}'.")
            else:
                logger.info(f"Transformation rule for '{attr}' is '{transformation_str}'. Generic application not implemented for this type.")


    return df_transformed



class ExcelRequest(BaseModel):
    customerName: str
    InstanceName: str
    parent: str  # Global BO
    filename: str  # Component Name (Excel file name)

@app.post("/excel")
async def get_excel_file_post(request_data: ExcelRequest):
    """
    Returns the requested Excel file from the uploads/Excel_Files directory using POST request with payload.
    """
    logger.info(f"Received Excel POST request with {request_data.customerName}/{request_data.InstanceName}/{request_data.parent}/{request_data.filename}")
    
    file_path = DIR / request_data.customerName / request_data.InstanceName / request_data.parent / request_data.filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")
    
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=request_data.filename
    )

# --- NEW VALIDATION ENDPOINT AND HELPER FUNCTIONS ---

class AttributeValidationData(BaseModel):
    Attributes: str
    required: bool
    LookUp_data: Optional[str] = None
    CodeName: Optional[str] = None
    Data_Transformation: Optional[str] = None
    # --- NEW: Add the includeInDatFileGeneration field to the Pydantic model ---
    includeInDatFileGeneration: bool = True # Default to True as per frontend logic

def required_field_validations(
    df: pd.DataFrame, required_columns: List[str]
) -> pd.DataFrame:
    """
    Performs validation to ensure required fields are present and not empty in the DataFrame.
    Adds a 'Reason for Failed' column to the DataFrame, detailing validation failures.

    This function operates on a copy of the DataFrame to prevent modifying the original.
    It identifies two types of failures for 'required' fields:
    1.  Missing required columns in the DataFrame itself.
    2.  Empty (or whitespace-only) values within existing required columns.

    Args:
        df (pd.DataFrame): The input DataFrame to validate.
        required_columns (List[str]): A list of column names that are mandatory.

    Returns:
        pd.DataFrame: A new DataFrame with an added 'Reason for Failed' column,
                      detailing validation issues for each row. Rows without issues
                      will have an empty string in this column.
    """
    if not isinstance(df, pd.DataFrame):
        logger.error("Input 'df' must be a pandas DataFrame.")
        raise TypeError("Input 'df' must be a pandas DataFrame.")
    if not isinstance(required_columns, list) or not all(isinstance(col, str) for col in required_columns):
        logger.error("Input 'required_columns' must be a list of strings.")
        raise TypeError("Input 'required_columns' must be a list of strings.")

    if df.empty and required_columns:
        logger.warning("Input DataFrame is empty, but required columns were specified. No validation performed on rows.")
        # Create a 'Reason for Failed' column if it doesn't exist, as per function contract
        if "Reason for Failed" not in df.columns:
            df["Reason for Failed"] = ""
        return df # Return empty DataFrame, as there are no rows to validate

    df_copy = df.copy()

    # Initialize a list of lists to store reasons for each row
    # This is more robust than string concatenation inside a loop
    all_row_reasons = [[] for _ in range(len(df_copy))]

    # Check for missing required columns in the DataFrame schema
    actual_columns = set(df_copy.columns)
    missing_cols_in_excel = [col for col in required_columns if col not in actual_columns]

    if missing_cols_in_excel:
        missing_cols_str = ', '.join(missing_cols_in_excel)
        logger.warning(f"Missing required columns in Excel file: {missing_cols_str}. Marking all rows as failed for these columns.")
        # Mark all rows as failed for missing columns
        for i in range(len(df_copy)):
            all_row_reasons[i].append(f"Missing required column(s): {missing_cols_str}")
        # Add the missing columns to the DataFrame so subsequent steps don't fail,
        # but they will contain NaN or empty strings, correctly failing validation.
        for col in missing_cols_in_excel:
            df_copy[col] = '' # Or pd.NA, depends on desired behavior for subsequent steps

    # Validate required values (non-empty) for existing required columns
    for col in required_columns:
        if col in df_copy.columns: # Only check if the column actually exists in the DataFrame
            # Identify rows where the required column is empty or contains only whitespace
            # Convert to string to handle mixed types and NaN safely
            empty_mask = df_copy[col].astype(str).str.strip() == ""
            
            # Append reason for each affected row
            for i in df_copy.index[empty_mask]:
                all_row_reasons[i].append(f"'{col}' is required and cannot be empty")
            logger.info(f"Validated required field '{col}'. Found {empty_mask.sum()} empty values.")
        # else: already logged in missing_cols_in_excel block or handled above.

    # Consolidate all reasons for each row into the 'Reason for Failed' column
    df_copy["Reason for Failed"] = ["; ".join(filter(None, reasons)).strip(" ;") for reasons in all_row_reasons]
    
    # Replace empty strings (from rows with no failures) with a proper empty string, if needed.
    # The filter(None, ...) and strip(";") already handles this for the most part.
    df_copy["Reason for Failed"] = df_copy["Reason for Failed"].replace("^$", "", regex=True)

    return df_copy

def lookup_validations(df: pd.DataFrame, all_lookups: Dict[str, List[Dict[str, str]]]) -> pd.DataFrame:
    """
    Performs lookup validations on the DataFrame.
    If a column has lookup values defined but a row's data for that column is empty,
    it will be ignored and not added to the failed list for lookup validation.
    """
    df_copy = df.copy()
    if "Reason for Failed" not in df_copy.columns:
        df_copy["Reason for Failed"] = ""
    else:
        df_copy["Reason for Failed"] = df_copy["Reason for Failed"].astype(str)

    for attribute, lookup_list in all_lookups.items():
        if attribute in df_copy.columns:
            # FIX: Use getattr to safely access 'Value' from Pydantic LookupItem objects
            valid_values = set(
                str(getattr(item, 'Value', '')).strip().lower()
                for item in lookup_list
                if getattr(item, 'Value', '') is not None and str(getattr(item, 'Value', '')).strip() != ''
            )
            
            # If no valid values are defined for a lookup, skip validation for this attribute entirely
            if not valid_values:
                logger.warning(f"No valid lookup values found for attribute '{attribute}'. Skipping lookup validation for this column.")
                continue

            is_empty_in_row_data = df_copy[attribute].astype(str).str.strip() == ""
            is_valid_lookup = df_copy[attribute].astype(str).str.strip().str.lower().isin(valid_values)
            
            # An invalid lookup occurs if the value is NOT empty AND is NOT a valid lookup value
            invalid_mask = (~is_empty_in_row_data) & (~is_valid_lookup)
            df_copy.loc[invalid_mask, "Reason for Failed"] += f"; Invalid lookup value for {attribute}"
            logger.info(f"Validated lookup for '{attribute}'. Found {invalid_mask.sum()} invalid values (excluding empty row data).")
        else:
            logger.warning(f"Lookup attribute '{attribute}' not found in DataFrame columns. Skipping lookup validation for this column.")
    
    # Clean up the "Reason for Failed" column
    df_copy["Reason for Failed"] = df_copy["Reason for Failed"].str.strip(" ;").replace("^$", "", regex=True)
    return df_copy


def transformation_for_validation(excel_file_io: BytesIO):
    """
    Applies data transformations based on the transformation file (Excel).
    Reads the Excel file from the BytesIO stream and applies transformations
    according to the mapping in TRANSFORMATION_ATTRIBUTES_FILE_PATH for all attributes present in the file.
    """
    # Load the mapping from the Excel file (same as apply_transformation_and_download)
    mapping_file_path = TRANSFORMATION_ATTRIBUTES_FILE_PATH
    mapping_df = pd.read_excel(mapping_file_path, engine="openpyxl")
    mapping_df.columns = [col.strip() for col in mapping_df.columns]
    mapping_dict = mapping_df[['Attributes for Transformation', 'Transformation Map Name']].set_index('Attributes for Transformation')['Transformation Map Name'].fillna('').to_dict()
    mapping_dict = {k: v.strip() for k, v in mapping_dict.items()}

    df = pd.read_excel(excel_file_io, engine='openpyxl')
    df.columns = [col.strip() for col in df.columns]
    df = df.fillna("")
    for col in df.columns:
        if df[col].dtype == 'object':
            if pd.isna(df[col]).any():
                df[col] = df[col].fillna('')
        elif pd.api.types.is_numeric_dtype(df[col]):
            if pd.isna(df[col]).any() or pd.isinf(df[col]).any():
                df[col] = df[col].replace({np.nan: 0, np.inf: 1e308, -np.inf: -1e308})
    # Apply transformations based on the mapping for each attribute in the DataFrame
    for attr in df.columns:
        transformation_str = mapping_dict.get(attr, "")
        if transformation_str:
            def transform_value(value):
                clean_value = str(value).strip().lower()
                # Example: handle Suffix transformation, can be extended for other rules
                if transformation_str.lower().startswith('suffix'):
                    if clean_value == 'senior':
                        return 'Sr'
                    elif clean_value == 'junior':
                        return 'Jr'
                    elif clean_value == '':
                        return ''
                return value
            df[attr] = df[attr].apply(transform_value)
            logger.info(f"Applied transformation '{transformation_str}' to column: '{attr}'.")
    if "Reason for Failed" not in df.columns:
        df["Reason for Failed"] = ""
    else:
        df["Reason for Failed"] = df["Reason for Failed"].astype(str)
    return df


def get_generic_filename(prefix: str, identifier: str, extension: str) -> str:
    """Generates a unique filename."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{identifier}_{timestamp}.{extension}"

# Helper function to convert base64 to DataFrame
def base64_to_dataframe(base64_string: str) -> pd.DataFrame:
    """Decodes a base64 string to a pandas DataFrame."""
    try:
        decoded_bytes = base64.b64decode(base64_string)
        # Try reading as CSV first, then Excel
        try:
            df = pd.read_csv(io.StringIO(decoded_bytes.decode('utf-8')))
        except Exception:
            df = pd.read_excel(io.BytesIO(decoded_bytes))
        return df
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error decoding base64 or reading file: {e}")




# Helper function to convert base64 to DataFrame
def base64_to_dataframe(base64_string: str) -> pd.DataFrame:
    """Decodes a base64 string to a pandas DataFrame."""
    try:
        decoded_bytes = base64.b64decode(base64_string)
        # Try reading as CSV first, then Excel
        try:
            df = pd.read_csv(io.StringIO(decoded_bytes.decode('utf-8')))
        except Exception:
            df = pd.read_excel(io.BytesIO(decoded_bytes))
        return df
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error decoding base64 or reading file: {e}")


def get_hdl_setup_validate_fetch(customer_name: str, instance_name: str) -> dict:
    try:
        setup_dir = Path("User/setup_files")

        filename = f"{customer_name.replace(' ', '_')}_{instance_name.replace(' ', '_')}_setup.json"
        filepath = setup_dir / filename
        logger.info(f"Looking for setup file at: {filepath}")
        if not filepath.exists():
            logger.info(f"Looking for setup file at: {filepath}")
            raise HTTPException(status_code=404, detail=f"Setup file not found. Searched at: {filepath}")

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        return data  # <-- just return dict, not JSONResponse

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching setup: {e}")


import pandas as pd
from typing import List

import pandas as pd
from typing import List

DATA_TYPE_MAPPING = {
    "VARCHAR": "string",
    "CHAR": "string",
    "TEXT": "string",
    "STRING": "string",
    "NUMBER": "float",
    "INTEGER": "integer",
    "INT": "integer",
    "FLOAT": "float",
    "DECIMAL": "float",
    "DATE": "date",
    "DATETIME": "date",
    "BOOLEAN": "boolean",
    "BOOL": "boolean",
    "TIMESTAMP": "timestamp",
}

def is_oracle_safe_date(value: str) -> bool:
    """
    Oracle allows years from 4712 BC to 9999 AD.
    Accepts both YYYY/MM/DD and MM/DD/YYYY formats.
    """
    try:
        # Try to extract year intelligently
        parts = re.split(r"[-/]", str(value))
        parts = [p for p in parts if p.strip()]

        if not parts:
            return False

        # Detect if first part is year or last part
        if len(parts[0]) == 4:  # YYYY/MM/DD
            year = int(parts[0])
        else:  # MM/DD/YYYY or DD/MM/YYYY
            year = int(parts[-1])

        return 4712 <= year <= 9999
    except Exception:
        return False


def validate_data_types(df: pd.DataFrame, attributes: List) -> pd.DataFrame:
    """
    Validate each column's values against its expected data_type.
    Updates 'Reason for Failed' column with errors.
    Keeps Oracle-safe dates (4712–9999) as strings to avoid NaT conversion.
    """
    for attr in attributes:
        col = attr.Attributes
        dtype_raw = attr.data_type.upper()
        expected_type = DATA_TYPE_MAPPING.get(dtype_raw)

        if not expected_type:
            for idx in df.index:
                existing_reason = df.at[idx, "Reason for Failed"]
                df.at[idx, "Reason for Failed"] = (
                    f"{existing_reason}; Unknown data_type '{attr.data_type}' specified for column."
                    if existing_reason else
                    f"Unknown data_type '{attr.data_type}' specified for column."
                )
            continue

        if col not in df.columns:
            continue

        for idx, val in df[col].items():
            if val == "" or pd.isna(val):
                continue
            reason = ""
            try:
                if expected_type == "string":
                    str(val)
                elif expected_type == "integer":
                    if not str(val).isdigit():
                        reason = f"Expected integer but got '{val}'"
                elif expected_type == "float":
                    float(val)
                elif expected_type == "boolean":
                    if str(val).strip().upper() not in ["TRUE", "FALSE", "1", "0"]:
                        reason = f"Expected boolean (TRUE/FALSE/1/0) but got '{val}'"
                elif expected_type in ["date", "timestamp"]:
                    # Oracle-safe date? keep it as raw string
                    if is_oracle_safe_date(str(val)):
                        df.at[idx, col] = str(val)
                    else:
                        # Only parse valid modern dates
                        parsed = pd.to_datetime(val, errors='raise')
                        df.at[idx, col] = parsed.strftime("%Y/%m/%d")
            except Exception:
                reason = f"Value '{val}' does not match expected data_type '{expected_type}'"

            if reason:
                existing_reason = df.at[idx, "Reason for Failed"]
                df.at[idx, "Reason for Failed"] = (
                    f"{existing_reason}; {reason}" if existing_reason else reason
                )

        # Ensure column stays as string to preserve Oracle special dates
        if expected_type in ["date", "timestamp"]:
            df[col] = df[col].astype(str)

    return df


def apply_workrelationship_rules(df: pd.DataFrame,
                                hire_actions: list,
                                rehire_actions: list,
                                gt_actions: list,
                                term_actions: list) -> pd.DataFrame:
    """
    HCM-style WorkRelationship rules:
    - Tracks multiple hires/rehire sequences after terminations.
    - ActualTerminationDate = next termination - 1 for the current hire/rehire/GT sequence.
    - Returns only hire/rehire/GT rows.
    """
    if df.empty:
        logger.info("Input DataFrame is empty. Returning as-is.")
        return df

    # Ensure necessary columns exist
    for col in ["PersonNumber", "ActionCode", "EffectiveStartDate"]:
        if col not in df.columns:
            logger.warning(f"Column '{col}' missing. Skipping WR rules.")
            return df

    # Normalize columns
    df["ActionCode"] = df["ActionCode"].astype(str).str.strip().str.upper()
    df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors="coerce")

    # Ensure ActualTerminationDate exists
    if "ActualTerminationDate" not in df.columns:
        df["ActualTerminationDate"] = None

    final_rows = []

    for person_number, group in df.groupby("PersonNumber"):
        # Sort chronologically
        group = group.sort_values("EffectiveStartDate").copy()
        last_term_date = None
        next_term_dates = group[group["ActionCode"].isin(term_actions)]["EffectiveStartDate"].tolist()
        term_idx = 0  # pointer for next termination

        for idx, row in group.iterrows():
            action = row["ActionCode"]

            # If termination, skip it (we’ll assign its date to prior hires)
            if action in term_actions:
                continue

            # Determine the next termination date for this hire/rehire/GT
            next_term_date = None
            while term_idx < len(next_term_dates) and next_term_dates[term_idx] <= row["EffectiveStartDate"]:
                term_idx += 1  # skip past terminations before or on this row

            if term_idx < len(next_term_dates):
                next_term_date = next_term_dates[term_idx] - pd.Timedelta(days=1)
                logger.info(f"For PersonNumber {person_number}, Action {action} on {row['EffectiveStartDate'].date()}, next termination is on {next_term_dates[term_idx].date()} -> setting ActualTerminationDate to {next_term_date.date()}")

            # Assign ActualTerminationDate
            if next_term_date:
                group.at[idx, "ActualTerminationDate"] = next_term_date.strftime("%Y/%m/%d")

        # Append only hire/rehire/GT rows
        final_rows.append(group[group["ActionCode"].isin(hire_actions + rehire_actions + gt_actions)])

    df_final = pd.concat(final_rows, ignore_index=True)

    # Ensure proper string formatting for ActualTerminationDate
    df_final["ActualTerminationDate"] = df_final["ActualTerminationDate"].apply(safe_format_date)


    return df_final

def safe_format_date(val):
    if pd.isna(val) or val is None:
        return None
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.strftime("%Y/%m/%d")
    except Exception:
        return None


def load_validation_module(file_path: Path):
    spec = importlib.util.spec_from_file_location(file_path.stem, file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def enforce_dtypes(df_to_cast, reference_dtypes):
    for col, dtype in reference_dtypes.items():
        if col not in df_to_cast.columns:
            continue
        try:
            if "datetime" in str(dtype):
                df_to_cast[col] = pd.to_datetime(df_to_cast[col], errors="coerce")
            elif "int" in str(dtype):
                df_to_cast[col] = pd.to_numeric(df_to_cast[col], errors="coerce").fillna(0).astype(int)
            elif "float" in str(dtype):
                df_to_cast[col] = pd.to_numeric(df_to_cast[col], errors="coerce")
            else:
                df_to_cast[col] = df_to_cast[col].astype(str)
        except Exception as e:
            logger.warning(f"Failed to enforce dtype for column {col}: {e}")
    return df_to_cast

def validate_delta_logic(df, delta_df, hire_actions, term_actions, rehire_actions, gt_actions):
    """
    Dynamic Validation for Assignment Component:
    - Checks if PersonNumber exists in Cloud (delta_df).
    - IF EXISTS (Delta Load = True): Performs Section B (Date continuity, Sequence checks).
    - IF NOT EXISTS (Delta Load = False): Performs Section A (Hire checks).
    - Handles 'WorkSequence' (EffectiveSequence) increment logic for same-day actions.
    """
    logger.info("Starting Dynamic Delta/New Hire Validation...")

    # --- 0. COLUMN NORMALIZATION HELPER ---
    canonical_cols = {
        "personnumber": "PersonNumber",
        "effectivestartdate": "EffectiveStartDate",
        "effectiveenddate": "EffectiveEndDate",
        "actioncode": "ActionCode",
        "effectivesequence": "EffectiveSequence",
        "worksequence": "EffectiveSequence", # Map worksequence to EffectiveSequence
        "assignmentsequence": "AssignmentSequence",
        "reason for failed": "Reason for Failed",
        "legalemployername": "LegalEmployerName",
        "assignmentstatustypecode": "AssignmentStatusTypeCode",
        "businessunitshortcode": "BusinessUnitShortCode"
    }

    def normalize_columns(dataframe, mapping):
        new_columns = {}
        for col in dataframe.columns:
            col_str = str(col).strip()
            # Handle BOM if present in the string itself
            col_str = col_str.replace('\ufeff', '')
            col_lower = col_str.lower()
            if col_lower in mapping:
                new_columns[col] = mapping[col_lower]
            else:
                new_columns[col] = col_str
        return dataframe.rename(columns=new_columns)

    # --- 1. Prepare Cloud Data ---
    delta_df = normalize_columns(delta_df, canonical_cols)
    
    if "PersonNumber" not in delta_df.columns:
        logger.error("Delta/Cloud file is missing 'PersonNumber' after normalization. Skipping Delta logic.")
        return df

    delta_df["EffectiveStartDate"] = pd.to_datetime(delta_df["EffectiveStartDate"], errors='coerce')
    delta_df["PersonNumber"] = delta_df["PersonNumber"].astype(str).str.strip()
    
    # numeric sequences
    for col in ["EffectiveSequence", "AssignmentSequence"]:
        if col in delta_df.columns:
            delta_df[col] = pd.to_numeric(delta_df[col], errors='coerce').fillna(1)

    # Sort to get LATEST row
    sort_cols = ["PersonNumber", "EffectiveStartDate"]
    if "EffectiveSequence" in delta_df.columns:
        sort_cols.append("EffectiveSequence")
    
    delta_df_sorted = delta_df.sort_values(by=sort_cols, ascending=[True, False, False])
    latest_cloud_records = delta_df_sorted.drop_duplicates(subset=["PersonNumber"], keep="first").set_index("PersonNumber")

    # --- 2. Prepare Input Data ---
    df = normalize_columns(df, canonical_cols)
    if "PersonNumber" not in df.columns:
        return df

    df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors='coerce')
    df["EffectiveEndDate"] = pd.to_datetime(df["EffectiveEndDate"], errors='coerce')
    df["PersonNumber"] = df["PersonNumber"].astype(str).str.strip()
    
    # Ensure EffectiveSequence column exists
    if "EffectiveSequence" not in df.columns:
        df["EffectiveSequence"] = 1

    if "Reason for Failed" not in df.columns:
        df["Reason for Failed"] = ""
    else:
        df["Reason for Failed"] = df["Reason for Failed"].astype(str)

    # Helper for Section A Validation (Reusable)
    def perform_section_a_checks(person_group, person_num):
        # Sort by Date then Sequence
        # Note: We rely on the caller's sort order, but verifying here helps
        # person_group = person_group.sort_values(by=["EffectiveStartDate", "EffectiveSequence"])
        
        # 1. Check First Row is HIRE (Only if it's the very first action in history, checked by logic below)
        # In this helper, we assume we are checking the whole group for a new hire
        if len(person_group) > 0:
            first_idx = person_group.index[0]
            first_action = str(person_group.loc[first_idx, "ActionCode"]).strip().upper()
            if first_action not in hire_actions:
                msg = f"Section A Error: First row must be 'HIRE', found '{first_action}'."
                df.at[first_idx, "Reason for Failed"] = f"{df.at[first_idx, 'Reason for Failed']}; {msg}"

        prev_row = None
        current_legal_emp = None
        
        for i in range(len(person_group)):
            idx = person_group.index[i]
            row = person_group.iloc[i]
            action = str(row.get("ActionCode", "")).strip().upper()
            eff_start = row["EffectiveStartDate"]
            eff_end = row["EffectiveEndDate"]
            legal_emp = str(row.get("LegalEmployerName", "")).strip()
            bu_code = str(row.get("BusinessUnitShortCode", "")).strip()
            status_code = str(row.get("AssignmentStatusTypeCode", "")).strip().upper()

            # 2) Legal Employer Consistency
            if action in hire_actions or action in rehire_actions or action in gt_actions:
                current_legal_emp = legal_emp # Reset/Set expected
            elif current_legal_emp and legal_emp != current_legal_emp:
                 df.at[idx, "Reason for Failed"] += f"; Legal Employer mismatch. Expected '{current_legal_emp}'"

            # 3) Global Transfer Logic
            if action in gt_actions and prev_row is not None:
                prev_legal = str(prev_row.get("LegalEmployerName", "")).strip()
                prev_bu = str(prev_row.get("BusinessUnitShortCode", "")).strip()
                if legal_emp == prev_legal and bu_code == prev_bu:
                     df.at[idx, "Reason for Failed"] += "; GT requires change in Legal Employer or BU"
                # Gap check
                if prev_row["EffectiveEndDate"] + timedelta(days=1) != eff_start:
                     df.at[idx, "Reason for Failed"] += "; Gap detected in GT"

            # 5) Effective End Date Logic
            if i < len(person_group) - 1:
                next_row = person_group.iloc[i+1]
                next_start = next_row["EffectiveStartDate"]
                if next_start > eff_start:
                    expected_end = next_start - timedelta(days=1)
                    if pd.notna(eff_end) and eff_end.date() != expected_end.date():
                         df.at[idx, "Reason for Failed"] += f"; Incorrect EndDate. Expected {expected_end.date()}"

            # 6) AssignmentStatusTypeCode
            if action in term_actions:
                if "INACTIVE" not in status_code:
                    df.at[idx, "Reason for Failed"] += "; Status must be INACTIVE for Termination"
            elif "SUSPEND" in action:
                 if "SUSPENDED" not in status_code:
                    df.at[idx, "Reason for Failed"] += "; Status must be SUSPENDED"
            else:
                if "ACTIVE" not in status_code and status_code != "":
                     df.at[idx, "Reason for Failed"] += "; Status must be ACTIVE"

            prev_row = row

    # --- 3. Iterate through Input Groups ---
    # Sort input by Person, Date, and potentially Sequence if provided, or index to maintain Excel order
    grouped = df.sort_values(by=["PersonNumber", "EffectiveStartDate", "EffectiveSequence"]).groupby("PersonNumber")

    for person_number, group in grouped:
        is_in_cloud = person_number in latest_cloud_records.index
        
        # --- State Initialization for Sequence Logic ---
        if is_in_cloud:
            latest_row = latest_cloud_records.loc[person_number]
            curr_date = latest_row["EffectiveStartDate"]
            curr_seq = int(latest_row.get("EffectiveSequence", 1))
            curr_action = str(latest_row.get("ActionCode", "")).strip().upper()
            curr_asg_seq = int(latest_row.get("AssignmentSequence", 1))
            curr_status = str(latest_row.get("STATUS", "")).strip().upper()
        else:
            # New Hire / Not in Cloud
            curr_date = None
            curr_seq = 0
            curr_action = ""
            curr_asg_seq = 1
            curr_status = ""

        # --- Process Rows for Sequence & Logic ---
        for idx, row in group.iterrows():
            inc_date = row["EffectiveStartDate"]
            inc_action = str(row.get("ActionCode", "")).strip().upper()

            # --- 1. Effective Sequence Logic (WorkSequence) ---
            new_seq = 1
            if pd.notna(inc_date):
                if curr_date is not None and inc_date == curr_date:
                    # Same date: Increment sequence if actions differ (or force increment for multiple rows)
                    if inc_action != curr_action:
                        new_seq = curr_seq + 1
                    else:
                        # Logic: If strict requirement "2 changes in action code... then sequence increment"
                        new_seq = curr_seq + 1
                elif curr_date is not None and inc_date < curr_date:
                    # Input date is older than current/cloud date -> Logic error or Correction
                    # For safety in this specific "Delta" logic, we might not auto-calculate correctly without full history.
                    # We leave it as 1 or rely on user input if provided, but here we reset to 1.
                    new_seq = 1
                    df.at[idx, "Reason for Failed"] += f"; Date {inc_date.date()} is before Cloud/Previous {curr_date.date()}"
                else:
                    # New future date
                    new_seq = 1
            
            # Apply calculated sequence
            df.at[idx, "EffectiveSequence"] = new_seq
            
            # Update state
            curr_date = inc_date
            curr_seq = new_seq
            curr_action = inc_action

            # --- 2. Assignment Sequence Logic ---
            if is_in_cloud:
                # Delta specific checks
                if inc_action in rehire_actions:
                    if "INACTIVE" not in curr_status and curr_action not in term_actions:
                         df.at[idx, "Reason for Failed"] += "; Error: Rehire allowed only if previous status is Terminated"
                    else:
                         curr_asg_seq += 1
                         df.at[idx, "AssignmentSequence"] = curr_asg_seq
                elif inc_action in gt_actions:
                    if "INACTIVE" in curr_status or curr_action in term_actions:
                         df.at[idx, "Reason for Failed"] += "; Error: GT not allowed on Terminated employee"
                    else:
                         curr_asg_seq += 1
                         df.at[idx, "AssignmentSequence"] = curr_asg_seq
            else:
                 # New Hire Assignment Sequence is usually 1, handled by default
                 pass

        # --- Perform Section A Checks (New Hire Logic) ---
        if is_in_cloud:
             # Check for logic specific to Delta (e.g., date mismatches on HIRE correction)
             first_incoming = group.iloc[0]
             first_inc_action = str(first_incoming.get("ActionCode", "")).strip().upper()
             
             if first_inc_action in hire_actions:
                 cloud_hire_date = latest_cloud_records.loc[person_number]["EffectiveStartDate"]
                 inc_start = first_incoming["EffectiveStartDate"]
                 if pd.notna(inc_start) and pd.notna(cloud_hire_date) and inc_start != cloud_hire_date:
                    df.at[first_incoming.name, "Reason for Failed"] += f"; Delta Error: HIRE date {inc_start.date()} must match Cloud HIRE date {cloud_hire_date.date()}"
                 
                 # Even if in cloud, if they are sending a HIRE row, we might want to run A-checks on the input stack
                 perform_section_a_checks(group, person_number)
        else:
            # Full New Hire Validation
            perform_section_a_checks(group, person_number)

    logger.info("Completed Dynamic Delta Validation with WorkSequence Logic.")
    return df

class AttributeConfig(BaseModel):
    Attributes: str
    required: bool
    keyValues: bool
    LookUp_data: str
    CodeName: str
    Data_Transformation: str
    data_type: str
    includeInDatFileGeneration: bool

# Pydantic model for lookup data structure
class LookupItem(BaseModel):
    CODE_Name: str
    Value: str
    Meaning: str
    Enabled_Flag: str
    Effective_Date: str

# Pydantic model for the entire validation payload
class ValidatePayload(BaseModel):
    pyFileName: Optional[str] = None
    componentName: str
    attributes: List[AttributeConfig]
    allLookups: Dict[str, List[LookupItem]]
    allMapping: Dict[str, Any] # Assuming mapping can be flexible
    excelFile: str # Base64 encoded Excel file content
    globalBoName: Optional[str] = None
    sourceKeys: Optional[Dict[str, str]] = None # Changed from List[str] to Dict[str, str]
    datColumnOrder: Optional[List[str]] = None
    hireActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    rehireActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    terminationActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    globalTransferActions: Optional[List[str]] = Field(default_factory=list) # Added with default
    customerName: Optional[str] = None # Added customerName
    InstanceName: Optional[str] = None # Added InstanceName
    DeltaLoad: bool = False # Added DeltaLoad flag

@app.post("/api/hdl/validate-data")
async def validate_data(payload: ValidatePayload):
    """
    Validates the uploaded Excel file against the provided attributes, lookups, and mappings.
    Returns a JSON response with validation results.
    Calls each function to check for its status and return the results.
    Fetches excel file from the static directory using the componentName and globalBoName.
    """
    component_name = payload.componentName
    global_bo_name = payload.globalBoName
    attributes_to_validate = payload.attributes
    all_lookups = payload.allLookups
    hire_actions = [act.upper().strip() for act in (payload.hireActions or [])]
    term_actions = [act.upper().strip() for act in (payload.terminationActions or [])]
    gt_actions = [act.upper().strip() for act in (payload.globalTransferActions or [])]
    rehire_actions = [act.upper().strip() for act in (payload.rehireActions or [])]
    customerName = payload.customerName or "default_customer"
    instanceName = payload.InstanceName or ""
    delta_load = payload.DeltaLoad 
    # ----------------------------------------------------------
    all_mapping = payload.allMapping
    excel_base64 = payload.excelFile
    dat_column_order = payload.datColumnOrder or []

    # -----------------------------------------------------------
    # use the setup data endpoint function to get these values for hire_actions, term_actions, gt_actions
    # Ensure these are lists even if None is passed

    if not hire_actions or not term_actions or not gt_actions or not rehire_actions:
        logger.info("Fetching missing action lists from setup data.")
        try:
            setup_data = get_hdl_setup_validate_fetch(customer_name=customerName, instance_name=instanceName)
            hire_actions = setup_data.get("hireActions", hire_actions)
            term_actions = setup_data.get("termActions", term_actions)
            gt_actions = setup_data.get("globalTransferActions", gt_actions)
            rehire_actions = setup_data.get("rehireActions", rehire_actions)
            logger.info(f"Fetched actions - Hire: {hire_actions}, Termination: {term_actions}, GT: {gt_actions}, Rehire: {rehire_actions}")
        except HTTPException as e:
            logger.error(f"Setup fetch error: {e.detail}")
            raise
        except Exception as e:
            logger.error(f"Unexpected error fetching setup: {e}")
            raise HTTPException(status_code=500, detail=f"Error fetching setup data: {str(e)}")

    # ----------------------------------------------------------
    if not excel_base64:
        raise HTTPException(status_code=400, detail="Excel file content is missing.")

    try:
        # Decode the base64 Excel file
        excel_bytes = base64.b64decode(excel_base64)
        excel_file_io = BytesIO(excel_bytes)
        excel_filename = f"{component_name}.xlsx"
        store_excel = DIR / (global_bo_name or component_name) / excel_filename
        # Ensure the directory exists
        store_excel.parent.mkdir(parents=True, exist_ok=True)
        # Save the Excel file to the determined path
        with open(store_excel, "wb") as f:
            f.write(excel_bytes)
        file_path_in_static = Path("uploads/Excel_Files") / (global_bo_name or component_name) / excel_filename
        # Load Excel file into DataFrame
        df = pd.read_excel(excel_file_io, engine='openpyxl')
        df.columns = [str(col).strip() for col in df.columns]
        df.columns = [col.strip() for col in df.columns] # Clean column names
        df = df.fillna("") # Fill NaN values with empty string for consistent validation

        # Initialize variables for Delta Logic
        delta_df = pd.DataFrame()
        delta_logic_executed = False

    # =========================================================================
    # 4. DYNAMIC DELTA / NEW HIRE VALIDATION BLOCK
    # =========================================================================
        if component_name.lower() == "assignment":
            delta_filename = f"{customerName}_{instanceName}_{component_name}_Report.csv"
            delta_file_path = Path("required_files") / delta_filename
            
            file_loaded = False

            if delta_file_path.exists():
                logger.info(f"Found Delta file at: {delta_file_path}")
                
                # Robust Load: Try Excel (openpyxl) -> CSV -> Excel (xlrd)
                if not file_loaded:
                    try:
                        delta_df = pd.read_excel(delta_file_path, engine='openpyxl')
                        file_loaded = True
                        logger.info("Successfully loaded Delta file using openpyxl.")
                    except Exception: pass

                if not file_loaded:
                    try:
                        delta_df = pd.read_csv(delta_file_path, sep=None, engine='python', encoding='utf-8-sig')
                        # Remove non-alphanumeric characters from column names
                        delta_df.columns = [col.replace('\ufeff', '').strip() for col in delta_df.columns]
                        file_loaded = True
                        logger.info("Successfully loaded Delta file as CSV.")
                    except Exception as e: 
                        logger.warning(f"CSV load failed: {e}")
                        pass

                if not file_loaded:
                        logger.error("CRITICAL: Delta file exists but could not be read. Validation will proceed assuming New Hires.")
            else:
                    logger.warning(f"Delta file not found at {delta_file_path}. Proceeding without delta logic (treat all as New Hires).")
                    logger.warning("To enable delta validation, fetch the Oracle report first via the Fetch Delta Report button.")
            # Execute Logic if Delta file is loaded
            if not delta_df.empty:
                    try:
                        # IMPORTANT: This function MUST be defined globally (outside this function)
                        df = validate_delta_logic(
                        df=df, 
                        delta_df=delta_df,
                        hire_actions=hire_actions,
                        term_actions=term_actions,
                        rehire_actions=rehire_actions,
                        gt_actions=gt_actions
                    )
                        delta_logic_executed = True
                    except Exception as logic_err:
                        logger.error(f"Error executing Delta Logic: {logic_err}", exc_info=True)
        # =========================================================================
        # END DELTA BLOCK
        # =========================================================================
        # Get the original columns from the uploaded Excel (before any additions like sourceKeys)
        original_excel_columns = df.columns.tolist()

        # Initialize 'Reason for Failed' column
        if "Reason for Failed" not in df.columns:
            df["Reason for Failed"] = ""
        else:
            df["Reason for Failed"] = df["Reason for Failed"].astype(str)

        all_errors = []

        # Remove only leading and trailing whitespaces from all string values
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)


        # 1. Perform Data Transformation/Mapping Validations (NOW FIRST)
        logger.info("Skipping data transformation step as requested.")
        # df = transformation_for_validation(excel_file_io) 
        # logger.info("Finished data transformation.")

        # 2. Perform Required Field Validations (after transformation)
        logger.info("Starting required field validations.")
        required_cols_from_payload = [attr.Attributes for attr in attributes_to_validate if attr.required]
        df = required_field_validations(df, required_cols_from_payload)
        logger.info("Finished required field validations.")

        # 2a. Perform Data Type Validation
        logger.info("Starting datatype validations.")
        df = validate_data_types(df, attributes_to_validate)
        logger.info("Finished datatype validations.")

        # 2b. Perform Key Values Uniqueness Validation
        logger.info("Starting key values uniqueness validation.")

        # EffectiveSequence validation
        #if the column exists 
        if "EffectiveSequence" in df.columns:
            logger.info("Starting WorkSequence auto-correction...")
            
            for person_number, group in df.groupby("PersonNumber"):
                # Sort ensures we process in the correct chronological order
                group = group.sort_values(by=["EffectiveStartDate", "EffectiveSequence"])
                indices = group.index
                
                # Initialize trackers with the first row of the group
                if len(indices) > 0:
                    previous_date = df.at[indices[0], "EffectiveStartDate"]
                    previous_sequence = df.at[indices[0], "EffectiveSequence"]

                # Iterate starting from the second row
                for i in range(1, len(indices)):
                    idx = indices[i]
                    current_date = df.at[idx, "EffectiveStartDate"]
                    current_seq = df.at[idx, "EffectiveSequence"]

                    if current_date == previous_date:
                        # If same day, ensure strictly increasing sequence
                        if current_seq <= previous_sequence:
                            new_sequence = previous_sequence + 1
                            df.at[idx, "EffectiveSequence"] = new_sequence
                            
                            # Log the correction if needed
                            logger.info(f"Auto-corrected Person {person_number} on {current_date.date()}: {current_seq} -> {new_sequence}")
                            
                            # Update the tracker to the new valid sequence
                            current_seq = new_sequence
                    
                    # Update trackers for the next iteration
                    previous_date = current_date
                    previous_sequence = current_seq

            logger.info("Finished WorkSequence validation and correction.")


        

        # Collect all attributes flagged as keyValues=True
        key_value_cols_from_payload = [attr.Attributes for attr in attributes_to_validate if attr.keyValues]

        if key_value_cols_from_payload:
            missing_cols = [col for col in key_value_cols_from_payload if col not in df.columns]
            if missing_cols:
                logger.warning(f"Key value validation skipped for missing columns: {', '.join(missing_cols)}")
            else:
                # Build composite key if multiple key columns exist
                df["_key_combo"] = df[key_value_cols_from_payload].astype(str).agg("|".join, axis=1)

                # Detect duplicates in that composite key
                dup_mask = df["_key_combo"].duplicated(keep=False)  # mark all duplicates, not just later ones
                if dup_mask.any():
                    for col in key_value_cols_from_payload:
                        df.loc[dup_mask, "Reason for Failed"] = df.loc[dup_mask, "Reason for Failed"].astype(str) + \
                            f"; Duplicate detected in key combination ({', '.join(key_value_cols_from_payload)})"

                # Drop helper column after validation
                df = df.drop(columns=["_key_combo"])

        logger.info("Finished key values uniqueness validation.")

        # --- VALIDATE START DATE BEFORE END DATE ---
        logger.info("Starting EffectiveStartDate <= EffectiveEndDate validation...")

        if "EffectiveStartDate" in df.columns and "EffectiveEndDate" in df.columns:
            df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors="coerce")
            df["EffectiveEndDate"] = pd.to_datetime(df["EffectiveEndDate"], errors="coerce")

            for idx, row in df.iterrows():
                start = row["EffectiveStartDate"]
                end = row["EffectiveEndDate"]

                if pd.isna(start) or pd.isna(end):
                    continue  

                if end < start:
                    logger.info(f"Row {idx} failed EffectiveStartDate <= EffectiveEndDate validation: {start} > {end}")
                    reason = f"EffectiveEndDate ({end.date()}) is before EffectiveStartDate ({start.date()})"
                else:
                    continue  

                existing_reason = df.at[idx, "Reason for Failed"] if "Reason for Failed" in df.columns else ""
                df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

        logger.info("Completed EffectiveStartDate <= EffectiveEndDate validation.")

        logger.info("Starting DateFrom <= DateTo validation...")

        if "DateFrom" in df.columns and "DateTo" in df.columns:
            df["DateFrom"] = pd.to_datetime(df["DateFrom"], errors="coerce")
            df["DateTo"] = pd.to_datetime(df["DateTo"], errors="coerce")

            for idx, row in df.iterrows():
                start = row["DateFrom"]
                end = row["DateTo"]

                if pd.isna(start) or pd.isna(end):
                    continue  

                if end < start:
                    logger.info(f"Row {idx} failed DateFrom <= DateTo validation: {start} > {end}")
                    reason = f"DateTo ({end.date()}) is before DateFrom ({start.date()})"
                else:
                    continue 

                existing_reason = df.at[idx, "Reason for Failed"] if "Reason for Failed" in df.columns else ""
                df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

        logger.info("Completed DateFrom <= DateTo validation.")


        logger.info("Started StartDate <= EndDate validation.")

        if "StartDate" in df.columns and "EndDate" in df.columns:
            df["StartDate"] = pd.to_datetime(df["StartDate"], errors="coerce")
            df["EndDate"] = pd.to_datetime(df["EndDate"], errors="coerce")

            for idx, row in df.iterrows():
                start = row["StartDate"]
                end = row["EndDate"]

                if pd.isna(start) or pd.isna(end):
                    continue  

                if end < start:
                    logger.info(f"Row {idx} failed StartDate <= EndDate validation: {start} > {end}")
                    reason = f"EndDate ({end.date()}) is before StartDate ({start.date()})"
                else:
                    continue 

                existing_reason = df.at[idx, "Reason for Failed"] if "Reason for Failed" in df.columns else ""
                df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

        logger.info("Completed StartDate <= EffectiveEndDate validation.")

        # 3. Perform Lookup Validations (after transformation and required field checks)
        logger.info("Starting lookup validations.")
        # `all_lookups` from payload is already in the correct format (attribute -> list of dicts)
        df = lookup_validations(df, all_lookups)
        logger.info("Finished lookup validations.")        

        
        # 4. Custom validation: First row for a Person Number must be 'HIRE' based on minimum Start Date
        # FIX: SKIP IF DELTA LOGIC HAS EXECUTED
        if not delta_logic_executed:
            logger.info("Starting custom validation: First row for Person Number must be 'HIRE'.")
            
            start_date_column = None
            if "EffectiveStartDate" in df.columns:
                start_date_column = "EffectiveStartDate"
            elif "DateStart" in df.columns:
                start_date_column = "DateStart"

            if "PersonNumber" in df.columns and "ActionCode" in df.columns and start_date_column:
                # Convert start_date_column to datetime, handling potential errors
                df[start_date_column] = pd.to_datetime(df[start_date_column], errors='coerce')

                # Drop rows where start_date_column could not be parsed, as they cannot be validated on date
                df_cleaned_dates = df.dropna(subset=[start_date_column]).copy()

                # Group by Person Number and find the row with the minimum DateStart
                idx = df_cleaned_dates.groupby('PersonNumber')[start_date_column].idxmin()
                first_rows_for_person = df_cleaned_dates.loc[idx]

                for _, row in first_rows_for_person.iterrows():
                    person_number = row["PersonNumber"]
                    action_code = str(row["ActionCode"]).strip().upper()

                    if action_code not in hire_actions:
                        # Find all original indices for this person_number to mark all their rows as failed
                        original_indices_for_person = df[df["PersonNumber"] == person_number].index
                        for i in original_indices_for_person:
                            # Append the reason, ensuring it's not None
                            current_reason = df.loc[i, "Reason for Failed"]
                            if current_reason:
                                df.loc[i, "Reason for Failed"] = f"{current_reason}; First action for PersonNumber '{person_number}' (based on minimum start date) must be 'HIRE', but was '{action_code}'"
                            else:
                                df.loc[i, "Reason for Failed"] = f"First action for PersonNumber '{person_number}' (based on minimum start date) must be 'HIRE', but was '{action_code}'"
            else: 
                missing_cols = []
                if "PersonNumber" not in df.columns: missing_cols.append("PersonNumber")
                if "ActionCode" not in df.columns: missing_cols.append("ActionCode")
                if not start_date_column: missing_cols.append("EffectiveStartDate or DateStart") # Updated warning message
                if missing_cols:
                    logger.warning(f"Skipping 'HIRE' validation due to missing column(s): {', '.join(missing_cols)}")
            logger.info("Finished custom validation: First row for Person Number must be 'HIRE'.")
        else:
             logger.info("Skipping generic 'First Row HIRE' validation because Delta Logic was executed.")


        # Legal Employer Name has to be consistant 
        logger.info("Starting validation: LegalEmployerName consistency check with reset on HIRE/TERMINATION/GT.")
        start_date_column = "EffectiveStartDate" if "EffectiveStartDate" in df.columns else "DateStart"
        required_cols = ["PersonNumber", "ActionCode", start_date_column, "LegalEmployerName"]
        if all(col in df.columns for col in required_cols):
            df[start_date_column] = pd.to_datetime(df[start_date_column], errors='coerce')
            df["ActionCode"] = df["ActionCode"].astype(str).str.strip().str.upper()
            df["LegalEmployerName"] = df["LegalEmployerName"].astype(str).str.strip()

            grouped = df.sort_values(by=["PersonNumber", start_date_column]).groupby("PersonNumber")

            for person, group in grouped:
                expected_legal = None
                for _, row in group.iterrows():
                    action = row["ActionCode"]
                    legal = row["LegalEmployerName"]
                    idx = row.name

                    if expected_legal is None:
                        expected_legal = legal
                    elif action in rehire_actions or action in gt_actions:
                        expected_legal = legal
                    else:
                        if legal != expected_legal:
                            existing_reason = df.at[idx, "Reason for Failed"]
                            reason = f"Inconsistent LegalEmployerName. Expected '{expected_legal}', but found '{legal}' and action code '{action}'."
                            df.at[idx, "Reason for Failed"] = f"{existing_reason}; {reason}" if existing_reason else reason

        else:
            logger.warning("Missing columns for LegalEmployerName consistency validation. Skipping.")

        logger.info("Completed validation: LegalEmployerName consistency check.")



        # --- GLOBAL TRANSFER validation: legal employer change + no employment gap ---
        logger.info("Starting GLOBAL TRANSFER validation (LegalEmployerName change + date continuity)...")

        required_cols = ["PersonNumber", "ActionCode", "EffectiveStartDate", "EffectiveEndDate", "LegalEmployerName"]
        if all(col in df.columns for col in required_cols):
            df["EffectiveStartDate"] = pd.to_datetime(df["EffectiveStartDate"], errors="coerce")
            df["EffectiveEndDate"] = pd.to_datetime(df["EffectiveEndDate"], errors="coerce")
            df["ActionCode"] = df["ActionCode"].astype(str).str.strip()
            df["LegalEmployerName"] = df["LegalEmployerName"].astype(str).str.strip()

            grouped = df.sort_values(by=["PersonNumber", "EffectiveStartDate"]).groupby("PersonNumber")

            for person_number, group in grouped:
                group = group.reset_index()

                for i in range(1, len(group)):
                    prev = group.loc[i - 1]
                    curr = group.loc[i]

                    if curr["ActionCode"].upper() in gt_actions:
                        legal_changed = curr["LegalEmployerName"] != prev["LegalEmployerName"]
                        expected_start = prev["EffectiveEndDate"] + pd.Timedelta(days=1)
                        actual_start = curr["EffectiveStartDate"]


                        reasons = []
                        if not legal_changed:
                            reasons.append("Legal Employer Name must change for change legal employer")
                        if pd.notna(expected_start) and pd.notna(actual_start) and actual_start != expected_start:
                            reasons.append(
                                f"EffectiveStartDate ({actual_start.date()}) must be the day after previous EffectiveEndDate ({prev['EffectiveEndDate'].date()})"
                            )

                        if reasons:
                            row_index = curr["index"]
                            existing_reason = df.at[row_index, "Reason for Failed"]
                            combined = "; ".join(reasons)
                            df.at[row_index, "Reason for Failed"] = f"{existing_reason}; {combined}" if existing_reason else combined
        else:
            missing = [c for c in required_cols if c not in df.columns]
            logger.warning(f"Skipping GLOBAL TRANSFER validation due to missing columns: {', '.join(missing)}")
        logger.info("Completed GLOBAL TRANSFER validation.")

        # --- WORKRELATIONSHIP specific rules ---
        if component_name.lower() == "workrelationship":
            logger.info("Applying WorkRelationship-specific rules.")
            df = apply_workrelationship_rules(
                df=df.copy(),
                hire_actions=hire_actions,
                rehire_actions=rehire_actions,
                gt_actions=gt_actions,
                term_actions=term_actions
            )
            logger.info("Completed WorkRelationship-specific rules.")

        
        
        # Filter passed and failed rows
        failed_df = df[df["Reason for Failed"] != ""].copy()
        passed_df = df[df["Reason for Failed"] == ""].copy()

        # =========================================================================
        # START: EFFECTIVE SEQUENCING LOGIC
        # =========================================================================
        if "PersonNumber" in passed_df.columns and "EffectiveStartDate" in passed_df.columns:
            logger.info("Calculating EffectiveSequence for same-day transactions.")
            
            # Create temp date for sorting
            passed_df["_tmp_eff_date"] = pd.to_datetime(passed_df["EffectiveStartDate"], errors='coerce')
            
            # Generate Sequence: Group by Person + Date
            passed_df["EffectiveSequence"] = passed_df.groupby(["PersonNumber", "_tmp_eff_date"]).cumcount() + 1
            
            # Cleanup
            passed_df.drop(columns=["_tmp_eff_date"], inplace=True)
            
            # Ensure the column is treated as part of the Excel columns so it gets exported
            if "EffectiveSequence" not in original_excel_columns:
                original_excel_columns.append("EffectiveSequence")
            
            logger.info("Effective Sequencing applied successfully.")

        # Save results to temporary files
        output_dir = VALIDATION_RESULTS_DIR
        
        # Ensure the output directory exists
        output_dir.mkdir(parents=True, exist_ok=True)

        validation_results = {}

        # Step 1 — Get all PersonNumbers with at least one failed record
        if "PersonNumber" not in df.columns:
            logger.error("PersonNumber column missing when trying to cascade fail. Available columns: %s", df.columns.tolist())
        else:
            failed_persons = df.loc[df["Reason for Failed"] != "", "PersonNumber"].unique()

            # Step 2 — For all rows of those PersonNumbers, mark as failed (if not already)
            for person_number in failed_persons:
                person_rows = df[df["PersonNumber"] == person_number].index
                for idx in person_rows:
                    if df.at[idx, "Reason for Failed"] == "":
                        df.at[idx, "Reason for Failed"] = "Failed due to other row(s) for this PersonNumber failing validation."


        # Save failed_df to an Excel file if it's not empty
        if not failed_df.empty:
            failed_file_name = get_generic_filename("failed", component_name, "xlsx")
            failed_file_path = output_dir / failed_file_name
            try:
                failed_df.to_excel(failed_file_path, index=False, engine='openpyxl')
                logger.info(f"Failed records saved to {failed_file_path}")
                validation_results["failed_file_url"] = f"/excel_validation_results/{failed_file_name}" # URL for frontend
                validation_results["failed_count"] = len(failed_df)
            except Exception as e:
                logger.error(f"Error saving failed_df to Excel: {e}")
                validation_results["failed_file_error"] = f"Failed to save error file: {str(e)}"
        else:
            logger.info("No failed records found. Skipping failed file generation.")
            validation_results["failed_count"] = 0

        # --- inside validate_data ---
        saved_code_file = UPLOAD_DIR / "saved_code" / f"{customerName}_{instanceName}_{component_name}.py"

        if saved_code_file.exists():
            try:
                module = load_validation_module(saved_code_file)
                logger.info(f"Loaded custom validation module: {saved_code_file}")

                if hasattr(module, "validate_row"):
                    logger.info(f"Found 'validate_row' function in {saved_code_file}. Applying custom row-level validation.")
                    # Initialize column to track row-level validation failures
                    if "RowValidationFailed" not in df.columns:
                        df["RowValidationFailed"] = False

                    for idx, row in df.iterrows():
                        try:
                            logger.debug(f"Applying custom validation to row {idx}")
                            is_valid, reason = module.validate_row(row.to_dict(), idx)
                            # Normalize outputs
                            is_valid = bool(is_valid)
                            reason = str(reason).strip() if reason else ""
                            
                            if not is_valid:
                                logger.debug(f"Row {idx} failed custom validation: {reason}")
                                # Append to existing "Reason for Failed"
                                existing_reason = df.at[idx, "Reason for Failed"] or ""
                                combined_reason = f"{existing_reason}; {reason}" if existing_reason else reason
                                df.at[idx, "Reason for Failed"] = combined_reason
                                df.at[idx, "RowValidationFailed"] = True
                        except Exception as row_err:
                            # Catch row-level exceptions, log but continue
                            logger.error(f"Custom validation failed for row {idx}: {row_err}", exc_info=True)
                            existing_reason = df.at[idx, "Reason for Failed"] or ""
                            df.at[idx, "Reason for Failed"] = f"{existing_reason}; Custom validation error: {row_err}" if existing_reason else f"Custom validation error: {row_err}"
                            df.at[idx, "RowValidationFailed"] = True

                else:
                    logger.warning(f"No 'validate_row' function found in {saved_code_file}")

            except Exception as mod_err:
                logger.error(f"Failed to load custom validation module: {mod_err}", exc_info=True)
        else:
            logger.info(f"No custom validation file found at: {saved_code_file}")


        # Filter passed and failed rows
        failed_df = df[df["Reason for Failed"] != ""].copy()
        passed_df = df[df["Reason for Failed"] == ""].copy()

        # Save results to temporary files
        output_dir = VALIDATION_RESULTS_DIR

        # --- Filter columns for DAT file generation based on includeInDatFileGeneration flag ---
        # Remove 'Reason for Failed' column before any further processing if it's not meant for the DAT output.
        if 'Reason for Failed' in passed_df.columns:
            passed_df = passed_df.drop(columns=['Reason for Failed'])
            
        # --- Prepare iterator values for passed_df ---
        actioncode_col = None
        personnumber_col = None
        for col in passed_df.columns:
            if col.strip().lower() == "actioncode":
                actioncode_col = col
            if col.strip().lower() == "personnumber":
                personnumber_col = col

        # --- FIX: Initialize Iterator from Cloud Max Sequence if Delta Logic Ran ---
        cloud_max_seq_map = {}
        if delta_logic_executed and not delta_df.empty:
             try:
                 # Assuming "AssignmentSequence" is the relevant column for iterator
                 seq_col = "AssignmentSequence" 
                 # Normalize column names in delta_df just in case
                 delta_df_norm = delta_df.copy()
                 delta_df_norm.columns = [str(c).replace(" ", "") for c in delta_df_norm.columns]
                 
                 if seq_col in delta_df_norm.columns:
                     # Ensure numeric and fill NaNs
                     delta_df_norm[seq_col] = pd.to_numeric(delta_df_norm[seq_col], errors='coerce').fillna(0)
                     # Get Max per person
                     if "PersonNumber" in delta_df_norm.columns:
                         delta_df_norm["PersonNumber"] = delta_df_norm["PersonNumber"].astype(str).str.strip()
                         cloud_max_seq_map = delta_df_norm.groupby("PersonNumber")[seq_col].max().to_dict()
             except Exception as e:
                 logger.warning(f"Failed to build iterator map from Cloud data: {e}")

        iterator_list = []
        iterator_counter = 0
        prev_person_number = None
        combined_actions_increment = gt_actions + rehire_actions # These definitely increment sequence
        hire_correction_actions = hire_actions # These increment ONLY if starting from 0

        for _, row in passed_df.iterrows():
            current_person_number = str(row[personnumber_col]).strip() if personnumber_col else None
            
            if personnumber_col and current_person_number != prev_person_number:
                # Reset or Initialize Iterator
                if delta_logic_executed:
                    # If person exists in cloud, start from their max sequence
                    iterator_counter = int(cloud_max_seq_map.get(current_person_number, 0))
                else:
                    iterator_counter = 0
                
                prev_person_number = current_person_number
            
            # Determine if we need to increment
            if actioncode_col:
                action = str(row[actioncode_col]).strip().upper()
                
                if action in combined_actions_increment:
                    iterator_counter += 1
                elif action in hire_correction_actions:
                    # Special Case: Hire only increments if we are at 0 (New Hire)
                    # If we are at 1+ (Existing), a HIRE is a correction and shouldn't increment.
                    if iterator_counter == 0:
                        iterator_counter = 1
                # Else (Promotion, Correction, etc) -> Do not increment, inherit current assignment ID
            else:
                # Fallback if no action code? Just increment?
                iterator_counter += 1 # Default behavior for unknown structures

            iterator_list.append(str(iterator_counter))


        # Insert sourceKeys columns at the front, using iterator_list for {Iterator} (including as part of a string)
        source_keys = getattr(payload, 'sourceKeys', None)
        # Store names of source keys and their original order as they appear in the payload.sourceKeys dict
        source_key_names_in_order = [] 
        if source_keys and isinstance(source_keys, dict):
            # Create a temporary DataFrame for just the source keys, maintaining their intended order
            source_keys_data = {}
            for key, val in source_keys.items():
                source_key_names_in_order.append(key) # Keep track of the order of source keys
                col_values = []
                for idx, (_, row) in enumerate(passed_df.iterrows()): # Iterate over the passed_df rows to substitute values
                    def replacer(match):
                        if match.group(1) == 'Iterator':
                            return str(iterator_list[idx]) if idx < len(iterator_list) else ''
                        col_name = match.group(1)
                        # Ensure we are getting values from the original passed_df (before sourceKey insertion or initial drop of 'Reason for Failed')
                        # For this, it's safer to get from the 'df' DataFrame (initial loaded Excel data) if possible,
                        # or ensure 'passed_df' still contains all original columns before dropping 'Reason for Failed' and adding source keys.
                        # As it is, `passed_df` right before this block *does* contain original columns (minus 'Reason for Failed' if it was there).
                        return str(row.get(col_name, ''))
                    
                    substituted_value = re.sub(r'\{([^}]+)\}', replacer, str(val))
                    col_values.append(substituted_value)
                source_keys_data[key] = col_values
            
            # Create a DataFrame for source keys and concatenate it to the passed_df
            # This ensures source keys are at the beginning in their specified order
            source_keys_df = pd.DataFrame(source_keys_data, index=passed_df.index)
            # Use concat to place source_keys_df at the start
            passed_df = pd.concat([source_keys_df, passed_df], axis=1)
            logger.info(f"Added source keys to passed_df. New columns added: {source_key_names_in_order}")

        # Now, construct the final columns for DAT output based on datColumnOrder or original Excel order.
        explicitly_excluded_from_dat = {
            attr_data.Attributes for attr_data in attributes_to_validate if not attr_data.includeInDatFileGeneration
        }

        final_dat_columns_ordered = []

        # 1. Add source keys first
        for sk_name in source_key_names_in_order:
            if sk_name in passed_df.columns: # Ensure it actually exists in the combined DF
                final_dat_columns_ordered.append(sk_name)

        if dat_column_order:
            logger.info(f"Using provided datColumnOrder: {dat_column_order}")
            # Then add columns from datColumnOrder, respecting exclusions and ensuring they exist in passed_df
            for col_name in dat_column_order:
                # IMPORTANT: Only add if NOT already added as a source key AND NOT explicitly excluded
                if col_name not in source_key_names_in_order and col_name not in explicitly_excluded_from_dat:
                    if col_name in passed_df.columns:
                        final_dat_columns_ordered.append(col_name)
        else:
            logger.info("datColumnOrder not provided. Falling back to original Excel column order with source keys prepended.")
            # Then add original Excel columns
            for col_name in original_excel_columns:
                # IMPORTANT: Only add if NOT already added as a source key AND NOT explicitly excluded
                if col_name not in source_key_names_in_order and col_name not in explicitly_excluded_from_dat:
                    if col_name in passed_df.columns:
                        final_dat_columns_ordered.append(col_name)

        # Remove any duplicates that might arise from edge cases, while preserving order as much as possible
        seen = set()
        deduplicated_final_dat_columns_ordered = []
        for col in final_dat_columns_ordered:
            if col not in seen:
                deduplicated_final_dat_columns_ordered.append(col)
                seen.add(col)

        # Apply the final ordering and filtering to the DataFrame
        # Ensure that all columns in deduplicated_final_dat_columns_ordered are actually in passed_df
        existing_cols_in_final_order = [col for col in deduplicated_final_dat_columns_ordered if col in passed_df.columns]

        passed_df_final_output = passed_df[existing_cols_in_final_order]
        logger.info(f"Final columns for DAT file generation (ordered): {passed_df_final_output.columns.tolist()}")

        # --- STEP: ENFORCE DATATYPE CONSISTENCY BEFORE SAVING ---
        logger.info("Enforcing consistent dtypes for passed and failed DataFrames...")
        # Capture reference dtypes before enforcing consistency
        reference_dtypes = df.dtypes.to_dict()

        # ------------------ START PATCH: Preserve Oracle-safe dates ------------------
        # helper: detect oracle-safe year (year >= 4712 up to 9999)
        def _is_oracle_safe_date_str(s: Any) -> bool:
            try:
                s = str(s).strip()
                # Accept formats like YYYY-MM-DD, YYYY/MM/DD, YYYYMMDD, YYYY-MM-DD HH:MM:SS etc.
                m = re.match(r'^(-?\d{4})', s)
                if not m:
                    return False
                year = int(m.group(1))
                return 4712 <= year <= 9999
            except Exception:
                return False

        # Build a mask map: for each date/datetime column from payload, mark rows that are oracle-safe strings.
        oracle_safe_mask = {}  # col -> boolean Series (index same as df)
        date_cols_from_payload = [attr.Attributes for attr in attributes_to_validate if str(attr.data_type).lower() in ("date","datetime","timestamp")]

        for col in date_cols_from_payload:
            if col in df.columns:
                # Build mask from original df (before we coerce to datetimes)
                oracle_safe_mask[col] = df[col].apply(lambda v: _is_oracle_safe_date_str(v))
            else:
                oracle_safe_mask[col] = pd.Series(False, index=df.index)

        # After split, ensure passed_df retains original string values for oracle-safe rows.
        # But first ensure passed_df exists (it does)
        # Fill missing EffectiveEndDate in passed_df with oracle default as string BEFORE dtype enforcement
        if "EffectiveEndDate" and component_name.lower() != "workrelationship"  in passed_df.columns:
            #if the action code column exists and that row data have termactions then we set the default end date as ""
            action_code_col = None
            for col in passed_df.columns:
                if col.strip().lower() == "actioncode":
                    action_code_col = col
            if action_code_col and any(act in term_actions for act in passed_df[action_code_col].astype(str).str.strip().str.upper()):
                passed_df["EffectiveEndDate"] = passed_df["EffectiveEndDate"].astype("object")
                passed_df.loc[
                    (passed_df["EffectiveEndDate"].astype(str).str.strip() == "") & (passed_df[action_code_col].astype(str).str.strip().str.upper().isin(term_actions)),
                    "EffectiveEndDate"
                ] = "" 
        elif "EffectiveEndDate" in passed_df.columns:
            passed_df["EffectiveEndDate"] = passed_df["EffectiveEndDate"].astype("object")
            passed_df.loc[
                passed_df["EffectiveEndDate"].astype(str).str.strip() == "", "EffectiveEndDate"
            ] = "4712/12/31"
        
        elif "DateTo" in passed_df.columns:
            passed_df["DateTo"] = passed_df["DateTo"].astype("object")
            passed_df.loc[
                passed_df["DateTo"].astype(str).str.strip() == "", "DateTo"
            ] = "4712/12/31"
            
        # Also if EffectiveStartDate blank and StartDate present, fallback
        if "EffectiveStartDate" in passed_df.columns:
            passed_df["EffectiveStartDate"] = passed_df["EffectiveStartDate"].replace({pd.NaT: "", None: ""})
            if "StartDate" in passed_df.columns:
                mask_missing_start = passed_df["EffectiveStartDate"].astype(str).str.strip() == ""
                passed_df.loc[mask_missing_start & passed_df["StartDate"].astype(str).str.strip().ne(""), "EffectiveStartDate"] = passed_df.loc[mask_missing_start & passed_df["StartDate"].astype(str).str.strip().ne(""), "StartDate"].astype(str)
            # if still empty, optionally leave as "" or set to 1900/01/01. skip here.

        # Now update df used downstream for dtype enforcement: we want to preserve oracle-safe strings in passed_df_final_output later.
        # The enforce functions below will reference `oracle_safe_mask` and will avoid coercing those cells.

        # Replace enforce_dtypes and enforce_payload_dtypes with oracle-aware versions

        def enforce_dtypes_oracle_aware(df_to_cast, reference_df=None):
            """
            Enforce data types from reference_df (usually the original dataframe) onto df_to_cast.
            Special handling for Oracle-safe date strings.
            """
            if reference_df is not None:
                reference_dtypes = reference_df.dtypes.to_dict()
            else:
                reference_dtypes = df_to_cast.dtypes.to_dict()  # fallback if nothing is passed

            for col, dtype in reference_dtypes.items():
                if col not in df_to_cast.columns:
                    continue

                # Skip dtype enforcement for Oracle-safe date strings
                if col in oracle_safe_mask and oracle_safe_mask[col].any():
                    mask = oracle_safe_mask[col]
                    temp_series = df_to_cast[col].copy()
                    try:
                        coerced = pd.to_datetime(temp_series, errors="coerce")
                        coerced = coerced.astype("datetime64[ns]")
                        coerced[mask] = temp_series[mask]  # restore oracle-safe strings
                        df_to_cast[col] = coerced
                    except Exception as e:
                        logger.warning(f"Could not safely enforce datetime dtype on {col}: {e}")
                else:
                    try:
                        df_to_cast[col] = df_to_cast[col].astype(dtype)
                    except Exception as e:
                        logger.warning(f"Could not enforce dtype {dtype} on {col}: {e}")

            return df_to_cast


        def enforce_payload_dtypes_oracle_aware(df_to_cast, column_type_map):
            for col, dtype in column_type_map.items():
                if col not in df_to_cast.columns:
                    continue
                try:
                    if dtype.lower() in ["date", "datetime", "timestamp"]:
                        # cell-by-cell: keep oracle-safe strings, parse others
                        new_vals = []
                        for idx, val in df_to_cast[col].items():
                            if _is_oracle_safe_date_str(val):
                                new_vals.append(str(val).replace('-', '/'))
                            else:
                                new_vals.append(pd.to_datetime(val, errors='coerce'))
                        df_to_cast[col] = pd.Series(new_vals, index=df_to_cast.index)
                    elif dtype.lower() in ["int", "integer"]:
                        df_to_cast[col] = pd.to_numeric(df_to_cast[col], errors="coerce").fillna(0).astype(int)
                    elif dtype.lower() in ["float", "double", "decimal"]:
                        df_to_cast[col] = pd.to_numeric(df_to_cast[col], errors="coerce")
                    else:
                        df_to_cast[col] = df_to_cast[col].astype(str)
                except Exception as e:
                    logger.warning(f"Failed to cast column {col} to {dtype}: {e}")
            return df_to_cast

        # Replace calls
        passed_df_final_output = enforce_dtypes_oracle_aware(passed_df_final_output)
        failed_df = enforce_dtypes_oracle_aware(failed_df)
        logger.info("Dtype enforcement completed (oracle-aware).")

        # Build column_type_map dynamically from payload
        column_type_map = {}
        for attr in attributes_to_validate:
            col_name = attr.Attributes
            dtype = str(attr.data_type).lower() if hasattr(attr, "data_type") else "string"
            column_type_map[col_name] = dtype
        column_type_map["EffectiveSequence"] = "int" 

        # Update payload-enforcement cast
        passed_df_final_output = enforce_payload_dtypes_oracle_aware(passed_df_final_output, column_type_map)
        failed_df = enforce_payload_dtypes_oracle_aware(failed_df, column_type_map)
        logger.info("Datatype enforcement from payload completed (oracle-aware).")

        # Fix format_row_for_dat: accept both datetime objects and oracle date-strings
        def format_row_for_dat(row, df, column_type_map):
            formatted = []
            for col, val in row.items():
                dtype = column_type_map.get(col, None)
                # If val is a string and looks like an oracle-safe date (YYYY/MM/DD or YYYY-MM-DD), just normalize and append
                if isinstance(val, str) and _is_oracle_safe_date_str(val):
                    s = val.replace('-', '/')
                    # if it's already YYYY/MM/DD or YYYY/MM/DD HH:MM:SS keep as needed
                    if re.match(r'^\d{4}/\d{2}/\d{2}$', s):
                        formatted.append(s)
                    else:
                        # try to normalize date portion only
                        formatted.append(s.split(' ')[0])
                    continue

                if dtype and dtype.lower() in ["date", "datetime", "timestamp"]:
                    if pd.isna(val):
                        formatted.append('')
                    else:
                        # If pandas Timestamp:
                        if isinstance(val, (pd.Timestamp, datetime)):
                            if dtype.lower() == "date":
                                formatted.append(val.strftime('%Y/%m/%d'))
                            else:
                                # datetime or timestamp
                                formatted.append(val.strftime('%Y/%m/%d %H:%M:%S'))
                        else:
                            # fallback: convert to string and try to normalize separators
                            s = str(val).replace('-', '/')
                            formatted.append(s)
                else:
                    formatted.append(str(val) if val is not None else '')
            return formatted

        # ------------------ END PATCH ------------------
        # ---------------- Oracle 4712-Year Safe Patch ----------------
        def enforce_oracle_safe_dates(df: pd.DataFrame) -> pd.DataFrame:
            """
            Ensures Oracle-compatible '4712' and null dates are preserved as strings.
            Prevents pandas datetime64 truncation and NaT display issues.
            """
            for col in df.columns:
                if "date" in col.lower():
                    # Convert full column to string before any replacements
                    df[col] = df[col].astype(str)

                    # Replace NaT or NaN with empty string
                    df[col] = df[col].replace(["NaT", "nan", "NaN"], "")

                    # Replace blanks with Oracle's max date
                    if col == "DateTo" or col == "EndDate" or col == "EffectiveEndDate":
                        df.loc[df[col].str.strip() == "", col] = df[col].apply(
                            lambda x: "4712/12/31" if x.strip() == "" else x
                        )

                    # Ensure consistent format
                    df[col] = df[col].apply(
                        lambda x: x.replace("-", "/") if "/" not in x else x
                    )
            return df

        # Apply before exporting both passed and failed dataframes
        passed_df_final_output = enforce_oracle_safe_dates(passed_df_final_output)
        failed_df = enforce_oracle_safe_dates(failed_df)
        # -------------------------------------------------------------
        # --- SAVE PASSED DATA ---
        passed_file_name = None  # define upfront

        if not passed_df_final_output.empty:
            passed_file_name = get_generic_filename(f"{component_name}_passed", "data", "dat")
            passed_file_path = output_dir / passed_file_name
            with open(passed_file_path, "w", encoding="utf-8") as f:
                f.write("|".join([str(col) for col in passed_df_final_output.columns]) + "\n")
                for _, row in passed_df_final_output.iterrows():
                    f.write("|".join(format_row_for_dat(row, passed_df_final_output, column_type_map)) + "\n")
            logger.info(f"Passed validation data saved to: {passed_file_path}")
            
        # --- SAVE FAILED DATA ---
        failed_file_name = None  # define upfront

        if not failed_df.empty:
            failed_file_name = get_generic_filename("failed", component_name, "xlsx")
            failed_file_path = output_dir / failed_file_name
            tmp_path_failed = failed_file_path.with_suffix(".xlsx.tmp")
            failed_df.to_excel(tmp_path_failed, index=False, engine='openpyxl')
            os.replace(tmp_path_failed, failed_file_path)
            logger.info(f"Failed validation data saved to: {failed_file_path}")



        # --- PREPARE RESPONSE ---
        response_content = {
            "message": "Validation complete.",
            "status": "success" if failed_df.empty else "failed",
            "passed_records_count": len(passed_df_final_output),
            "failed_records_count": len(failed_df),
            "delta_logic_executed": delta_logic_executed,
            "passed_file_url": f"http://localhost:8000/validation_results/{passed_file_name}" if passed_file_name else None,
            "failed_file_url": f"http://localhost:8000/validation_results/{failed_file_name}" if failed_file_name else None,
            "errors": all_errors
        }

        return JSONResponse(content=response_content)


    except FileNotFoundError as e:
        logger.error(f"File not found during validation: {e}")
        raise HTTPException(status_code=404, detail=f"Required file not found: {e}")
    except Exception as e:
        logger.exception("An error occurred during data validation.")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred during validation: {str(e)}")


class DeltaLoadPayload(BaseModel):
    customerName: str
    instanceName: str




VALIDATION_RESULTS_DIR = "validation_results"
os.makedirs(VALIDATION_RESULTS_DIR, exist_ok=True)
# Mount the directory for serving validation results. It's now ensured to exist at startup.
app.mount("/validation_results", StaticFiles(directory=VALIDATION_RESULTS_DIR), name="validation_results")

@app.post("/api/validate-personname")
async def validate_personname(
    user_id: str = Form(...),
    excel_file_id: str = Form(...),
    dat_file_id: str = Form(...),
    required_columns: str = Form(...),
    lookup_json: str = Form(None),
    mapping_json: str = Form(None)
):
    """
    Validate the uploaded Excel file for required columns, lookups, and mapping values in one go.
    Returns file paths for failed and passed validations.
    """
    try:
        # COMPLETED_FOLDER is now defined globally and created at startup
        # UPLOAD_DIR is now defined globally and created at startup
        excel_path = UPLOAD_DIR / excel_file_id
        dat_path = UPLOAD_DIR / dat_file_id

        if not excel_path.exists() or not dat_path.exists():
            logging.error(f"Excel or DAT file not found: {excel_path}, {dat_path}")
            raise HTTPException(status_code=404, detail="Excel or DAT file not found.")

        # Load Excel
        df = pd.read_excel(excel_path)
        df.columns = df.columns.str.strip().str.replace(" ", "")
        df = df.fillna("")
        required_cols = [col.strip().replace(" ", "") for col in required_columns.split(",") if col.strip()]
        errors = []

        # Ensure Reason for Failed column exists and is string type
        if "Reason for Failed" not in df.columns:
            df["Reason for Failed"] = ""
        else:
            df["Reason for Failed"] = df["Reason for Failed"].astype(str)

        # 1. Validate required columns
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            for col in missing_cols:
                df["Reason for Failed"] += f"; Missing required column: {col}"

        # 2. Validate required values (non-empty)
        for col in required_cols:
            if col in df.columns:
                empty_mask = df[col].astype(str).str.strip() == ""
                df.loc[empty_mask, "Reason for Failed"] += f"; {col} is required"

        # 3. Validate lookups (if provided)
        if lookup_json:
            try:
                lookup_data = json.loads(lookup_json)
                for lookup in lookup_data:
                    col = lookup.get("column")
                    valid_values = set(str(v).strip().lower() for v in lookup.get("valid_values", []))
                    if col and col in df.columns:
                        invalid_mask = ~df[col].astype(str).str.strip().str.lower().isin(valid_values)
                        df.loc[invalid_mask, "Reason for Failed"] += f"; Invalid lookup for {col}"
            except Exception as e:
                logging.error(f"Lookup validation error: {str(e)}")
                errors.append(f"Lookup validation error: {str(e)}")

        # 4. Validate mappings (if provided)
        if mapping_json:
            try:
                mapping_data = json.loads(mapping_json)
                for mapping in mapping_data:
                    col = mapping.get("column")
                    mapping_dict = {str(m["source"]).strip().lower(): m["target"] for m in mapping.get("mappings", [])}
                    if col and col in df.columns:
                        mapped = df[col].astype(str).str.strip().str.lower().map(mapping_dict)
                        invalid_mask = mapped.isnull()
                        df.loc[invalid_mask, "Reason for Failed"] += f"; Mapping failed for {col}"
                        # Optionally, update column with mapped values where valid
                        df.loc[~invalid_mask, col] = mapped[~invalid_mask]
            except Exception as e:
                logging.error(f"Mapping validation error: {str(e)}")
                errors.append(f"Mapping validation error: {str(e)}")

        
        # Clean up Reason for Failed column (remove leading/trailing semicolons and whitespace)
        df["Reason for Failed"] = df["Reason for Failed"].str.strip(" ;")

        # Split passed/failed
        failed_df = df[df["Reason for Failed"] != ""]
        passed_df = df[df["Reason for Failed"] == ""]

        # Save failed and passed files
        failed_file_path = None
        passed_file_path = None
        if not failed_df.empty:
            failed_file_path = COMPLETED_FOLDER / get_generic_filename(user_id, "validation_errors", "xlsx")
            # Use atomic write
            tmp_path = failed_file_path.with_suffix(".xlsx.tmp")
            failed_df.to_excel(tmp_path, index=False)
            os.replace(tmp_path, failed_file_path)
            logging.info(f"Failed rows saved at {failed_file_path}.")
        if not passed_df.empty:
            passed_file_path = COMPLETED_FOLDER / get_generic_filename(user_id, "validation_passed", "dat")
            # Write as .dat (pipe-separated)
            with open(passed_file_path, "w", encoding="utf-8") as f:
                f.write("|".join(passed_df.columns) + "\n")
                for _, row in passed_df.iterrows():
                    f.write("|".join([str(x) for x in row]) + "\n")
            logging.info(f"Passed rows saved at {passed_file_path}.")

        return JSONResponse(content={
            "status": "success" if failed_df.empty and not errors else "error",
            "failed_file_path": str(failed_file_path.relative_to(UPLOAD_DIR.parent)) if failed_file_path else None,
            "passed_file_path": str(passed_file_path.relative_to(UPLOAD_DIR.parent)) if passed_file_path else None,
            "error_messages": errors
        })
    except Exception as e:
        logging.error(f"Error in unified validation: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error in unified validation: {str(e)}")
    


# --- Define the base directory for uploads, matching the bulk upload script ---
UPLOAD_DIR = Path("uploads/Excel_Files")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True) # Ensure the base directory exists

# --- Define the request body model using Pydantic for automatic validation ---
class ValidationPayload(BaseModel):
    parent_name: str
    component_files: Dict[str, str]

# Helper function to clean dictionary for JSON serialization
def clean_dict_for_json(data):
    """
    Recursively cleans a dictionary or list for JSON serialization by converting
    NaN and NaT values to None.
    """
    if isinstance(data, dict):
        return {k: clean_dict_for_json(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [clean_dict_for_json(elem) for elem in data]
    elif pd.isna(data): # Checks for both NaN and NaT from pandas
        return None
    elif isinstance(data, float) and math.isnan(data): # Checks for pure float NaNs
        return None
    elif isinstance(data, datetime): # Optional: Convert datetime objects to ISO format string
        return data.isoformat()
    else:
        return data


@app.post("/api/hdl/bulk/cross-file/personNumber/validate")
async def validate_person_numbers(
    parent_name: str = Body(...),
    component_files: Dict[str, str] = Body(...),
    all_mandatory_objects: List[str] = Body(..., embed=True),
    all_non_mandatory_objects: List[str] = Body(..., embed=True),
    export_as_excel: bool = Query(False)
):
    logging.info(f"Starting validation for parent: {parent_name}")

    all_person_numbers: Set[str] = set()
    common_person_numbers: Optional[Set[str]] = None
    all_rows_list: List[pd.DataFrame] = []
    person_numbers_per_component: Dict[str, Set[str]] = {}
    structured_failed_person_numbers: List[Dict] = []

    # Load mandatory components
    for component_name, file_name in component_files.items():
        file_path = UPLOAD_DIR / parent_name / file_name
        if not file_path.exists():
            logging.warning(f"File not found: {file_path}")
            continue

        try:
            df = pd.read_excel(file_path, header=1)
            if "PersonNumber" not in df.columns:
                logging.warning(f"Missing 'PersonNumber' in {file_path.name}, skipping.")
                continue

            df["PersonNumber"] = df["PersonNumber"].astype(str).str.strip().str.replace(r"\\.0$", "", regex=True)
            person_nums = set(df["PersonNumber"].dropna().unique())
            all_person_numbers.update(person_nums)

            if component_name in all_mandatory_objects:
                common_person_numbers = person_nums if common_person_numbers is None else common_person_numbers & person_nums

            person_numbers_per_component[component_name] = person_nums
            df["Component"] = component_name
            all_rows_list.append(df)

        except Exception as e:
            logging.error(f"Error reading file {file_name}: {e}", exc_info=True)
            continue

    if common_person_numbers is None:
        common_person_numbers = set()

    failed_person_numbers_set = all_person_numbers - common_person_numbers

    for p_num in failed_person_numbers_set:
        missing = [comp for comp in all_mandatory_objects if p_num not in person_numbers_per_component.get(comp, set())]
        structured_failed_person_numbers.append({
            "person_number": p_num,
            "missing_components": missing,
            "description": f"'{p_num}' missing in: {', '.join(missing)}"
        })

    # Load non-mandatory components directly from disk
    mandatory_person_numbers = set()
    for comp in all_mandatory_objects:
        mandatory_person_numbers.update(person_numbers_per_component.get(comp, set()))

    for non_mand_comp in all_non_mandatory_objects:
        file_name = f"{non_mand_comp}.xlsx"
        file_path = UPLOAD_DIR / parent_name / file_name

        if not file_path.exists():
            logging.info(f"Skipping non-mandatory file not found: {file_path}")
            continue

        try:
            df = pd.read_excel(file_path, header=1)
            if "PersonNumber" not in df.columns:
                logging.warning(f"'PersonNumber' column missing in {file_name}")
                continue

            df["PersonNumber"] = df["PersonNumber"].astype(str).str.strip().str.replace(r"\\.0$", "", regex=True)
            non_mand_pnums = set(df["PersonNumber"].dropna().unique())
            person_numbers_per_component[non_mand_comp] = non_mand_pnums

            for p_num in non_mand_pnums:
                if p_num not in mandatory_person_numbers:
                    structured_failed_person_numbers.append({
                        "person_number": p_num,
                        "missing_components": ["[Not Found in Mandatory]"],
                        "description": f"'{p_num}' is in non-mandatory '{non_mand_comp}' but not in any mandatory component"
                    })
                    failed_person_numbers_set.add(p_num)

            df["Component"] = non_mand_comp
            all_rows_list.append(df)

        except Exception as e:
            logging.error(f"Error reading non-mandatory file {file_name}: {e}", exc_info=True)
            continue

    structured_failed_person_numbers.sort(key=lambda x: x["person_number"])

    validation_failed_df = pd.DataFrame()
    if all_rows_list:
        full_df = pd.concat(all_rows_list, ignore_index=True)
        validation_failed_df = full_df[full_df["PersonNumber"].isin(failed_person_numbers_set)].copy()

        cols_to_check = [col for col in validation_failed_df.columns if col not in ["PersonNumber", "Component"]]
        mask = pd.Series(True, index=validation_failed_df.index)
        for col in cols_to_check:
            mask &= validation_failed_df["PersonNumber"] != validation_failed_df[col].astype(str)
        validation_failed_df = validation_failed_df[mask].reset_index(drop=True)

    exported_excel_filename = None
    export_status = None
    if export_as_excel and not validation_failed_df.empty:
        filename = f"validation_failed_{uuid.uuid4()}.xlsx"
        filepath = VALIDATION_RESULTS_DIR / filename
        validation_failed_df.to_excel(filepath, index=False)
        exported_excel_filename = filename
        export_status = "Excel file generated successfully."
    elif export_as_excel:
        export_status = "No failed records to export."

    response = {
        "validation_summary": {
            "total_unique_person_numbers": len(all_person_numbers),
            "passed_validation_count": len(common_person_numbers),
            "failed_validation_count": len(failed_person_numbers_set),
            "excel_export_status": export_status
        },
        "passed_person_numbers": sorted(list(common_person_numbers)),
        "failed_person_numbers": structured_failed_person_numbers,
        "validation_failed_details": validation_failed_df.to_dict(orient="records"),
        "exported_excel_filename": exported_excel_filename
    }

    return JSONResponse(content=clean_dict_for_json(response), status_code=200)


@app.post("/api/hdl/bulk/cross-file/personNumber/remove-failed-values")
async def remove_failed_person_numbers(
    parent_name: str = Body(..., description="The parent (L4) node name for the uploaded Excel."),
    component_files: Dict[str, str] = Body(..., description="Dictionary of component_name: file_path_in_uploads."),
    person_numbers_to_remove: List[str] = Body(..., description="List of person numbers to remove from the component files.")

):
    """
    Removes rows containing specified person numbers from the given component Excel files.
    This operation directly modifies the Excel files on disk.
    """
    logging.info(f"Received request to remove failed person numbers for parent: {parent_name}")
    logging.info(f"Component files to process: {component_files}")
    logging.info(f"Person numbers to remove: {person_numbers_to_remove}")

    removal_summary: Dict[str, Dict] = {}

    if not component_files:
        logging.warning("No component files provided for removal operation.")
        return JSONResponse(
            content={"message": "No component files provided for removal.", "removal_summary": {}},
            status_code=200
        )

    for component_name, file_name_only in component_files.items():
        file_path = UPLOAD_DIR / parent_name / file_name_only
        logging.info(f"Attempting to process file for removal: {file_path} (Component: {component_name})")

        if not file_path.exists():
            logging.warning(f"File not found for component {component_name}: {file_path}. Skipping removal for this file.")
            removal_summary[component_name] = {"status": "skipped", "reason": "File not found", "rows_removed": 0}
            continue

        try:
            df = pd.read_excel(file_path, header=1)
            logging.info(f"Successfully read file for removal: {file_path}")

            if 'PersonNumber' not in df.columns:
                logging.warning(f"No 'PersonNumber' column found in {file_path}. Skipping removal for this file.")
                removal_summary[component_name] = {"status": "skipped", "reason": "No 'PersonNumber' column", "rows_removed": 0}
                continue

            # Convert PersonNumber column to string for consistent comparison
            df['PersonNumber'] = df['PersonNumber'].astype(str)

            # Get the initial number of rows
            initial_rows = len(df)

            # Filter out rows where 'PersonNumber' is in the 'person_numbers_to_remove' list
            # The ~ operator negates the boolean Series, keeping rows NOT in the set
            df_cleaned = df[~df['PersonNumber'].isin(person_numbers_to_remove)].reset_index(drop=True)

            # Calculate the number of rows removed
            rows_removed = initial_rows - len(df_cleaned)

            if rows_removed > 0:
                # Save the modified DataFrame back to the original file
                df_cleaned.to_excel(file_path, index=False)
                logging.info(f"Removed {rows_removed} rows from {file_path}. File updated successfully.")
                removal_summary[component_name] = {"status": "success", "rows_removed": rows_removed, "reason": None}
            else:
                logging.info(f"No rows removed from {file_path} as no matching person numbers were found.")
                removal_summary[component_name] = {"status": "no_change", "rows_removed": 0, "reason": "No matching person numbers found"}

        except Exception as e:
            logging.error(f"Error processing file {file_path} for removal: {e}", exc_info=True)
            removal_summary[component_name] = {"status": "failed", "reason": str(e), "rows_removed": 0}

    cleaned_summary = clean_dict_for_json(removal_summary)
    return JSONResponse(
        content={
            "message": "Removal operation completed.",
            "removal_summary": cleaned_summary
        },
        status_code=200
    )



@app.get("/api/hdl/download-excel/{filename}")
async def download_excel_file(filename: str):
    """
    Allows downloading a previously generated Excel file by its filename.
    """
    file_path = VALIDATION_RESULTS_DIR / filename
    if not file_path.is_file():
        logging.error(f"File not found for download: {file_path}")
        raise HTTPException(status_code=404, detail="File not found.")

    try:
        return FileResponse(
            path=file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )
    except Exception as e:
        logging.error(f"Error serving file {filename}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to serve file: {e}")



@app.post("/api/hdl/bulk/cross-file/legalEmployer/validate")
async def validate_workterms_file(
    file: UploadFile = File(...),
    hire_action_codes: Optional[str] = Form(None),
    termination_action_codes: Optional[str] = Form(None),
    allowed_le_change_action_codes: Optional[str] = Form(None)
):
    """
    Uploads an Excel file (e.g., WorkTerms.xlsx) and validates consistency
    of 'Legal Employer Name' across all records for a 'Person Number'.
    
    Rules:
    1. A 'Legal Employer Name' change for a 'Person Number' is only valid if it's
       preceded by a defined termination action and followed by a defined hire/rehire action,
       OR if the change itself is triggered by a specified 'allowed_le_change_action_code'
       (e.g., GLOBAL_TRANSFER).
    2. If an active employment period exists, the Legal Employer Name must remain consistent,
       unless an explicitly allowed LE change action occurs.
    3. Termination records should ideally have a preceding hire.
    """
    logger.info(f"Received file: {file.filename}")
    logger.info(f"Received Hire Action Codes from Form: '{hire_action_codes}'")
    logger.info(f"Received Termination Action Codes from Form: '{termination_action_codes}'")
    logger.info(f"Received Allowed LE Change Action Codes from Form: '{allowed_le_change_action_codes}'")


    if not file.filename.lower().endswith(('.xls', '.xlsx', '.csv')):
        logger.warning(f"Invalid file type uploaded: {file.filename}")
        raise HTTPException(status_code=400, detail="Invalid file type. Only Excel (.xls, .xlsx) or CSV files are allowed.")

    try:
        file_content = await file.read()
        
        if file.filename.lower().endswith(('.xls', '.xlsx')):
            df = pd.read_excel(BytesIO(file_content))
            logger.info("File read as Excel.")
        elif file.filename.lower().endswith('.csv'):
            df = pd.read_csv(BytesIO(file_content))
            logger.info("File read as CSV.")
        else:
            logger.error(f"Unsupported file type after initial check: {file.filename}")
            raise HTTPException(status_code=400, detail="Unsupported file type.")

        # Normalize column names to upper case and strip spaces for robust handling
        df.columns = df.columns.str.strip().str.replace(' ', '').str.upper()

        required_columns = ['PERSONNUMBER', 'ACTIONCODE', 'LEGALEMPLOYERNAME', 'EFFECTIVESTARTDATE']
        for col in required_columns:
            if col not in df.columns:
                logger.error(f"Required column '{col}' is missing in the file: {file.filename}. Available columns: {df.columns.tolist()}")
                raise HTTPException(status_code=400, detail=f"Required column '{col}' is missing in the file.")
        
        # Drop rows where critical columns are missing
        df.dropna(subset=required_columns, inplace=True)
        if df.empty:
            logger.warning(f"File {file.filename} is empty or contains no valid data after filtering for required columns.")
            raise HTTPException(status_code=400, detail="File contains no valid data after filtering for required columns or all rows have missing critical data.")

        logger.info(f"Columns in file (after normalization): {df.columns.tolist()}")

        # Ensure relevant columns are correctly typed and cleaned
        df['ACTIONCODE'] = df['ACTIONCODE'].astype(str).str.strip().str.upper()
        df['LEGALEMPLOYERNAME'] = df['LEGALEMPLOYERNAME'].astype(str).str.strip()
        df['PERSONNUMBER'] = df['PERSONNUMBER'].astype(str).str.strip()
        df['EFFECTIVESTARTDATE'] = pd.to_datetime(df['EFFECTIVESTARTDATE'])

        hire_actions_list = []
        if hire_action_codes:
            hire_actions_list = [code.strip().upper() for code in hire_action_codes.split(',') if code.strip()]
        logger.info(f"Effective Hire Action Codes for validation: {hire_actions_list}")
        
        termination_actions_list = [] 
        if termination_action_codes:
            termination_actions_list = [code.strip().upper() for code in termination_action_codes.split(',') if code.strip()]
        logger.info(f"Effective Termination Action Codes for validation: {termination_actions_list}")
        allowed_le_change_actions_list = []
        if allowed_le_change_action_codes:
            allowed_le_change_actions_list = [code.strip().upper() for code in allowed_le_change_action_codes.split(',') if code.strip()]
        logger.info(f"Effective Allowed LE Change Action Codes: {allowed_le_change_actions_list}")

        inconsistent_records = []

        for person_number, group in df.groupby('PERSONNUMBER'):
            logger.debug(f"Processing PersonNumber: {person_number}")
            
            # Sort by EffectiveStartDate to ensure chronological processing
            group = group.sort_values(by='EFFECTIVESTARTDATE').reset_index(drop=True)

            current_le = None
            last_action_was_termination = False # Flag to indicate if the immediate previous action was a termination

            for i, row in group.iterrows():
                action_code = row['ACTIONCODE']
                legal_employer_name = row['LEGALEMPLOYERNAME']
                effective_start_date = row['EFFECTIVESTARTDATE']
                
                # Special handling for null/empty LegalEmployerName
                if pd.isna(legal_employer_name) or legal_employer_name == '':
                    inconsistent_records.append({
                        'PersonNumber': person_number,
                        'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                        'ActionCode': action_code,
                        'LegalEmployerName': legal_employer_name,
                        'Scenario': 'Missing Legal Employer Name',
                        'Status': 'Legal Employer Name is missing or null for this record.'
                    })
                    logger.warning(f"Person {person_number} (Effectivity {effective_start_date}): Missing Legal Employer Name.")
                    continue # Skip further checks for this row as LE is invalid

                logger.debug(f"Person {person_number} (Effective {effective_start_date}) - Action: {action_code}, LE: '{legal_employer_name}', Current LE state: '{current_le}', Last Action was Term: {last_action_was_termination}")


                if action_code in hire_actions_list:
                    # Case 1: Hire/Rehire action
                    if current_le is not None:
                        if not last_action_was_termination:
                            pass 
                        if current_le != legal_employer_name:
                            if not last_action_was_termination:
                                inconsistent_records.append({
                                    'PersonNumber': person_number,
                                    'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                                    'ActionCode': action_code,
                                    'LegalEmployerName': legal_employer_name,
                                    'PreviousLegalEmployer': current_le,
                                    'Scenario': 'Legal Employer Change with HIRE/REHIRE without prior Termination',
                                    'Status': f"Legal Employer changed to '{legal_employer_name}' with '{action_code}' but previous employment period was not terminated. Previous LE was '{current_le}'."
                                })
                                logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): LE changed with '{action_code}' but no prior termination.")
                            else:
                                logger.debug(f"Person {person_number} (Effectivity {effective_start_date}): Valid REHIRE with new LE '{legal_employer_name}' after termination.")
                    
                    current_le = legal_employer_name 
                    last_action_was_termination = False 

                elif action_code in termination_actions_list:
                    if current_le is None:
                        inconsistent_records.append({
                            'PersonNumber': person_number,
                            'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                            'ActionCode': action_code,
                            'LegalEmployerName': legal_employer_name,
                            'Scenario': 'Termination without preceding active employment',
                            'Status': 'Termination record found without an active prior hire record for this person.'
                        })
                        logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): Termination without prior active employment.")
                    if current_le is not None and current_le != legal_employer_name:
                        inconsistent_records.append({
                            'PersonNumber': person_number,
                            'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                            'ActionCode': action_code,
                            'LegalEmployerName': legal_employer_name,
                            'PreviousLegalEmployer': current_le,
                            'Scenario': 'Legal Employer Name Mismatch at Termination',
                            'Status': f"Legal Employer '{legal_employer_name}' at termination does not match last active LE '{current_le}'."
                        })
                        logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): LE mismatch at termination.")

                    current_le = None 
                    last_action_was_termination = True 
                    logger.debug(f"Person {person_number} (Effectivity {effective_start_date}): TERMINATION '{action_code}'. Active period ended.")

                else: 
                    if current_le is None:
                        # Action without an active employment period (e.g., after a termination but before a rehire)
                        # This could be legitimate if it's a "post-employment" record, or an error.
                        # It depends on how your data is structured. For now, flag if it's not a hire/term.
                        inconsistent_records.append({
                            'PersonNumber': person_number,
                            'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                            'ActionCode': action_code,
                            'LegalEmployerName': legal_employer_name,
                            'Scenario': 'Action without active employment period',
                            'Status': f"Record found with non-hire/termination action '{action_code}' and LE '{legal_employer_name}' outside an active employment period."
                        })
                        logger.info(f"Inconsistency for {person_number} (Effectivity {effective_start_date}): Action '{action_code}' found without active employment.")
                    else:
                        # An active employment period exists. Check if LE changed.
                        if current_le != legal_employer_name:
                            # LE changed, but it's not a hire/rehire or termination.
                            # Is this an allowed LE change action (e.g., GLOBAL_TRANSFER)?
                            if action_code in allowed_le_change_actions_list:
                                logger.info(f"Person {person_number} (Effectivity {effective_start_date}): Valid LE change from '{current_le}' to '{legal_employer_name}' via allowed action '{action_code}'.")
                                current_le = legal_employer_name # Update current LE as this is an allowed change
                            else:
                                # This is an inconsistency: LE changed mid-employment without a proper trigger.
                                inconsistent_records.append({
                                    'PersonNumber': person_number,
                                    'EffectiveStartDate': effective_start_date.strftime('%Y-%m-%d'),
                                    'ActionCode': action_code,
                                    'LegalEmployerName': legal_employer_name,
                                    'PreviousLegalEmployer': current_le,
                                    'Scenario': 'Legal Employer Name changed mid-employment without proper action',
                                    'Status': f"Legal Employer Name changed from '{current_le}' to '{legal_employer_name}' with action '{action_code}'. This action is not a hire, termination, or an explicitly allowed LE change action."
                                })
                                logger.info(f"Inconsistency found for {person_number} (Effectivity {effective_start_date}): LE changed from '{current_le}' to '{legal_employer_name}' with invalid action '{action_code}'.")
                                current_le = legal_employer_name # Still update for subsequent checks

                    last_action_was_termination = False # Reset flag

        # --- Post-processing checks for hire/termination alignment (your original scenario 1 & 2) ---
        # These checks might duplicate some of what the chronological loop found, but can catch
        # broader discrepancies in the dataset where a hire-termination pair exists.
        
        # We need a unique list of inconsistencies, converting dicts to hashable tuples
        unique_inconsistent_records_final = []
        seen_inconsistencies = set()
        
        # Add the inconsistencies found by the chronological scan
        for rec in inconsistent_records:
            rec_tuple = tuple(sorted(rec.items()))
            if rec_tuple not in seen_inconsistencies:
                unique_inconsistent_records_final.append(rec)
                seen_inconsistencies.add(rec_tuple)

        # Now, perform the original first-hire vs first-termination check
        for person_number, group in df.groupby('PERSONNUMBER'):
            hire_rows_sorted = group[group['ACTIONCODE'].isin(hire_actions_list)].sort_values(by='EFFECTIVESTARTDATE')
            termination_rows_sorted = group[group['ACTIONCODE'].isin(termination_actions_list)].sort_values(by='EFFECTIVESTARTDATE')

            if not hire_rows_sorted.empty and not termination_rows_sorted.empty:
                first_hire_le = hire_rows_sorted['LEGALEMPLOYERNAME'].iloc[0]
                first_termination_le = termination_rows_sorted['LEGALEMPLOYERNAME'].iloc[0]
                first_hire_date = hire_rows_sorted['EFFECTIVESTARTDATE'].iloc[0]
                first_termination_date = termination_rows_sorted['EFFECTIVESTARTDATE'].iloc[0]

                # Only check if first hire occurs BEFORE first termination
                if first_hire_date < first_termination_date:
                    if first_hire_le != first_termination_le:
                        rec_data = {
                            'PersonNumber': person_number,
                            'Scenario': 'First Hire vs First Termination LE Mismatch',
                            'FirstHireLegalEmployer': first_hire_le,
                            'FirstTerminationLegalEmployer': first_termination_le,
                            'Status': 'Legal Employer of initial hire does not match Legal Employer of first termination.'
                        }
                        rec_tuple = tuple(sorted(rec_data.items()))
                        if rec_tuple not in seen_inconsistencies:
                            unique_inconsistent_records_final.append(rec_data)
                            seen_inconsistencies.add(rec_tuple)
                            logger.info(f"Inconsistency found for {person_number} (First Hire vs First Termination LE): '{first_hire_le}' != '{first_termination_le}'")
            elif not termination_rows_sorted.empty and hire_rows_sorted.empty:
                rec_data = {
                    'PersonNumber': person_number,
                    'Scenario': 'Termination without any Hire',
                    'Status': f"One or more {', '.join(termination_actions_list)} record(s) found but no {', '.join(hire_actions_list)} record(s) exist for this person."
                }
                rec_tuple = tuple(sorted(rec_data.items()))
                if rec_tuple not in seen_inconsistencies:
                    unique_inconsistent_records_final.append(rec_data)
                    seen_inconsistencies.add(rec_tuple)
                    logger.info(f"Inconsistency for {person_number} (Termination without Hire).")
        
        # Final response
        if unique_inconsistent_records_final:
            logger.info(f"Validation complete for {file.filename}. Inconsistencies found for {len(unique_inconsistent_records_final)} unique records.")
            return JSONResponse(
                content={
                    "message": "Validation complete. Inconsistencies found.",
                    "inconsistent_records": unique_inconsistent_records_final
                },
                status_code=200
            )
        else:
            logger.info(f"Validation complete for {file.filename}. All 'Legal Employer Name' records are consistent as per rules.")
            return JSONResponse(
                content={
                    "message": "Validation complete. All 'Legal Employer Name' records are consistent as per rules."
                },
                status_code=200
            )

    except KeyError as ke:
        logger.error(f"KeyError: Required column missing - {str(ke)} in file {file.filename}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"Required column missing in the file: {str(ke)}. Please ensure 'PersonNumber', 'ActionCode', 'LegalEmployerName', and 'EffectiveStartDate' columns exist (case-insensitive).")
    except pd.errors.EmptyDataError:
        logger.error(f"Pandas EmptyDataError: The file {file.filename} is empty or malformed.")
        raise HTTPException(status_code=400, detail="The uploaded file is empty or contains no data.")
    except pd.errors.ParserError:
        logger.error(f"Pandas ParserError: Could not parse the file {file.filename}. It might be malformed or not in the expected format.", exc_info=True)
        raise HTTPException(status_code=400, detail="Could not parse the file. Please ensure it is a valid Excel or CSV file.")
    except Exception as e:
        logger.critical(f"An unhandled error occurred during file processing or validation for {file.filename}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"An internal server error occurred during file processing or validation: {str(e)}")


BASE_DIR = Path(__file__).resolve().parent
VALIDATION_RESULTS_DIR = BASE_DIR / "validation_results"
BUNDLE_DEPOT_ZONE = BASE_DIR / "BUNDLE_DEPOT_ZONE"

# Create the funky folder if it doesn't exist
BUNDLE_DEPOT_ZONE.mkdir(parents=True, exist_ok=True)


class FileBundleRequest(BaseModel):
    """
    Pydantic model for the request body, expecting a list of filenames to bundle.
    """
    files: List[str]

@app.post("/api/hdl/download-bundle")
async def bundle_dat_files_in_order(payload: FileBundleRequest):
    """
    Bundles .dat files by dynamically parsing the 'W{integer}' value from the
    'SourceSystemId' column of *each data record*.
    All records identified as 'W1' or 'ungrouped' will be combined into a single 'W1' bundle.
    Other unique W{integer} (e.g., W2, W3) will result in separate bundled .dat files.
    Each bundle will include the original METADATA header from each component file
    before its corresponding row data, with a line break between components.
    The bundle filename will include the W-group, a list of original component names (without _passed_data_timestamp), and a timestamp.
    Returns a list of URLs for the generated bundle files.
    """
    logger.info(f"Received bundle request for files: {payload.files}")
    try:
        if not payload.files:
            logger.warning("No filenames provided in the payload.")
            raise HTTPException(status_code=400, detail="No filenames provided.")

        # Initialize bundle_results here
        bundle_results = []
        
        # Recursively find all .dat files in VALIDATION_RESULTS_DIR
        all_dat_paths = {f.name: f for f in VALIDATION_RESULTS_DIR.rglob("*.dat")}
        logger.info(f"Found {len(all_dat_paths)} .dat files in VALIDATION_RESULTS_DIR.")

        # Dictionary to hold individual data records grouped by 'W{integer}'
        # Key: "W1", "W2", etc. Value: List of tuples (data_line_string, original_filename_string)
        grouped_records: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
        
        # Store headers for each file, keyed by filename
        all_file_headers: Dict[str, str] = {}

        # To store the set of original filenames that contribute to each group
        group_source_components: Dict[str, Set[str]] = defaultdict(set)

        # Regex to capture the "W{integer}" pattern from the content.
        content_pattern = re.compile(r"(W\d+)", re.IGNORECASE)
        logger.debug(f"Using content regex pattern: {content_pattern.pattern}")

        # Regex to remove '_passed_data_YYYYMMDD_HHMMSS' from filenames
        filename_cleanup_pattern = re.compile(r'_passed_data_\d{8}_\d{6}')

        for fname in payload.files:
            logger.debug(f"Processing file: {fname}")
            dat_path = all_dat_paths.get(fname)
            if not dat_path:
                logger.warning(f"File '{fname}' not found in validation results directory. Skipping.")
                continue

            try:
                with open(dat_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                    
                    if not lines:
                        logger.warning(f"File '{fname}' is empty. Skipping.")
                        continue

                    # Log raw lines to debug what's actually being read
                    logger.debug(f"File '{fname}': Raw lines[0]: '{lines[0].strip()}'")
                    if len(lines) > 1:
                        logger.debug(f"File '{fname}': Raw lines[1]: '{lines[1].strip()}'")

                    # The actual header line is the one starting with "METADATA|"
                    current_file_metadata_line = lines[0].strip()
                    
                    if not current_file_metadata_line.startswith("METADATA|"):
                        logger.error(f"File '{fname}': First line does not start with 'METADATA|'. Expected header format not found. Treating all lines as ungrouped data.")
                        # If the first line isn't the expected METADATA header,
                        # we can't reliably find SourceSystemId. Group all its data as ungrouped.
                        # These will now go into the 'W1' bundle as per the new logic.
                        for data_line in lines: # All lines are potentially data if header is missing
                            grouped_records["W1"].append((data_line.strip(), fname)) # Group into W1
                            group_source_components["W1"].add(fname)
                        continue # Skip to next file

                    # Store this file's header for later use in bundling
                    all_file_headers[fname] = current_file_metadata_line
                    logger.debug(f"Stored header for '{fname}': '{current_file_metadata_line}'")
                    
                    # Determine SourceSystemId column index for *this specific file's header*
                    # Use the METADATA line itself to get the headers
                    current_headers = [h.strip() for h in current_file_metadata_line.split('|')]
                    current_source_system_id_col_index = -1
                    for i, header in enumerate(current_headers):
                        if "SourceSystemId" in header: # Checks for "SourceSystemId" or "PersonId(SourceSystemId)"
                            current_source_system_id_col_index = i
                            break
                    
                    if current_source_system_id_col_index == -1:
                        logger.warning(f"File '{fname}': 'SourceSystemId' column not found in its METADATA/HEADER line. All data records from this file will be grouped as 'W1'.")
                        # If SourceSystemId is not found in this file's header, all its data lines go to 'W1'
                        for line_num, data_line in enumerate(lines[1:], start=2): # Iterate from the first data line
                            grouped_records["W1"].append((data_line.strip(), fname))
                            group_source_components["W1"].add(fname)
                        continue # Move to next file

                    # Process data lines (starting from the second line, as first is METADATA/HEADER)
                    for line_num, data_line in enumerate(lines[1:], start=2):
                        data_line_stripped = data_line.strip()
                        if not data_line_stripped: # Skip empty data lines
                            continue

                        data_fields = data_line_stripped.split('|')
                        
                        if len(data_fields) > current_source_system_id_col_index:
                            source_system_id_value = data_fields[current_source_system_id_col_index]
                            logger.debug(f"File '{fname}', Line {line_num}: Extracted 'SourceSystemId' value: '{source_system_id_value}'")
                            match = content_pattern.search(source_system_id_value)
                            if match:
                                group_key = match.group(1) # Extract the captured "W{integer}"
                                # If the group key is W1, or if it was originally 'ungrouped', map it to 'W1'
                                if group_key.upper() == "W1":
                                    grouped_records["W1"].append((data_line_stripped, fname))
                                    group_source_components["W1"].add(fname)
                                    logger.debug(f"File '{fname}', Line {line_num}: Found '{group_key}'. Appended to group 'W1'.")
                                else:
                                    grouped_records[group_key].append((data_line_stripped, fname))
                                    group_source_components[group_key].add(fname)
                                    logger.debug(f"File '{fname}', Line {line_num}: Found '{group_key}'. Appended to group '{group_key}'.")
                            else:
                                # If no W{integer} pattern found, append to 'W1' group
                                grouped_records["W1"].append((data_line_stripped, fname))
                                group_source_components["W1"].add(fname)
                                logger.debug(f"File '{fname}', Line {line_num}: No 'W{{integer}}' pattern found in 'SourceSystemId' ('{source_system_id_value}'). Appended to 'W1'.")
                        else:
                            # If SourceSystemId column value is not accessible, append to 'W1' group
                            grouped_records["W1"].append((data_line_stripped, fname))
                            group_source_components["W1"].add(fname)
                            logger.warning(f"File '{fname}', Line {line_num}: Data line has fewer columns ({len(data_fields)}) than expected for 'SourceSystemId' index ({current_source_system_id_col_index}). Appended to 'W1'.")

            except Exception as file_process_error:
                logger.error(f"Error processing file '{fname}': {file_process_error}", exc_info=True)
                # If an error occurs during processing a file, its records might be lost or misgrouped.
                # For robustness, we could try to append remaining lines to 'W1' group.
                # For now, just log and skip this file's further processing.


        # Final check if any records were grouped at all
        if not grouped_records:
            logger.warning("No data records were successfully grouped from any of the provided files.")
            raise HTTPException(
                status_code=404,
                detail="No data records found or processed from the provided files for bundling."
            )
        
        # Ensure that all necessary headers were captured from the input files.
        # If any file was processed and had data, its header should be in all_file_headers.
        if not all_file_headers:
            logger.critical("No valid METADATA/Header lines could be extracted from any input file. Cannot form valid bundles.")
            raise HTTPException(
                status_code=500,
                detail="Failed to extract essential METADATA/Header information from input files. Cannot proceed with bundling."
            )

        # Log the final state of grouped_records and group_source_components before bundle generation
        logger.info(f"Final grouped records before bundle generation (showing first 50 chars of each record and source file):")
        for k, v in grouped_records.items():
            logger.info(f"   Group '{k}' has {len(v)} records. Example: {[f'{rec[:50]}... (from {src})' for rec, src in v[:min(3, len(v))]]}")
        logger.info(f"Final group source components: {json.dumps({k: list(v) for k, v in group_source_components.items()}, indent=2)}")


        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        logger.info(f"Starting bundle generation for {len(grouped_records)} groups.")

        # Generate bundles for each group
        for group_key, records_with_source_list in grouped_records.items():
            if not records_with_source_list:
                logger.debug(f"Skipping empty group: '{group_key}'.")
                continue # Skip empty groups (e.g., if 'ungrouped' is empty)

            # Build the content for the bundle, adding specific headers and line breaks
            bundle_content_lines = []
            previous_source_fname = None
            
            # Sort records by original filename to ensure components are grouped together within the bundle
            records_with_source_list.sort(key=lambda x: x[1])

            for record_line, source_fname in records_with_source_list:
                # If the source file changes, add a line break and the new component's header
                if previous_source_fname is None or source_fname != previous_source_fname:
                    if previous_source_fname is not None: # Don't add a blank line before the very first header in the bundle
                        bundle_content_lines.append("") # Add a blank line for separation
                        logger.debug(f"Added line break in '{group_key}' bundle: source changed from '{previous_source_fname}' to '{source_fname}'")
                    
                    # Retrieve the header for the current component (source_fname)
                    header_for_current_component = all_file_headers.get(source_fname)
                    if header_for_current_component:
                        bundle_content_lines.append(header_for_current_component)
                        logger.debug(f"Added header for component '{source_fname}' to '{group_key}' bundle.")
                    else:
                        logger.warning(f"Header for source file '{source_fname}' not found in all_file_headers. Skipping header for this block in '{group_key}' bundle.")

                bundle_content_lines.append(record_line)
                previous_source_fname = source_fname
            
            # Construct the full content for the bundle
            combined_content = "\n".join(bundle_content_lines)
            
            # Prepend the required command to the combined content
            final_bundle_content = "SET PURGE_FUTURE_CHANGES N\n\n" + combined_content

            # Construct filename for the bundle based on the group key and component list
            # Get sorted unique component names for this group
            component_names_for_bundle = sorted(list(group_source_components[group_key]))
            # Remove the '_passed_data_timestamp' pattern and '.dat' extension for cleaner filename
            clean_component_names = []
            for name in component_names_for_bundle:
                cleaned_name = filename_cleanup_pattern.sub('', name) # Remove _passed_data_timestamp
                cleaned_name = cleaned_name.replace('.dat', '') # Remove .dat extension
                clean_component_names.append(cleaned_name)
            
            components_string = "_".join(clean_component_names)
            
            # Ensure filename doesn't get too long (optional, but good practice)
            max_filename_len = 200 # Max length for the component string part
            if len(components_string) > max_filename_len:
                components_string = components_string[:max_filename_len] + "_TRUNCATED"
                logger.warning(f"Component string for filename truncated for group '{group_key}'.")

            bundle_filename = f"{group_key}_{components_string}_{timestamp}.dat"
            
            # Define the full path where the bundled file will be saved
            save_path = BUNDLE_DEPOT_ZONE / bundle_filename

            try:
                with open(save_path, "w", encoding="utf-8") as f:
                    f.write(final_bundle_content) # Write the content with the prepended command
                
                logger.info(f"Successfully generated bundle for group '{group_key}': {save_path}")
                
                # Add the bundle information to the results list
                bundle_results.append({
                    "group": group_key,
                    "filename": bundle_filename,
                    "url": f"/bundle_depot_zone/{bundle_filename}"
                })
            except Exception as bundle_write_error:
                logger.error(f"Error writing bundle file for group '{group_key}' at '{save_path}': {bundle_write_error}", exc_info=True)
                bundle_results.append({
                    "group": group_key,
                    "filename": bundle_filename,
                    "url": None,
                    "error": f"Failed to save bundle: {bundle_write_error}"
                })


        # Sort bundle results for consistent output (e.g., W1, W2, then other Ws)
        def sort_key(item):
            group = item['group']
            # 'W1' should come first as it now includes 'ungrouped'
            if group == 'W1':
                return (0, 0) # Highest priority for W1
            match = re.match(r'W(\d+)', group)
            if match:
                try:
                    return (1, int(match.group(1))) # Sorts other W numbers numerically
                except ValueError:
                    return (2, group) # Fallback for non-numeric W groups
            return (2, group) # Fallback for other unexpected group names (shouldn't happen with this logic)

        bundle_results.sort(key=sort_key)
        logger.info(f"Bundling process completed. Generated {len(bundle_results)} bundles.")
        return JSONResponse(content={"bundles": bundle_results}, status_code=200)

    except HTTPException as e:
        logger.error(f"HTTPException during bundling: {e.detail}", exc_info=True)
        raise e
    except Exception as e:
        logger.critical(f"A critical unexpected error occurred during .dat file bundling: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error during bundling: {str(e)}")


# Mount the BUNDLE_DEPOT_ZONE for serving generated bundles
app.mount("/bundle_depot_zone", StaticFiles(directory=BUNDLE_DEPOT_ZONE), name="bundle_depot_zone")

@app.post("/api/hdl/zip-dat")
async def zip_dat_file_by_name(
    fileName: str = Form(...),
    componentName: str = Form(...),
    group: str= Form(...)
):
    """
    ✅ Searches recursively for `fileName` inside bundle folders,
    renames it as `<componentName>.dat`,
    zips it as `<componentName>.zip`,
    and returns the zip file to the frontend.
    """
    try:
        if not fileName.endswith(".dat"):
            raise HTTPException(status_code=400, detail="File must be a .dat file")

        logger.info(f"📂 Looking for {fileName} under {BUNDLE_DEPOT_ZONE}...")
        matches = list(BUNDLE_DEPOT_ZONE.rglob(fileName))

        if not matches:
            logger.warning(f"❌ {fileName} not found anywhere inside bundle folders.")
            raise HTTPException(status_code=404, detail="Original .dat file not found.")
        
        new_fileName = fileName.replace(".dat", f".zip")
        original_path = matches[0]
        new_dat_path = BUNDLE_DEPOT_ZONE / f"{componentName}.dat"
        new_zip_path = BUNDLE_DEPOT_ZONE / f"{new_fileName}"

        logger.info(f"🔁 Copying {original_path.name} → {new_dat_path.name}")
        shutil.copy(original_path, new_dat_path)

        logger.info(f"🗜️ Zipping {new_dat_path.name} → {new_zip_path.name}")
        with zipfile.ZipFile(new_zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(new_dat_path, arcname=new_dat_path.name)

        # Optional: delete renamed .dat after zip to avoid clutter
        # os.remove(new_dat_path)

        logger.info(f"✅ Successfully zipped {componentName}.zip — sending response.")

        return FileResponse(
            new_zip_path,
            media_type="application/zip",
            filename=new_zip_path.name
        )

    except Exception as e:
        logger.error(f"🔥 Zipping failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Zip failed: {str(e)}")

@app.post("/api/hdl/zip-to-base64-by-name")
async def zip_to_base64_from_name(
    fileName: str = Form(...)
):
    zip_path = BUNDLE_DEPOT_ZONE / fileName

    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="ZIP file not found.")

    try:
        with open(zip_path, "rb") as f:
            file_bytes = f.read()

        encoded = base64.b64encode(file_bytes).decode("utf-8")

        return {
            "fileName": fileName,
            "content": encoded
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Encoding failed: {e}")
    

    
class HDLUploadRequest(BaseModel):
    file_name: str


class HDLTriggerRequest(BaseModel):
    customerName: str
    instanceName: str
    contentId: str

    # Optional Oracle fields
    deleteSourceFileFlag: Optional[str] = None
    dataSetName: Optional[str] = None
    loadConcurrentThreads: Optional[int] = None
    importMaxErrorPercentage: Optional[int] = None
    verificationKey: Optional[str] = None
    loadMaxErrorPercentage: Optional[int] = None
    fileEncryption: Optional[str] = None
    fileAction: Optional[str] = None
    importConcurrentThreads: Optional[int] = None

    model_config = ConfigDict(extra="ignore")


def load_oracle_credentials(customer_name: str, instance_name: str):
    load_dotenv(dotenv_path="./.env")  # make sure the path is right

    base_key = f"{customer_name}_{instance_name}".replace(" ", "").upper()
    
    oracle_env = os.getenv(f"{base_key}_ORACLE_URL")
    username = os.getenv(f"{base_key}_ORACLE_USERNAME")
    password = os.getenv(f"{base_key}_ORACLE_PASSWORD")

    if not all([oracle_env, username, password]):
        raise ValueError(f"Missing Oracle credentials for {base_key}")

    return oracle_env, username, password

class OracleUploadRequest(BaseModel):
    content: str
    fileName: str
    contentId: None
    fileEncryption: str
    customerName: str
    instanceName: str

@app.post("/api/hdl/upload-to-oracle")
async def upload_zip_to_oracle(req: OracleUploadRequest):
    try:
        customerName = req.customerName
        instanceName = req.instanceName
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/action/uploadFile"

        # ✅ Build payload strictly per schema
        payload = {
            "fileName": req.fileName,
            "contentId": req.contentId,
            "content": req.content,  # base64-encoded zip
            "fileEncryption": req.fileEncryption or "NONE"
        }

        headers = {"Content-Type": "application/vnd.oracle.adf.action+json"}

        logger.info("Uploading HDL file to Oracle")
        logger.debug("Final Oracle payload: %s", json.dumps(payload, indent=2))

        res = requests.post(url, json=payload, auth=(username, password), headers=headers)
        res.raise_for_status()
        response = {
            "status_code": res.status_code,
            "response_text": res.json(),
            "env_creds": {
                "oracle_env": oracle_env,
                "username": username,
                "password": password,
            }
        }
        return response

    except Exception as e:
        logger.exception("HDL upload failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/hdl/trigger-oracle-job")
async def trigger_hdl_job(req: HDLTriggerRequest):
    try:
        # ✅ Extract values properly from req, not the class
        customerName = req.customerName
        instanceName = req.instanceName
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/action/createFileDataSet"

        # ✅ Build payload strictly according to schema
        payload = {
            "deleteSourceFileFlag": req.deleteSourceFileFlag,
            "dataSetName": req.dataSetName,
            "contentId": req.contentId,
            "loadConcurrentThreads": req.loadConcurrentThreads,
            "importMaxErrorPercentage": req.importMaxErrorPercentage,
            "verificationKey": req.verificationKey,
            "loadMaxErrorPercentage": req.loadMaxErrorPercentage,
            "fileEncryption": req.fileEncryption or "NONE",
            "fileAction": req.fileAction or "IMPORT_AND_LOAD",
            "importConcurrentThreads": req.importConcurrentThreads,
        }

        # 🔒 Remove keys with None (Oracle rejects unknown/empty props)
        payload = {k: v for k, v in payload.items() if v is not None}

        headers = {
            "Content-Type": "application/vnd.oracle.adf.action+json",
            "Accept": "application/json"
        }

        logger.info("Triggering HDL job | DataSet=%s | ContentId=%s", 
                    payload.get("dataSetName"), payload.get("contentId"))
        logger.debug("Final Oracle payload: %s", json.dumps(payload, indent=2))

        res = requests.post(url, auth=(username, password), headers=headers, json=payload)

        if not res.ok:
            logger.warning("Oracle HDL job trigger failed | Status=%s | Response=%s",
                           res.status_code, res.text)
            raise HTTPException(status_code=res.status_code,
                                detail=f"Trigger HDL job failed: {res.text}")

        data = res.json()
        request_id = (
            data.get("RequestId")
            or data.get("requestId")
            or data.get("result", {}).get("RequestId")
            or "UNKNOWN"
        )

        return {
            "message": "Trigger job request submitted successfully.",
            "RequestId": request_id,
        }

    except Exception as e:
        logger.exception("HDL job trigger failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hdl/status/{customerName}/{instanceName}/{request_id}")
def get_status_by_request_id(customerName: str, instanceName: str, request_id: str):
    """
    Retrieves the status of an Oracle HDL data load request by its request ID.

    Args:
        customerName (str): Customer name to resolve Oracle credentials.
        instanceName (str): Instance name (e.g., Prod, Test).
        request_id (str): The ID of the HDL data load request.

    Returns:
        dict: A dictionary containing the full Oracle response.

    Raises:
        HTTPException: If there's an error during the API call or environment variable retrieval.
    """
    try:
        # Resolve environment + credentials from helper
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        if not oracle_env or not username or not password:
            logging.error(f"❌ Missing Oracle credentials for {customerName}/{instanceName}")
            raise HTTPException(
                status_code=500,
                detail=f"Server configuration error: Oracle credentials missing for {customerName}/{instanceName}."
            )

        # Ensure oracle_env does not end with a slash
        oracle_env = oracle_env.rstrip('/')

        # Build Oracle status check URL
        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/{request_id}"
        logging.info(f"🔍 Fetching HDL status from Oracle for RequestId={request_id}, URL={url}")

        headers = {
            "Accept": "application/json"
        }

        # Make the GET request to Oracle
        res = requests.get(url, auth=(username, password), headers=headers)

        if not res.ok:
            logging.error(f"❌ Oracle HDL Status Failed: HTTP {res.status_code} - {res.text}")
            raise HTTPException(status_code=res.status_code, detail=res.text)

        data = res.json()
        logging.info(f"✅ Successfully retrieved HDL status for RequestId={request_id}")

        return {
            "requestId": request_id,
            "oracle_response": data
        }

    except requests.exceptions.ConnectionError as ce:
        logging.error(f"🌐 Connection Error to Oracle: {ce}")
        raise HTTPException(status_code=503, detail=f"Failed to connect to Oracle environment: {ce}")
    except requests.exceptions.Timeout as te:
        logging.error(f"⏳ Timeout while connecting to Oracle: {te}")
        raise HTTPException(status_code=504, detail=f"Request to Oracle timed out: {te}")
    except requests.exceptions.RequestException as req_e:
        logging.error(f"⚠️ Unexpected request error: {req_e}")
        raise HTTPException(status_code=500, detail=f"Unexpected Oracle request error: {req_e}")
    except HTTPException:
        raise
    except Exception as e:
        logging.critical(f"🔥 Unhandled error checking HDL status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to check HDL status due to internal error: {e}")




@app.get("/api/hdl/status/byContentId/{content_id}/{oracle_env}/{username}/{password}")
def get_status_by_content_id(content_id: str):
    try:
        oracle_env = os.getenv("ORACLE_ENV")
        username = os.getenv("ORACLE_USERNAME")
        password = os.getenv("ORACLE_PASSWORD")

        url = f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadSubmissions?q=ContentId={content_id}"
        headers = { "Accept": "application/json" }

        res = requests.get(url, auth=(username, password), headers=headers)

        # ✅ If Oracle says 404 – don't crash, return gracefully
        if res.status_code == 404:
            return {
                "status": "NOT_FOUND",
                "requestId": "UNKNOWN"
            }

        if not res.ok:
            logger.info("❌ Oracle HDL Status Failed:", res.status_code, res.text)
            raise HTTPException(status_code=500, detail="Failed to fetch Oracle job status.")

        data = res.json()
        if not data.get("items"):
            return {
                "status": "NOT_FOUND",
                "requestId": "UNKNOWN"
            }

        job = data["items"][0]

        return {
            "status": job.get("Status") or "UNKNOWN",
            "requestId": job.get("RequestId") or "UNKNOWN",
            "submissionRef": job.get("SubmissionReference") or "UNKNOWN"
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/hdl/errors/{customerName}/{instanceName}/{request_id}")
async def get_oracle_errors(customerName: str, instanceName: str, request_id: str):
    """
    Retrieves the error messages for a given Oracle HDL data load request.

    Args:
        customerName (str): Customer name to resolve Oracle credentials.
        instanceName (str): Instance name (e.g., Prod, Test).
        request_id (str): The ID of the HDL data load request.

    Returns:
        dict: Oracle error messages response.
    """
    try:
        # Resolve environment + credentials
        oracle_env, username, password = load_oracle_credentials(customerName, instanceName)

        if not oracle_env or not username or not password:
            logging.error(f"❌ Missing Oracle credentials for {customerName}/{instanceName}")
            raise HTTPException(
                status_code=500,
                detail=f"Server configuration error: Oracle credentials missing for {customerName}/{instanceName}."
            )

        # Ensure no trailing slash in URL
        oracle_env = oracle_env.rstrip('/')

        # Oracle error messages API endpoint
        url = (
            f"{oracle_env}/hcmRestApi/resources/11.13.18.05/dataLoadDataSets/{request_id}/child/messages"
            "?totalResults=true"
            "&orderBy=DatFileName,FileLine"
            "&fields=DatFileName,BusinessObjectDiscriminator,OriginatingProcessCode,FileLine,"
            "ConcatenatedUserKey,SourceSystemOwner,SourceSystemId,SourceReference001,"
            "MessageTypeCode,MessageText,MessageUserDetails"
            "&onlyData=true"
        )
        logging.info(f"🔍 Fetching Oracle HDL error messages for RequestId={request_id}, URL={url}")

        headers = {"Accept": "application/json"}
        res = requests.get(url, auth=(username, password), headers=headers)

        if not res.ok:
            logging.error(f"❌ Oracle HDL Errors Fetch Failed: HTTP {res.status_code} - {res.text}")
            raise HTTPException(status_code=res.status_code, detail=res.text)

        data = res.json()
        logging.info(f"✅ Successfully retrieved HDL error messages for RequestId={request_id}")

        return {
            "requestId": request_id,
            "oracle_response": data
        }

    except requests.exceptions.ConnectionError as ce:
        logging.error(f"🌐 Connection Error to Oracle: {ce}")
        raise HTTPException(status_code=503, detail=f"Failed to connect to Oracle environment: {ce}")
    except requests.exceptions.Timeout as te:
        logging.error(f"⏳ Timeout while fetching Oracle errors: {te}")
        raise HTTPException(status_code=504, detail=f"Request to Oracle timed out: {te}")
    except requests.exceptions.RequestException as req_e:
        logging.error(f"⚠️ Unexpected Oracle request error: {req_e}")
        raise HTTPException(status_code=500, detail=f"Unexpected Oracle request error: {req_e}")
    except HTTPException:
        raise
    except Exception as e:
        logging.critical(f"🔥 Unhandled error fetching Oracle HDL errors: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error: {e}")


class InstanceModel(BaseModel):
    instanceName: str
    oracleUrl: str
    oracleUsername: str
    oraclePassword: str



class CustomerModel(BaseModel):
    customerName: str
    instances: List[InstanceModel]

DATA_EXCEL_FILE_PATH = "Required_files/Hiearchy_data.xlsx"

@app.post("/api/customers")
def sync_hierarchy_with_customers(customers: List[CustomerModel]):
    
    """
    Synchronizes customer and instance hierarchy with the main Excel file.
    For new Level-1 (customerName) and Level-2 (instanceName) pairs,
    it duplicates Level-1 entries from Hiearchy_data.xlsx as Level-3 entries,
    and subsequent levels (Level-2 to Level-8) from Hiearchy_data.xlsx
    are mapped to Level-4 to Level-10 respectively in the main file.
    """
    try:
        logger.info("Starting hierarchy synchronization process.")

        # --- File Existence Checks ---
        # Explicitly convert to Path object before calling .exists() to ensure type correctness
        if not Path(EXCEL_FILE_PATH).exists():
            logger.error(f"Main hierarchy file not found: {EXCEL_FILE_PATH}")
            raise HTTPException(status_code=404, detail=f"Main hierarchy file not found at {EXCEL_FILE_PATH}")
        if not Path(DATA_EXCEL_FILE_PATH).exists():
            logger.error(f"Hierarchy data source file not found: {DATA_EXCEL_FILE_PATH}")
            raise HTTPException(status_code=404, detail=f"Hierarchy data source file not found at {DATA_EXCEL_FILE_PATH}")

        # --- Load DataFrames ---
        df = pd.read_excel(EXCEL_FILE_PATH)
        dataf = pd.read_excel(DATA_EXCEL_FILE_PATH)
        logger.info("Successfully loaded main hierarchy and data source files.")

        added_pairs = []
        skipped_pairs = []
        all_new_rows_to_add = [] # This list will collect all new rows before concatenating them to df

        # Normalize existing Level-1 and Level-2 pairs from the main hierarchy DataFrame
        # Using a set for efficient lookup
        existing_hierarchy_pairs = set(
            (str(row.get("Level-1", "")).strip(), str(row.get("Level-2", "")).strip())
            for _, row in df.iterrows()
        )
        logger.info(f"Loaded {len(existing_hierarchy_pairs)} existing Level-1/Level-2 pairs.")

        # The full dataf DataFrame will be used for duplication, not just unique Level-1s
        # This ensures all levels from dataf are considered for mapping
        logger.info(f"Using full data source for duplication, containing {len(dataf)} rows.")

        # --- Process Incoming Customer Data ---
        for customer in customers:
            for instance in customer.instances:
                current_l1 = customer.customerName.strip()
                current_l2 = instance.instanceName.strip()
                current_pair = (current_l1, current_l2)

                # Check if the Level-1/Level-2 pair already exists in the main hierarchy
                # Note: This check is for the specific Level-1/Level-2 pair.
                # If a pair exists, no new blank row or duplicated hierarchy will be added for it.
                if current_pair in existing_hierarchy_pairs:
                    skipped_pairs.append(current_pair)
                    logger.debug(f"Skipping existing pair: {current_pair}")
                    continue  # Move to the next instance

                # If it's a new pair, add it to our tracking and prepare for insertion
                existing_hierarchy_pairs.add(current_pair) # Add to set to prevent duplicate processing in same run
                added_pairs.append(current_pair)
                logger.info(f"Found new pair to add: {current_pair}")

                # --- Step 1: Add the initial blank row for the new Level-1/Level-2 pair ---
                # This row represents the new customer/instance entry itself.
                blank_row_data = {col: "" for col in df.columns}
                blank_row_data["Level-1"] = current_l1
                blank_row_data["Level-2"] = current_l2
                all_new_rows_to_add.append(blank_row_data)
                logger.debug(f"Prepared initial blank row for {current_pair}.")

                # --- Step 2: Duplicate full hierarchy data from dataf for this new pair ---
                # Iterate through each row of the hierarchy data source (dataf)
                for _, data_row in dataf.iterrows():
                    # Create a new row, ensuring it has all columns present in the main df
                    duplicated_row_data = {col: "" for col in df.columns}

                    # Populate Template Name and File Name from the source hierarchy data (dataf)
                    duplicated_row_data["Template Name"] = data_row.get("Template Name", "")
                    duplicated_row_data["File Name"] = data_row.get("File Name", "")

                    # Set Level-1 and Level-2 to the newly added customer/instance pair
                    duplicated_row_data["Level-1"] = current_l1
                    duplicated_row_data["Level-2"] = current_l2

                    # Map Level-1 from dataf to Level-3 in the main df, and so on.
                    # Hiearchy_data.xlsx has levels up to Level-8.
                    # Main HDL_BO_Hierarchy_All_Objects_Charlie.xlsx has levels up to Level-10.
                    # So, Level-X from dataf maps to Level-(X+2) in the main df.
                    for i in range(1, 9): # Iterate for Level-1 to Level-8 from dataf
                        source_level_col = f"Level-{i}"
                        target_level_col = f"Level-{i+2}" # Map to Level-3, Level-4, ..., Level-10

                        if source_level_col in data_row and target_level_col in df.columns:
                            duplicated_row_data[target_level_col] = data_row.get(source_level_col, "")

                    # Populate Mandatory_Objects from the source hierarchy data (dataf) if it exists
                    if "Mandatory_Objects" in data_row and "Mandatory_Objects" in df.columns:
                        duplicated_row_data["Mandatory_Objects"] = data_row.get("Mandatory_Objects", False) # Default to False

                    all_new_rows_to_add.append(duplicated_row_data)
                logger.debug(f"Prepared duplicated full hierarchy rows for {current_pair}.")

        # --- Append All New Rows and Save ---
        if all_new_rows_to_add:
            # Concatenate all collected new rows to the main DataFrame
            df = pd.concat([df, pd.DataFrame(all_new_rows_to_add)], ignore_index=True)
            logger.info(f"Appended {len(all_new_rows_to_add)} new rows to the main hierarchy DataFrame.")
        else:
            logger.info("No new rows to add to the main hierarchy DataFrame.")

        # Save the updated DataFrame back to the Excel file
        # Use index=False to prevent writing the DataFrame index as a column in the Excel file
        df.to_excel(EXCEL_FILE_PATH, index=False)
        logger.info(f"Successfully saved updated hierarchy to {EXCEL_FILE_PATH}.")
        # Build env-friendly payload
        env_payload = []
        for cust in customers:
            env_payload.append({
                "customerName": str(cust.customerName),
                "instances": [
                    {
                        "instanceName": str(inst.instanceName),
                        "oracleUrl": str(inst.oracleUrl),
                        "oracleUsername": str(inst.oracleUsername),
                        "oraclePassword": str(inst.oraclePassword),
                    }
                    for inst in cust.instances
                ]
            })

        merge_env_files(customers=env_payload)
        load_dotenv(".env", override=True)

        # --- Return Response ---
        return {
            "message": "Customer hierarchy sync complete.",
            "added": [f"{l1} - {l2}" for l1, l2 in added_pairs],
            "skipped": [f"{l1} - {l2}" for l1, l2 in skipped_pairs],
            "total_added": len(added_pairs),
            "total_skipped": len(skipped_pairs),
            "output_file": str(EXCEL_FILE_PATH)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"An unexpected error occurred during hierarchy sync: {e}")
        raise HTTPException(status_code=500, detail=f"An unexpected error occurred during hierarchy sync: {e}")


class CustomerUpdateModel(BaseModel):
    old_customerName: str
    old_instanceName: str
    new_customerName: str
    new_instanceName: str

@app.put("/api/customers")
def update_customer_instance(update: CustomerUpdateModel):
    try:
        if not EXCEL_FILE_PATH.exists():
            raise HTTPException(status_code=404, detail="Main Excel file not found.")

        df = pd.read_excel(EXCEL_FILE_PATH)
        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty.")

        mask = (
            df["Level-1"].astype(str).str.strip() == update.old_customerName.strip()
        ) & (
            df["Level-2"].astype(str).str.strip() == update.old_instanceName.strip()
        )

        if not mask.any():
            raise HTTPException(status_code=404, detail="Customer/Instance pair not found.")

        df.loc[mask, "Level-1"] = update.new_customerName.strip()
        df.loc[mask, "Level-2"] = update.new_instanceName.strip()

        df.to_excel(EXCEL_FILE_PATH, index=False)
        logger.info(f"Updated ({update.old_customerName}, {update.old_instanceName}) to ({update.new_customerName}, {update.new_instanceName})")

        return {
            "message": "Customer and instance updated successfully.",
            "updated_rows": int(mask.sum())
        }

    except Exception as e:
        logger.error(f"Update failed: {e}")
        raise HTTPException(status_code=500, detail=f"Update failed: {str(e)}")


class CustomerDeleteModel(BaseModel):
    customerName: str
    instanceName: str

@app.delete("/api/customers")
def delete_customer_instance(delete_req: CustomerDeleteModel):
    try:
        if not EXCEL_FILE_PATH.exists():
            raise HTTPException(status_code=404, detail="Main Excel file not found.")

        df = pd.read_excel(EXCEL_FILE_PATH)
        if df.empty:
            raise HTTPException(status_code=400, detail="Excel file is empty.")

        original_count = len(df)
        mask = ~(
            (df["Level-1"].astype(str).str.strip() == delete_req.customerName.strip()) &
            (df["Level-2"].astype(str).str.strip() == delete_req.instanceName.strip())
        )

        new_df = df[mask]
        removed_count = original_count - len(new_df)

        if removed_count == 0:
            raise HTTPException(status_code=404, detail="No matching customer/instance pair found to delete.")

        new_df.to_excel(EXCEL_FILE_PATH, index=False)
        logger.info(f"Deleted {removed_count} rows for customer {delete_req.customerName} and instance {delete_req.instanceName}")

        return {
            "message": "Customer/Instance pair deleted successfully.",
            "deleted_rows": removed_count
        }

    except Exception as e:
        logger.error(f"Delete failed: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")







#Env population
# -------------------- 🚀 Helper Function ---------------------
def format_env_filename(customer: str, instance: str):
    # Remove spaces and special chars
    safe_customer = "".join(e for e in customer if e.isalnum())
    safe_instance = "".join(e for e in instance if e.isalnum())
    return f"{safe_customer}_{safe_instance}.env"

def write_env_file(file_path: Path, env_data: dict):
    with file_path.open("w") as f:
        for key, val in env_data.items():
            f.write(f'{key}="{val}"\n')  # Safely quote values

# -------------------- 🎯 API Endpoint ---------------------
def sanitize(text: str) -> str:
    return "".join(e for e in text if e.isalnum()).upper()

def reload_env():
    load_dotenv(dotenv_path=".env", override=True)

def write_env_file_for_customers(customers: List[CustomerModel], env_path=".env"):
    env_lines = []
    
    for customer in customers:
        customer_key = customer.customerName.replace(" ", "").upper()
        for instance in customer.instances:
            instance_key = instance.instanceName.replace(" ", "").upper()
            prefix = f"{customer_key}_{instance_key}"
            # Remove the last "/" if present in URL
            instance.oracleUrl = instance.oracleUrl.rstrip("/") if instance.oracleUrl else ""
            # Only write if all fields are present
            if instance.oracleUrl and instance.oracleUsername and instance.oraclePassword:
                env_lines.extend([
                    f"{prefix}_ORACLE_URL={instance.oracleUrl.strip()}",
                    f"{prefix}_ORACLE_USERNAME={instance.oracleUsername.strip()}",
                    f"{prefix}_ORACLE_PASSWORD={instance.oraclePassword.strip()}",
                ])
            else:
                print(f"⚠️ Missing credentials for {prefix}. Skipping...")

    # Write to .env file
    env_file = Path(env_path).resolve()
    with env_file.open("w", encoding="utf-8") as f:
        f.write("\n".join(env_lines))

    print(f"✅ .env saved at: {env_file}")
    load_dotenv(dotenv_path=env_file, override=True)

    # Print sample env check
    if env_lines:
        sample_key = env_lines[0].split("=")[0]
        print(f"🔍 Sample env check: {sample_key} = {os.getenv(sample_key)}")


@app.post("/api/save-env")
def save_envs(customers: List[CustomerModel]):
    try:
        env_lines = []

        for customer in customers:
            customer_key = customer.customerName.replace(" ", "").upper()
            for instance in customer.instances:
                instance_key = instance.instanceName.replace(" ", "").upper()
                prefix = f"{customer_key}_{instance_key}"

                # Skip if any critical field is missing
                if not instance.oracleUrl or not instance.oracleUsername or not instance.oraclePassword:
                    continue
                # Remove trailing slash from URL if present
                instance.oracleUrl = instance.oracleUrl.rstrip("/")
                env_lines.append(f"{prefix}_ORACLE_URL=\"{instance.oracleUrl}\"")
                env_lines.append(f"{prefix}_ORACLE_USERNAME=\"{instance.oracleUsername}\"")
                env_lines.append(f"{prefix}_ORACLE_PASSWORD=\"{instance.oraclePassword}\"")

        if not env_lines:
            raise HTTPException(status_code=400, detail="No valid instance data to write to .env")

        env_path = Path(".env")
        with env_path.open("w", encoding="utf-8") as f:
            f.write("\n".join(env_lines) + "\n")

        load_dotenv(dotenv_path=env_path, override=True)

        return {
            "message": "✅ Environment variables saved",
            "lines_written": len(env_lines),
            "env_path": str(env_path.resolve())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"❌ Failed to save .env file: {str(e)}")

@app.get("/api/customers")
def get_customers_from_root_env():
    try:
        env_path = Path(".env")
        if not env_path.exists():
            raise HTTPException(status_code=404, detail=".env file not found at root")

        env_data = dotenv_values(env_path)
        print("🌱 Loaded .env data:", env_data)

        customer_map = {}

        for full_key, value in env_data.items():
            print(f"🔍 Parsing key: {full_key} = {value}")
            if "_ORACLE_" not in full_key:
                continue

            try:
                prefix, field = full_key.split("_ORACLE_", 1)
                field = field.lower()

                parts = prefix.split("_")
                if len(parts) < 2:
                    print(f"⚠️ Skipping malformed key: {full_key}")
                    continue

                customer = "_".join(parts[:-1])
                instance = parts[-1]

                if customer not in customer_map:
                    customer_map[customer] = {}

                if instance not in customer_map[customer]:
                    customer_map[customer][instance] = {
                        "instanceName": instance
                    }

                customer_map[customer][instance][f"oracle{field.capitalize()}"] = value

            except Exception as inner_e:
                print(f"🚨 Error parsing key: {full_key} — {inner_e}")
                continue

        print("✅ Final parsed customer map:", customer_map)

        response_data = []
        for customer, instances in customer_map.items():
            instance_list = list(instances.values())
            response_data.append({
                "customerName": customer,
                "instances": instance_list
            })

        return response_data

    except Exception as e:
        print("❌ Fatal error in get_customers_from_root_env:", e)
        raise HTTPException(status_code=500, detail="Could not load customers from .env")

def read_excel_sheet_with_dynamic_header(file_stream: io.BytesIO, sheet_name: str, header_row_index: int = 0) -> pd.DataFrame:
    """
    Reads a specific sheet from an Excel file stream, inferring the header from a specific row index (0-based).
    Skips rows above the header and uses the specified row as column names.
    It also cleans column names by stripping whitespace and replacing special characters.

    Args:
        file_stream (io.BytesIO): The byte stream of the Excel file.
        sheet_name (str): The name of the sheet to read.
        header_row_index (int): The 0-based index of the row to be used as the header.

    Returns:
        pandas.DataFrame: The loaded and cleaned DataFrame, or an empty DataFrame if an error occurs.
    """
    try:
        logging.info(f"Attempting to read sheet '{sheet_name}' from file stream with header at row {header_row_index}.")
        # Ensure the stream is at the beginning before reading each sheet
        file_stream.seek(0)
        df = pd.read_excel(file_stream, sheet_name=sheet_name, skiprows=header_row_index, header=0, engine='openpyxl')
        
        # Drop any completely empty columns that might result from uneven header rows
        original_cols = df.columns.tolist()
        df = df.dropna(axis=1, how='all')
        if len(df.columns) < len(original_cols):
            logging.warning(f"Dropped {len(original_cols) - len(df.columns)} empty columns from sheet '{sheet_name}'.")

        # Clean column names: strip whitespace, replace special characters
        # This regex replaces non-alphanumeric characters (except underscore) and then spaces with underscores.
        df.columns = df.columns.str.strip().str.replace(r'[^a-zA-Z0-9_]', '', regex=True).str.replace(' ', '_', regex=False)
        logging.info(f"Successfully loaded sheet '{sheet_name}'. Columns: {df.columns.tolist()}")
        return df
    except KeyError:
        logging.error(f"Error: Sheet '{sheet_name}' not found in the uploaded Excel file. Please check the sheet name.")
        return pd.DataFrame()
    except Exception as e:
        logging.error(f"An unexpected error occurred while reading sheet '{sheet_name}' from file stream: {e}", exc_info=True)
        return pd.DataFrame()

def format_date(date_series: pd.Series) -> pd.Series:
    """
    Converts a pandas Series to datetime objects and formats them as 'YYYY-MM-DD' strings.
    Invalid dates are coerced to NaT (Not a Time) and then replaced with empty strings.

    Args:
        date_series (pandas.Series): The series containing date values.

    Returns:
        pandas.Series: The series with dates formatted as 'YYYY-MM-DD' strings, or empty strings.
    """
    if date_series.empty:
        return pd.Series([], dtype='object')
    return pd.to_datetime(date_series, errors='coerce').dt.strftime('%Y-%m-%d').replace({np.nan: ''})

def validate_columns(df: pd.DataFrame, required_columns: list, df_name: str) -> bool:
    """
    Checks if all required columns are present in the DataFrame.
    Logs a warning if any required column is missing.

    Args:
        df (pandas.DataFrame): The DataFrame to validate.
        required_columns (list): A list of column names that must be present.
        df_name (str): The name of the DataFrame for logging purposes.

    Returns:
        bool: True if all required columns are present, False otherwise.
    """
    if df.empty:
        logging.warning(f"DataFrame '{df_name}' is empty. Skipping column validation.")
        return False
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.warning(f"Missing required columns in '{df_name}': {', '.join(missing_columns)}. This might affect conversion accuracy.")
        return False
    return True

# --- Conversion Functions ---

def convert_worker_data(person_df: pd.DataFrame, work_relationship_df: pd.DataFrame) -> pd.DataFrame:
    required_person_cols = ['PersonNumber', 'EarliestHireDate', 'EffectiveEndDate', 'DateOfBirth', 'Country']
    required_work_rel_cols = ['PersonNumber', 'ActionCode', 'ReasonCode', 'DateStart']

    if not validate_columns(person_df, required_person_cols, 'Person') or \
       not validate_columns(work_relationship_df, required_work_rel_cols, 'WorkRelationship'):
        logging.warning("Skipping Worker data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'End Date', 'Hire Date', 'Action Code', 'Reason Code', 'Birth Date', 'Country of Birth', 'Region of Birth', 'Town of Birth', 'Correspondence Language', 'Blood Type'])

    person_df = person_df.rename(columns={'PersonNumber': 'Employee Number'})
    work_relationship_df = work_relationship_df.rename(columns={'PersonNumber': 'Employee Number'})

    hire_work_rel = work_relationship_df[work_relationship_df['ActionCode'] == 'HIRE'].copy()
    if not hire_work_rel.empty:
        hire_work_rel['DateStart'] = pd.to_datetime(hire_work_rel['DateStart'], errors='coerce')
        hire_work_rel = hire_work_rel.sort_values(by=['Employee Number', 'DateStart']).drop_duplicates(subset=['Employee Number'], keep='first')
    else:
        logging.warning("No 'HIRE' records found in WorkRelationship data. Worker 'Action Code' and 'Reason Code' might be incomplete.")

    worker_df = pd.merge(person_df, hire_work_rel[['Employee Number', 'ActionCode', 'ReasonCode']], on='Employee Number', how='left')

    worker_df_converted = pd.DataFrame()
    worker_df_converted['Employee Number'] = worker_df['Employee Number']
    worker_df_converted['Start Date'] = format_date(worker_df['EarliestHireDate'])
    worker_df_converted['End Date'] = format_date(worker_df['EffectiveEndDate'])
    worker_df_converted['Hire Date'] = format_date(worker_df['EarliestHireDate'])
    worker_df_converted['Action Code'] = worker_df['ActionCode'].fillna('HIRE')
    worker_df_converted['Reason Code'] = worker_df['ReasonCode'].fillna('')
    worker_df_converted['Birth Date'] = format_date(worker_df['DateOfBirth'])
    
    worker_df_converted['Country of Birth'] = ''
    worker_df_converted['Region of Birth'] = ''
    worker_df_converted['Town of Birth'] = ''
    worker_df_converted['Correspondence Language'] = ''
    worker_df_converted['Blood Type'] = ''

    return worker_df_converted

def convert_person_name_data(person_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EarliestHireDate', 'LastName', 'FirstName', 'NameSuffix', 'MiddleNames', 'KnownAs', 'PreviousLastName', 'Title', 'Country']
    if not validate_columns(person_df, required_cols, 'Person'):
        logging.warning("Skipping PersonName data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Last Name', 'First Name', 'Suffix', 'Middle Names', 'Known As', 'Previous Last Name', 'Title', 'Country Code', 'Name Type'])

    person_name_df = person_df.copy()
    person_name_df_converted = pd.DataFrame()

    person_name_df_converted['Employee Number'] = person_name_df['PersonNumber']
    person_name_df_converted['Start Date'] = format_date(person_name_df['EarliestHireDate'])
    person_name_df_converted['Last Name'] = person_name_df['LastName']
    person_name_df_converted['First Name'] = person_name_df['FirstName']
    person_name_df_converted['Suffix'] = person_name_df['NameSuffix'].fillna('')
    person_name_df_converted['Middle Names'] = person_name_df['MiddleNames'].fillna('')
    person_name_df_converted['Known As'] = person_name_df['KnownAs'].fillna('')
    person_name_df_converted['Previous Last Name'] = person_name_df['PreviousLastName'].fillna('')
    person_name_df_converted['Title'] = person_name_df['Title'].fillna('')
    person_name_df_converted['Country Code'] = person_name_df['Country'].fillna('')
    person_name_df_converted['Name Type'] = 'GLOBAL'

    return person_name_df_converted

def convert_work_relationship_data(work_relationship_df: pd.DataFrame, assignment_df: pd.DataFrame) -> pd.DataFrame:
    required_work_rel_cols = ['PersonNumber', 'DateStart', 'ActualTerminationDate', 'ActionCode', 'ReasonCode', 'LegalEmployerName', 'WorkerType', 'PrimaryFlag', 'RehireRecommendationFlag']
    required_assignment_cols = ['PersonNumber', 'EffectiveStartDate', 'EffectiveEndDate', 'BusinessUnitShortCode', 'JobCode', 'GradeCode', 'LocationCode', 'AssignmentStatusType', 'AssignmentType', 'HourlySalariedCode', 'ManagerFlag', 'WorkingatHome', 'WorkerPeriodType', 'NormalHours', 'Frequency', 'PrimaryAssignmentFlag', 'DepartmentName', 'UserPersonType']

    if not validate_columns(work_relationship_df, required_work_rel_cols, 'WorkRelationship') or \
       not validate_columns(assignment_df, required_assignment_cols, 'Assignment'):
        logging.warning("Skipping WorkRelationship data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'End Date', 'Effective Sequence', 'Latest Change Flag', 'Action Code', 'Reason Code', 'Legal Employer Name', 'Business Unit Short Code', 'Department Name', 'Job Code', 'Grade Code', 'Location Code', 'Assignment Status Type Code', 'Assignment Type', 'Hourly Salaried Code', 'Manager Flag', 'Work At Home Flag', 'Permanent Temporary', 'Full Part Time', 'Normal Hours', 'Frequency', 'Primary Assignment Flag', 'Worker Type', 'Person Type Code', 'System Person Type', 'Start Date', 'Actual Termination Date', 'Rehire Recommendation Flag', 'Global Transfer Flag', 'PrimaryFlag'])

    work_rel_df = work_relationship_df.rename(columns={'PersonNumber': 'Employee Number'})
    asg_df = assignment_df.rename(columns={'PersonNumber': 'Employee Number'})

    work_rel_df['DateStart'] = pd.to_datetime(work_rel_df['DateStart'], errors='coerce')
    work_rel_df['ActualTerminationDate'] = pd.to_datetime(work_rel_df['ActualTerminationDate'], errors='coerce')
    asg_df['EffectiveStartDate'] = pd.to_datetime(asg_df['EffectiveStartDate'], errors='coerce')
    asg_df['EffectiveEndDate'] = pd.to_datetime(asg_df['EffectiveEndDate'], errors='coerce')

    work_rel_df = work_rel_df.sort_values(by=['Employee Number', 'DateStart'])
    asg_df_sorted = asg_df.sort_values(by=['Employee Number', 'EffectiveStartDate'])

    merged_df = pd.merge_asof(
        work_rel_df,
        asg_df_sorted,
        left_on='DateStart',
        right_on='EffectiveStartDate',
        by='Employee Number',
        direction='backward',
        suffixes=('_x', '_y')
    )

    merged_df['Effective Sequence'] = merged_df.groupby('Employee Number').cumcount() + 1
    merged_df['Latest Change Flag'] = merged_df.groupby('Employee Number')['DateStart'].transform(lambda x: x == x.max()).map({True: 'Y', False: 'N'})

    work_rel_converted = pd.DataFrame()
    work_rel_converted['Employee Number'] = merged_df['Employee Number']
    work_rel_converted['Start Date'] = format_date(merged_df['DateStart'])
    work_rel_converted['End Date'] = format_date(merged_df['ActualTerminationDate'])
    work_rel_converted['Effective Sequence'] = merged_df['Effective Sequence']
    work_rel_converted['Latest Change Flag'] = merged_df['Latest Change Flag']
    work_rel_converted['Action Code'] = merged_df['ActionCode_x'].fillna('')
    work_rel_converted['Reason Code'] = merged_df['ReasonCode_x'].fillna('')
    work_rel_converted['Legal Employer Name'] = merged_df['LegalEmployerName_x'].fillna('')
    work_rel_converted['Business Unit Short Code'] = merged_df['BusinessUnitShortCode'].fillna('')
    work_rel_converted['Department Name'] = merged_df['DepartmentName'].fillna('') 
    work_rel_converted['Job Code'] = merged_df['JobCode'].fillna('')
    work_rel_converted['Grade Code'] = merged_df['GradeCode'].fillna('')
    work_rel_converted['Location Code'] = merged_df['LocationCode'].fillna('')
    work_rel_converted['Assignment Status Type Code'] = merged_df['AssignmentStatusType'].fillna('')
    work_rel_converted['Assignment Type'] = merged_df['AssignmentType'].fillna('')
    work_rel_converted['Hourly Salaried Code'] = merged_df['HourlySalariedCode'].fillna('')
    work_rel_converted['Manager Flag'] = merged_df['ManagerFlag'].fillna('')
    work_rel_converted['Work At Home Flag'] = merged_df['WorkingatHome'].fillna('')
    work_rel_converted['Permanent Temporary'] = merged_df['WorkerPeriodType'].fillna('') # Mapping WorkerPeriodType to Permanent Temporary
    work_rel_converted['Full Part Time'] = '' # No direct mapping in snippet, leaving empty
    work_rel_converted['Normal Hours'] = merged_df['NormalHours'].fillna('')
    work_rel_converted['Frequency'] = merged_df['Frequency'].fillna('')
    work_rel_converted['Primary Assignment Flag'] = merged_df['PrimaryAssignmentFlag_y'].fillna('')
    work_rel_converted['Worker Type'] = merged_df['WorkerType_x'].fillna('')
    work_rel_converted['Person Type Code'] = merged_df.get('UserPersonType', '').fillna('')
    work_rel_converted['System Person Type'] = merged_df.get('SystemPersonType', '').fillna('')
    work_rel_converted['Start Date'] = format_date(merged_df['DateStart'])
    work_rel_converted['Actual Termination Date'] = format_date(merged_df['ActualTerminationDate'])
    work_rel_converted['Rehire Recommendation Flag'] = merged_df['RehireRecommendationFlag'].fillna('')
    work_rel_converted['Global Transfer Flag'] = ''
    work_rel_converted['PrimaryFlag'] = merged_df['PrimaryFlag_x'].fillna('')

    return work_rel_converted

def convert_work_terms_data(work_relationship_df: pd.DataFrame, assignment_df: pd.DataFrame) -> pd.DataFrame:
    # WorkTerms has the same columns as WorkRelationship in the target format
    return convert_work_relationship_data(work_relationship_df, assignment_df)

def convert_assignment_data(assignment_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EffectiveStartDate', 'EffectiveEndDate', 'ActionCode', 'ReasonCode', 'LegalEmployerName', 'BusinessUnitShortCode', 'DepartmentName', 'JobCode', 'GradeCode', 'LocationCode', 'AssignmentStatusType', 'AssignmentType', 'HourlySalariedCode', 'ManagerFlag', 'WorkingatHome', 'WorkerPeriodType', 'NormalHours', 'Frequency', 'PrimaryAssignmentFlag', 'UserPersonType']
    if not validate_columns(assignment_df, required_cols, 'Assignment'):
        logging.warning("Skipping Assignment data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'End Date', 'Effective Sequence', 'Latest Change Flag', 'Action Code', 'Reason Code', 'Legal Employer Name', 'Business Unit Short Code', 'Department Name', 'Job Code', 'Grade Code', 'Location Code', 'Assignment Status Type Code', 'Assignment Type', 'Hourly Salaried Code', 'Manager Flag', 'Work At Home Flag', 'Permanent Temporary', 'Full Part Time', 'Normal Hours', 'Frequency', 'Primary Assignment Flag', 'Worker Type', 'Person Type Code', 'System Person Type'])

    asg_df = assignment_df.rename(columns={'PersonNumber': 'Employee Number'})

    asg_df['EffectiveStartDate'] = pd.to_datetime(asg_df['EffectiveStartDate'], errors='coerce')
    asg_df['EffectiveEndDate'] = pd.to_datetime(asg_df['EffectiveEndDate'], errors='coerce')
    
    asg_df = asg_df.sort_values(by=['Employee Number', 'EffectiveStartDate']).reset_index(drop=True)

    asg_df['Effective Sequence'] = asg_df.groupby('Employee Number').cumcount() + 1
    asg_df['Latest Change Flag'] = asg_df.groupby('Employee Number')['EffectiveStartDate'].transform(lambda x: x == x.max()).map({True: 'Y', False: 'N'})

    asg_converted = pd.DataFrame()
    asg_converted['Employee Number'] = asg_df['Employee Number']
    asg_converted['Start Date'] = format_date(asg_df['EffectiveStartDate'])
    asg_converted['End Date'] = format_date(asg_df['EffectiveEndDate'])
    asg_converted['Effective Sequence'] = asg_df['Effective Sequence']
    asg_converted['Latest Change Flag'] = asg_df['Latest Change Flag']
    asg_converted['Action Code'] = asg_df['ActionCode'].fillna('')
    asg_converted['Reason Code'] = asg_df['ReasonCode'].fillna('')
    asg_converted['Legal Employer Name'] = asg_df['LegalEmployerName'].fillna('')
    asg_converted['Business Unit Short Code'] = asg_df['BusinessUnitShortCode'].fillna('')
    asg_converted['Department Name'] = asg_df['DepartmentName'].fillna('')
    asg_converted['Job Code'] = asg_df['JobCode'].fillna('')
    asg_converted['Grade Code'] = asg_df['GradeCode'].fillna('')
    asg_converted['Location Code'] = asg_df['LocationCode'].fillna('')
    asg_converted['Assignment Status Type Code'] = asg_df['AssignmentStatusType'].fillna('')
    asg_converted['Assignment Type'] = asg_df['AssignmentType'].fillna('')
    asg_converted['Hourly Salaried Code'] = asg_df['HourlySalariedCode'].fillna('')
    asg_converted['Manager Flag'] = asg_df['ManagerFlag'].fillna('')
    asg_converted['Work At Home Flag'] = asg_df['WorkingatHome'].fillna('')
    asg_converted['Permanent Temporary'] = asg_df['WorkerPeriodType'].fillna('')
    asg_converted['Full Part Time'] = '' # No direct mapping in snippet, leaving empty
    asg_converted['Normal Hours'] = asg_df['NormalHours'].fillna('')
    asg_converted['Frequency'] = asg_df['Frequency'].fillna('')
    asg_converted['Primary Assignment Flag'] = asg_df['PrimaryAssignmentFlag'].fillna('')
    asg_converted['Worker Type'] = asg_df['WorkerType'].fillna('') # Corrected to 'WorkerType' from HDL header
    asg_converted['Person Type Code'] = asg_df.get('UserPersonType', '').fillna('')
    asg_converted['System Person Type'] = asg_df.get('SystemPersonType', '').fillna('')

    return asg_converted

def convert_contract_data(assignment_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EffectiveStartDate', 'ActionCode']
    if not validate_columns(assignment_df, required_cols, 'Assignment'):
        logging.warning("Skipping Contract data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Contract Type', 'Duration', 'Duration Units', 'Action Code'])

    contract_df = assignment_df.rename(columns={'PersonNumber': 'Employee Number'})
    
    contract_df['EffectiveStartDate'] = pd.to_datetime(contract_df['EffectiveStartDate'], errors='coerce')
    contract_df = contract_df.sort_values(by=['Employee Number', 'EffectiveStartDate']).reset_index(drop=True)

    contract_converted = pd.DataFrame()
    contract_converted['Employee Number'] = contract_df['Employee Number']
    contract_converted['Start Date'] = format_date(contract_df['EffectiveStartDate'])
    contract_converted['Contract Type'] = ''
    contract_converted['Duration'] = ''
    contract_converted['Duration Units'] = ''
    contract_converted['Action Code'] = contract_df['ActionCode'].fillna('')

    return contract_converted

def convert_national_id_data(person_df: pd.DataFrame, nat_id_multi_df: pd.DataFrame) -> pd.DataFrame:
    nat_id_converted = pd.DataFrame()

    multi_required_cols = ['PersonNumber', 'LegislationCode', 'NationalIdentifierType', 'NationalIdentifierNumber']
    person_required_cols = ['PersonNumber', 'Country', 'NationalIdentifierType', 'NationalIdentifierNumber'] # Updated based on Person HDL Header

    if not nat_id_multi_df.empty and validate_columns(nat_id_multi_df, multi_required_cols, 'Nat. ID Multi'):
        nat_id_df = nat_id_multi_df.copy()
        nat_id_converted['Employee Number'] = nat_id_df['PersonNumber'].fillna('')
        nat_id_converted['Legislation Code'] = nat_id_df['LegislationCode'].fillna('')
        nat_id_converted['National Identifier Type'] = nat_id_df['NationalIdentifierType'].fillna('')
        nat_id_converted['National Identifier Number'] = nat_id_df['NationalIdentifierNumber'].fillna('')
    elif not person_df.empty and validate_columns(person_df, person_required_cols, 'Person'):
        nat_id_df = person_df.rename(columns={'PersonNumber': 'Employee Number'})
        nat_id_converted['Employee Number'] = nat_id_df['Employee Number']
        nat_id_converted['Legislation Code'] = nat_id_df['Country'].fillna('')
        nat_id_converted['National Identifier Type'] = nat_id_df['NationalIdentifierType'].fillna('')
        nat_id_converted['National Identifier Number'] = nat_id_df['NationalIdentifierNumber'].fillna('')
    else:
        logging.warning("Skipping National ID data conversion due to missing critical source columns in both 'Nat. ID Multi' and 'Person' dataframes.")
        return pd.DataFrame(columns=['Employee Number', 'Legislation Code', 'National Identifier Type', 'National Identifier Number'])

    return nat_id_converted

def convert_person_religion_data(person_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'Country']
    if not validate_columns(person_df, required_cols, 'Person'):
        logging.warning("Skipping PersonReligion data creation due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Legislation Code', 'Religion', 'Primary Flag'])

    religion_df = person_df.rename(columns={'PersonNumber': 'Employee Number'})
    religion_converted = pd.DataFrame()
    religion_converted['Employee Number'] = religion_df['Employee Number']
    religion_converted['Legislation Code'] = religion_df['Country'].fillna('US')
    religion_converted['Religion'] = 'Christianity'
    religion_converted['Primary Flag'] = 'Y'

    return religion_converted

def convert_person_address_data(address_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'EffectiveStartDate', 'AddressType', 'Country', 'AddressLine1', 'AddressLine2', 'TownOrCity', 'Region2', 'PostalCode', 'Region1']
    if not validate_columns(address_df, required_cols, 'Address'):
        logging.warning("Skipping PersonAddress data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Address Type', 'Country', 'Address Line 1', 'Address Line 2', 'Town Or City', 'State', 'Zip Code', 'County'])

    address_df = address_df.rename(columns={'PersonNumber': 'Employee Number'})
    address_converted = pd.DataFrame()

    address_converted['Employee Number'] = address_df['Employee Number']
    address_converted['Start Date'] = format_date(address_df['EffectiveStartDate'])
    address_converted['Address Type'] = address_df['AddressType'].fillna('')
    address_converted['Country'] = address_df['Country'].fillna('')
    address_converted['Address Line 1'] = address_df['AddressLine1'].fillna('')
    address_converted['Address Line 2'] = address_df['AddressLine2'].fillna('')
    address_converted['Town Or City'] = address_df['TownOrCity'].fillna('')
    address_converted['State'] = address_df['Region2'].fillna('')
    address_converted['Zip Code'] = address_df['PostalCode'].fillna('')
    address_converted['County'] = address_df['Region1'].fillna('')

    return address_converted

def convert_person_citizenship_data(citizenship_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'DateFrom', 'LegislationCode', 'CitizenshipStatus']
    if not validate_columns(citizenship_df, required_cols, 'Citizenship'):
        logging.warning("Skipping PersonCitizenship data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Country Code', 'Status'])

    citizenship_df = citizenship_df.rename(columns={'PersonNumber': 'Employee Number'})
    citizenship_converted = pd.DataFrame()

    citizenship_converted['Employee Number'] = citizenship_df['Employee Number']
    citizenship_converted['Start Date'] = format_date(citizenship_df['DateFrom'])
    citizenship_converted['Country Code'] = citizenship_df['LegislationCode'].fillna('')
    citizenship_converted['Status'] = citizenship_df['CitizenshipStatus'].fillna('')

    return citizenship_converted

def convert_person_email_data(email_df: pd.DataFrame) -> pd.DataFrame:
    required_cols = ['PersonNumber', 'DateFrom', 'EmailAddress', 'PrimaryFlag', 'EmailType']
    if not validate_columns(email_df, required_cols, 'Email'):
        logging.warning("Skipping PersonEmail data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start date', 'Email Address', 'Primary Flag', 'Email Type'])

    email_df = email_df.rename(columns={'PersonNumber': 'Employee Number'})
    email_converted = pd.DataFrame()

    email_converted['Employee Number'] = email_df['Employee Number']
    email_converted['Start date'] = format_date(email_df['DateFrom'])
    email_converted['Email Address'] = email_df['EmailAddress'].fillna('')
    email_converted['Primary Flag'] = email_df['PrimaryFlag'].fillna('')
    email_converted['Email Type'] = email_df['EmailType'].fillna('')

    return email_converted

def convert_person_phone_data(phone_df: pd.DataFrame) -> pd.DataFrame:
    """
    Converts data to the 'PersonPhone.csv' format.
    """
    required_cols = ['PersonNumber', 'DateFrom', 'PhoneType', 'LegislationCode', 'AreaCode', 'PhoneNumber', 'PrimaryFlag', 'DateTo']
    if not validate_columns(phone_df, required_cols, 'Phone'):
        logging.warning("Skipping PersonPhone data conversion due to missing critical source columns.")
        return pd.DataFrame(columns=['Employee Number', 'Start Date', 'Phone Type', 'Legislation Code', 'Area Code', 'Phone Number', 'Primary Flag', 'End Date'])

    phone_df = phone_df.rename(columns={'PersonNumber': 'Employee Number'})
    phone_converted = pd.DataFrame()

    phone_converted['Employee Number'] = phone_df['Employee Number']
    phone_converted['Start Date'] = format_date(phone_df['DateFrom'])
    phone_converted['Phone Type'] = phone_df['PhoneType'].fillna('')
    phone_converted['Legislation Code'] = phone_df['LegislationCode'].fillna('')
    phone_converted['Area Code'] = phone_df['AreaCode'].fillna('')
    phone_converted['Phone Number'] = phone_df['PhoneNumber'].fillna('')
    phone_converted['Primary Flag'] = phone_df['PrimaryFlag'].fillna('')
    phone_converted['End Date'] = format_date(phone_df['DateTo'])

    return phone_converted

def convert_person_ethnicity_data(person_df: pd.DataFrame, multi_diversity_df: pd.DataFrame) -> pd.DataFrame:
    ethnicity_converted = pd.DataFrame()

    multi_diversity_required_cols = ['EMPLID', 'LEGISLATIONCODE', 'ETHNICITY', 'PRIMARY_FLAG2']
    person_required_cols = ['PersonNumber', 'Country', 'PER_ETHNICITY', 'PrimaryFlag1'] # Updated to PrimaryFlag1

    if not multi_diversity_df.empty and validate_columns(multi_diversity_df, multi_diversity_required_cols, 'MultiDiversity'):
        ethnicity_df = multi_diversity_df.copy()
        ethnicity_converted['PersonNumber'] = ethnicity_df['EMPLID'].fillna('')
        ethnicity_converted['LegislationCode'] = ethnicity_df['LEGISLATIONCODE'].fillna('')
        ethnicity_converted['Ethnicity'] = ethnicity_df['ETHNICITY'].fillna('')
        ethnicity_converted['PrimaryFlag'] = ethnicity_df['PRIMARY_FLAG2'].fillna('')
    elif not person_df.empty and validate_columns(person_df, person_required_cols, 'Person'):
        ethnicity_df = person_df.rename(columns={'PersonNumber': 'PersonNumber'})
        ethnicity_converted['PersonNumber'] = ethnicity_df['PersonNumber']
        ethnicity_converted['LegislationCode'] = ethnicity_df['Country'].fillna('')
        ethnicity_converted['Ethnicity'] = ethnicity_df['PER_ETHNICITY'].fillna('')
        ethnicity_converted['PrimaryFlag'] = person_df.get('PrimaryFlag1', 'Y').fillna('Y') # Use .get for robustness, default to 'Y'
    else:
        logging.warning("Skipping PersonEthnicity data conversion due to missing critical source columns in both 'MultiDiversity' and 'Person' dataframes.")
        return pd.DataFrame(columns=['PersonNumber', 'LegislationCode', 'Ethnicity', 'PrimaryFlag'])

    return ethnicity_converted

# --- FastAPI Endpoint ---

@app.post("/convert-excel/", summary="Convert Excel Employee Data", response_description="Converted Excel file")
async def convert_excel_endpoint(excel_file: UploadFile = File(..., description="The Excel file (EmployeeWithHistory-Template.xlsx) to convert.")):
    """
    Receives an Excel file, processes its sheets, converts the data format,
    and returns a new Excel file with the converted data.
    """
    if not excel_file.filename.endswith('.xlsx'):
        logging.error(f"Invalid file type uploaded: {excel_file.filename}")
        raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx files are allowed.")

    try:
        # Read the incoming Excel file content into a BytesIO object
        file_content = await excel_file.read()
        file_stream = io.BytesIO(file_content)

        # Configuration for source Excel sheets (header rows are fixed based on template)
        sheets_config = {
            'Person': {'sheet_name': 'Person', 'header_row_index': 6}, 
            'WorkRelationship': {'sheet_name': 'WorkRelationship', 'header_row_index': 7}, 
            'Assignment': {'sheet_name': 'Assignment', 'header_row_index': 6}, 
            'MultiDiversity': {'sheet_name': 'MultiDiversity', 'header_row_index': 0}, 
            'Address': {'sheet_name': 'Address', 'header_row_index': 6}, 
            'Nat. ID Multi': {'sheet_name': 'Nat. ID Multi', 'header_row_index': 5}, 
            'Phone': {'sheet_name': 'Phone', 'header_row_index': 5}, 
            'Email': {'sheet_name': 'Email', 'header_row_index': 5}, 
            'Citizenship': {'sheet_name': 'Citizenship', 'header_row_index': 4}, 
        }

        source_dfs = {}
        for name, sheet_config in sheets_config.items():
            source_dfs[name] = read_excel_sheet_with_dynamic_header(
                file_stream,
                sheet_config['sheet_name'],
                sheet_config['header_row_index']
            )

        output_dfs = {}
        logging.info("Performing data transformations via API.")

        output_dfs['Worker'] = convert_worker_data(
            source_dfs.get('Person', pd.DataFrame()),
            source_dfs.get('WorkRelationship', pd.DataFrame())
        )

        output_dfs['PersonName'] = convert_person_name_data(
            source_dfs.get('Person', pd.DataFrame())
        )

        output_dfs['WorkRelationship'] = convert_work_relationship_data(
            source_dfs.get('WorkRelationship', pd.DataFrame()),
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['WorkTerms'] = convert_work_terms_data(
            source_dfs.get('WorkRelationship', pd.DataFrame()),
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['Assignment'] = convert_assignment_data(
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['Contract'] = convert_contract_data(
            source_dfs.get('Assignment', pd.DataFrame())
        )

        output_dfs['PersonNationalIdentifier'] = convert_national_id_data(
            source_dfs.get('Person', pd.DataFrame()),
            source_dfs.get('Nat. ID Multi', pd.DataFrame())
        )

        output_dfs['PersonReligion'] = convert_person_religion_data(
            source_dfs.get('Person', pd.DataFrame())
        )

        output_dfs['PersonAddress'] = convert_person_address_data(
            source_dfs.get('Address', pd.DataFrame())
        )

        output_dfs['PersonCitizenship'] = convert_person_citizenship_data(
            source_dfs.get('Citizenship', pd.DataFrame())
        )

        output_dfs['PersonEmail'] = convert_person_email_data(
            source_dfs.get('Email', pd.DataFrame())
        )
        
        output_dfs['PersonPhone'] = convert_person_phone_data(
            source_dfs.get('Phone', pd.DataFrame())
        )

        output_dfs['PersonEthnicity'] = convert_person_ethnicity_data(
            source_dfs.get('Person', pd.DataFrame()),
            source_dfs.get('MultiDiversity', pd.DataFrame())
        )

        # Write converted dataframes to a BytesIO object
        output_stream = io.BytesIO()
        with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
            for sheet_name, df in output_dfs.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logging.info(f"Prepared sheet '{sheet_name}' for output.")
                else:
                    logging.warning(f"Sheet '{sheet_name}' is empty, skipping writing to output Excel.")
        output_stream.seek(0) # Rewind to the beginning of the stream

        logging.info("Excel conversion successful. Sending file back.")
        return StreamingResponse(
            output_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=Converted_Employee_Data.xlsx"}
        )

    except Exception as e:
        logging.error(f"An unexpected error occurred during Excel processing: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal server error during processing: {e}")




#------------- Setup data storage ------------------#
class HDLSetupPayload(BaseModel):
    customerName: str
    instanceName: str
    hireActions: List[str]
    rehireActions: List[str]
    termActions: List[str]
    globalTransferActions: List[str]
    statusTypes: List[str]  # Must be length 6
    assignmentStatusRules: List[Dict[str, str]]

@app.post("/api/hdl/save-setup")
def save_hdl_setup(data: HDLSetupPayload):
    try:
        # Construct directory and filename
        setup_dir = Path("User/setup_files")
        setup_dir.mkdir(parents=True, exist_ok=True)

        # Normalize file name
        filename = f"{data.customerName.replace(' ', '_')}_{data.instanceName.replace(' ', '_')}_setup.json"
        filepath = setup_dir / filename

        # Save JSON
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data.dict(), f, indent=4)

        return {
            "message": "✅ Setup saved successfully",
            "file": str(filepath.resolve())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving setup: {e}")
    

@app.get("/api/hdl/get-setup/{customer_name}/{instance_name}")
def get_hdl_setup(
    customer_name: str,
    instance_name: str 
):
    try:
        setup_dir = Path("User/setup_files")
        filename = f"{customer_name.replace(' ', '_')}_{instance_name.replace(' ', '_')}_setup.json"
        filepath = setup_dir / filename

        if not filepath.exists():
            raise HTTPException(status_code=404, detail="Setup file not found.")

        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)

        return JSONResponse(content=data)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching setup: {e}")


class AttributeMappingPayload(BaseModel):
    customerName: str
    instanceName: str
    componentName: str
    mappedAttributes: Dict[str, str]

@app.post("/api/hdl/save-attribute-mapping")
def save_hdl_attribute_mapping(data: AttributeMappingPayload):
    try:
        # 🔐 Normalize folder and file path
        safe_customer = data.customerName.replace(" ", "_")
        safe_instance = data.instanceName.replace(" ", "_")
        safe_component = data.componentName.replace(" ", "_")

        mapping_dir = Path(f"User/attribute_mappings")
        mapping_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{safe_customer.replace(' ', '_')}_{safe_instance.replace(' ', '_')}_{safe_component.replace(' ', '_')}_attributes.json"
        file_path = mapping_dir / filename

        # 💾 Save the mapping to JSON
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data.mappedAttributes, f, indent=4)

        return {
            "message": "✅ Attribute mapping saved successfully.",
            "file": str(file_path.resolve())
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving attribute mapping: {e}")


@app.get("/api/hdl/get-attribute-mapping/{customer_name}/{instance_name}/{component_name}")
def get_hdl_attribute_mapping(customer_name: str, instance_name: str, component_name: str):
    try:
        safe_customer = customer_name.replace(" ", "_")
        safe_instance = instance_name.replace(" ", "_")
        safe_component = component_name.replace(" ", "_")
        filename = f"{safe_customer.replace(' ', '_')}_{safe_instance.replace(' ', '_')}_{safe_component.replace(' ', '_')}_attributes.json"
        file_path = Path(f"User/attribute_mappings/{filename}")

        if not file_path.exists():
            return {
                "success": 200,
                "create" : "create New"
            }

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        return {
            "customerName": customer_name,
            "instanceName": instance_name,
            "componentName": component_name,
            "mappedAttributes": data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading mapping: {e}")



class LookupDataAPIOracle(BaseModel):
    customerName: str
    instanceName: str

@app.post("/api/hdl/oracle_fetch/lookupdataload")
async def LookupDataLoading(req: LookupDataAPIOracle):
    customerName = req.customerName
    instanceName = req.instanceName

    oracle_env, username, password = [x.strip() for x in load_oracle_credentials(customerName, instanceName)]
    logger.warning(f"oracle credentials are, {oracle_env}, {username}, {password}")
    
    # Define SOAP save zone path
    soap_save_zone = Path(f"{customerName}/{instanceName}/soap_temp_storage")
    soap_save_zone.mkdir(parents=True, exist_ok=True)  # create dirs if they don't exist
    LookupData_Directory = Path(f"Required_files/{customerName}_{instanceName}_LookupData.xlsx")
    LookupData_Directory.parent.mkdir(parents=True, exist_ok=True)

    # SOAP service URL
    SOAP_URL = f"{oracle_env}/xmlpserver/services/ExternalReportWSSService?wsdl"

    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": "",
    }

    # Read from soap_request.xml
    with open("soap_request_Lookup_Data.xml", "r", encoding="utf-8") as file:
        soap_body = file.read().strip()

    # Save a copy of the SOAP request XML to the soap_save_zone
    try:
        saved_file_path = soap_save_zone / "soap_request_Lookup_Data_saved.xml"
        with open(saved_file_path, "w", encoding="utf-8") as f:
            f.write(soap_body)
        logger.info(f"SOAP request saved at: {saved_file_path}")
    except Exception as save_err:
        logger.error(f"Failed to save SOAP XML: {save_err}")

    # Now make the actual SOAP call
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            #log everything
            logger.info(f"Making SOAP request to {SOAP_URL} with user {username}")
            response = await client.post(SOAP_URL, data=soap_body, headers=headers, auth=(username, password))
            response.raise_for_status()
            logger.info(f"SOAP request successful with status code {response.status_code}")
            # log all the response code as well as content
            logger.warning(f"Response content: {response.text[:500]}...")  # Log first 500 chars for brevity
            saved_file_path = soap_save_zone / "soap_response_saved.xml"
            with open(saved_file_path, "w", encoding="utf-8") as f:
                f.write(response.text)
            logger.info(f"SOAP request saved at: {saved_file_path}")

            excel_path = parse_soap_response_to_excel(saved_file_path, LookupData_Directory)
            return {
                "status": "success",
                "results": "Lookup data loaded successfully"
            }
    except httpx.HTTPError as e:
        logger.error(f"SOAP call failed: {str(e)}")
        if e.response:
            logger.error(f"Response content: {e.response.text}")
        raise HTTPException(status_code=500, detail=f"SOAP call failed: {str(e)}")

class mandatoryFieldsReqOracle(BaseModel):
    customerName: str
    instanceName: str

@app.post("/api/hdl/oracle_fetch/mandatoryFields")
async def MandatoryfieldsLoading(req: mandatoryFieldsReqOracle):
    customerName = req.customerName
    instanceName = req.instanceName

    oracle_env, username, password = [x.strip() for x in load_oracle_credentials(customerName, instanceName)]
    logger.warning(f"oracle credentials are, {oracle_env}, {username}, {password}")
    
    # Define SOAP save zone path
    soap_save_zone = Path(f"{customerName}/{instanceName}/soap_temp_storage")
    soap_save_zone.mkdir(parents=True, exist_ok=True)  # create dirs if they don't exist
    Mandatory_field_Directory = Path(f"Required_files/{customerName}_{instanceName}_MandatoryFields.xlsx")
    Mandatory_field_Directory.parent.mkdir(parents=True, exist_ok=True)

    # SOAP service URL
    SOAP_URL = f"{oracle_env}/xmlpserver/services/ExternalReportWSSService?wsdl"

    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
        "SOAPAction": "",
    }

    # Read from soap_request.xml
    with open("soap_request_Mandatory_Fields.xml", "r", encoding="utf-8") as file:
        soap_body = file.read().strip()

    # Save a copy of the SOAP request XML to the soap_save_zone
    try:
        saved_file_path = soap_save_zone / "soap_request_Mandatory_Fields_saved.xml"
        with open(saved_file_path, "w", encoding="utf-8") as f:
            f.write(soap_body)
        logger.info(f"SOAP request saved at: {saved_file_path}")
    except Exception as save_err:
        logger.error(f"Failed to save SOAP XML: {save_err}")

    # Now make the actual SOAP call
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            #log everything
            logger.info(f"Making SOAP request to {SOAP_URL} with user {username}")
            response = await client.post(SOAP_URL, data=soap_body, headers=headers, auth=(username, password))
            response.raise_for_status()
            logger.info(f"SOAP request successful with status code {response.status_code}")
            # log all the response code as well as content
            logger.warning(f"Response content: {response.text[:500]}...")  # Log first 500 chars for brevity
            saved_file_path = soap_save_zone / "soap_response_saved.xml"
            with open(saved_file_path, "w", encoding="utf-8") as f:
                f.write(response.text)
            logger.info(f"SOAP request saved at: {saved_file_path}")

            excel_path = parse_soap_response_to_excel(saved_file_path, Mandatory_field_Directory)
            return {
                "status": "success",
                "results": "Mandatory Fields loaded successfully"
            }
    except httpx.HTTPError as e:
        logger.error(f"SOAP call failed: {str(e)}")
        if e.response:
            logger.error(f"Response content: {e.response.text}")
        raise HTTPException(status_code=500, detail=f"SOAP call failed: {str(e)}")

# Person report fetching 
# query parameter as customername and instance name 
@app.get("/api/load/delta_report")
async def Report_Loading(customerName: str, instanceName: str, componentName: str):
    
    # Load credentials
    try:
        oracle_env, username, password = [x.strip() for x in load_oracle_credentials(customerName, instanceName)]
    except Exception as e:
        logger.error(f"Credential load error: {e}")
        return {"status": "error", "message": "Failed to load Oracle credentials"}
    if componentName != "Assignment":
        return {"status": "sucess", "message": "Delta load only available for Assignment component"}
    # Define Paths
    soap_save_zone = Path(f"{customerName}/{instanceName}/soap_temp_storage")
    soap_save_zone.mkdir(parents=True, exist_ok=True)
    
    # CORRECTED: Changed extension from .xlsx to .csv to match the parser output
    Assignment_report_Directory = Path(f"required_files/{customerName}_{instanceName}_{componentName}_Report.csv")
    Assignment_report_Directory.parent.mkdir(parents=True, exist_ok=True)

    SOAP_URL = f"{oracle_env}/xmlpserver/services/ExternalReportWSSService?wsdl"

    headers = {
        "Content-Type": "application/soap+xml; charset=utf-8",
    }

    # Read SOAP XML Template
    soap_template_path = Path(f"soap_request_{componentName}_Report.xml")
    if not soap_template_path.exists():
        return {"status": "error", "message": f"SOAP Request Template 'soap_request_{componentName}_Report.xml' not found."}

    try:
        with open(soap_template_path, "r", encoding="utf-8") as file:
            soap_body = file.read().strip()
    except Exception as e:
        return {"status": "error", "message": f"Failed to read SOAP template: {e}"}

    # Make SOAP call
    try:
        async with httpx.AsyncClient(timeout=120) as client: # Increased timeout for reports
            logger.info(f"Fetching Delta Report from: {SOAP_URL}")
            
            response = await client.post(
                SOAP_URL, 
                data=soap_body, 
                headers=headers, 
                auth=(username, password)
            )
            
            if response.status_code != 200:
                logger.error(f"SOAP Failed: {response.status_code} - {response.text}")
                return {"status": "error", "message": f"Oracle Error: {response.status_code}"}

            # Save Raw SOAP response (for debugging)
            response_save_path = soap_save_zone / "soap_response_raw.xml"
            with open(response_save_path, "w", encoding="utf-8") as f:
                f.write(response.text)

            # --- CRITICAL STEP ---
            # Parse the XML, extract Base64, decode, and save as CLEAN CSV
            # The parser now ensures the file at Assignment_report_Directory is a valid CSV
            final_path = parse_soap_response_to_csv(str(response_save_path), str(Assignment_report_Directory))

            return {
                "status": "success",
                "message": "Assignment report loaded and converted successfully",
                "file_path": str(final_path)
            }

    except Exception as e:
        logger.exception("Error in Delta Report Loading Endpoint")
        return {
            "status": "error",
            "message": f"Processing failed: {str(e)}",
            "file_path": None
        }
def parse_soap_response_to_excel(xml_path: str, output_excel_path: str = "output.xlsx", customerName: str = "", instanceName: str = "") -> str:
    """
    Parses SOAP response XML, extracts embedded base64 data, 
    CONVERTS it (from CSV or Excel) to a valid .xlsx structure, and saves it.
    """
    ns = {
        'env': 'http://www.w3.org/2003/05/soap-envelope',
        'ns2': 'http://xmlns.oracle.com/oxp/service/PublicReportService'
    }

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Navigate to the reportBytes element
        # Using .// to find it anywhere in the structure to be safe
        report_bytes_elem = root.find('.//ns2:reportBytes', ns)
        
        if report_bytes_elem is None or not report_bytes_elem.text:
            raise ValueError("reportBytes not found or empty in the response.")

        # Decode base64 content
        decoded_bytes = base64.b64decode(report_bytes_elem.text)

        # --- CRITICAL FIX START ---
        # Do not just write bytes. Read them into Pandas to standardize format.
        df = None
        
        # Attempt 1: Try reading as CSV (most likely format for data integration)
        try:
            df = pd.read_csv(io.BytesIO(decoded_bytes))
            logger.info("SOAP content identified as CSV. Converting to Excel...")
        except Exception:
            # Attempt 2: Try reading as Excel (if Oracle actually returned a binary xls/xlsx)
            try:
                df = pd.read_excel(io.BytesIO(decoded_bytes))
                logger.info("SOAP content identified as valid Excel.")
            except Exception as e:
                # If both fail, the data is likely corrupt or HTML
                raise ValueError(f"Decoded data is neither valid CSV nor Excel. First 50 chars: {decoded_bytes[:50]}")

        # Save as a guaranteed clean Excel file using openpyxl engine
        output_path = Path(output_excel_path)
        
        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        df.to_excel(output_path, index=False, engine='openpyxl')
        # --- CRITICAL FIX END ---

        logger.info(f"Assignment report successfully saved to: {output_path}")
        return str(output_path.resolve())

    except Exception as e:
        logger.error(f"Failed to extract/convert Excel from SOAP XML: {e}")
        raise RuntimeError(f"Failed to extract Excel from SOAP XML: {e}")


def parse_soap_response_to_csv(xml_path: str, output_csv_path: str = "output.csv", customerName: str = "", instanceName: str = "") -> str:
    """
    Parses SOAP response XML, extracts embedded base64 data, 
    validates it, and saves it as a CLEAN CSV file.
    """
    ns = {
        'env': 'http://www.w3.org/2003/05/soap-envelope',
        'ns2': 'http://xmlns.oracle.com/oxp/service/PublicReportService'
    }

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Navigate to the reportBytes element
        report_bytes_elem = root.find('.//ns2:reportBytes', ns)
        
        if report_bytes_elem is None or not report_bytes_elem.text:
            raise ValueError("reportBytes not found or empty in the response.")

        # Decode base64 content
        decoded_bytes = base64.b64decode(report_bytes_elem.text)

        # --- LOAD DATA (Validation & Standardization) ---
        df = None
        
        # Attempt 1: Try reading as CSV (Standardize format)
        try:
            df = pd.read_csv(io.BytesIO(decoded_bytes))
            logger.info("SOAP content identified as CSV.")
        except Exception:
            # Attempt 2: Try reading as Excel (In case report was actually xlsx)
            try:
                df = pd.read_excel(io.BytesIO(decoded_bytes))
                logger.info("SOAP content identified as Excel. Converting to CSV...")
            except Exception as e:
                raise ValueError(f"Decoded data is neither valid CSV nor Excel. First 50 chars: {decoded_bytes[:50]}")

        # --- SAVE AS CSV ---
        output_path = Path(output_csv_path)
        
        # Ensure parent directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Write to CSV 
        # index=False removes the pandas row numbers
        # encoding='utf-8-sig' ensures special characters work in Excel if opened manually later
        df.to_csv(output_path, index=False, encoding='utf-8-sig')

        logger.info(f"Assignment report successfully saved to: {output_path}")
        return str(output_path.resolve())

    except Exception as e:
        logger.error(f"Failed to extract/save CSV from SOAP XML: {e}")
        raise RuntimeError(f"Failed to extract CSV from SOAP XML: {e}")


@app.get("/api/lookupdata/available")
def get_available_lookupdata_files(customerName: str = "", instanceName: str = ""):
    try:
        lookupdata_dir = Path("Required_files")
        pattern = f"{customerName}_{instanceName}_LookupData.xlsx" if customerName and instanceName else "*_LookupData.xlsx"
        files = list(lookupdata_dir.glob(pattern))
        file_list = [f.name for f in files]
        return {
            "available": len(file_list) > 0,  # ✅ return availability flag
            "files": file_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")

@app.get("/api/mandatoryfields/available")
def get_available_mandatoryfields_files(customerName: str = "", instanceName: str = ""):
    try:
        mandatoryfields_dir = Path("Required_files")
        pattern = f"{customerName}_{instanceName}_MandatoryFields.xlsx" if customerName and instanceName else "*_MandatoryFields.xlsx"
        files = list(mandatoryfields_dir.glob(pattern))
        file_list = [f.name for f in files]
        return {
            "available": len(file_list) > 0,  # ✅ return availability flag
            "files": file_list
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to list files: {str(e)}")
        


ENV_DIR = Path("Required_files/env_store")
ENV_DIR.mkdir(parents=True, exist_ok=True)


# ✅ Get customers from env files
@app.get("/api/env/customers")
def get_all_env_customers():
    try:
        all_customers = []
        for env_file in ENV_DIR.glob("*.env"):
            env_data = dotenv_values(env_file)
            grouped = {}

            for key, value in env_data.items():
                if not "_" in key:
                    continue
                parts = key.split("_")
                if len(parts) < 3:
                    continue
                customer = parts[0]
                instance = parts[1]
                field = "_".join(parts[2:])

                if customer not in grouped:
                    grouped[customer] = {}
                if instance not in grouped[customer]:
                    grouped[customer][instance] = {"instanceName": instance}

                grouped[customer][instance][field.lower()] = value

            for customer, instances in grouped.items():
                instance_list = list(instances.values())
                all_customers.append({
                    "customerName": customer,
                    "assigned_instances": instance_list
                })

        return {"data": all_customers}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load customers: {str(e)}")


def merge_env_files(customers: List[Dict]):
    try:
        env_path = Path(".env")
        existing_env: dict = {}
        if env_path.exists():
            existing_env = dict(dotenv_values(env_path))

        for customer in customers:
            customer_name = customer["customerName"].strip().upper().replace(" ", "_")
            for instance in customer["instances"]:
                inst_name = instance["instanceName"].strip().upper().replace(" ", "_")
                prefix = f"{customer_name}_{inst_name}"
                existing_env[f"{prefix}_ORACLE_URL"] = instance["oracleUrl"].strip()
                existing_env[f"{prefix}_ORACLE_USERNAME"] = instance["oracleUsername"].strip()
                existing_env[f"{prefix}_ORACLE_PASSWORD"] = instance["oraclePassword"].strip()

        with open(".env", "w") as f:
            f.write("\n".join(f"{k}={v}" for k, v in existing_env.items()))

        return {"message": "All envs merged", "total_added": len(customers)}

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Merge failed: {str(e)}")



ENV_FILE = Path(".env")
# ✅ Delete customer
@app.delete("/api/customers/{customer_name}")
@app.delete("/api/customers/{customer_name}/{instance_name}")
def delete_customer(customer_name: str, instance_name: str = None):
    try:
        
        if not ENV_FILE.exists():
            raise HTTPException(status_code=404, detail="Env file not found")

        lines = ENV_FILE.read_text().splitlines()

        if instance_name:
            # Delete only that specific instance
            key_prefix = f"{customer_name.upper()}_{instance_name.upper()}_"
        else:
            # Delete all instances of that customer
            key_prefix = f"{customer_name.upper()}_"

        # Filter out matching lines
        new_lines = [line for line in lines if not line.startswith(key_prefix)]

        if len(lines) == len(new_lines):
            if instance_name:
                raise HTTPException(
                    status_code=404,
                    detail=f"Customer instance '{customer_name}_{instance_name}' not found"
                )
            else:
                raise HTTPException(
                    status_code=404,
                    detail=f"Customer '{customer_name}' not found"
                )

        # Write back updated file
        ENV_FILE.write_text("\n".join(new_lines) + "\n")

        if instance_name:
            return {"message": f"Customer instance '{customer_name}_{instance_name}' deleted successfully."}
        else:
            return {"message": f"All instances of customer '{customer_name}' deleted successfully."}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Delete failed: {str(e)}")


# ✅ Delete a specific instance from a customer
@app.delete("/api/customers/{customer_name}/instances/{instance_name}")
def delete_instance(customer_name: str, instance_name: str):
    try:
        customer_name = customer_name.upper().replace(" ", "_")
        instance_name = instance_name.upper().replace(" ", "_")
        env_file = ENV_DIR / f"{customer_name}.env"

        if not env_file.exists():
            raise HTTPException(status_code=404, detail="Customer not found")

        lines_to_keep = []
        with open(env_file, "r") as f:
            for line in f:
                if not line.startswith(f"{customer_name}_{instance_name}_"):
                    lines_to_keep.append(line.strip())

        with open(env_file, "w") as f:
            f.write("\n".join(lines_to_keep))

        return {"message": f"Instance '{instance_name}' deleted from customer '{customer_name}'."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Instance delete failed: {str(e)}")

@app.delete("/api/customers")
def delete_all_customers():
    try:
        deleted_files = []
        for env_file in ENV_DIR.glob("*.env"):
            env_file.unlink()
            deleted_files.append(env_file.name)

        return {"message": "All customer .env files deleted.", "deleted": deleted_files}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete all customers: {str(e)}")


@app.post("/api/admin/reset-system")
async def reset_system(
    reset_type: str = Query(..., description="Type of reset: 'soft' (keeps configs) or 'hard' (complete reset)"),
    admin_token: str = Query(..., description="Admin authentication token")
):
    """
    Reset the system for deployment/production. Cleans up temporary files and data.
    Use with caution as this will delete user data and temporary files.
    """
    # Simple admin token check (you should use a more secure method in production)
    expected_token = os.getenv("ADMIN_RESET_TOKEN", "reset123")
    if admin_token != expected_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid admin token"
        )

    if reset_type not in ["soft", "hard"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Reset type must be 'soft' or 'hard'"
        )

    try:
        reset_log = []
        
        # Define directories to clean
        directories_to_clean = [
            UPLOAD_DIR,
            VALIDATION_RESULTS_DIR,
            COMPLETED_FOLDER,
            BUNDLE_DEPOT_ZONE,
            Path("uploads/user")  # User saved component data
        ]

        # Files to keep (configuration files)
        files_to_keep = [
            EXCEL_FILE_PATH,
            TRANSFORMATION_ATTRIBUTES_FILE_PATH,
            USER_EXCEL_FILE_PATH,
            Path("Required_files/Mandatory Fields.xlsx"),
            Path("Required_files/Available_NLP.xlsx"),
            Path(".env")  # Keep .env file structure but clear sensitive data in hard reset
        ]

        if reset_type == "soft":
            # Soft reset - keep configuration files, remove temporary data
            reset_log.append("🔄 Performing SOFT reset...")
            
            for directory in directories_to_clean:
                if directory.exists():
                    # Remove contents but keep directory structure
                    for item in directory.iterdir():
                        if item.is_file():
                            item.unlink()
                            reset_log.append(f"🗑️ Deleted file: {item}")
                        elif item.is_dir():
                            shutil.rmtree(item)
                            reset_log.append(f"🗑️ Deleted directory: {item}")
                    reset_log.append(f"✅ Cleaned directory: {directory}")

            # Clear in-memory data
            global pass_df, fail_df, USER_DB
            pass_df = pd.DataFrame()
            fail_df = pd.DataFrame()
            USER_DB = load_user_data(USER_EXCEL_FILE_PATH)  # Reload user data
            
            reset_log.append("✅ Cleared in-memory data structures")

        elif reset_type == "hard":
            # Hard reset - remove everything except essential configuration
            reset_log.append("🔥 Performing HARD reset...")
            combos = extract_customer_instance_names_from_env(env_path=Path(".env"))
            customerName = combos[0][0] if combos else ""
            instanceName = combos[0][1] if combos else ""
            reset_log.append(f"Identified customer: {customerName}, instance: {instanceName}")
            #Remove all user-specific LookupData files
            lookupdata_pattern = f"{customerName}_{instanceName}_LookupData.xlsx"
            mandatory_pattern = f"{customerName}_{instanceName}_MandatoryFields.xlsx"
            try: 
                for file in Path("Required_files").glob(lookupdata_pattern):
                    file.unlink()
                    reset_log.append(f"🗑️ Deleted LookupData file: {file}")
            except Exception as e:
                reset_log.append(f"⚠️ Failed to delete LookupData files: {e}")
            
            try:
                for file in Path("Required_files").glob(mandatory_pattern):
                    file.unlink()
                    reset_log.append(f"🗑️ Deleted MandatoryFields file: {file}")
            except Exception as e:
                reset_log.append(f"⚠️ Failed to delete MandatoryFields files: {e}")
            

            # Clear environment variables from memory
            reset_log.append("🧹 Clearing environment variables from memory...")
            env_vars_to_preserve = {
                'ADMIN_RESET_TOKEN': os.getenv('ADMIN_RESET_TOKEN'),
                'PYTHONPATH': os.getenv('PYTHONPATH'),
                'PATH': os.getenv('PATH')
            }
            
            # Clear all environment variables in current process
            os.environ.clear()
            
            # Restore essential variables
            for key, value in env_vars_to_preserve.items():
                if value:
                    os.environ[key] = value
            
            reset_log.append("✅ Cleared environment variables from memory")

            # Reset .env file to default template (keep structure but clear sensitive data)
            env_file = Path(".env")
            if env_file.exists():
                # Create a minimal .env template
                minimal_env_content = """"""
                env_file.write_text(minimal_env_content)
                reset_log.append("🔧 Reset .env file to minimal template")

            # Clean directories
            for directory in directories_to_clean:
                if directory.exists():
                    shutil.rmtree(directory)
                    directory.mkdir(parents=True, exist_ok=True)
                    reset_log.append(f"🔥 Recreated directory: {directory}")

            # Recreate essential directories
            essential_dirs = [
                UPLOAD_DIR,
                VALIDATION_RESULTS_DIR,
                COMPLETED_FOLDER,
                BUNDLE_DEPOT_ZONE,
                DAT_FILES_DIR,
                Path("uploads/user")
            ]
            
            for dir_path in essential_dirs:
                dir_path.mkdir(parents=True, exist_ok=True)
                reset_log.append(f"📁 Ensured directory exists: {dir_path}")

            # Clear all in-memory data
            USER_DB = {}  # Clear user database
            
            # Reload essential configuration files if they exist
            if USER_EXCEL_FILE_PATH.exists():
                USER_DB = load_user_data(USER_EXCEL_FILE_PATH)
            
            reset_log.append("✅ Cleared all in-memory data structures")

        # Log system information
        reset_log.append(f"📊 System reset completed at: {datetime.now().isoformat()}")
        reset_log.append(f"🔧 Reset type: {reset_type}")
        reset_log.append(f"🐍 Python version: {sys.version}")
        
        # Log directory sizes for monitoring
        for dir_path in [UPLOAD_DIR, VALIDATION_RESULTS_DIR, BUNDLE_DEPOT_ZONE]:
            if dir_path.exists():
                size = sum(f.stat().st_size for f in dir_path.rglob('*') if f.is_file())
                reset_log.append(f"💾 {dir_path.name} size: {size / (1024*1024):.2f} MB")

        return {
            "status": "success",
            "message": f"System reset completed successfully ({reset_type} reset)",
            "reset_type": reset_type,
            "timestamp": datetime.now().isoformat(),
            "details": reset_log
        }

    except Exception as e:
        logger.error(f"System reset failed: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"System reset failed: {str(e)}"
        )


@app.get("/api/admin/system-status")
async def get_system_status(admin_token: str = Query(...)):
    """
    Get current system status and disk usage information.
    """
    expected_token = os.getenv("ADMIN_RESET_TOKEN", "reset123")
    if admin_token != expected_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid admin token"
        )

    try:
        status_info = {
            "timestamp": datetime.now().isoformat(),
            "system_directories": {},
            "file_counts": {},
            "total_disk_usage_mb": 0,
            "important_files_exist": {}
        }

        # Check important directories
        important_dirs = {
            "uploads": UPLOAD_DIR,
            "validation_results": VALIDATION_RESULTS_DIR,
            "bundle_depot": BUNDLE_DEPOT_ZONE,
            "completed_files": COMPLETED_FOLDER,
            "dat_files": DAT_FILES_DIR
        }

        for name, dir_path in important_dirs.items():
            if dir_path.exists():
                files = list(dir_path.rglob("*"))
                file_count = len([f for f in files if f.is_file()])
                total_size = sum(f.stat().st_size for f in files if f.is_file())
                
                status_info["system_directories"][name] = {
                    "exists": True,
                    "file_count": file_count,
                    "size_mb": total_size / (1024 * 1024),
                    "path": str(dir_path)
                }
                status_info["total_disk_usage_mb"] += total_size / (1024 * 1024)
            else:
                status_info["system_directories"][name] = {
                    "exists": False,
                    "file_count": 0,
                    "size_mb": 0,
                    "path": str(dir_path)
                }

        # Check important configuration files
        important_files = {
            "hierarchy_excel": EXCEL_FILE_PATH,
            "transformation_attributes": TRANSFORMATION_ATTRIBUTES_FILE_PATH,
            "user_database": USER_EXCEL_FILE_PATH,
            "mandatory_fields": Path("Required_files/Mandatory Fields.xlsx"),
            "nlp_rules": Path("Required_files/Available_NLP.xlsx"),
            "environment_file": Path(".env")
        }

        for name, file_path in important_files.items():
            status_info["important_files_exist"][name] = {
                "exists": file_path.exists(),
                "path": str(file_path),
                "size_mb": file_path.stat().st_size / (1024 * 1024) if file_path.exists() else 0
            }

        # Add memory usage info
        import psutil
        process = psutil.Process()
        memory_info = process.memory_info()
        status_info["memory_usage_mb"] = memory_info.rss / (1024 * 1024)
        
        # Add system info
        status_info["python_version"] = sys.version
        status_info["platform"] = sys.platform

        return status_info

    except Exception as e:
        logger.error(f"Failed to get system status: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to get system status: {str(e)}"
        )
    
#------------- HDL Job Management ------------------#
DATA_FILE = Path("hdl_jobs.json")


class HDLJob(BaseModel):
    id: Optional[int] = None
    component: str
    fileName: str
    timeCreated: Optional[datetime] = None
    contentId: Optional[str] = None
    requestId: Optional[str] = None
    status: str = "-"
    oracleJobSummary: Optional[str] = None


# -----------------------------
# Persistence Layer
# -----------------------------
def load_data():
    """Safely load jobs data from JSON file"""
    if not DATA_FILE.exists() or DATA_FILE.stat().st_size == 0:
        return {}
    with open(DATA_FILE, "r") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return {}


def save_data(data):
    """Persist jobs data to JSON file"""
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2, default=str)


# -----------------------------
# API Endpoints
# -----------------------------
@app.get("/api/hdl/fetchdata/{customer}/{instance}", response_model=List[HDLJob])
async def fetch_data(customer: str, instance: str):
    """Fetch jobs for given customer + instance"""
    data = load_data()
    jobs = data.get(customer, {}).get(instance, [])
    return [HDLJob(**j) for j in jobs]


@app.post("/api/hdl/getdata/{customer}/{instance}")
async def add_data(customer: str, instance: str, job: HDLJob):
    """Add new job for customer+instance"""
    data = load_data()

    if customer not in data:
        data[customer] = {}
    if instance not in data[customer]:
        data[customer][instance] = []

    jobs = data[customer][instance]

    # Auto-assign ID if not provided
    if not job.id:
        max_id = max((j.get("id", 0) for j in jobs), default=0)
        job.id = max_id + 1

    # Add creation timestamp if missing
    if not job.timeCreated:
        job.timeCreated = datetime.utcnow()

    jobs.append(job.dict())
    save_data(data)
    return {"message": "Job added", "job": job}


@app.put("/api/hdl/updatedata/{customer}/{instance}/{job_id}")
async def update_data(customer: str, instance: str, job_id: int, job_update: HDLJob):
    """Update existing job by ID"""
    data = load_data()

    if customer not in data or instance not in data[customer]:
        raise HTTPException(status_code=404, detail="Customer/Instance not found")

    jobs = data[customer][instance]
    for i, j in enumerate(jobs):
        if j["id"] == job_id:
            # Preserve ID if missing from update payload
            if job_update.id is None:
                job_update.id = job_id
            jobs[i] = job_update.dict()
            save_data(data)
            return {"message": "Job updated", "job": job_update}

    raise HTTPException(status_code=404, detail="Job not found")


# ✅ Path to your JSON file
JSON_FILE_PATH = os.path.join(os.getcwd(), "Required_files", "oracle_value_checks.json")

# ✅ Request model
class OracleValueRequest(BaseModel):
    componentName: Optional[str] = None


@app.post("/api/oracle/value-check")
def oracle_value_check(request: OracleValueRequest):
    try:
        # ✅ Ensure JSON file exists
        if not os.path.exists(JSON_FILE_PATH):
            raise HTTPException(status_code=404, detail="Checklist file not found")

        # ✅ Load JSON data
        with open(JSON_FILE_PATH, "r", encoding="utf-8") as f:
            all_data = json.load(f)

        # ✅ Collect all available checkpoints across all components
        all_available_checks = set()
        for item in all_data:
            for check in item.get("availableChecks", []):
                all_available_checks.add(check)
        all_available_checks = sorted(list(all_available_checks))

        # ✅ If no component name is provided → return all components + full checklist
        if not request.componentName:
            return {
                "components": [item["componentName"] for item in all_data],
                "allAvailableChecks": all_available_checks
            }

        # ✅ Find specific component data
        component_data = next(
            (item for item in all_data if item.get("componentName", "").lower() == request.componentName.lower()),
            None
        )

        # ✅ If component not found → return empty checklist but still show global checks
        if not component_data:
            return {
                "componentName": request.componentName,
                "availableChecks": [],
                "allAvailableChecks": all_available_checks
            }

        # ✅ Return component-specific + global checks
        return {
            "componentName": component_data["componentName"],
            "availableChecks": component_data.get("availableChecks", []),
            "allAvailableChecks": all_available_checks
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
        


def detect_value_type(value, defined_type: str):
    """Parse the value based on defined data type."""
    if pd.isna(value) or str(value).strip() == "":
        return "", "string"

    t = defined_type.lower()
    try:
        if t in ["integer", "int"]:
            return int(float(value)), "integer"
        elif t in ["float", "number", "decimal"]:
            return float(value), "float"
        elif t in ["boolean", "bool"]:
            str_val = str(value).strip().lower()
            if str_val in ["true", "1", "yes", "y"]:
                return True, "boolean"
            elif str_val in ["false", "0", "no", "n"]:
                return False, "boolean"
            return "", "boolean"
        elif t in ["date", "datetime"]:
            try:
                parsed_date = pd.to_datetime(value, errors="coerce")
                if pd.isna(parsed_date):
                    return str(value), "string"
                return parsed_date.strftime("%Y-%m-%d"), "date"
            except Exception:
                return str(value), "string"
        elif t == "timestamp":
            try:
                parsed_date = pd.to_datetime(value, errors="coerce")
                if pd.isna(parsed_date):
                    return str(value), "string"
                return parsed_date.strftime("%Y-%m-%d %H:%M:%S"), "timestamp"
            except Exception:
                return str(value), "string"
        else:
            return str(value), "string"
    except Exception:
        return str(value), "string"

@app.post("/api/parse-file")
async def parse_file(
    file: UploadFile,
    columnDataTypes: str = Form("{}")
):
    """
    Reads an uploaded Excel or DAT file, parses each column value based on data types,
    and returns structured JSON with rows and columns.
    """
    try:
        # 🧠 Convert JSON string to Python dict safely
        try:
            columnDataTypes = json.loads(columnDataTypes)
        except json.JSONDecodeError:
            columnDataTypes = {}

        content = await file.read()
        filename = file.filename.lower()

        # Read Excel or DAT
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df = pd.read_excel(io.BytesIO(content))
        elif filename.endswith(".dat"):
            df = pd.read_csv(io.BytesIO(content), sep="|")
        else:
            return JSONResponse(content={"error": "Unsupported file type"}, status_code=400)

        df.columns = [str(c).strip() for c in df.columns]

        parsed_rows = []
        for idx, row in df.iterrows():
            row_obj = {"id": idx + 1}
            for col in df.columns:
                defined_type = columnDataTypes.get(col, "string").upper()
                mapped_type = DATA_TYPE_MAPPING.get(defined_type, "string")
                value, val_type = detect_value_type(row[col], mapped_type)
                row_obj[col] = value
                row_obj[f"{col}_type"] = val_type
            parsed_rows.append(row_obj)

        parsed_columns = [
            {
                "field": col,
                "headerName": col,
                "type": DATA_TYPE_MAPPING.get(columnDataTypes.get(col, "string").upper(), "string"),
            }
            for col in df.columns
        ]

        return {"columns": parsed_columns, "rows": parsed_rows}

    except Exception as e:
        return JSONResponse(content={"error": f"Failed to parse file: {str(e)}"}, status_code=500)


@app.get("/api/hdl/precheck/list")
async def available_precheck_values():
    """Fetch available pre-check validation rules from JSON file."""
    CHECKLIST_FILE = Path("Required_files/precheck_validation_rules.json")
    try:
        if not CHECKLIST_FILE.exists():
            raise HTTPException(status_code=404, detail="Pre-check validation rules file not found")

        with open(CHECKLIST_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        return {"preCheckValidations": data}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load pre-check validations: {str(e)}")
    



class PrecheckReportRequest(BaseModel):
    customerName: str
    instanceName: str
    componentId: str


@app.post("/api/hdl/precheck/reports/fetch/{componentId}")
async def fetch_precheck_report(componentId: str, req: PrecheckReportRequest):

    customerName = req.customerName.strip()
    instanceName = req.instanceName.strip()

    CHECKLIST_FILE = Path("Required_files/precheck_validation_rules.json")

    if not CHECKLIST_FILE.exists():
        raise HTTPException(status_code=404, detail="Precheck JSON not found")

    with open(CHECKLIST_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    component = next(
        (comp for comp in data.get("Components", []) if comp.get("id") == componentId),
        None
    )
    logger.error(f"component: {component}")
    component_name = component.get("Name", "Unknown").replace(" ", "_")
    if not component:
        raise HTTPException(status_code=404, detail="Component ID not found")

    report_url = component.get("ReportURL")
    if not report_url:
        raise HTTPException(status_code=400, detail="Missing ReportURL for component")

    oracle_env, username, password = load_oracle_credentials(customerName, instanceName)
    logger.warning(f"Using Oracle Credentials: {oracle_env}, {username}, *****")

    SOAP_URL = f"{oracle_env}/xmlpserver/services/PublicReportService"

    # ✅ Save folder
    save_dir = Path(f"Required_files/precheck_reports/{customerName}/{instanceName}")
    save_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    saved_file = save_dir / f"{customerName}_{instanceName}_{component_name}_Precheck_.xlsx"

    soap_body = f"""
        <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                xmlns:pub="http://xmlns.oracle.com/oxp/service/PublicReportService">
            <soapenv:Header/>
            <soapenv:Body>
                <pub:runReport>
                    <pub:reportRequest>
                        <pub:reportAbsolutePath>{report_url}</pub:reportAbsolutePath>
                        <pub:attributeFormat>xlsx</pub:attributeFormat>
                        <pub:sizeOfDataChunkDownload>-1</pub:sizeOfDataChunkDownload>
                        <pub:parameterNameValues>
                            <pub:item>
                                <pub:name>P_CUSTOMER_NAME</pub:name>
                                <pub:values>
                                    <pub:item>{customerName}</pub:item>
                                </pub:values>
                            </pub:item>
                            <pub:item>
                                <pub:name>P_INSTANCE_NAME</pub:name>
                                <pub:values>
                                    <pub:item>{instanceName}</pub:item>
                                </pub:values>
                            </pub:item>
                        </pub:parameterNameValues>
                    </pub:reportRequest>
                    <pub:userID>{username}</pub:userID>
                    <pub:password>{password}</pub:password>
                </pub:runReport>
            </soapenv:Body>
        </soapenv:Envelope>
    """

    headers = {
        "Content-Type": "text/xml;charset=UTF-8",
    }

    try:
        async with httpx.AsyncClient(timeout=180) as client:
            logger.info(f"📡 Calling: {SOAP_URL}")
            resp = await client.post(SOAP_URL, data=soap_body.encode("utf-8"), headers=headers)

        if resp.status_code == 401:
            logger.error("❌ Unauthorized — Creds Incorrect or Endpoint Secured!")
            raise HTTPException(status_code=401, detail="Unauthorized Oracle Credentials")

        logger.info(f"✅ Oracle Status: {resp.status_code}")

        # ✅ Parse and extract reportBytes
        xml_tree = ET.fromstring(resp.text)
        namespace = {"pub": "http://xmlns.oracle.com/oxp/service/PublicReportService"}

        report_bytes_elem = xml_tree.find(".//pub:reportBytes", namespace)
        if report_bytes_elem is None:
            raise HTTPException(status_code=500, detail="Missing reportBytes in response")

        report_bytes = base64.b64decode(report_bytes_elem.text)

        with open(saved_file, "wb") as f:
            f.write(report_bytes)

        # ✅ Update JSON
        component["Fetch_Status"] = "Fetched"
        component["Last Uploaded Excel File Path"] = str(saved_file)
        component["Last Fetched"] = datetime.now().isoformat()

        with open(CHECKLIST_FILE, "w", encoding="utf-8") as fw:
            json.dump(data, fw, indent=2)

        return {
            "success": True,
            "componentId": componentId,
            "filePath": str(saved_file),
            "message": "✅ Excel Report Generated Successfully"
        }

    except Exception as e:
        logger.error(f"❌ Fetch Failed: {e}")
        component["Fetch_Status"] = "Failed"
        with open(CHECKLIST_FILE, "w", encoding="utf-8") as fw:
            json.dump(data, fw, indent=2)
        raise HTTPException(status_code=500, detail=str(e))


CHECKLIST_FILE = Path("Required_files/precheck_validation_rules.json")

@app.post("/api/hdl/precheck/reports/upload/{id}")
async def upload_precheck_userdata(
    id: str,
    file: UploadFile = File(...),
    customerName: str = Form(...),
    instanceName: str = Form(...),
):
    # ✅ A little sanitization
    customerName = customerName.strip()
    instanceName = instanceName.strip()

    if not CHECKLIST_FILE.exists():
        raise HTTPException(status_code=404, detail="Precheck JSON not found")

    # ✅ Load JSON
    with open(CHECKLIST_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    component = next(
        (comp for comp in data.get("Components", []) if comp.get("id") == id),
        None
    )

    if not component:
        raise HTTPException(status_code=404, detail="Component ID not found")

    # ✅ Create save directory
    save_dir = Path(f"Required_files/precheck_reports/User_Uploads/{customerName}/{instanceName}/precheck_reports")
    save_dir.mkdir(parents=True, exist_ok=True)

    # ✅ Filename based on component name
    comp_name = component.get("Name", "Component").replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    file_path = save_dir / f"{comp_name}_Uploaded_{timestamp}.xlsx"

    # ✅ Save the uploaded file
    try:
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to save uploaded file")

    with open(CHECKLIST_FILE, "w", encoding="utf-8") as fw:
        json.dump(data, fw, indent=2)

    return {
        "status": "✅ Success",
        "componentId": id,
        "filePath": str(file_path),
        "message": "Excel uploaded Sucessfully"
    }

class ValidationRequest(BaseModel):
    customerName: str
    instanceName: str
    componentId: str
    uploadedFileRef: str | None = None

@app.post("/api/hdl/precheck/reports/validate/{id}")
async def precheck_validation_global(id: str, payload: ValidationRequest):
    logger.info("called global validation...")
    logger.info(f"Validation payload: {payload}")


@app.post("/api/excel/sheets")
async def get_excel_sheets(
    file: UploadFile = File(...),
    customerName: str = Form(...),
    instanceName: str = Form(...)
):
    """
    Upload an Excel file and return:
      - List of sheet names
      - Columns from each sheet
    Also saves the Excel file inside Required_files/Post-Validation_Excels
    using an iterator (1, 2, 3...) for duplicate uploads.
    """
    try:
        # Read file bytes
        content = await file.read()
        filename = file.filename.lower()

        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            return JSONResponse(
                content={"error": "Unsupported file type. Upload .xlsx or .xls only."},
                status_code=400,
            )

        # Read Excel content once
        excel_file = pd.ExcelFile(io.BytesIO(content))
        sheet_names = excel_file.sheet_names

        # Extract columns for each sheet
        columns_dict = {}
        for sheet in sheet_names:
            df = excel_file.parse(sheet)
            columns_dict[sheet] = df.columns.tolist()

        # Prepare save directory
        save_dir = Path("Required_files/Post-Validation_Excels")
        save_dir.mkdir(parents=True, exist_ok=True)

        # File name pattern: Customer_Instance_#.xlsx
        base_name = f"{customerName}_{instanceName}"
        existing_files = list(save_dir.glob(f"{base_name}_*.xlsx"))

        # Determine next iterator number
        next_index = len(existing_files) + 1
        save_excel_name = f"{base_name}_{next_index}.xlsx"
        save_path = save_dir / save_excel_name

        # Save the file
        with open(save_path, "wb") as f:
            f.write(content)

        # Return structured response
        return {
            "sheets": sheet_names,
            "columns": columns_dict,
            "saved_as": save_excel_name,
            "message": "Excel uploaded and processed successfully."
        }

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to read Excel file: {str(e)}"},
            status_code=500,
        )
    
GOOGLE_API_KEY = os.environ.get('GEMINI_API_KEY', '')
if genai is None:
    logging.info("GenAI client not installed; GenAI features are disabled.")
elif not GOOGLE_API_KEY:
    logging.info("GEMINI_API_KEY not set; GenAI features are disabled until configured.")

def get_smart_mapping_from_gemini(legacy_cols: List[str], oracle_cols: List[str]):
    """
    Updated to return Mapping + Data Type Detection
    """
    fallback_result = {
        "mapping": {},
        "date_columns": [],
        "timestamp_columns": []
    }
    
    # Simple exact match fallback
    oracle_cols_lower = {col.lower(): col for col in oracle_cols}
    for l_col in legacy_cols:
        if l_col.lower() in oracle_cols_lower:
            fallback_result["mapping"][l_col] = oracle_cols_lower[l_col.lower()]

    if not GOOGLE_API_KEY or genai is None:
        return fallback_result

    try:
        client = genai.Client(api_key=GOOGLE_API_KEY)

        prompt = f"""
            You are a Data Integration Expert.
            Legacy Columns: {json.dumps(legacy_cols)}
            Oracle Columns: {json.dumps(oracle_cols)}

            Task:
            1. Map 'Legacy Columns' to semantically similar 'Oracle Columns'.
               If a legacy column has NO suitable semantic match in the Oracle columns,
               map it to an empty string "" (blank). Do NOT force a mapping.
            2. Identify which 'Legacy Columns' likely contain DATES (e.g., DOB, Start Date).
            3. Identify which 'Legacy Columns' likely contain TIMESTAMPS (e.g., Created At, Last Update).

            IMPORTANT: Every legacy column MUST appear as a key in the mapping.
            If there is no good match, set its value to "" (empty string).

            Return JSON format:
            {{
                "mapping": {{ "LegacyCol": "OracleCol or empty string", ... }},
                "date_columns": [ "LegacyCol1", ... ],
                "timestamp_columns": [ "LegacyCol2", ... ]
            }}
        """

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config=genai.types.GenerateContentConfig(
                response_mime_type="application/json"
            )
        )
        text = response.text
        # response_mime_type is application/json so text is already JSON;
        # fall back to regex extraction in case of extra surrounding text
        try:
            ai_data = json.loads(text)
        except json.JSONDecodeError:
            match = re.search(r'\{.*\}', text, re.DOTALL)
            if not match:
                return fallback_result
            ai_data = json.loads(match.group(0))

        # Validate AI suggestions — only keep targets that actually exist in oracle_cols.
        # Gemini sometimes hallucinates column names; discard those entirely.
        oracle_set       = set(oracle_cols)
        oracle_lower_map = {c.lower(): c for c in oracle_cols}
        validated_ai = {}
        for src, tgt in ai_data.get("mapping", {}).items():
            if not tgt:
                continue  # AI said no match — let fallback decide
            if tgt in oracle_set:
                validated_ai[src] = tgt
            elif tgt.lower() in oracle_lower_map:
                validated_ai[src] = oracle_lower_map[tgt.lower()]
            # else: hallucinated name — discard

        # Merge: start with fallback exact-matches, then layer validated AI on top.
        # AI only overrides when it found a real oracle column.
        final_mapping = {**fallback_result["mapping"], **validated_ai}
        return {
            "mapping": final_mapping,
            "date_columns": ai_data.get("date_columns", []),
            "timestamp_columns": ai_data.get("timestamp_columns", [])
        }

    except Exception as e:
        print(f"❌ Gemini Error: {e}")
        return fallback_result


@app.post("/api/hdl/gemini-map")
def gemini_map(payload: Dict = Body(...)):
    """Return a smart mapping between legacy (source) and oracle (target) columns.

    Expected payload shape (example):
    {
      "legacy_columns": [...],
      "oracle_columns": [...],
      "suggested_mapping": {"LegacyCol": "OracleCol"},
      "date_columns": [ ... ]
    }

    The endpoint will call `get_smart_mapping_from_gemini` when configured, fall
    back to exact-matches otherwise, then merge any `suggested_mapping` values
    supplied by the caller (these take precedence).
    """
    try:
        legacy_cols = payload.get("legacy_columns") or payload.get("source_columns") or []
        oracle_cols = payload.get("oracle_columns") or payload.get("target_columns") or []
        client_suggested = payload.get("suggested_mapping", {}) or {}
        client_date_cols = payload.get("date_columns", []) or []

        if not isinstance(legacy_cols, list) or not isinstance(oracle_cols, list):
            raise HTTPException(status_code=400, detail="`legacy_columns` and `oracle_columns` must be lists")

        ai_result = get_smart_mapping_from_gemini(legacy_cols, oracle_cols)

        # Merge AI mapping with client suggestions. Client suggestions override.
        final_mapping = dict(ai_result.get("mapping", {}))
        for k, v in client_suggested.items():
            if k in legacy_cols:
                final_mapping[k] = v

        # Build response in the format the frontend expects
        response = {
            "legacy_columns": legacy_cols,
            "oracle_columns": oracle_cols,
            "suggested_mapping": final_mapping,
            "date_columns": sorted(list(set(ai_result.get("date_columns", []) + client_date_cols))),
            "timestamp_columns": sorted(list(set(ai_result.get("timestamp_columns", []))))
        }

        return JSONResponse(content=response, status_code=200)

    except HTTPException:
        raise
    except Exception as e:
        logging.exception("Error in gemini_map endpoint")
        raise HTTPException(status_code=500, detail=str(e))


import re
from dateutil import parser

ORACLE_MIN_YEAR = -4712
ORACLE_MAX_YEAR = 9999

def normalize_dates(df, explicit_cols):
    def safe_parse(val):
        if pd.isna(val) or str(val).strip() == "":
            return ""

        val_str = str(val).strip()

        # Excel serial date handling
        if re.fullmatch(r"\d{5,}", val_str):
            try:
                return pd.to_datetime(float(val), unit="d", origin="1899-12-30") \
                         .strftime("%Y/%m/%d")
            except:
                return ""

        try:
            dt = parser.parse(
                val_str,
                dayfirst=True,
                yearfirst=False,
                fuzzy=True
            )

            year = dt.year

            # Oracle year boundaries
            if year < ORACLE_MIN_YEAR or year > ORACLE_MAX_YEAR:
                return ""

            return f"{year:04d}/{dt.month:02d}/{dt.day:02d}"

        except Exception:
            # Final fallback: keep original string (Oracle may still accept)
            return val_str

    for col in df.columns:
        if col in explicit_cols or pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].apply(safe_parse)

    return df



# ── Mapping job tracker (same pattern as validation jobs) ────────────────
import threading as _threading
_mapping_jobs: Dict[str, Dict[str, Any]] = {}
_mapping_jobs_lock = _threading.Lock()

def _compute_smooth_eta(job: dict) -> Optional[float]:
    """
    Compute a smooth, realistic ETA using an EMA of the progress rate.

    Fixes vs. prior version:
    - alpha=0.5: ensures the NEWEST rate always gets the most weight regardless
      of history length (alpha=0.35 gave oldest rate 42% vs newest 35% for 3 samples).
    - Stall detection: when progress hasn't changed for 5+ seconds, a near-zero
      implied rate is injected so the ETA rises instead of staying falsely low.
    - prev_eta smoothing reduced (50/50 instead of 40/60): converges faster when
      the rate changes dramatically (e.g., fast Polars → slow openpyxl write).
    - Linear-rate blend reduced (15% instead of 30%): early fast stages no longer
      pull the ETA too low during later slow stages.
    """
    prog = job.get("progress", 0)
    if prog >= 100:
        return 0.0
    if prog <= 0:
        return None

    now = time.time()
    started = job.get("started_at")
    if not started:
        return None

    elapsed = now - started
    if elapsed < 1.5:
        return None  # too early, not enough data

    # Track progress history: list of (timestamp, progress) snapshots
    history = job.setdefault("_eta_history", [])
    if not history or history[-1][1] != prog:
        history.append((now, prog))
    if len(history) > 20:
        job["_eta_history"] = history[-20:]
        history = job["_eta_history"]

    if len(history) < 2:
        rate = prog / elapsed
        raw_eta = (100 - prog) / rate
        return round(min(raw_eta, 7200), 1)

    # Compute rates between consecutive snapshots (only for intervals with progress)
    rates = []
    for i in range(1, len(history)):
        dt = history[i][0] - history[i - 1][0]
        dp = history[i][1] - history[i - 1][1]
        if dt > 0.1 and dp > 0:
            rates.append(dp / dt)

    # Stall detection: if progress hasn't moved for 5+ seconds, inject a
    # near-zero implied rate so the ETA climbs instead of staying falsely optimistic.
    dt_stalled = now - history[-1][0]
    if dt_stalled > 5.0 and history[-1][1] == prog and rates:
        # implied rate: assume 0.05% would take dt_stalled seconds at this pace
        rates.append(0.05 / dt_stalled)

    if not rates:
        rate = prog / elapsed
        raw_eta = (100 - prog) / max(rate, 1e-6)
        return round(min(raw_eta, 7200), 1)

    # EMA seeded from the NEWEST rate so recent rates dominate.
    # With alpha=0.5 and n rates: newest gets 50%, second-newest 25%, etc.
    # (alpha=0.35 inverted this for short histories: oldest got 42%, newest 35%)
    alpha = 0.5
    ema_rate = rates[-1]
    for r in rates[-2::-1]:
        ema_rate = alpha * ema_rate + (1 - alpha) * r

    # Blend EMA with the overall linear rate — minimal linear weight (15%) to
    # avoid early-stage fast rates pulling the ETA too low in slow later stages.
    linear_rate = prog / elapsed
    blended_rate = 0.85 * ema_rate + 0.15 * linear_rate

    if blended_rate <= 1e-6:
        return round(min((100 - prog) / max(linear_rate, 1e-6), 7200), 1)

    raw_eta = (100 - prog) / blended_rate

    # Smooth against previous ETA — 50/50 blend converges twice as fast as 40/60
    # when the rate changes dramatically between stages.
    prev_eta = job.get("eta_seconds")
    if prev_eta is not None and prev_eta > 0:
        smoothed = 0.5 * raw_eta + 0.5 * prev_eta
    else:
        smoothed = raw_eta

    return round(max(0, min(smoothed, 7200)), 1)


def _mapping_job_update(job_id: str, **kwargs):
    with _mapping_jobs_lock:
        if job_id in _mapping_jobs:
            _mapping_jobs[job_id].update(kwargs)
            job = _mapping_jobs[job_id]
            job["eta_seconds"] = _compute_smooth_eta(job)

def _mapping_job_get(job_id: str) -> Optional[Dict]:
    with _mapping_jobs_lock:
        return _mapping_jobs.get(job_id, {}).copy()


def _read_file_bytes(file_bytes: bytes, filename: str = ""):
    """Read Excel or CSV from bytes, return DataFrame."""
    ext = os.path.splitext(filename)[1].lower() if filename else ""
    if ext == ".csv":
        return pd.read_csv(io.BytesIO(file_bytes))
    else:
        return pd.read_excel(io.BytesIO(file_bytes))


def _run_mapping_job(job_id: str, legacy_bytes: bytes, oracle_bytes: bytes,
                     legacy_filename: str = "", oracle_filename: str = ""):
    """Background thread: read files → call Gemini → store result."""
    logger.info(f"[Mapping] Starting job {job_id[:8]}, legacy_file='{legacy_filename}', oracle_file='{oracle_filename}'")
    try:
        # ── Stage 1: Read Legacy File ──
        _mapping_job_update(job_id, progress=10, stage="Reading source file")
        try:
            legacy_df = _read_file_bytes(legacy_bytes, legacy_filename)
            legacy_columns = legacy_df.columns.astype(str).str.strip().tolist()
        except Exception as e:
            _mapping_job_update(job_id, status="failed", error=f"Failed to read Legacy file: {str(e)}")
            return

        # ── Stage 2: Read Oracle File ──
        _mapping_job_update(job_id, progress=25, stage="Reading target file")
        try:
            oracle_df = _read_file_bytes(oracle_bytes, oracle_filename)
            oracle_columns = oracle_df.columns.astype(str).str.strip().tolist()
        except Exception as e:
            _mapping_job_update(job_id, status="failed", error=f"Failed to read Target file: {str(e)}")
            return

        # ── Stage 3: Gemini AI Mapping ──
        _mapping_job_update(job_id, progress=40, stage="Gemini analyzing columns")
        ai_result = get_smart_mapping_from_gemini(legacy_columns, oracle_columns)
        _mapping_job_update(job_id, progress=80, stage="Parsing AI response")

        # ── Stage 4: Build result ──
        _mapping_job_update(job_id, progress=95, stage="Finalizing mapping")
        result = {
            "legacy_columns": legacy_columns,
            "oracle_columns": oracle_columns,
            "suggested_mapping": ai_result["mapping"],
            "date_columns": ai_result.get("date_columns", []),
            "timestamp_columns": ai_result.get("timestamp_columns", []),
            "message": "Columns analyzed successfully."
        }

        _mapping_job_update(job_id, status="complete", progress=100, stage="Done", result=result)

    except Exception as e:
        _mapping_job_update(job_id, status="failed", error=f"Processing Error: {str(e)}")


@app.post("/api/excel/columns/mapping")
async def get_excel_columns_mapping_async(
    background_tasks: BackgroundTasks,
    legacyFile: UploadFile = File(...),
    oracleFile: UploadFile = File(...),
    legacySheet: str = Form(default=""),
    oracleSheet: str = Form(default=""),
    customerName: str = Form(...),
    instanceName: str = Form(...)
):
    """
    Async job-based mapping. Returns { job_id } immediately.
    Poll GET /api/excel/columns/mapping/status/{job_id} for progress.
    """
    job_id = str(uuid.uuid4())

    legacy_bytes = await legacyFile.read()
    oracle_bytes = await oracleFile.read()

    with _mapping_jobs_lock:
        _mapping_jobs[job_id] = {
            "status": "running",
            "progress": 0,
            "stage": "Uploading files",
            "error": None,
            "result": None,
            "started_at": time.time(),
            "eta_seconds": None,
        }

    legacy_fname = legacyFile.filename or ""
    oracle_fname = oracleFile.filename or ""
    logger.info(f"[Mapping] Submitting job {job_id[:8]}, legacy='{legacy_fname}', oracle='{oracle_fname}'")

    background_tasks.add_task(
        _run_mapping_job, job_id, legacy_bytes, oracle_bytes,
        legacy_fname, oracle_fname
    )

    return {"job_id": job_id}


@app.get("/api/excel/columns/mapping/status/{job_id}")
async def get_mapping_status(job_id: str):
    """
    Returns current progress of a mapping job.
    When status == 'complete', the result payload is included.
    """
    job = _mapping_job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Mapping job not found")

    resp = {
        "status": job.get("status"),
        "progress": job.get("progress", 0),
        "stage": job.get("stage", ""),
        "error": job.get("error"),
        "eta_seconds": job.get("eta_seconds"),
    }

    if job.get("status") == "complete" and job.get("result"):
        resp["result"] = job["result"]
        # Clean up the job after delivering result
        with _mapping_jobs_lock:
            _mapping_jobs.pop(job_id, None)

    return resp


# --- OPTIMIZED VALIDATION LOGIC STARTS HERE ---

# ═══════════════════════════════════════════════════════════════════════════
# POLARS-NATIVE HIGH-PERFORMANCE ENGINE  (10M+ row optimization)
# ═══════════════════════════════════════════════════════════════════════════

def _polars_read_file(file_path: str, sheet_name=None):
    """Read CSV/Excel into Polars LazyFrame. Avoids pandas entirely.
    For CSV: uses scan_csv (lazy, memory-mapped, multi-threaded).
    For Excel: uses calamine (Rust, 5-10x faster than openpyxl).
    """
    import polars as pl
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return pl.scan_csv(
            file_path, infer_schema_length=0,
            try_parse_dates=False, null_values=[""]
        )

    # ── Excel handling ──
    sheet_id_param = None
    sheet_name_param = None

    if sheet_name is None:
        sheet_id_param = 1          # Polars is 1-indexed
    elif isinstance(sheet_name, int):
        sheet_id_param = max(sheet_name, 1)
    elif isinstance(sheet_name, str):
        stripped = sheet_name.strip()
        if not stripped:
            sheet_id_param = 1
        else:
            try:
                idx = int(stripped)
                sheet_id_param = max(idx + 1, 1) if idx == 0 else idx
            except ValueError:
                sheet_name_param = stripped

    # calamine (Rust) only — xlsx2csv intentionally removed: it's 10-100x slower on
    # large files and will hang for 10+ minutes on 60 MB workbooks.
    df = None
    try:
        kw: dict = {"engine": "calamine"}
        if sheet_id_param is not None:
            kw["sheet_id"] = sheet_id_param
        elif sheet_name_param is not None:
            kw["sheet_name"] = sheet_name_param
        df = pl.read_excel(file_path, **kw)
    except Exception as _cal_err:
        logger.warning(f"calamine failed ({type(_cal_err).__name__}: {_cal_err}), falling back to openpyxl")

    if df is None:
        sp = 0
        if sheet_name_param:
            sp = sheet_name_param
        elif sheet_id_param:
            sp = max(sheet_id_param - 1, 0)
        pdf = pd.read_excel(file_path, sheet_name=sp, dtype=str, engine="openpyxl")
        df = pl.from_pandas(pdf)

    # Only cast non-Utf8 columns (calamine emits Int64/Float64 for numeric cells)
    _non_utf8 = {c: pl.Utf8 for c in df.columns if df.schema[c] != pl.Utf8}
    if _non_utf8:
        df = df.cast(_non_utf8)
    renames = {c: c.strip() for c in df.columns if c != c.strip()}
    if renames:
        df = df.rename(renames)
    return df.lazy()


_COMMON_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%d-%m-%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
]


def _detect_date_fmt(values, formats=_COMMON_DATE_FORMATS):
    """Sample values to auto-detect date format. Returns format string or None."""
    vals = [str(v).strip() for v in values[:30]
            if v is not None and str(v).strip() not in ("", "nan", "None", "NaN", "NaT")]
    if not vals:
        return None
    for fmt in formats:
        ok = 0
        for v in vals[:15]:
            try:
                datetime.strptime(v, fmt)
                ok += 1
            except Exception:
                pass
        if ok >= max(len(vals[:15]) * 0.6, 1):
            return fmt
    return None


# Compiled once at module load — used by _pl_is_date_like_column for fast date detection
_DATE_LIKE_RE = re.compile(
    r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}'       # YYYY-MM-DD, YYYY/MM/DD
    r'|^\d{1,2}[-/]\d{1,2}[-/]\d{4}'       # DD-MM-YYYY, MM/DD/YYYY
    r'|^\d{1,2}[-/]\d{1,2}[-/]\d{2}\b'     # DD-MM-YY
    r'|^\d{4}\d{2}\d{2}$'                   # YYYYMMDD
    r'|^\d{4}[-/]\d{1,2}[-/]\d{1,2}[T ]',  # ISO-8601 with time
    re.ASCII,
)


def _pl_is_date_like_column(df, col_name, sample_size=50):
    """Check if a Polars column likely contains date values by sampling (regex, no dateutil)."""
    import polars as pl
    if col_name not in df.columns:
        return False
    sample = (df[col_name].cast(pl.Utf8).fill_null("")
              .head(sample_size)
              .to_list())
    vals = [v.strip() for v in sample if v.strip() not in ('', 'nan', 'None', 'NaN', 'NaT')]
    if not vals:
        return False
    parsed = sum(1 for v in vals if _DATE_LIKE_RE.match(v))
    return parsed >= max(len(vals) * 0.6, 1)


def _pl_apply_date_normalization(df, date_cols, all_cols, auto_detect=False, compare_cols=None):
    """Apply date normalization on an eager Polars DataFrame — batched, single pass.

    vs. prior version:
    - ONE head(50) across all candidate columns instead of N individual head() calls
    - ONE head(50) across all date columns for format sampling instead of N per-column calls
    - ONE with_columns(date_exprs) for all date columns instead of N separate materializations
    - Polars-native str.to_date() fallback instead of pandas/dateutil round-trip
    """
    import polars as pl
    _SENTINELS = {'', 'nan', 'None', 'NaN', 'NaT'}

    actual_date_cols = set(date_cols) if date_cols else set()

    # Auto-detect: ONE head(50) for all candidate columns — replaces N individual head() calls
    if auto_detect and compare_cols:
        candidate_cols = [c for c in compare_cols if c not in actual_date_cols and c in df.columns]
        if candidate_cols:
            sample_df = df.select(
                [pl.col(c).cast(pl.Utf8).fill_null("") for c in candidate_cols]
            ).head(50)
            for col in candidate_cols:
                vals = [v.strip() for v in sample_df[col].to_list() if v.strip() not in _SENTINELS]
                if vals and sum(1 for v in vals if _DATE_LIKE_RE.match(v)) >= max(len(vals) * 0.6, 1):
                    actual_date_cols.add(col)
                    logger.info(f"  Auto-detected date column (Polars): {col}")

    valid_date_cols = [c for c in actual_date_cols if c in all_cols and c in df.columns]
    if not valid_date_cols:
        return df

    # ONE head(50) for all date columns — replaces N per-column sample calls
    format_sample = df.select(
        [pl.col(c).cast(pl.Utf8).fill_null("") for c in valid_date_cols]
    ).head(50)

    # Build all strptime expressions; apply in a SINGLE with_columns() — one frame scan
    date_exprs = []
    for col in valid_date_cols:
        sample = [v.strip() for v in format_sample[col].to_list() if v.strip() not in _SENTINELS]
        fmt = _detect_date_fmt(sample)
        if fmt:
            if "H" in fmt or "M" in fmt:
                expr = (
                    pl.col(col).cast(pl.Utf8).fill_null("")
                    .str.strptime(pl.Datetime, fmt, strict=False)
                    .dt.strftime("%Y/%m/%d").fill_null("")
                )
            else:
                expr = (
                    pl.col(col).cast(pl.Utf8).fill_null("")
                    .str.strptime(pl.Date, fmt, strict=False)
                    .dt.strftime("%Y/%m/%d").fill_null("")
                )
        else:
            # No format detected — Polars native auto-parse; preserves original for non-date values
            expr = (
                pl.col(col).cast(pl.Utf8).fill_null("")
                .str.to_date(format=None, strict=False)
                .dt.strftime("%Y/%m/%d")
                .fill_null(pl.col(col).cast(pl.Utf8).fill_null(""))
            )
        date_exprs.append(expr.alias(col))

    if date_exprs:
        try:
            df = df.with_columns(date_exprs)  # single frame scan for ALL date columns
        except Exception as _e:
            logger.debug(f"Batch date normalization failed ({_e}), retrying per-column")
            for single_expr in date_exprs:
                try:
                    df = df.with_columns([single_expr])
                except Exception:
                    pass

    return df


def _pl_detect_numeric_cols(df_l, df_o, cols_to_compare):
    """Detect numeric columns — one vectorized select() across all columns (single frame scan)."""
    import polars as pl
    numeric_columns = set()
    sentinel = ["", "nan", "None", "NaN"]

    valid_cols = [c for c in cols_to_compare if c in df_l.columns and c in df_o.columns]
    if not valid_cols:
        return numeric_columns

    l_sample = df_l.select([pl.col(c).cast(pl.Utf8).alias(c) for c in valid_cols]).head(2500)
    o_sample = df_o.select([pl.col(c).cast(pl.Utf8).alias(c) for c in valid_cols]).head(2500)
    combined = pl.concat([l_sample, o_sample])

    n = len(valid_cols)
    # ONE select() — N×3 expressions run in Polars' parallel Rust executor; replaces N×3 serial ops
    try:
        row = combined.select(
            # numeric count: non-sentinel AND castable to Float64
            [
                (
                    ~pl.col(c).is_in(sentinel) &
                    pl.col(c).str.replace_all("[%,]", "").cast(pl.Float64, strict=False).is_not_null()
                ).sum().alias(f"_nc_{i}")
                for i, c in enumerate(valid_cols)
            ] +
            # total non-sentinel count (denominator)
            [
                (~pl.col(c).is_in(sentinel)).sum().alias(f"_tot_{i}")
                for i, c in enumerate(valid_cols)
            ] +
            # dash count: non-sentinel AND contains "-" (guards against date-like strings)
            [
                (~pl.col(c).is_in(sentinel) & pl.col(c).str.contains(r"\-")).sum().alias(f"_dc_{i}")
                for i, c in enumerate(valid_cols)
            ]
        ).row(0)
    except Exception:
        return numeric_columns

    for i, col in enumerate(valid_cols):
        tot = row[n + i]
        if tot == 0:
            continue
        nc = row[i]
        dc = row[2 * n + i]
        if (nc / tot) >= 0.8 and (dc / tot) < 0.2:
            numeric_columns.add(col)

    return numeric_columns


def _pl_clean_str_expr(col_name, case_sensitive=True):
    """Polars expression: clean string for comparison."""
    import polars as pl
    expr = (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.replace_all(r"\s+", " ")
        .str.strip_chars()
        .str.replace_all(",", "")
        .str.replace_all(r"\.0+$", "")
        .str.replace_all("(?i)^nan$", "")
        .str.replace_all("(?i)^none$", "")
        .fill_null("")
    )
    if not case_sensitive:
        expr = expr.str.to_lowercase()
    return expr


def _pl_clean_num_expr(col_name):
    """Polars expression: clean & cast to Float64 for numeric comparison."""
    import polars as pl
    return (
        pl.col(col_name).cast(pl.Utf8).fill_null("")
        .str.replace_all("%", "")
        .str.replace_all(",", "")
        .str.strip_chars()
        .str.replace_all("^$", "NaN")
        .cast(pl.Float64, strict=False)
        .round(4)
    )

def _pl_normalize_key_cols(df, key_cols):
    """Normalize key columns in Polars: strip, remove .0 suffix, blank out null sentinels."""
    import polars as pl
    exprs = []
    for col in key_cols:
        if col not in df.columns:
            continue
        exprs.append(
            pl.col(col).cast(pl.Utf8).fill_null("")
            .str.strip_chars()
            .str.replace(r"\.0+$", "", literal=False)
            .str.replace_all(r"^(?:nan|None|NaN)$", "")
            .alias(col)
        )
    return df.with_columns(exprs) if exprs else df


def _pl_gen_composite_key(key_cols):
    """Return a Polars expr that concatenates key columns with '|' separator."""
    import polars as pl
    return pl.concat_str(
        [pl.col(c).cast(pl.Utf8).fill_null("") for c in key_cols],
        separator="|",
    )


def _pl_add_positional_key(df, key_col):
    """Append row-position-within-group suffix to key_col to handle duplicate keys safely."""
    import polars as pl
    return (
        df.with_row_index("_global_rn")
        .with_columns(
            (pl.col("_global_rn") - pl.col("_global_rn").min().over(key_col))
            .cast(pl.Utf8)
            .alias("_rn")
        )
        .with_columns(
            (pl.col(key_col) + "|_rn=" + pl.col("_rn")).alias(key_col)
        )
        .drop(["_global_rn", "_rn"])
    )


def _polars_write_source_target_csv(joined, leg_only, orc_only, cols_to_compare, csv_path, internal_key):
    """Write full source/target data directly from Polars to CSV. Zero pandas.
    Column order: all PS (_S) columns first, then all OC (_T) columns, then Record Status."""
    import polars as pl
    # All PS columns first, then all OC columns (not interleaved)
    s_exprs = [pl.col(col).cast(pl.Utf8).fill_null("").alias(f"{col}_S") for col in cols_to_compare]
    t_exprs = [pl.col(f"{col}_T").cast(pl.Utf8).fill_null("").alias(f"{col}_T") for col in cols_to_compare]
    matched = joined.select(s_exprs + t_exprs + [pl.lit("MATCHED").alias("Record Status")])
    parts = [matched]

    if len(leg_only) > 0:
        l_s = [
            (pl.col(col).cast(pl.Utf8).fill_null("") if col in leg_only.columns else pl.lit("")).alias(f"{col}_S")
            for col in cols_to_compare
        ]
        l_t = [pl.lit("").alias(f"{col}_T") for col in cols_to_compare]
        parts.append(leg_only.select(l_s + l_t + [pl.lit("MISSING_IN_TARGET").alias("Record Status")]))

    if len(orc_only) > 0:
        o_s = [pl.lit("").alias(f"{col}_S") for col in cols_to_compare]
        o_t = [
            (pl.col(col).cast(pl.Utf8).fill_null("") if col in orc_only.columns else pl.lit("")).alias(f"{col}_T")
            for col in cols_to_compare
        ]
        parts.append(orc_only.select(o_s + o_t + [pl.lit("MISSING_IN_SOURCE").alias("Record Status")]))

    full = pl.concat(parts)
    full.write_csv(csv_path)
    return len(full)


def _is_date_like_column(series: pd.Series, sample_size: int = 50) -> bool:
    """
    Heuristic: check if a column likely contains date values by sampling.
    Returns True if >= 60% of non-empty sampled values parse as dates.
    """
    non_empty = series.dropna().astype(str).str.strip()
    non_empty = non_empty[~non_empty.isin(['', 'nan', 'None', 'NaN', 'NaT'])]
    if len(non_empty) == 0:
        return False
    sample = non_empty.head(sample_size).tolist()
    parsed = 0
    for val in sample:
        # Skip pure numbers (could be IDs, amounts etc.)
        try:
            float(val)
            # If it's a pure number with no date separators, skip
            if not any(sep in val for sep in ['-', '/', '.']):
                continue
        except ValueError:
            pass
        # Must contain date-like separators
        if not any(sep in val for sep in ['-', '/', '.', 'T']):
            continue
        try:
            from dateutil import parser as du_parser
            dt = du_parser.parse(val, fuzzy=False)
            if 1900 <= dt.year <= 2100:
                parsed += 1
        except Exception:
            pass
    return parsed >= max(len(sample) * 0.6, 1)


def _smart_parse_dates(series: pd.Series) -> pd.Series:
    """
    Parse a date column trying multiple strategies:
    1. pd.to_datetime default (handles YYYY-MM-DD, ISO, etc.)
    2. pd.to_datetime with dayfirst=True (handles DD-MM-YYYY, DD/MM/YYYY)
    3. dateutil.parser as final fallback for remaining unparsed values
    Returns series formatted as YYYY/MM/DD strings.
    """
    original = series.copy()
    str_vals = original.astype(str).str.strip()
    str_vals = str_vals.replace({'nan': '', 'None': '', 'NaN': '', 'NaT': ''})

    # Track which values are non-empty
    non_empty_mask = str_vals != ''
    result = pd.Series('', index=series.index)

    if non_empty_mask.sum() == 0:
        return result

    # Strategy 1: Try default pd.to_datetime (handles YYYY-MM-DD, ISO)
    parsed1 = pd.to_datetime(str_vals[non_empty_mask], errors='coerce')
    success1 = parsed1.notna()
    result.loc[success1[success1].index] = parsed1[success1].dt.strftime('%Y/%m/%d')

    # Strategy 2: For remaining unparsed, try dayfirst=True (DD-MM-YYYY)
    remaining = non_empty_mask & ~success1.reindex(series.index, fill_value=False)
    if remaining.sum() > 0:
        parsed2 = pd.to_datetime(str_vals[remaining], errors='coerce', dayfirst=True)
        success2 = parsed2.notna()
        result.loc[success2[success2].index] = parsed2[success2].dt.strftime('%Y/%m/%d')
        remaining = remaining & ~success2.reindex(series.index, fill_value=False)

    # Strategy 3: dateutil.parser for any still remaining
    if remaining.sum() > 0:
        from dateutil import parser as du_parser
        for idx in remaining[remaining].index:
            try:
                dt = du_parser.parse(str(str_vals.loc[idx]), dayfirst=True, fuzzy=True)
                result.loc[idx] = dt.strftime('%Y/%m/%d')
            except Exception:
                result.loc[idx] = str_vals.loc[idx]  # Keep original if unparseable

    return result


def fast_normalize_dates(df: pd.DataFrame, explicit_cols: Set[str]) -> pd.DataFrame:
    """
    Vectorized date normalization with smart multi-format parsing.
    Handles DD-MM-YYYY, MM-DD-YYYY, YYYY-MM-DD, and other formats.
    """
    # 1. Identify date columns (Explicit + Auto-detected)
    target_cols = [col for col in df.columns if col in explicit_cols or pd.api.types.is_datetime64_any_dtype(df[col])]

    for col in target_cols:
        df[col] = _smart_parse_dates(df[col])

    return df

def fast_generate_key(df: pd.DataFrame, columns: List[str]) -> pd.Series:
    """
    Vectorized key generation. 
    Replaces slow .agg('|'.join, axis=1) with vectorized string addition.
    """
    if not columns:
        return pd.Series([""] * len(df), index=df.index)
    res = df[columns[0]].astype(str).fillna("")
    for col in columns[1:]:
        res = res + "|" + df[col].astype(str).fillna("")
        
    return res



MAX_COLUMNS_PER_SHEET = 450

def enforce_sheet_column_limit(df: pd.DataFrame, sheet_name: str):
    col_count = df.shape[1]
    if col_count > MAX_COLUMNS_PER_SHEET:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Sheet '{sheet_name}' exceeds column limit. "
                f"Found {col_count} columns, maximum allowed is {MAX_COLUMNS_PER_SHEET}. "
                f"Please reduce mappings or included columns."
            )
        )

EXCEL_MAX_ROWS = 1_048_000

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    for ch in r'\/*?:[]':
        name = name.replace(ch, '_')
    return name[:max_len]

def write_df_excel_paginated(
    writer,
    df: pd.DataFrame,
    base_sheet_name: str,
    max_rows: int = EXCEL_MAX_ROWS
):
    # Use 28-char base so paginated variants (_N, _NN) stay within Excel's 31-char limit.
    safe_base = _safe_sheet_name(base_sheet_name, max_len=28)

    if len(df) <= max_rows:
        df.to_excel(writer, index=False, sheet_name=safe_base)
        return [safe_base]

    sheet_names = []
    for idx, start in enumerate(range(0, len(df), max_rows), start=1):
        sheet_name = f"{safe_base}_{idx}"
        df.iloc[start:start + max_rows].to_excel(
            writer,
            index=False,
            sheet_name=sheet_name
        )
        sheet_names.append(sheet_name)

    return sheet_names


@app.post("/api/excel/post_validation/validate")
async def post_validation_excel(
    background_tasks: BackgroundTasks,
    legacyFile: UploadFile = File(...),
    oracleFile: UploadFile = File(...),
    customerName: str = Form(...),
    instanceName: str = Form(...),
    mappings: str = Form(...),
    keyColumns: str = Form(...),
    includedColumns: str = Form(default="[]"),
    dateColumns: str = Form(default="[]"),
    timestampColumns: str = Form(default="[]"),
    dateColumnstarget: str = Form(default="[]"),
    timestampColumnstarget: str = Form(default="[]"),
    legacySheet: str = Form(default=None),
    oracleSheet: str = Form(default=None),
    includeSourceTargetFiles: bool = Form(default=False),
    sourceLabel: str = Form(default="Source"),
    targetLabel: str = Form(default="Target"),
    caseSensitive: bool = Form(default=True)
):
    """
    OPTIMIZED Post-validation endpoint with FAST STYLING.
    Fixes: 
    1. Comment columns styled Orange with Borders.
    2. ROBUST String comparison (ignores 100 vs 100.0 differences).
    3. Consistent Excel Engine usage.
    4. PRE-NORMALIZES Key Columns to ensure correct row merging across systems.
    5. IMPROVED NUMERIC DETECTION for sparse columns (prevents false positive string mismatches).
    """
    logger.info("Starting ULTRA-OPTIMIZED validation (Polars-native pipeline)...")
    start_time = time.time()
    src_label = (sourceLabel or "Source").strip() or "Source"
    tgt_label = (targetLabel or "Target").strip() or "Target"
    case_sensitive = caseSensitive

    temp_dir = tempfile.mkdtemp()
    main_output_path = os.path.join(temp_dir, "PostValidation_Main.xlsx")
    source_target_csv_path = os.path.join(temp_dir, "SourceTarget_Data.csv")

    INTERNAL_KEY = "_derived_key"

    try:
        # --- [1. Parse Inputs] ---
        mappings_dict: Dict[str, str] = json.loads(mappings)
        included_cols_list: List[str] = json.loads(includedColumns)
        key_cols_list: List[str] = json.loads(keyColumns)
        
        # --- [1.1 Build Configuration DataFrame (New Tab Logic)] ---
        # Reconstructs the UI table state based on the inputs received
        config_rows = []
        
        # Combine date lists for easy lookup
        legacy_date_list = json.loads(dateColumns) + json.loads(timestampColumns)
        target_date_list = json.loads(dateColumnstarget) + json.loads(timestampColumnstarget)
        legacy_date_set = set(legacy_date_list)
        target_date_set = set(target_date_list)
        
        for source_col, target_col in mappings_dict.items():
            is_key = source_col in key_cols_list
            # Check if either source or target is marked as date
            is_date = (source_col in legacy_date_set) or (target_col in target_date_set)
            is_included = source_col in included_cols_list
            
            config_rows.append({
                f"{src_label} Column": source_col,
                f"{tgt_label} Column": target_col,
                "Is Key?": "Yes" if is_key else "No",
                "Is Date?": "Yes" if is_date else "No",
                "Validate": "Yes", # Implicitly Yes since it is in the mappings dict
                "Include in Report": "Yes" if is_included else "No"
            })
        
        config_df = pd.DataFrame(config_rows)
        # Enforce column limits on this new DF
        if len(config_df.columns) > 0:
             enforce_sheet_column_limit(config_df, "Configuration")

        legacy_date_cols = set(json.loads(dateColumns) + json.loads(timestampColumns))
        target_date_cols = set(json.loads(dateColumnstarget) + json.loads(timestampColumnstarget))

        if not mappings_dict: raise ValueError("Mappings is empty")
        if not key_cols_list: raise ValueError("Key columns list is empty")

        # --- [2. Read Files — Polars Native Engine (5-10x faster)] ---
        import polars as pl
        import gc
        t_io = time.time()
        legacy_path = os.path.join(temp_dir, f"src_{legacyFile.filename}")
        oracle_path = os.path.join(temp_dir, f"tgt_{oracleFile.filename}")
        legacy_sheet_param = legacySheet if legacySheet and legacySheet.strip() else None
        oracle_sheet_param = oracleSheet if oracleSheet and oracleSheet.strip() else None

        # Stream uploads to disk (8MB chunks — memory efficient)
        for upload, dest in [(legacyFile, legacy_path), (oracleFile, oracle_path)]:
            with open(dest, "wb") as f:
                while True:
                    chunk = await upload.read(8 * 1024 * 1024)
                    if not chunk: break
                    f.write(chunk)

        # Polars native read (multi-threaded, zero-copy)
        logger.info(f"Reading files with Polars native... Legacy Sheet: {legacySheet}, Oracle Sheet: {oracleSheet}")
        _lf_legacy = _polars_read_file(legacy_path, legacy_sheet_param)
        _lf_oracle = _polars_read_file(oracle_path, oracle_sheet_param)
        legacy_df = _lf_legacy.collect()
        oracle_df = _lf_oracle.collect()
        del _lf_legacy, _lf_oracle; gc.collect()
        legacy_row_count = len(legacy_df)
        oracle_row_count = len(oracle_df)
        logger.info(f"Files read in {time.time() - t_io:.2f}s — Legacy: {legacy_row_count:,}, Oracle: {oracle_row_count:,} rows")

        # --- [2.1] Align Oracle columns to Legacy column space (case-insensitive) ---
        oracle_to_legacy_map = {v: k for k, v in mappings_dict.items()}
        # Build case-insensitive lookup for oracle rename
        oracle_to_legacy_lower = {v.strip().lower(): k for k, v in mappings_dict.items()}

        cols_to_rename = {}
        for col in oracle_df.columns:
            if col in oracle_to_legacy_map:
                cols_to_rename[col] = oracle_to_legacy_map[col]
            elif col.strip().lower() in oracle_to_legacy_lower:
                cols_to_rename[col] = oracle_to_legacy_lower[col.strip().lower()]

        oracle_renamed = oracle_df.rename(cols_to_rename) if cols_to_rename else oracle_df
        del oracle_df

        # Also normalise legacy column names to match mapping keys (case-insensitive)
        legacy_mapped_lower = {k.strip().lower(): k for k in mappings_dict.keys()}
        legacy_cols_rename = {}
        for col in legacy_df.columns:
            lk = col.strip().lower()
            if lk in legacy_mapped_lower and col != legacy_mapped_lower[lk]:
                legacy_cols_rename[col] = legacy_mapped_lower[lk]
        if legacy_cols_rename:
            legacy_df = legacy_df.rename(legacy_cols_rename)

        # Resolve key_cols_list against actual oracle columns (case-insensitive)
        oracle_col_lower = {c.strip().lower(): c for c in oracle_renamed.columns}
        legacy_col_lower = {c.strip().lower(): c for c in legacy_df.columns}
        key_cols_list = [
            oracle_col_lower.get(k.strip().lower(), k) for k in key_cols_list
        ]
        # Rebuild mappings_dict to match actual column names
        resolved_mappings = {}
        for l_col, o_col in mappings_dict.items():
            actual_l = legacy_col_lower.get(l_col.strip().lower(), l_col)
            actual_o = oracle_col_lower.get(l_col.strip().lower(), l_col)
            resolved_mappings[actual_l] = o_col
        mappings_dict = resolved_mappings

        missing_keys = [k for k in key_cols_list if k not in oracle_renamed.columns]
        if missing_keys:
            raise HTTPException(
                status_code=400,
                detail=f"Key columns missing in Oracle after rename: {missing_keys}"
            )

        if len(legacy_df.columns) > 450 or len(oracle_renamed.columns) > 450:
             raise HTTPException(status_code=400, detail="File has too many columns. Max allowed is 450.")

        # Validate columns existence
        missing = []
        for l, o in mappings_dict.items():
            if l not in legacy_df.columns: missing.append(f"{src_label} column '{l}' not found in sheet '{legacy_sheet_param}'")
            if l not in oracle_renamed.columns: missing.append(f"{tgt_label} column '{o}' not found in sheet '{oracle_sheet_param}'")
        
        if missing:
            raise HTTPException(status_code=400, detail={"errors": missing})

        # Resolve included_cols_list and date column sets to match actual names
        included_cols_list = [
            oracle_col_lower.get(c.strip().lower(), legacy_col_lower.get(c.strip().lower(), c))
            for c in included_cols_list
        ]
        legacy_date_cols = {
            legacy_col_lower.get(c.strip().lower(), c) for c in legacy_date_cols
        }
        oracle_to_legacy_lower_map = {v.strip().lower(): k for k, v in
                                       {k: v for k, v in zip(mappings_dict.keys(), [oracle_to_legacy_map.get(k, k) for k in mappings_dict.values()])}.items()}
        target_date_cols_resolved = set()
        for c in target_date_cols:
            # target dates are in oracle-original space; map to legacy then resolve
            mapped = oracle_to_legacy_map.get(c, c)
            actual = oracle_col_lower.get(mapped.strip().lower(), legacy_col_lower.get(mapped.strip().lower(), mapped))
            target_date_cols_resolved.add(actual)
        target_date_cols = target_date_cols_resolved

        # --- [POLARS-NATIVE] Normalize Key Columns ---
        logger.info("Normalizing Key Columns (Polars)...")
        key_norm_exprs = [
            pl.col(col).cast(pl.Utf8).fill_null("")
            .str.strip_chars()
            .str.replace_all(r"\.0+$", "")
            .str.replace_all("^nan$", "")
            .str.replace_all("^None$", "")
            .str.replace_all("^NaN$", "")
            .alias(col)
            for col in key_cols_list
            if col in legacy_df.columns
        ]
        if key_norm_exprs:
            legacy_df = legacy_df.with_columns(key_norm_exprs)
            oracle_renamed = oracle_renamed.with_columns(key_norm_exprs)

        # --- [POLARS-NATIVE] Date Normalization (parallel — both sides at once) ---
        logger.info("Normalizing dates (Polars, parallel, auto-detection)...")
        _pre_cols_to_compare = list(mappings_dict.keys())
        _pre_cols_to_compare = [c for c in _pre_cols_to_compare if c in legacy_df.columns and c in oracle_renamed.columns]
        with _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="date-norm") as _dn_pool:
            _f_l_dn = _dn_pool.submit(
                _pl_apply_date_normalization, legacy_df, legacy_date_cols,
                list(legacy_df.columns), True, _pre_cols_to_compare
            )
            _f_o_dn = _dn_pool.submit(
                _pl_apply_date_normalization, oracle_renamed, target_date_cols,
                list(oracle_renamed.columns), True, _pre_cols_to_compare
            )
            legacy_df = _f_l_dn.result()
            oracle_renamed = _f_o_dn.result()

        # --- [POLARS-NATIVE] Key Generation ---
        logger.info("Generating keys (Polars)...")
        key_expr = pl.col(key_cols_list[0]).cast(pl.Utf8).fill_null("")
        for c in key_cols_list[1:]:
            key_expr = key_expr + pl.lit("|") + pl.col(c).cast(pl.Utf8).fill_null("")
        legacy_df = legacy_df.with_columns(key_expr.alias(INTERNAL_KEY))
        oracle_renamed = oracle_renamed.with_columns(key_expr.alias(INTERNAL_KEY))

        # --- [4.5 Positional row-numbering for duplicate keys] ---
        # When multiple rows share the same composite key (e.g. GL journal lines),
        # add a row number within each key group so row-1 of key "A" in legacy
        # matches row-1 of key "A" in oracle.  Prevents cross-join explosion.
        _leg_dup_count = legacy_df.select(pl.col(INTERNAL_KEY).is_duplicated().sum()).item()
        _orc_dup_count = oracle_renamed.select(pl.col(INTERNAL_KEY).is_duplicated().sum()).item()
        if _leg_dup_count > 0 or _orc_dup_count > 0:
            logger.info(f"Duplicate keys found (Legacy: {_leg_dup_count:,}, Oracle: {_orc_dup_count:,}) — adding positional row numbers")
            legacy_df = legacy_df.with_columns(
                pl.int_range(pl.len()).over(INTERNAL_KEY).cast(pl.Utf8).alias("_rn")
            )
            oracle_renamed = oracle_renamed.with_columns(
                pl.int_range(pl.len()).over(INTERNAL_KEY).cast(pl.Utf8).alias("_rn")
            )
            legacy_df = legacy_df.with_columns(
                (pl.col(INTERNAL_KEY) + pl.lit("|_rn=") + pl.col("_rn")).alias(INTERNAL_KEY)
            ).drop("_rn")
            oracle_renamed = oracle_renamed.with_columns(
                (pl.col(INTERNAL_KEY) + pl.lit("|_rn=") + pl.col("_rn")).alias(INTERNAL_KEY)
            ).drop("_rn")
            logger.info(f"Positional keys assigned — Legacy: {len(legacy_df):,}, Oracle: {len(oracle_renamed):,}")

        # --- [5. Comparison Logic — POLARS MULTI-THREADED ENGINE] ---
        logger.info("Comparing data (Polars multi-threaded engine)...")
        t_compare = time.time()

        cols_to_compare = list(mappings_dict.keys())
        cols_to_compare = [c for c in cols_to_compare if c in legacy_df.columns and c in oracle_renamed.columns]

        # Already Polars DataFrames — select only needed columns for comparison
        pl_legacy = legacy_df.select([INTERNAL_KEY] + cols_to_compare)
        pl_oracle = oracle_renamed.select([INTERNAL_KEY] + cols_to_compare)

        # Free pandas DataFrames for comparison columns (keep originals for missing records)
        gc.collect()

        # --- Detect numeric columns (Polars-native, no pandas) ---
        numeric_columns = _pl_detect_numeric_cols(pl_legacy, pl_oracle, cols_to_compare)
        logger.info(f"Numeric columns detected: {numeric_columns}")

        # --- INNER JOIN on _derived_key (Polars hash join — O(n)) ---
        joined = pl_legacy.join(pl_oracle, on=INTERNAL_KEY, suffix="_T", how="inner")
        del pl_oracle; gc.collect()
        logger.info(f"Polars join complete: {len(joined):,} matched rows in {time.time() - t_compare:.2f}s")

        # --- Build clean + diff expressions for ALL columns at once (multi-threaded) ---
        diff_exprs = []
        for col in cols_to_compare:
            l_col = col
            o_col = f"{col}_T"
            if col in numeric_columns:
                diff_exprs.append(_pl_clean_num_expr(l_col).alias(f"__ln_{col}"))
                diff_exprs.append(_pl_clean_num_expr(o_col).alias(f"__on_{col}"))
            else:
                diff_exprs.append(_pl_clean_str_expr(l_col, case_sensitive).alias(f"__ls_{col}"))
                diff_exprs.append(_pl_clean_str_expr(o_col, case_sensitive).alias(f"__os_{col}"))

        # Materialize all cleaned columns in ONE pass (multi-threaded)
        joined_clean = joined.with_columns(diff_exprs)

        # --- Build boolean diff mask columns ---
        mask_exprs = []
        for col in cols_to_compare:
            if col in numeric_columns:
                ln = f"__ln_{col}"
                on = f"__on_{col}"
                mask_exprs.append(
                    (
                        (pl.col(ln).is_not_null() & pl.col(on).is_not_null()
                         & ((pl.col(ln) - pl.col(on)).abs() > 0.0001))
                        | (pl.col(ln).is_null() ^ pl.col(on).is_null())
                    ).alias(f"__diff_{col}")
                )
            else:
                ls = f"__ls_{col}"
                os_ = f"__os_{col}"
                mask_exprs.append(
                    (pl.col(ls).fill_null("") != pl.col(os_).fill_null("")).alias(f"__diff_{col}")
                )

        joined_diffs = joined_clean.with_columns(mask_exprs)

        # --- [6. Extract Discrepancies — Polars unpivot] ---
        diff_col_names = [f"__diff_{col}" for col in cols_to_compare]
        # Check if ANY diffs exist (fast boolean reduce)
        has_any_diff = joined_diffs.select(diff_col_names).sum().row(0)
        total_diff_count = sum(has_any_diff)

        if total_diff_count > 0:
            # Extract discrepancies by building union of per-column filters
            disc_parts = []
            for col in cols_to_compare:
                diff_flag = f"__diff_{col}"
                col_label = f"{col} - {mappings_dict.get(col, col)}"
                filtered = joined_diffs.filter(pl.col(diff_flag))
                if len(filtered) == 0:
                    continue
                part = filtered.select([
                    pl.col(INTERNAL_KEY),
                    pl.lit(col_label).alias("Column Name"),
                    pl.col(col).cast(pl.Utf8).fill_null("").alias(f"{src_label} Value"),
                    pl.col(f"{col}_T").cast(pl.Utf8).fill_null("").alias(f"{tgt_label} Value"),
                ])
                disc_parts.append(part)

            if disc_parts:
                discrepancies_pl = pl.concat(disc_parts)
            else:
                discrepancies_pl = pl.DataFrame({
                    INTERNAL_KEY: [], "Column Name": [],
                    f"{src_label} Value": [], f"{tgt_label} Value": []
                })

            # Add context columns (key + included) — Polars native
            context_cols = list(dict.fromkeys(key_cols_list + included_cols_list))
            valid_context_cols = [c for c in context_cols if c in legacy_df.columns]
            if valid_context_cols:
                ctx_pl = legacy_df.select([INTERNAL_KEY] + valid_context_cols).unique(subset=[INTERNAL_KEY])
                discrepancies_pl = discrepancies_pl.join(ctx_pl, on=INTERNAL_KEY, how="left")
                del ctx_pl

            # Convert to pandas for Excel output
            validation_df = discrepancies_pl.drop(INTERNAL_KEY).to_pandas()

            # Order columns
            final_report_cols = key_cols_list + [c for c in included_cols_list if c not in key_cols_list] + ["Column Name", f"{src_label} Value", f"{tgt_label} Value"]
            final_report_cols = [c for c in final_report_cols if c in validation_df.columns]
            validation_df = validation_df[final_report_cols]
        else:
            validation_df = pd.DataFrame([{"Status": "All mapped columns matched perfectly"}])
        
        # --- SORT Data Discrepancies BEFORE pagination ---
        if "Column Name" in validation_df.columns and not validation_df.empty:
            sort_cols = ["Column Name"] + [
                c for c in key_cols_list if c in validation_df.columns
            ]
            validation_df = validation_df.sort_values(
                by=sort_cols,
                kind="mergesort"
            ).reset_index(drop=True)

        # --- [NEW] Add Comment Columns to Discrepancies ---
        comment_cols = ["Mythics Comments", "Oracle Comments", "ParkView Comments"]
        for col in comment_cols:
            validation_df[col] = ""

        # --- [6.1] Discrepancy Count Per Column (for Summary) ---
        column_discrepancy_counts = []
        logger.info("Calculating discrepancy counts per column...")
        if "Status" not in validation_df.columns and not validation_df.empty:
            col_counts = (
                validation_df["Column Name"]
                .value_counts()
                .sort_values(ascending=False)
            )
            for col_name, count in col_counts.items():
                column_discrepancy_counts.append(["", col_name, int(count), "", "", ""])

        # --- [7. Missing Records — Polars ANTI JOIN (O(n) hash join)] ---
        t_miss = time.time()
        pl_matched_keys = pl.DataFrame({INTERNAL_KEY: joined_diffs[INTERNAL_KEY]})

        # Anti-join directly on Polars DataFrames (zero conversion overhead)
        legacy_only_pl = legacy_df.join(pl_matched_keys, on=INTERNAL_KEY, how="anti").drop(INTERNAL_KEY)
        oracle_only_pl = oracle_renamed.join(pl_matched_keys, on=INTERNAL_KEY, how="anti").drop(INTERNAL_KEY)
        del pl_matched_keys

        # Convert small missing-record sets to pandas for Excel output
        legacy_only_df = legacy_only_pl.to_pandas()
        oracle_only_df = oracle_only_pl.to_pandas()
        logger.info(f"Missing records: {len(legacy_only_df):,} in Oracle, {len(oracle_only_df):,} in PS — {time.time() - t_miss:.2f}s")

        # --- [NEW] Add Comment Columns to Missing Record DFs ---
        for col in comment_cols:
            legacy_only_df[col] = ""
            oracle_only_df[col] = ""

        # --- [8. Full Data Report — CSV directly from Polars (no pandas overhead)] ---
        t_csv = time.time()

        # Convert small missing-record DFs to Polars for the CSV writer
        _leg_only_pl = legacy_only_pl
        _orc_only_pl = oracle_only_pl

        _polars_write_source_target_csv(
            joined_diffs, _leg_only_pl, _orc_only_pl,
            cols_to_compare, source_target_csv_path, INTERNAL_KEY
        )
        logger.info(f"Full data CSV written in {time.time() - t_csv:.2f}s")

        # Free large Polars DataFrames (row counts already saved)
        del joined, joined_diffs, joined_clean, pl_legacy
        del _leg_only_pl, _orc_only_pl
        del legacy_df, oracle_renamed
        gc.collect()

        # --- [9. Generate Summary] ---
        total_discrepancies = len(validation_df) if "Status" not in validation_df.columns else 0
        count_missing_ps = len(legacy_only_df)
        count_missing_oc = len(oracle_only_df)
        grand_total = count_missing_ps + count_missing_oc + total_discrepancies

        summary_data = [
            ["", "Comparison Statistics", "", "", "", ""],
            ["", f"{src_label} File Name", legacyFile.filename, "", "", ""],
            ["", f"{src_label} Records Count", legacy_row_count, "", "", ""],
            ["", f"{tgt_label} File Name", oracleFile.filename, "", "", ""],
            ["", f"{tgt_label} Records Count", oracle_row_count, "", "", ""],
            ["", "Validation DateTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", ""],
            ["", "", "", "", "", ""],

            ["", "Missing Records Summary", "", "Mythics Comments", "Oracle Comments", "ParkView Comments"],
            ["", f"Records Missing in {src_label}", count_missing_oc, "", "", ""],
            ["", f"Records Missing in {tgt_label}", count_missing_ps, "", "", ""],
            ["", "Total Missing Records", count_missing_ps + count_missing_oc, "", "", ""],
            ["", "", "", "", "", ""],

            ["", "Data Discrepancies Summary", "", "Mythics Comments", "Oracle Comments", "ParkView Comments"],
            *column_discrepancy_counts,
            ["", "Total Data Discrepancies", total_discrepancies, "", "", ""],
            ["", "", "", "", "", ""],

            ["", "Total Validation Issues", grand_total, "", "", ""]
        ]

        summary_df = pd.DataFrame(summary_data)

        # --- [10. Write to Excel with Perfect Styling] ---


        logger.info("Writing and Styling Excel...")
        

        def style_configuration_sheet(workbook, sheet_name="Configuration"):
            if sheet_name not in workbook.sheetnames:
                return

            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"

            # Header styling
            for cell in ws[1]:
                cell.fill = fill_green
                cell.font = font_white
                cell.alignment = align_center
                cell.border = border_thin

            # Body styling
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = font_normal
                    cell.border = border_thin

                    # Center align Yes / No / Validate columns
                    if str(cell.value).strip().lower() in {"yes", "no"}:
                        cell.alignment = align_center

            # Auto column width
            for col in ws.iter_cols(max_row=ws.max_row):
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 45)

        sheet_missing_ps = _safe_sheet_name(f"Missing in {src_label}")
        sheet_missing_oc = _safe_sheet_name(f"Missing in {tgt_label}")
        sheet_discrepancies = "Data Discrepancies"
        sheet_full_data = _safe_sheet_name(f"{src_label} - {tgt_label} Data", max_len=28)

        enforce_sheet_column_limit(summary_df, "Summary")
        enforce_sheet_column_limit(validation_df, "Data Discrepancies")
        enforce_sheet_column_limit(legacy_only_df, f"Missing in {tgt_label}")
        enforce_sheet_column_limit(oracle_only_df, f"Missing in {src_label}")

        # Define Styling Functions (same as before but included for completeness)
        font_white = Font(name="Calibri", size=8, color="FFFFFF", bold=True)
        font_black = Font(name="Calibri", size=8, color="000000", bold=True)
        font_bold_black = Font(name="Calibri", size=8, bold=True)
        font_normal = Font(name="Calibri", size=8)
        fill_header_ps = PatternFill("solid", fgColor="1F497D")
        fill_header_oc = PatternFill("solid", fgColor="31869B")
        fill_header_err = PatternFill("solid", fgColor="C0504D")
        fill_green = PatternFill("solid", fgColor="00B050")
        fill_grey = PatternFill("solid", fgColor="D9D9D9")
        fill_orange = PatternFill("solid", fgColor="FF9900")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal="center", vertical="center")

        def style_full_data_header_only(ws, num_comparison_cols):
            border_ps_last  = Border(left=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="medium"))
            border_oc_first = Border(left=Side(style="medium"), top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"))
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = font_white
                cell.alignment = align_center
                if cell.column <= num_comparison_cols:
                    cell.fill = fill_header_ps
                    cell.border = border_ps_last if cell.column == num_comparison_cols else border_thin
                else:
                    cell.fill = fill_header_oc
                    cell.border = border_oc_first if cell.column == num_comparison_cols + 1 else border_thin
            for col in ws.iter_cols(max_row=50):
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                if col[0].value: max_length = len(str(col[0].value))
                for cell in col[1:]:
                    if cell.value: max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max((max_length + 2) * 1.2, 10), 60)

        def style_sheet_header(workbook, sheet_name, fill_color):
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                ws.freeze_panes = "A2"
                comment_col_indices = []
                for cell in ws[1]:
                    cell.fill = fill_color
                    cell.font = font_white
                    cell.alignment = align_center
                    cell.border = border_thin
                    if cell.value in comment_cols:
                        cell.fill = fill_orange
                        cell.font = font_black
                        comment_col_indices.append(cell.column)
                # Body font (capped at 5000 rows for large sheets)
                body_cap = min(ws.max_row, 5000)
                for body_row in ws.iter_rows(min_row=2, max_row=body_cap):
                    for cell in body_row:
                        cell.font = font_normal
                # Auto-width: only sample first 50 rows for speed
                for col in ws.iter_cols(max_row=min(50, ws.max_row)):
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    if col[0].value: max_length = len(str(col[0].value))
                    for cell in col[1:]:
                        if cell.value: max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = min(max((max_length + 2) * 1.2, 10), 60)
                # Comment column borders — cap at 5000 rows to avoid O(n) styling
                if comment_col_indices:
                    border_limit = min(ws.max_row, 5000)
                    for col_idx in comment_col_indices:
                        for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=border_limit):
                            for cell in col_cells:
                                cell.border = border_thin

        # Generate Files — optimized: full data as CSV, styled report as xlsx
        with pd.ExcelWriter(main_output_path, engine="openpyxl") as main_writer:
            summary_df.to_excel(main_writer, index=False, header=False, sheet_name="Summary")
            config_df.to_excel(main_writer, index=False, sheet_name="Configuration")
            oracle_only_df.to_excel(main_writer, index=False, sheet_name=sheet_missing_ps)
            legacy_only_df.to_excel(main_writer, index=False, sheet_name=sheet_missing_oc)

            write_df_excel_paginated(main_writer, validation_df, sheet_discrepancies)
            main_workbook = main_writer.book
            if includeSourceTargetFiles:
                # Load from CSV (capped at 1M rows for Excel)
                try:
                    _st_df = pd.read_csv(source_target_csv_path, dtype=str, nrows=EXCEL_MAX_ROWS)
                    write_df_excel_paginated(main_writer, _st_df, sheet_full_data)
                    del _st_df
                except Exception as _csv_err:
                    logger.warning(f"Could not include source/target data: {_csv_err}")

            # Apply Styles
            style_sheet_header(main_workbook, sheet_missing_ps, fill_header_ps)
            style_sheet_header(main_workbook, sheet_missing_oc, fill_header_oc)
            for sheet in main_workbook.sheetnames:
                if sheet.startswith(sheet_discrepancies):
                    style_sheet_header(main_workbook, sheet, fill_header_err)
            style_configuration_sheet(main_workbook, "Configuration")
            main_workbook["Configuration"].sheet_state = "hidden"
            if includeSourceTargetFiles:
                for sheet_name in main_workbook.sheetnames:
                    if sheet_name.startswith(sheet_full_data):
                        style_full_data_header_only(main_workbook[sheet_name], len(cols_to_compare))

            # Summary Styling
            ws_sum = main_workbook["Summary"]
            ws_sum.sheet_view.showGridLines = False
            ws_sum.column_dimensions['A'].width = 2
            ws_sum.column_dimensions['B'].width = 45
            for c_char in ['C', 'D', 'E', 'F']: ws_sum.column_dimensions[c_char].width = 25
            
            for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row):
                if len(row) < 3: continue
                for cell in row:
                    cell.font = font_normal
                cell_b = row[1]
                if not cell_b.value: continue
                cell_b.border = border_thin
                row[2].border = border_thin
                comment_cells = [ws_sum.cell(row=cell_b.row, column=c) for c in [4, 5, 6]]

                if cell_b.value in ["Missing Records Summary", "Data Discrepancies Summary"]:
                    ws_sum.merge_cells(start_row=cell_b.row, start_column=2, end_row=cell_b.row, end_column=3)
                    cell_b.fill = fill_green
                    cell_b.font = font_white
                    cell_b.alignment = align_center
                    ws_sum.row_dimensions[cell_b.row].height = 20
                    for c in comment_cells:
                        c.fill = fill_orange
                        c.font = font_black
                        c.alignment = align_center
                        c.border = border_thin
                elif "Total" in str(cell_b.value) or "Comparison Statistics" in str(cell_b.value):
                    if "Comparison Statistics" in str(cell_b.value):
                         ws_sum.merge_cells(start_row=cell_b.row, start_column=2, end_row=cell_b.row, end_column=3)
                         cell_b.alignment = align_center
                         ws_sum.row_dimensions[cell_b.row].height = 20
                         cell_b.fill = fill_green
                         cell_b.font = font_white
                    else:
                        cell_b.fill = fill_grey
                        row[2].fill = fill_grey
                        cell_b.font = font_bold_black
                        row[2].font = font_bold_black
                        row[2].alignment = align_center
                else:
                    cell_b.fill = fill_green
                    cell_b.font = font_white
                    row[2].alignment = align_center
                    for c in comment_cells: c.border = border_thin

        logger.info(f"Process completed in {time.time() - start_time:.2f} seconds.")

        def _clean(path):
            try: shutil.rmtree(path, ignore_errors=True)
            except: pass

        background_tasks.add_task(_clean, temp_dir)
        report_ts = datetime.now().strftime('%Y%m%d_%H%M%S')

        return FileResponse(
            main_output_path,
            filename=f"MythicsValidationResults_{report_ts}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Processing Error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# =====================================================================================
# LARGE-SCALE POST-VALIDATION ENDPOINT (Polars Engine — supports 10M+ rows)
# =====================================================================================
# Async job-based architecture:
#   1. POST /validate_large  → saves files, returns { job_id } immediately
#   2. GET  /status/{job_id} → returns { progress, stage, status, error? }
#   3. GET  /download/{job_id} → streams the result .xlsx once status == "complete"
# =====================================================================================


# ── In-memory job tracker ────────────────────────────────────────────────
import threading as _threading
import concurrent.futures as _cf

_validation_jobs: Dict[str, Dict[str, Any]] = {}
_validation_jobs_lock = _threading.Lock()

# Bounded executor: prevents OOM from concurrent large validation jobs.
# Configurable via VALIDATION_MAX_WORKERS env var (default: 2).
_VALIDATION_EXECUTOR = _cf.ThreadPoolExecutor(
    max_workers=int(os.environ.get("VALIDATION_MAX_WORKERS", "2")),
    thread_name_prefix="validation",
)

def _job_update(job_id: str, **kwargs):
    """Thread-safe update of a job's progress dict."""
    with _validation_jobs_lock:
        if job_id in _validation_jobs:
            _validation_jobs[job_id].update(kwargs)
            job = _validation_jobs[job_id]
            job["eta_seconds"] = _compute_smooth_eta(job)

def _job_get(job_id: str) -> Optional[Dict]:
    with _validation_jobs_lock:
        return _validation_jobs.get(job_id, {}).copy()


# ── Stale job cleanup (runs on every new job submission) ─────────────
_STALE_JOB_MAX_AGE = 60 * 60  # 1 hour

def _cleanup_stale_jobs():
    """Remove completed/failed jobs older than 1 hour to prevent memory leaks."""
    now = time.time()
    stale_ids = []
    with _validation_jobs_lock:
        for jid, job in _validation_jobs.items():
            if job.get("status") in ("complete", "failed"):
                age = now - job.get("started_at", now)
                if age > _STALE_JOB_MAX_AGE:
                    stale_ids.append(jid)
        for jid in stale_ids:
            job = _validation_jobs.pop(jid, {})
            td = job.get("temp_dir")
            if td and os.path.exists(td):
                shutil.rmtree(td, ignore_errors=True)
    if stale_ids:
        logger.info(f"Cleaned up {len(stale_ids)} stale validation job(s)")

@app.post("/api/excel/post_validation/validate_large")
async def post_validation_large_scale(
    background_tasks: BackgroundTasks,
    legacyFile: UploadFile = File(...),
    oracleFile: UploadFile = File(...),
    customerName: str = Form(...),
    instanceName: str = Form(...),
    mappings: str = Form(...),
    keyColumns: str = Form(...),
    includedColumns: str = Form(default="[]"),
    dateColumns: str = Form(default="[]"),
    timestampColumns: str = Form(default="[]"),
    dateColumnstarget: str = Form(default="[]"),
    timestampColumnstarget: str = Form(default="[]"),
    legacySheet: str = Form(default=None),
    oracleSheet: str = Form(default=None),
    includeSourceTargetFiles: bool = Form(default=False),
    sourceLabel: str = Form(default="Source"),
    targetLabel: str = Form(default="Target"),
    caseSensitive: bool = Form(default=True)
):
    """
    Accepts files & config, returns { job_id } immediately.
    Processing runs in a background thread with live progress tracking.
    Poll GET /api/excel/post_validation/status/{job_id} for progress.
    """
    if not POLARS_AVAILABLE:
        raise HTTPException(status_code=501, detail="Polars is not installed. Run: pip install polars")

    # Housekeeping: clean up stale jobs from previous runs
    _cleanup_stale_jobs()

    # ── Generate job ID & save files to disk ────────────────────────────
    job_id = str(uuid.uuid4())
    temp_dir = tempfile.mkdtemp()

    legacy_path = os.path.join(temp_dir, f"src_{legacyFile.filename}")
    oracle_path = os.path.join(temp_dir, f"tgt_{oracleFile.filename}")

    for upload, dest in [(legacyFile, legacy_path), (oracleFile, oracle_path)]:
        with open(dest, "wb") as fh:
            while True:
                chunk = await upload.read(8 * 1024 * 1024)
                if not chunk:
                    break
                fh.write(chunk)

    # ── Register job ────────────────────────────────────────────────────
    with _validation_jobs_lock:
        _validation_jobs[job_id] = {
            "status": "running",
            "progress": 0,
            "stage": "Uploading files",
            "error": None,
            "temp_dir": temp_dir,
            "zip_path": None,
            "zip_filename": None,  # legacy key names (now stores .xlsx path)
            "legacy_filename": legacyFile.filename,
            "oracle_filename": oracleFile.filename,
            "started_at": time.time(),
        }

    # ── Capture all form values (strings) for the background thread ────
    job_params = {
        "legacy_path": legacy_path,
        "oracle_path": oracle_path,
        "legacy_filename": legacyFile.filename,
        "oracle_filename": oracleFile.filename,
        "mappings": mappings,
        "keyColumns": keyColumns,
        "includedColumns": includedColumns,
        "dateColumns": dateColumns,
        "timestampColumns": timestampColumns,
        "dateColumnstarget": dateColumnstarget,
        "timestampColumnstarget": timestampColumnstarget,
        "legacySheet": legacySheet,
        "oracleSheet": oracleSheet,
        "includeSourceTargetFiles": includeSourceTargetFiles,
        "sourceLabel": sourceLabel,
        "targetLabel": targetLabel,
        "caseSensitive": caseSensitive,
        "temp_dir": temp_dir,
    }

    # Launch in the bounded executor — queues excess jobs instead of OOMing the server
    _VALIDATION_EXECUTOR.submit(_run_validation_job, job_id, job_params)

    return {"job_id": job_id}

# --- Data Transformation for the Source Code 
@app.post("/api/excel/post_validation/data_mapping")
async def post_validation_data_mapping(
    legacyFile: UploadFile = File(...),
    oracleFile: UploadFile = File(...),
    mappingFile: UploadFile = File(None),
    customerName: str = Form(...),
    instanceName: str = Form(...),
    legacySheet: str = Form(default=None),
    oracleSheet: str = Form(default=None),
):
    """Simple synchronous mapping endpoint used by the frontend.
    Returns column lists and a suggested mapping dict. If a mappingFile (JSON)
    is provided the server will return it as the suggested mapping.
    """
    try:
        legacy_bytes = await legacyFile.read()
        oracle_bytes = await oracleFile.read()

        # Use helper to parse into pandas DataFrame
        try:
            df_legacy = _read_file_bytes(legacy_bytes, legacyFile.filename)
            df_oracle = _read_file_bytes(oracle_bytes, oracleFile.filename)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read uploaded files: {e}")

        legacy_cols = [str(c) for c in list(df_legacy.columns)]
        oracle_cols = [str(c) for c in list(df_oracle.columns)]

        suggested = {}
        # If mappingFile provided and is Excel/CSV, try to read mapping pairs from it
        if mappingFile is not None:
            try:
                mf_bytes = await mappingFile.read()
                # Parse using existing helper which reads csv/xlsx
                df_map = _read_file_bytes(mf_bytes, mappingFile.filename)
                # Normalize column names
                cols_lc = {c.lower(): c for c in df_map.columns}
                # Common header name candidates for source and target
                src_cands = ["source", "legacy", "from", "src", "source_column", "sourcecol"]
                tgt_cands = ["target", "oracle", "to", "tgt", "target_column", "targetcol"]

                src_col = None
                tgt_col = None
                for cand in src_cands:
                    if cand in cols_lc:
                        src_col = cols_lc[cand]
                        break
                for cand in tgt_cands:
                    if cand in cols_lc:
                        tgt_col = cols_lc[cand]
                        break

                # If headers not found, but there are at least two columns, use first two
                if src_col is None or tgt_col is None:
                    if len(df_map.columns) >= 2:
                        src_col = src_col or df_map.columns[0]
                        tgt_col = tgt_col or df_map.columns[1]

                if src_col and tgt_col:
                    for _, row in df_map[[src_col, tgt_col]].dropna(how='all').iterrows():
                        s = str(row[src_col]).strip()
                        t = str(row[tgt_col]).strip()
                        if s and t:
                            suggested[s] = t
            except Exception:
                # ignore parse errors and fall back to heuristic
                suggested = {}

        # If no mapping file was provided, use Gemini AI for smart column mapping
        date_columns = []
        if not suggested:
            ai_result = get_smart_mapping_from_gemini(legacy_cols, oracle_cols)
            suggested = ai_result.get("mapping", {})
            date_columns = ai_result.get("date_columns", [])

        # Ensure mapping entries only apply to the uploaded source (legacy) columns.
        legacy_lower_map = {c.lower(): c for c in legacy_cols}
        filtered = {}
        for s, t in suggested.items():
            if s in legacy_cols:
                filtered[s] = t
            else:
                sl = s.lower()
                if sl in legacy_lower_map:
                    filtered[legacy_lower_map[sl]] = t
                # otherwise ignore mappings that don't reference the source file

        return {
            "legacy_columns": legacy_cols,
            "oracle_columns": oracle_cols,
            "suggested_mapping": filtered,
            "date_columns": date_columns,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- Value-Level Transformation Endpoint
@app.post("/api/excel/post_validation/transform")
async def post_validation_transform(
    sourceFile:  UploadFile = File(...),
    mappingFile: UploadFile = File(...),
):
    """
    Apply value-level transformations to sourceFile using rules defined in mappingFile.

    Mapping file column detection (case-insensitive, positional fallback):
      3+ cols → column_name | old_value | new_value  (targeted per source column)
      2  cols → old_value   | new_value              (applied to ALL source columns)

    Returns a transformed .xlsx blob.  Transform statistics are reported in
    response headers so the frontend can display them without parsing the file.
    """
    try:
        source_bytes  = await sourceFile.read()
        mapping_bytes = await mappingFile.read()

        # ── Parse source file ──────────────────────────────────────────────
        try:
            df = _read_file_bytes(source_bytes, sourceFile.filename)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read source file: {e}")

        # ── Parse mapping/rules file ───────────────────────────────────────
        try:
            df_map = _read_file_bytes(mapping_bytes, mappingFile.filename)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to read mapping file: {e}")

        # ── Detect which columns represent column_name / old_value / new_value ──
        col_name_cands  = ["column_name", "column", "col", "field", "attribute"]
        old_value_cands = ["old_value", "from", "old", "original", "source_value"]
        new_value_cands = ["new_value", "to", "new", "replacement", "target_value"]

        cols_lc = {str(c).strip().lower(): str(c) for c in df_map.columns}

        def _pick(candidates):
            for cand in candidates:
                if cand in cols_lc:
                    return cols_lc[cand]
            return None

        num_map_cols  = len(df_map.columns)
        col_name_col  = None
        old_value_col = None
        new_value_col = None
        apply_all_cols = False   # True → 2-col mode, replace in every source column

        if num_map_cols >= 3:
            col_name_col  = _pick(col_name_cands)
            old_value_col = _pick(old_value_cands)
            new_value_col = _pick(new_value_cands)
            # Positional fallback when headers are not recognized
            if not col_name_col or not old_value_col or not new_value_col:
                col_name_col  = df_map.columns[0]
                old_value_col = df_map.columns[1]
                new_value_col = df_map.columns[2]
        elif num_map_cols == 2:
            apply_all_cols = True
            old_value_col  = df_map.columns[0]
            new_value_col  = df_map.columns[1]
        else:
            raise HTTPException(
                status_code=400,
                detail="Mapping file must have at least 2 columns (old_value, new_value)."
            )

        # ── Build rules list ───────────────────────────────────────────────
        rules = []  # each rule: { col: str|None, old: str, new: str }
        for _, row in df_map.dropna(subset=[old_value_col, new_value_col], how="any").iterrows():
            old_val = str(row[old_value_col]).strip()
            new_val = str(row[new_value_col]).strip()
            if not old_val:
                continue
            if apply_all_cols:
                target_col = None
            else:
                raw_col = row[col_name_col]
                target_col = str(raw_col).strip() if pd.notna(raw_col) else None
                if not target_col or target_col.lower() == "nan":
                    target_col = None
            rules.append({"col": target_col, "old": old_val, "new": new_val})

        # ── Apply transformations ──────────────────────────────────────────
        df_work      = df.copy().astype(object)   # object dtype avoids coercion errors
        df_cols_set  = set(df_work.columns.astype(str))

        rules_applied  = 0
        cells_changed  = 0
        cols_changed   = set()
        total_rules    = len(rules)

        for rule in rules:
            old_val    = rule["old"]
            new_val    = rule["new"]
            rule_col   = rule["col"]
            rule_hit   = False

            target_cols = (
                [rule_col] if rule_col and rule_col in df_cols_set
                else ([] if rule_col else list(df_work.columns))
            )

            for col in target_cols:
                series    = df_work[col].astype(str)
                mask      = series == old_val
                hit_count = int(mask.sum())
                if hit_count > 0:
                    df_work.loc[mask, col] = new_val
                    cells_changed += hit_count
                    cols_changed.add(str(col))
                    rule_hit = True

            if rule_hit:
                rules_applied += 1

        # ── Serialize result to xlsx bytes ─────────────────────────────────
        out_buf = BytesIO()
        df_work.to_excel(out_buf, index=False)
        out_buf.seek(0)

        stem         = Path(sourceFile.filename).stem if sourceFile.filename else "source"
        out_filename = f"{stem}_transformed.xlsx"

        return StreamingResponse(
            out_buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":          f'attachment; filename="{out_filename}"',
                "X-Transform-Rules-Applied":    str(rules_applied),
                "X-Transform-Cells-Changed":    str(cells_changed),
                "X-Transform-Columns-Changed":  str(len(cols_changed)),
                "X-Transform-Total-Rules":      str(total_rules),
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error in post_validation_transform")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/excel/post_validation/status/{job_id}")
async def get_validation_status(job_id: str):
    """
    Returns current progress of a validation job.
    Response: { status, progress, stage, error? }
      - status: "running" | "complete" | "failed"
      - progress: 0–100
      - stage: human-readable description of current step
    """
    job = _job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {
        "status": job.get("status"),
        "progress": job.get("progress", 0),
        "stage": job.get("stage", ""),
        "error": job.get("error"),
        "eta_seconds": job.get("eta_seconds"),
    }



@app.get("/api/excel/post_validation/download/{job_id}")
async def download_validation_result(job_id: str, background_tasks: BackgroundTasks):
    """
    Downloads the validation result (.xlsx) once the job is complete.
    After download, schedules cleanup of temp files.
    """
    job = _job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job["status"] == "running":
        raise HTTPException(status_code=202, detail="Job still running")
    if job["status"] == "failed":
        raise HTTPException(status_code=500, detail=job.get("error", "Unknown error"))

    zip_path = job.get("zip_path")
    if not zip_path or not os.path.exists(zip_path):
        raise HTTPException(status_code=410, detail="Result file no longer available")

    temp_dir = job.get("temp_dir")

    def _cleanup(path, jid):
        try:
            shutil.rmtree(path, ignore_errors=True)
        except Exception:
            pass
        with _validation_jobs_lock:
            _validation_jobs.pop(jid, None)

    background_tasks.add_task(_cleanup, temp_dir, job_id)

    # Detect file type from extension to support both .xlsx and .zip outputs
    dl_filename = job.get("zip_filename", "MythicsValidationResults.xlsx")
    if zip_path.endswith(".zip"):
        media = "application/zip"
    else:
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return FileResponse(
        zip_path,
        filename=dl_filename,
        media_type=media,
    )



# ── Background Processing Function ──────────────────────────────────────
def _run_validation_job(job_id: str, p: dict):
    """
    Runs the full Polars validation pipeline in a background thread.
    Updates _validation_jobs[job_id] with progress at each stage.
    """
    # Enable Rust backtraces for diagnosing Polars native crashes
    os.environ.setdefault("RUST_BACKTRACE", "1")
    os.environ.setdefault("POLARS_VERBOSE", "1")

    temp_dir = p["temp_dir"]
    src_label = (p.get("sourceLabel") or "Source").strip() or "Source"
    tgt_label = (p.get("targetLabel") or "Target").strip() or "Target"
    case_sensitive = p.get("caseSensitive", True)
    main_output_path = os.path.join(temp_dir, "PostValidation_Main.xlsx")
    full_data_csv_path = os.path.join(temp_dir, "PeopleSoft_OracleCloud_FullData.csv")
    discrepancies_csv_path = os.path.join(temp_dir, "DataDiscrepancies_Full.csv")
    missing_oracle_csv_path = os.path.join(temp_dir, "Missing_In_Oracle.csv")
    missing_ps_csv_path = os.path.join(temp_dir, "Missing_In_PeopleSoft.csv")
    INTERNAL_KEY = "_derived_key"

    try:
        start_time = time.time()
        import gc
        import polars as pl

        # ── Stage 1: Parse Inputs (2%) ──────────────────────────────────
        _job_update(job_id, progress=2, stage="Parsing configuration")
        logger.info(f"[{job_id[:8]}] Parsing inputs...")

        mappings_dict: Dict[str, str] = json.loads(p["mappings"])
        included_cols_list: List[str] = json.loads(p["includedColumns"])
        key_cols_list: List[str] = json.loads(p["keyColumns"])
        legacy_date_set = set(json.loads(p["dateColumns"]) + json.loads(p["timestampColumns"]))
        target_date_set = set(json.loads(p["dateColumnstarget"]) + json.loads(p["timestampColumnstarget"]))

        if not mappings_dict:
            raise ValueError("Mappings dictionary is empty")
        if not key_cols_list:
            raise ValueError("Key columns list is empty")

        config_rows = []
        for source_col, target_col in mappings_dict.items():
            config_rows.append({
                f"{src_label} Column": source_col,
                f"{tgt_label} Column": target_col,
                "Is Key?": "Yes" if source_col in key_cols_list else "No",
                "Is Date?": "Yes" if (source_col in legacy_date_set or target_col in target_date_set) else "No",
                "Validate": "Yes",
                "Include in Report": "Yes" if source_col in included_cols_list else "No",
            })
        config_df = pd.DataFrame(config_rows)

        # ── Stage 2: Load Files — parallel load, stay in Polars, no pandas ─
        _job_update(job_id, progress=8, stage="Reading source and target files (parallel)")
        logger.info(f"[{job_id[:8]}] Loading both files in parallel (calamine engine)...")

        def _load_frame(path, sheet):
            return _polars_read_file(path, sheet).collect()

        _t_load = time.perf_counter()
        with _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="file-load") as _load_pool:
            _f_legacy = _load_pool.submit(_load_frame, p["legacy_path"], p["legacySheet"])
            _f_oracle = _load_pool.submit(_load_frame, p["oracle_path"], p["oracleSheet"])
            pl_legacy = _f_legacy.result()
            pl_oracle = _f_oracle.result()
        logger.info(f"[{job_id[:8]}] Both files loaded in {time.perf_counter()-_t_load:.2f}s")

        legacy_count, oracle_count = len(pl_legacy), len(pl_oracle)
        _job_update(job_id, progress=20,
                    stage=f"Files loaded — {legacy_count:,} + {oracle_count:,} rows")
        logger.info(f"[{job_id[:8]}] Loaded — Legacy: {legacy_count:,}, Oracle: {oracle_count:,}")

        # Evaluate include_src_tgt early — skips expensive CSV write when unneeded
        include_src_tgt = p.get("includeSourceTargetFiles", False)

        # ── Stage 3: Column Alignment & Validation (25%) — pure Polars ──
        _job_update(job_id, progress=22, stage="Aligning column names")

        oracle_to_legacy = {v: k for k, v in mappings_dict.items()}
        oracle_to_legacy_lower = {v.strip().lower(): k for k, v in mappings_dict.items()}

        # Build oracle rename map using Polars column list
        cols_to_rename_pl: Dict[str, str] = {}
        for col in pl_oracle.columns:
            if col in oracle_to_legacy:
                cols_to_rename_pl[col] = oracle_to_legacy[col]
            elif col.strip().lower() in oracle_to_legacy_lower:
                cols_to_rename_pl[col] = oracle_to_legacy_lower[col.strip().lower()]
        if cols_to_rename_pl:
            pl_oracle = pl_oracle.rename(cols_to_rename_pl)

        # Normalise legacy column names (case-insensitive)
        legacy_mapped_lower = {k.strip().lower(): k for k in mappings_dict.keys()}
        legacy_rename_pl: Dict[str, str] = {}
        for col in pl_legacy.columns:
            lk = col.strip().lower()
            if lk in legacy_mapped_lower and col != legacy_mapped_lower[lk]:
                legacy_rename_pl[col] = legacy_mapped_lower[lk]
        if legacy_rename_pl:
            pl_legacy = pl_legacy.rename(legacy_rename_pl)

        legacy_cols = pl_legacy.columns
        oracle_cols = pl_oracle.columns

        # Resolve key_cols_list and mappings_dict against actual column names
        oracle_col_lower = {c.strip().lower(): c for c in oracle_cols}
        legacy_col_lower = {c.strip().lower(): c for c in legacy_cols}
        key_cols_list = [oracle_col_lower.get(k.strip().lower(), k) for k in key_cols_list]

        resolved_mappings: Dict[str, str] = {}
        for l_col, o_col in mappings_dict.items():
            actual_legacy = legacy_col_lower.get(l_col.strip().lower(), l_col)
            actual_oracle = oracle_col_lower.get(l_col.strip().lower(), l_col)
            common_name = actual_oracle if actual_oracle in oracle_cols else actual_legacy
            resolved_mappings[common_name] = o_col
        mappings_dict = resolved_mappings

        included_cols_list = [
            oracle_col_lower.get(c.strip().lower(), legacy_col_lower.get(c.strip().lower(), c))
            for c in included_cols_list
        ]
        legacy_date_set = {legacy_col_lower.get(c.strip().lower(), c) for c in legacy_date_set}
        target_date_set_resolved = set()
        for c in target_date_set:
            mapped = oracle_to_legacy.get(c, c)
            target_date_set_resolved.add(oracle_col_lower.get(mapped.strip().lower(), mapped))
        target_date_set = target_date_set_resolved

        missing_keys = [k for k in key_cols_list if k not in pl_oracle.columns]
        if missing_keys:
            raise ValueError(f"Key columns missing in target after rename: {missing_keys}")

        col_errors = []
        for l_col, o_col in mappings_dict.items():
            if l_col not in pl_legacy.columns:
                col_errors.append(f"{src_label} column '{l_col}' not found")
            if l_col not in pl_oracle.columns:
                col_errors.append(f"{tgt_label} column '{o_col}' not found after rename")
        if col_errors:
            raise ValueError(f"Column errors: {'; '.join(col_errors)}")

        cols_to_compare = [k for k in mappings_dict if k in pl_legacy.columns and k in pl_oracle.columns]
        num_comparison_cols = len(cols_to_compare)
        _job_update(job_id, progress=25, stage=f"Validating {num_comparison_cols} mapped columns")

        # ── Stage 4: Normalise keys & dates — Polars-native ──────────────
        _job_update(job_id, progress=28, stage="Normalising keys & dates (Polars)")

        # Key normalisation: strip whitespace, drop .0 suffix, blank sentinels
        pl_legacy = _pl_normalize_key_cols(pl_legacy, key_cols_list)
        pl_oracle = _pl_normalize_key_cols(pl_oracle, key_cols_list)

        # Date normalisation: explicit + auto-detected via Polars sampling
        legacy_date_cols_explicit = {c for c in legacy_date_set if c in pl_legacy.columns}
        target_date_cols_explicit = set()
        for col in target_date_set:
            mapped = oracle_to_legacy.get(col, col)
            if mapped in pl_oracle.columns:
                target_date_cols_explicit.add(mapped)

        logger.info(f"[{job_id[:8]}] Applying date normalisation (parallel, auto-detect enabled)...")
        _t_date = time.perf_counter()
        with _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="date-norm") as _dn_pool:
            _f_l_dn = _dn_pool.submit(
                _pl_apply_date_normalization,
                pl_legacy, legacy_date_cols_explicit, list(pl_legacy.columns),
                True, cols_to_compare
            )
            _f_o_dn = _dn_pool.submit(
                _pl_apply_date_normalization,
                pl_oracle, target_date_cols_explicit, list(pl_oracle.columns),
                True, cols_to_compare
            )
            pl_legacy = _f_l_dn.result()
            pl_oracle = _f_o_dn.result()
        logger.info(f"[{job_id[:8]}] Date normalisation complete in {time.perf_counter()-_t_date:.2f}s")

        # Composite key generation — both sides in parallel
        _job_update(job_id, progress=32, stage="Generating composite keys")
        _key_expr = _pl_gen_composite_key(key_cols_list)
        with _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="key-gen") as _kg_pool:
            _f_l_kg = _kg_pool.submit(
                lambda df: df.with_columns(_key_expr.alias(INTERNAL_KEY)), pl_legacy
            )
            _f_o_kg = _kg_pool.submit(
                lambda df: df.with_columns(_key_expr.alias(INTERNAL_KEY)), pl_oracle
            )
            pl_legacy = _f_l_kg.result()
            pl_oracle = _f_o_kg.result()
        _job_update(job_id, progress=35, stage="Keys generated — starting comparison")
        logger.info(f"[{job_id[:8]}] Normalisation complete")

        # ── Stage 5: Numeric detection — vectorized single-pass ──────────
        _job_update(job_id, progress=37, stage="Detecting column data types")
        numeric_cols = _pl_detect_numeric_cols(pl_legacy, pl_oracle, cols_to_compare)
        logger.info(f"[{job_id[:8]}] Numeric columns detected: {numeric_cols}")

        # ── Stage 5b: Duplicate key handling — parallel detection ─────────
        _job_update(job_id, progress=38, stage="Handling duplicate keys")
        _ik = INTERNAL_KEY
        with _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="dup-check") as _dup_pool:
            _f_l_dup = _dup_pool.submit(
                lambda df: df.select(pl.col(_ik).is_duplicated().sum()).item(), pl_legacy
            )
            _f_o_dup = _dup_pool.submit(
                lambda df: df.select(pl.col(_ik).is_duplicated().sum()).item(), pl_oracle
            )
            legacy_key_dupes = _f_l_dup.result()
            oracle_key_dupes = _f_o_dup.result()
        logger.info(f"[{job_id[:8]}] Key duplicates — Legacy: {legacy_key_dupes:,}, Oracle: {oracle_key_dupes:,}")

        if legacy_key_dupes > 0 or oracle_key_dupes > 0:
            logger.info(f"[{job_id[:8]}] Adding positional row number within each key group")
            with _cf.ThreadPoolExecutor(max_workers=2, thread_name_prefix="pos-key") as _pk_pool:
                _f_l_pk = _pk_pool.submit(_pl_add_positional_key, pl_legacy, INTERNAL_KEY)
                _f_o_pk = _pk_pool.submit(_pl_add_positional_key, pl_oracle, INTERNAL_KEY)
                pl_legacy = _f_l_pk.result()
                pl_oracle = _f_o_pk.result()
            logger.info(f"[{job_id[:8]}] Positional keys assigned")

        # ── Stage 6: Polars join (40–60%) — zero Pandas↔Polars conversion ─
        _job_update(job_id, progress=40, stage="Joining source & target (Polars)")
        logger.info(f"[{job_id[:8]}] Starting inner join on '{INTERNAL_KEY}'...")

        compare_cols_needed = [INTERNAL_KEY] + cols_to_compare
        pl_l_cmp = pl_legacy.select([c for c in compare_cols_needed if c in pl_legacy.columns])
        pl_o_cmp = pl_oracle.select([c for c in compare_cols_needed if c in pl_oracle.columns])

        try:
            joined = pl_l_cmp.join(pl_o_cmp, on=INTERNAL_KEY, suffix="_T", how="inner").rechunk()
        except Exception as join_err:
            logger.error(f"[{job_id[:8]}] Polars join failed: {join_err}")
            raise
        del pl_l_cmp, pl_o_cmp
        gc.collect()

        matched_count = len(joined)
        _job_update(job_id, progress=42,
                    stage=f"Joined {matched_count:,} matched rows — comparing")
        logger.info(f"[{job_id[:8]}] Polars join: {matched_count:,} matched rows")

        # Safety guard: abort if join exploded (non-unique keys)
        max_expected = max(legacy_count, oracle_count) * 3
        if matched_count > max_expected:
            raise ValueError(
                f"Join produced {matched_count:,} rows (source has {legacy_count:,}). "
                f"Key columns may not be unique. Please verify key column selection."
            )

        # Matched keys kept as a Polars frame — used for anti-join in Stage 9
        matched_keys_pl = joined.select(INTERNAL_KEY).unique()

        # Write full-data matched CSV only when the user requested it
        if include_src_tgt:
            _job_update(job_id, progress=44, stage="Writing matched data to CSV")
            logger.info(f"[{job_id[:8]}] Writing full matched data CSV...")
            full_select = (
                [pl.col(c).cast(pl.Utf8).fill_null("").alias(f"{c}_S") for c in cols_to_compare]
                + [pl.col(f"{c}_T").cast(pl.Utf8).fill_null("").alias(f"{c}_T") for c in cols_to_compare]
                + [pl.lit("MATCHED").alias("Record Status")]
            )
            joined.select(full_select).write_csv(full_data_csv_path)
            logger.info(f"[{job_id[:8]}] Matched CSV written")
        gc.collect()

        # ── Stage 6.2 / 7: Single-scan flag pass + mismatched-row extraction ──
        # Pass 1: ONE scan of `joined` → boolean flags only (tiny output: ~100 MB for 1M rows).
        # Pass 2: filter `joined` to mismatched rows only → per-column work is tiny.
        # This replaces N independent scans (one per comparison column).
        _job_update(job_id, progress=46, stage="Computing diff flags (single scan)")
        _t_disc = time.perf_counter()

        diff_flag_exprs = []
        valid_compare_cols = []
        for col in cols_to_compare:
            o_col = f"{col}_T"
            if o_col not in joined.columns:
                continue
            valid_compare_cols.append(col)
            if col in numeric_cols:
                cn = _pl_clean_num_expr(col)
                co_expr = _pl_clean_num_expr(o_col)
                flag = (
                    (cn.is_not_null() & co_expr.is_not_null() & ((cn - co_expr).abs() > 0.0001))
                    | (cn.is_null() ^ co_expr.is_null())
                )
            else:
                cs = _pl_clean_str_expr(col, case_sensitive)
                co_expr = _pl_clean_str_expr(o_col, case_sensitive)
                flag = cs.fill_null("") != co_expr.fill_null("")
            diff_flag_exprs.append(flag.alias(f"__diff_{col}"))

        # ONE scan: produce INTERNAL_KEY + N boolean flags — keeps memory flat
        flags_df = joined.lazy().select(
            [pl.col(INTERNAL_KEY)] + diff_flag_exprs
        ).collect()

        cols_with_diffs = [c for c in valid_compare_cols if flags_df[f"__diff_{c}"].any()]
        logger.info(
            f"[{job_id[:8]}] Flag scan: {time.perf_counter()-_t_disc:.2f}s — "
            f"{len(cols_with_diffs)}/{len(valid_compare_cols)} cols differ"
        )

        _job_update(job_id, progress=55, stage="Extracting discrepancies")

        _EMPTY_DISC = {
            INTERNAL_KEY: pl.Series([], dtype=pl.Utf8),
            "Column Name": pl.Series([], dtype=pl.Utf8),
            f"{src_label} Value": pl.Series([], dtype=pl.Utf8),
            f"{tgt_label} Value": pl.Series([], dtype=pl.Utf8),
        }

        if cols_with_diffs:
            # Build "any diff" mask (OR across only the differing columns)
            any_diff_expr = pl.col(f"__diff_{cols_with_diffs[0]}")
            for _c in cols_with_diffs[1:]:
                any_diff_expr = any_diff_expr | pl.col(f"__diff_{_c}")
            any_diff_mask = flags_df.select(any_diff_expr.alias("_any"))["_any"]

            # Pass 2: filter joined to mismatched rows; select only needed value columns
            needed_val_cols = (
                [INTERNAL_KEY]
                + [c for c in cols_with_diffs if c in joined.columns]
                + [f"{c}_T" for c in cols_with_diffs if f"{c}_T" in joined.columns]
            )
            mismatched_vals = joined.filter(any_diff_mask).select(needed_val_cols)

            # Attach per-column flags for mismatched rows (aligned by same mask)
            flags_mismatch = flags_df.filter(any_diff_mask).select(
                [f"__diff_{c}" for c in cols_with_diffs]
            )
            work = pl.concat([mismatched_vals, flags_mismatch], how="horizontal")
            del mismatched_vals, flags_mismatch, flags_df, any_diff_mask
            gc.collect()

            # Per-column filter on `work` — fast, work is only mismatched rows
            disc_parts = []
            for col in cols_with_diffs:
                col_label = f"{col} - {mappings_dict.get(col, col)}"
                part = work.filter(pl.col(f"__diff_{col}")).select([
                    pl.col(INTERNAL_KEY),
                    pl.lit(col_label).alias("Column Name"),
                    pl.col(col).cast(pl.Utf8).fill_null("").alias(f"{src_label} Value"),
                    pl.col(f"{col}_T").cast(pl.Utf8).fill_null("").alias(f"{tgt_label} Value"),
                ])
                if len(part) > 0:
                    disc_parts.append(part)
            del work

            discrepancies_pl = pl.concat(disc_parts) if disc_parts else pl.DataFrame(_EMPTY_DISC)
            del disc_parts
        else:
            del flags_df
            discrepancies_pl = pl.DataFrame(_EMPTY_DISC)
        gc.collect()

        # Context columns: joined from pl_legacy (no pandas conversion)
        context_cols_order = list(dict.fromkeys(key_cols_list + included_cols_list))
        valid_ctx = [c for c in context_cols_order if c in pl_legacy.columns and c != INTERNAL_KEY]
        if valid_ctx:
            ctx_pl = pl_legacy.select([INTERNAL_KEY] + valid_ctx).unique(subset=[INTERNAL_KEY])
            discrepancies_pl = discrepancies_pl.join(ctx_pl, on=INTERNAL_KEY, how="left")

        total_discrepancies = len(discrepancies_pl)

        # Always write CSV for large result sets (full fidelity)
        _EXCEL_ROW_CAP = 100_000
        if total_discrepancies > _EXCEL_ROW_CAP:
            discrepancies_pl.drop(INTERNAL_KEY).write_csv(discrepancies_csv_path)
            logger.info(f"[{job_id[:8]}] Discrepancy CSV written ({total_discrepancies:,} rows)")

        if total_discrepancies > 0:
            # Cap rows written to Excel — full data is in the CSV
            disc_for_excel = (
                discrepancies_pl.head(_EXCEL_ROW_CAP) if total_discrepancies > _EXCEL_ROW_CAP
                else discrepancies_pl
            )
            validation_df = disc_for_excel.drop(INTERNAL_KEY).to_pandas()
            del discrepancies_pl, disc_for_excel
            gc.collect()

            final_report_cols = (
                key_cols_list
                + [c for c in included_cols_list if c not in key_cols_list]
                + ["Column Name", f"{src_label} Value", f"{tgt_label} Value"]
            )
            final_report_cols = [c for c in final_report_cols if c in validation_df.columns]
            validation_df = validation_df[final_report_cols]
        else:
            del discrepancies_pl
            validation_df = pd.DataFrame([{"Status": "All mapped columns matched perfectly"}])

        _job_update(job_id, progress=62, stage=f"Found {total_discrepancies:,} discrepancies")
        logger.info(f"[{job_id[:8]}] Total discrepancies: {total_discrepancies:,}")

        if "Column Name" in validation_df.columns and not validation_df.empty:
            sort_cols = ["Column Name"] + [c for c in key_cols_list if c in validation_df.columns]
            validation_df = validation_df.sort_values(by=sort_cols, kind="mergesort").reset_index(drop=True)

        comment_cols = ["Mythics Comments", "Oracle Comments", "ParkView Comments"]
        for col in comment_cols:
            validation_df[col] = ""

        # ── Stage 8: Discrepancy Counts ───────────────────────────────
        column_discrepancy_counts = []
        if "Status" not in validation_df.columns and not validation_df.empty:
            col_counts = validation_df["Column Name"].value_counts().sort_values(ascending=False)
            for col_name, count in col_counts.items():
                column_discrepancy_counts.append(["", col_name, int(count), "", "", ""])

        # ── Stage 9: Missing Records — Polars anti-join on existing frames ─
        _job_update(job_id, progress=64, stage="Finding missing records")

        # Anti-join directly against matched_keys_pl — no pandas.isin() needed
        legacy_only_pl = pl_legacy.join(matched_keys_pl, on=INTERNAL_KEY, how="anti")
        oracle_only_pl = pl_oracle.join(matched_keys_pl, on=INTERNAL_KEY, how="anti")

        count_missing_oracle = len(legacy_only_pl)
        count_missing_ps = len(oracle_only_pl)
        _job_update(job_id, progress=68,
                    stage=f"Missing: {count_missing_oracle:,} in {tgt_label}, {count_missing_ps:,} in {src_label}")

        # Write CSV for large missing sets; cap Excel to _EXCEL_ROW_CAP rows
        if count_missing_oracle > _EXCEL_ROW_CAP:
            legacy_only_pl.drop(INTERNAL_KEY).write_csv(missing_oracle_csv_path)
        if count_missing_ps > _EXCEL_ROW_CAP:
            oracle_only_pl.drop(INTERNAL_KEY).write_csv(missing_ps_csv_path)

        # Convert missing DFs to pandas — capped for Excel
        legacy_for_excel = (
            legacy_only_pl.head(_EXCEL_ROW_CAP) if count_missing_oracle > _EXCEL_ROW_CAP
            else legacy_only_pl
        )
        oracle_for_excel = (
            oracle_only_pl.head(_EXCEL_ROW_CAP) if count_missing_ps > _EXCEL_ROW_CAP
            else oracle_only_pl
        )
        legacy_only_df = legacy_for_excel.drop(INTERNAL_KEY).to_pandas()
        oracle_only_df = oracle_for_excel.drop(INTERNAL_KEY).to_pandas()
        del legacy_for_excel, oracle_for_excel
        for col in comment_cols:
            legacy_only_df[col] = ""
            oracle_only_df[col] = ""

        # ── Stage 10: Full Data CSV — Polars append for missing rows ─────
        _job_update(job_id, progress=78, stage="Appending missing records to full data CSV")

        if include_src_tgt and os.path.exists(full_data_csv_path):
            # Append missing rows by opening the existing file in binary-append mode.
            # Polars write_csv accepts IO[bytes] so this avoids any intermediate pandas copy.
            if count_missing_oracle > 0:
                l_s = [(pl.col(c).cast(pl.Utf8).fill_null("") if c in legacy_only_pl.columns else pl.lit("")).alias(f"{c}_S")
                       for c in cols_to_compare]
                l_t = [pl.lit("").alias(f"{c}_T") for c in cols_to_compare]
                with open(full_data_csv_path, "ab") as _f:
                    legacy_only_pl.select(
                        l_s + l_t + [pl.lit("MISSING_IN_TARGET").alias("Record Status")]
                    ).write_csv(_f, include_header=False)
            if count_missing_ps > 0:
                o_s = [pl.lit("").alias(f"{c}_S") for c in cols_to_compare]
                o_t = [(pl.col(c).cast(pl.Utf8).fill_null("") if c in oracle_only_pl.columns else pl.lit("")).alias(f"{c}_T")
                       for c in cols_to_compare]
                with open(full_data_csv_path, "ab") as _f:
                    oracle_only_pl.select(
                        o_s + o_t + [pl.lit("MISSING_IN_SOURCE").alias("Record Status")]
                    ).write_csv(_f, include_header=False)

        # Lazy-load full data CSV for Excel — capped at _EXCEL_ROW_CAP (same as all other sheets).
        # Loading more rows would cause MemoryError when openpyxl builds its in-memory DOM.
        # Full-fidelity data is always available in the CSV file.
        full_data_for_excel = None
        if include_src_tgt and os.path.exists(full_data_csv_path):
            try:
                full_data_for_excel = pd.read_csv(full_data_csv_path, dtype=str, nrows=_EXCEL_ROW_CAP)
            except Exception as csv_err:
                logger.warning(f"[{job_id[:8]}] Could not reload full data CSV for Excel: {csv_err}")

        # Free all large Polars frames — pandas results are now held in small DFs
        del joined, pl_legacy, pl_oracle, matched_keys_pl, legacy_only_pl, oracle_only_pl
        gc.collect()
        logger.info(f"[{job_id[:8]}] Polars processing done in {time.time() - start_time:.2f}s")

        # ── Stage 13: Generate Summary (82%) ───────────────────────────
        _job_update(job_id, progress=82, stage="Generating summary statistics")

        grand_total = count_missing_oracle + count_missing_ps + total_discrepancies
        summary_data = [
            ["", "Comparison Statistics", "", "", "", ""],
            ["", f"{src_label} File Name", p["legacy_filename"], "", "", ""],
            ["", f"{src_label} Records Count", f"{legacy_count:,}", "", "", ""],
            ["", f"{tgt_label} File Name", p["oracle_filename"], "", "", ""],
            ["", f"{tgt_label} Records Count", f"{oracle_count:,}", "", "", ""],
            ["", "Validation DateTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", "", ""],
            ["", "", "", "", "", ""],
            ["", "Missing Records Summary", "", "Mythics Comments", "Oracle Comments", "ParkView Comments"],
            ["", f"Records Missing in {src_label}", f"{count_missing_ps:,}", "", "", ""],
            ["", f"Records Missing in {tgt_label}", f"{count_missing_oracle:,}", "", "", ""],
            ["", "Total Missing Records", f"{count_missing_oracle + count_missing_ps:,}", "", "", ""],
            ["", "", "", "", "", ""],
            ["", "Data Discrepancies Summary", "", "Mythics Comments", "Oracle Comments", "ParkView Comments"],
            *column_discrepancy_counts,
            ["", "Total Data Discrepancies", f"{total_discrepancies:,}", "", "", ""],
            ["", "", "", "", "", ""],
            ["", "Total Validation Issues", f"{grand_total:,}", "", "", ""],
        ]
        summary_df = pd.DataFrame(summary_data)

        # ── Stage 14: Write Styled Excel (85–95%) ─────────────────────
        _job_update(job_id, progress=85, stage="Writing styled Excel report")

        font_white = Font(name="Calibri", size=8, color="FFFFFF", bold=True)
        font_black = Font(name="Calibri", size=8, color="000000", bold=True)
        font_bold_black = Font(name="Calibri", size=8, bold=True)
        font_normal = Font(name="Calibri", size=8)
        fill_header_ps = PatternFill("solid", fgColor="1F497D")
        fill_header_oc = PatternFill("solid", fgColor="31869B")
        fill_header_err = PatternFill("solid", fgColor="C0504D")
        fill_green = PatternFill("solid", fgColor="00B050")
        fill_grey = PatternFill("solid", fgColor="D9D9D9")
        fill_orange = PatternFill("solid", fgColor="FF9900")
        border_thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        align_center = Alignment(horizontal="center", vertical="center")

        sheet_missing_ps = _safe_sheet_name(f"Missing in {src_label}")
        sheet_missing_oc = _safe_sheet_name(f"Missing in {tgt_label}")
        sheet_discrepancies = "Data Discrepancies"

        def _style_header(wb, sn, fill):
            if sn not in wb.sheetnames:
                return
            ws = wb[sn]
            ws.freeze_panes = "A2"
            comment_idxs = []
            for cell in ws[1]:
                cell.fill, cell.font = fill, font_white
                cell.alignment, cell.border = align_center, border_thin
                if cell.value in comment_cols:
                    cell.fill, cell.font = fill_orange, font_black
                    comment_idxs.append(cell.column)
            body_cap = min(ws.max_row, 500)
            for body_row in ws.iter_rows(min_row=2, max_row=body_cap):
                for cell in body_row:
                    cell.font = font_normal
            for col in ws.iter_cols(max_row=min(50, ws.max_row)):
                ml = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max((ml + 2) * 1.2, 10), 60)
            for ci in comment_idxs:
                border_cap = min(ws.max_row, 5000)
                for cells in ws.iter_cols(min_col=ci, max_col=ci, min_row=2, max_row=border_cap):
                    for c in cells:
                        c.border = border_thin

        def _style_config(wb, sn="Configuration"):
            if sn not in wb.sheetnames:
                return
            ws = wb[sn]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill, cell.font = fill_green, font_white
                cell.alignment, cell.border = align_center, border_thin
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 200)):
                for cell in row:
                    cell.font = font_normal
                    cell.border = border_thin
                    if str(cell.value).strip().lower() in {"yes", "no"}:
                        cell.alignment = align_center
            for col in ws.iter_cols(max_row=ws.max_row):
                ml = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 2, 12), 45)

        def _style_full_data_header(ws, n_compare_cols):
            """Style the full-data sheet with dual-colour headers (PS blue | OC teal)
            and a medium border separating the two groups."""
            border_ps_last  = Border(left=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="medium"))
            border_oc_first = Border(left=Side(style="medium"), top=Side(style="thin"), bottom=Side(style="thin"), right=Side(style="thin"))
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = font_white
                cell.alignment = align_center
                if cell.column <= n_compare_cols:
                    cell.fill = fill_header_ps
                    cell.border = border_ps_last if cell.column == n_compare_cols else border_thin
                else:
                    cell.fill = fill_header_oc
                    cell.border = border_oc_first if cell.column == n_compare_cols + 1 else border_thin
            for col_cells in ws.iter_cols(max_row=min(50, ws.max_row)):
                ml = max((len(str(c.value)) if c.value else 0) for c in col_cells)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
                    max((ml + 2) * 1.2, 10), 60
                )

        _job_update(job_id, progress=88, stage="Styling Excel sheets")

        sheet_full_data = _safe_sheet_name(f"{src_label} - {tgt_label} Data", max_len=28)

        # ── Phase 1: fast data write with xlsxwriter (small sheets only) ────
        # full_data_for_excel is intentionally excluded here: for large datasets
        # (600k rows × N cols) the BytesIO would be hundreds of MB and
        # openpyxl.load_workbook() would build a DOM of tens-of-millions of cell
        # objects → MemoryError. Full data is appended in Phase 3 instead.
        _job_update(job_id, progress=88, stage="Writing Excel data (xlsxwriter)")
        _t_xls = time.perf_counter()
        _xls_buf = BytesIO()
        with pd.ExcelWriter(_xls_buf, engine="xlsxwriter") as writer:
            summary_df.to_excel(writer, index=False, header=False, sheet_name="Summary")
            config_df.to_excel(writer, index=False, sheet_name="Configuration")
            oracle_only_df.to_excel(writer, index=False, sheet_name=sheet_missing_ps)
            legacy_only_df.to_excel(writer, index=False, sheet_name=sheet_missing_oc)
            write_df_excel_paginated(writer, validation_df, sheet_discrepancies)
        _xls_buf.seek(0)
        logger.info(f"[{job_id[:8]}] xlsxwriter data write: {time.perf_counter()-_t_xls:.2f}s")

        # ── Phase 2: open with openpyxl just for styling (small workbook) ─
        _job_update(job_id, progress=92, stage="Styling Excel sheets (openpyxl)")
        _t_style = time.perf_counter()
        wb = openpyxl.load_workbook(_xls_buf)
        del _xls_buf

        _style_header(wb, sheet_missing_ps, fill_header_ps)
        _style_header(wb, sheet_missing_oc, fill_header_oc)
        for sn in wb.sheetnames:
            if sn.startswith(sheet_discrepancies):
                _style_header(wb, sn, fill_header_err)
        _style_config(wb)
        wb["Configuration"].sheet_state = "hidden"

        # Summary styling
        ws_sum = wb["Summary"]
        ws_sum.sheet_view.showGridLines = False
        ws_sum.column_dimensions["A"].width = 2
        ws_sum.column_dimensions["B"].width = 45
        for ch in ("C", "D", "E", "F"):
            ws_sum.column_dimensions[ch].width = 25

        for row in ws_sum.iter_rows(min_row=1, max_row=ws_sum.max_row):
            if len(row) < 3:
                continue
            for cell in row:
                cell.font = font_normal
            cell_b = row[1]
            if not cell_b.value:
                continue
            cell_b.border = border_thin
            row[2].border = border_thin
            cc = [ws_sum.cell(row=cell_b.row, column=c) for c in (4, 5, 6)]

            if cell_b.value in ("Missing Records Summary", "Data Discrepancies Summary"):
                ws_sum.merge_cells(start_row=cell_b.row, start_column=2,
                                   end_row=cell_b.row, end_column=3)
                cell_b.fill, cell_b.font = fill_green, font_white
                cell_b.alignment = align_center
                ws_sum.row_dimensions[cell_b.row].height = 20
                for c in cc:
                    c.fill, c.font = fill_orange, font_black
                    c.alignment, c.border = align_center, border_thin
            elif "Total" in str(cell_b.value) or "Comparison Statistics" in str(cell_b.value):
                if "Comparison Statistics" in str(cell_b.value):
                    ws_sum.merge_cells(start_row=cell_b.row, start_column=2,
                                       end_row=cell_b.row, end_column=3)
                    cell_b.alignment = align_center
                    ws_sum.row_dimensions[cell_b.row].height = 20
                    cell_b.fill, cell_b.font = fill_green, font_white
                else:
                    cell_b.fill = fill_grey
                    row[2].fill = fill_grey
                    cell_b.font = font_bold_black
                    row[2].font = font_bold_black
                    row[2].alignment = align_center
            else:
                cell_b.fill, cell_b.font = fill_green, font_white
                row[2].alignment = align_center
                for c in cc:
                    c.border = border_thin

        wb.save(main_output_path)
        logger.info(f"[{job_id[:8]}] openpyxl styling + save: {time.perf_counter()-_t_style:.2f}s")
        del wb

        # ── Phase 3: Append full-data sheet (openpyxl append mode) ───────
        # Written separately so the Phase 2 BytesIO stays small (avoids MemoryError).
        # Header styling is applied inside the same context to avoid a second load.
        if include_src_tgt and full_data_for_excel is not None and not full_data_for_excel.empty:
            _job_update(job_id, progress=94, stage="Appending source/target data sheet")
            _t_fd = time.perf_counter()
            try:
                with pd.ExcelWriter(
                    main_output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
                ) as app_writer:
                    _fd_names = write_df_excel_paginated(app_writer, full_data_for_excel, sheet_full_data)
                    # Style full-data header while the workbook is still open (no second load)
                    _wb_app = app_writer.book
                    for _sn in (_fd_names or []):
                        if _sn in _wb_app.sheetnames:
                            _style_full_data_header(_wb_app[_sn], num_comparison_cols)
                logger.info(f"[{job_id[:8]}] Full-data sheet appended in {time.perf_counter()-_t_fd:.2f}s")
            except Exception as _fd_err:
                logger.warning(f"[{job_id[:8]}] Could not append full-data sheet: {_fd_err}")

        # ── Stage 14b: Cleanup auxiliary data ──────────────────────────
        if full_data_for_excel is not None:
            del full_data_for_excel
            gc.collect()

        # ── Stage 15: Finalize output (96%) ────────────────────────────
        _job_update(job_id, progress=96, stage="Finalizing output")

        report_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"MythicsValidationResults_{report_ts}.xlsx"

        elapsed = time.time() - start_time
        logger.info(f"[{job_id[:8]}] === COMPLETE in {elapsed:.2f}s | "
                     f"{legacy_count + oracle_count:,} total rows ===")

        # ── Mark Complete ──────────────────────────────────────────────
        _job_update(job_id, progress=100, stage="Validation complete",
                    status="complete", zip_path=main_output_path,
                    zip_filename=output_filename)

    except Exception as e:
        logger.error(f"[{job_id[:8]}] Validation failed: {str(e)}", exc_info=True)
        _job_update(job_id, status="failed", error=str(e),
                    stage=f"Failed: {str(e)[:100]}")
    except BaseException as e:
        # Catch SystemExit, KeyboardInterrupt, MemoryError, etc.
        logger.critical(f"[{job_id[:8]}] CRITICAL FAILURE (BaseException): {type(e).__name__}: {str(e)}", exc_info=True)
        _job_update(job_id, status="failed", error=f"{type(e).__name__}: {str(e)}",
                    stage=f"Critical failure: {type(e).__name__}")
    finally:
        import gc
        gc.collect()
        # Flush all log handlers
        for handler in logging.root.handlers:
            handler.flush()
        # Clean up temp files if job failed (successful jobs are cleaned after download)
        try:
            job = _job_get(job_id)
            if job and job.get("status") == "failed":
                shutil.rmtree(temp_dir, ignore_errors=True)
                logger.info(f"[{job_id[:8]}] Cleaned up temp dir after failure")
        except Exception:
            pass

@app.post("/get-sheets")
async def get_sheets(file: UploadFile = File(...)):
    """
    Returns a list of all worksheet names in the Excel file.
    """
    try:
        contents = await file.read()
        # Use ExcelFile to read metadata without loading the entire dataframe
        with io.BytesIO(contents) as bio:
            xls = pd.ExcelFile(bio, engine='openpyxl')
            return {"sheets": xls.sheet_names}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading sheets: {str(e)}")

@app.post("/extract-headers")
async def extract_headers(
    file: UploadFile = File(...), 
    sheet_name: str = Form(None) # Updated to accept sheet_name
):
    """
    Reads headers from a specific sheet. 
    If sheet_name is not provided, defaults to the first sheet.
    """
    try:
        contents = await file.read()
        with io.BytesIO(contents) as bio:
            # Read only the first few rows for performance
            if sheet_name:
                df = pd.read_excel(bio, sheet_name=sheet_name, nrows=5, engine='openpyxl')
            else:
                df = pd.read_excel(bio, nrows=5, engine='openpyxl')
                
        headers = df.columns.tolist()
        return {"headers": headers}
    except ValueError as ve:
         raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading headers: {str(e)}")

@app.post("/convert-excel")
async def convert_excel(
    structure_file: UploadFile = File(...),
    target_file: UploadFile = File(...),
    # New Sheet Parameters
    structure_sheet: str = Form(None),
    target_sheet: str = Form(None),
    # Existing Mapping Parameters
    legacy_col: str = Form(...),
    oracle_col: str = Form(...),
    attribute_col: str = Form(None),
    target_col: str = Form(None),
):
    """
    Transforms data using specific sheets for Structure and Target files.
    """
    try:
        # 1. Read Structure File (Mapping Logic)
        structure_content = await structure_file.read()
        try:
            # Load specific sheet if provided, else default (0)
            sheet_arg = structure_sheet if structure_sheet else 0
            df_structure = pd.read_excel(io.BytesIO(structure_content), sheet_name=sheet_arg, engine='openpyxl')
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Structure file: {str(e)}")

        # Validate structure columns
        if legacy_col not in df_structure.columns or oracle_col not in df_structure.columns:
            raise HTTPException(status_code=400, detail="Selected mapping columns not found in Structure sheet.")

        # 2. Read Target File (Data to Transform)
        target_content = await target_file.read()
        try:
            sheet_arg = target_sheet if target_sheet else 0
            df_target = pd.read_excel(io.BytesIO(target_content), sheet_name=sheet_arg, engine='openpyxl')
        except Exception as e:
             raise HTTPException(status_code=400, detail=f"Error reading Target file: {str(e)}")

        transform_log = []

        # --- LOGIC BRANCHING ---
        if attribute_col and attribute_col in df_structure.columns:
            # === DYNAMIC MODE ===
            # Normalize attribute column for matching
            df_structure[attribute_col] = df_structure[attribute_col].astype(str).str.strip()
            
            for col in df_target.columns:
                col_name = str(col).strip()
                
                # Check if this column name is defined in the structure file
                if col_name in df_structure[attribute_col].values:
                    # Filter structure rules for this specific column
                    subset = df_structure[df_structure[attribute_col] == col_name]
                    
                    # Create Map: Legacy -> Oracle
                    mapping = dict(zip(subset[legacy_col], subset[oracle_col]))
                    
                    # Apply Map
                    df_target[col] = df_target[col].map(mapping).fillna(df_target[col])
                    transform_log.append(col_name)

        else:
            # === SINGLE COLUMN MODE ===
            if not target_col:
                raise HTTPException(status_code=400, detail="Target column must be specified if Attribute column is not used.")
            
            if target_col in df_target.columns:
                mapping_df = df_structure[[legacy_col, oracle_col]].drop_duplicates(subset=[legacy_col])
                mapping_dict = dict(zip(mapping_df[legacy_col], mapping_df[oracle_col]))
                df_target[target_col] = df_target[target_col].map(mapping_dict).fillna(df_target[target_col])
            else:
                 raise HTTPException(status_code=400, detail=f"Column '{target_col}' not found in Target sheet.")

        # 3. Save to Buffer
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_target.to_excel(writer, index=False)
        
        output.seek(0)

        filename_prefix = "Dynamic_Transformed" if attribute_col else f"Transformed_{target_col}"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename_prefix}.xlsx"'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Transformation failed: {str(e)}")
    


# --- PeopleSoft Finance
@app.get("/api/utils/finance/menu-items")
async def get_finance_menu_items():
    """
    Returns a list of predefined menu items for PeopleSoft Finance.
    returns the json file content as is.
    """
    try:
        menu_file_path = Path("Required_files/finance_menu_items.json")
        with open(menu_file_path, "r") as f:
            menu_items = json.load(f)
            final = {
                "status": 200,
                "menu_items": menu_items
            }
        return final
    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to load menu items: {str(e)}"},
            status_code=500,
        )
        
